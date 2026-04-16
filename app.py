import streamlit as st
import pandas as pd
import numpy as np
import os
import pickle
import glob
import json
from io import BytesIO
from datetime import datetime
from dateutil.relativedelta import relativedelta
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import OneHotEncoder

# --- SCENARIOS DIRECTORY ---
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
SCENARIOS_DIR = os.path.join(_APP_DIR, "scenarios")
os.makedirs(SCENARIOS_DIR, exist_ok=True)

# --- PARAMS CONFIG DIRECTORY ---
PARAMS_CONFIG_DIR = os.path.join(_APP_DIR, "params_configs")
os.makedirs(PARAMS_CONFIG_DIR, exist_ok=True)

# --- AI PREDICTION CONSTANTS ---
ROLE_HIERARCHY = {
    'Accountant I':       'Accountant II',
    'Accountant II':      'General Accountant',
    'General Accountant': 'Sr. Accountant',
    'Sr. Accountant':     'Sr. Accountant',
}
BASELINE_NETWORK_DAYS = 21.0

COLS_T0_NUM = ['mrr']
COLS_T1_NUM = ['mrr']
COLS_T1_CAT = ['pms']
COLS_T2_NUM = ['mrr', 'res_doors', 'res_prop']
COLS_T2_CAT = ['pms']
COLS_T3_NUM = ['mrr', 'res_doors', 'res_prop', 'commercial_properties', 'commercial_doors']
COLS_T3_CAT = ['pms']
COLS_T4_NUM = ['mrr', 'res_doors', 'res_prop', 'commercial_properties', 'commercial_doors', 'sqft_commercial']
COLS_T4_CAT = ['pms', 'corp_books']
ALL_NUM_COLS = COLS_T4_NUM
ALL_CAT_COLS = COLS_T4_CAT

def _clean_cols(df):
    """Lowercase, strip, spaces→underscore, remove special chars."""
    df = df.copy()
    df.columns = (
        df.columns.str.lower().str.strip()
        .str.replace(r'[\s\-]+', '_', regex=True)
        .str.replace(r'[^a-z0-9_]', '', regex=True)
    )
    return df

def _get_pms_opts(include_all=False, include_unknown=True):
    """Return PMS options from session state (populated from volume file). Fragment-safe."""
    import streamlit as _st
    opts = list(_st.session_state.get('_lista_pms', []))
    if not opts:
        # fallback if file not yet loaded
        opts = ['AppFolio', 'Buildium', 'Propertyware', 'Yardi', 'RealPage', 'ResMan', 'MRI', 'Other']
    if include_unknown and 'Unknown' not in opts:
        opts = opts + ['Unknown']
    if include_all:
        return ['', 'All'] + opts
    return opts

def _build_vol_aht_task_df(pms_filter=None):
    """
    Build a task editor DataFrame from df_clean.
    Columns: Type, Subtype, Proc Role, Rev Role, Proc AHT (min), Rev AHT (min), Volume, QC %
    AHTs: median for the selected PMS; fallback to overall median if no data for that PMS.
    """
    import streamlit as _st
    if 'df_clean' not in _st.session_state:
        return pd.DataFrame()

    df = _st.session_state.df_clean.copy()
    # Find AHT columns
    _p_aht_col = next((c for c in df.columns if '>>> FINAL' in c and 'proc' in c.lower() and 'aht' in c.lower()), None)
    _r_aht_col = next((c for c in df.columns if '>>> FINAL' in c and 'rev' in c.lower()  and 'aht' in c.lower()), None)
    if _p_aht_col is None or _r_aht_col is None:
        return pd.DataFrame()

    # Ensure numeric AHTs
    df[_p_aht_col] = pd.to_numeric(df[_p_aht_col], errors='coerce')
    df[_r_aht_col] = pd.to_numeric(df[_r_aht_col], errors='coerce')

    # Keep only valid task rows
    df = df.dropna(subset=['type', 'subtype', _p_aht_col])
    df['_type_s']    = df['type'].astype(str).str.strip()
    df['_subtype_s'] = df['subtype'].astype(str).str.strip()
    df['_pms_s']     = df['PMS'].astype(str).str.strip().str.lower() if 'PMS' in df.columns else ''

    # Role per task (mode of Ideal Proc / Ideal Rev)
    _role_gb = df.groupby(['_type_s', '_subtype_s'])
    def _mode(s):
        v = s.dropna()
        return v.mode().iloc[0] if len(v) > 0 else ''

    _proc_role_map = _role_gb['Ideal Proc'].agg(_mode).to_dict() if 'Ideal Proc' in df.columns else {}
    _rev_role_map  = _role_gb['Ideal Rev' ].agg(_mode).to_dict() if 'Ideal Rev'  in df.columns else {}

    # Overall median AHT per task
    _overall_p = _role_gb[_p_aht_col].median().to_dict()
    _overall_r = _role_gb[_r_aht_col].median().to_dict()

    # PMS-specific median AHT per task (if filter given)
    _pms_p = {}
    _pms_r = {}
    if pms_filter:
        _pms_low = str(pms_filter).strip().lower()
        _df_pms  = df[df['_pms_s'] == _pms_low]
        if not _df_pms.empty:
            _pms_gb  = _df_pms.groupby(['_type_s', '_subtype_s'])
            _pms_p   = _pms_gb[_p_aht_col].median().to_dict()
            _pms_r   = _pms_gb[_r_aht_col].median().to_dict()

    # Build one row per unique (type, subtype)
    tasks = (
        df[['_type_s', '_subtype_s']].drop_duplicates()
        .sort_values(['_type_s', '_subtype_s'])
        .reset_index(drop=True)
    )
    rows = []
    for _, r in tasks.iterrows():
        key = (r['_type_s'], r['_subtype_s'])
        p_aht = round(float(_pms_p.get(key, _overall_p.get(key, 15.0)) or 15.0), 2)
        r_aht = round(float(_pms_r.get(key, _overall_r.get(key,  5.0)) or  5.0), 2)
        rows.append({
            'Type':         r['_type_s'],
            'Subtype':      r['_subtype_s'],
            'Proc Role':    _proc_role_map.get(key, 'Accountant I'),
            'Rev Role':     _rev_role_map.get(key,  'Sr. Accountant'),
            'Proc AHT (min)': p_aht,
            'Rev AHT (min)':  r_aht,
            'Volume':       0,
            'QC %':         100.0,
        })
    return pd.DataFrame(rows)


def _make_ai_prediction_fragment(pfx, add_to_scenario, add_to_baseline=False):
    @st.fragment
    def _ai_frag():
        _n_clients = st.session_state.df_clean['client_name'].nunique() if 'client_name' in st.session_state.df_clean.columns else '?'
        st.info(f"📊 Training on **{_n_clients}** existing clients from Step 1.")

        if add_to_scenario:
            # Mode selector only shown in Step 4
            _ai_mode = st.radio(
                "Purpose",
                ["➕ Add to Step 4 Scenario", "🔍 Just calculate prediction"],
                horizontal=True,
                key=f"{pfx}_ai_mode"
            )
            if _ai_mode == "➕ Add to Step 4 Scenario":
                st.caption(
                    "Predicted hours will be added to the scenario's Required Hours per role "
                    "when you click **Apply to Scenario** after running the prediction."
                )
            else:
                st.caption("Prediction only — results will not affect the scenario.")
        else:
            _ai_mode = "🔍 Just calculate prediction"
            st.caption("Prediction only — results will not affect the scenario.")

        # Input method
        _queued_nc = st.session_state.get('pipeline_new_clients', pd.DataFrame())
        _has_queued = len(_queued_nc) > 0
        if _has_queued:
            st.success(f"🤖 **{len(_queued_nc)} new clients queued** from HubSpot update.")

        _ai_input = st.radio(
            "Input method",
            ["📂 File Upload", "✏️ Manual Entry"],
            horizontal=True,
            key=f"{pfx}_ai_input_mode"
        )

        # Example file download
        _ex_buf = BytesIO()
        _ex_df = pd.DataFrame([{
            "Company Name": "Acme Properties", "POD": "POD A", "Go Live Date": "2026-04-01",
            "MRR ($)": 3500, "PMS": "AppFolio", "Res Doors": 250, "Res Properties": 80,
            "Comm Doors": 0, "Comm Properties": 0, "SQFT Commercial": 0, "Corp Books": "No"
        }, {
            "Company Name": "Blue Sky Mgmt", "POD": "POD B", "Go Live Date": "2026-05-01",
            "MRR ($)": 5200, "PMS": "Buildium", "Res Doors": 420, "Res Properties": 140,
            "Comm Doors": 20, "Comm Properties": 5, "SQFT Commercial": 0, "Corp Books": "No"
        }])
        with pd.ExcelWriter(_ex_buf, engine='openpyxl') as _ex_xw:
            _ex_df.to_excel(_ex_xw, sheet_name='New Clients', index=False)
        st.download_button(
            "📄 Download Example File",
            _ex_buf.getvalue(),
            file_name="NewClients_Example.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{pfx}_ai_ex_dl"
        )

        _ai_col1, _ai_col2 = st.columns(2)
        with _ai_col1:
            _ai_month = st.selectbox(
                "Projection month:",
                options=list(range(6)),
                format_func=lambda i: meses_proyeccion[i],
                key=f"{pfx}_ai_month_idx"
            )
        with _ai_col2:
            if _ai_input == "📂 File Upload":
                _ai_nc_lbl = "Override queued clients" if _has_queued else "Upload New Clients file (needs 'New Clients' sheet)"
                _ai_nc_file = st.file_uploader(_ai_nc_lbl, type=["xlsx","xls"], key=f"{pfx}_ai_nc_upload")
            else:
                _ai_nc_file = None

        if _ai_input == "✏️ Manual Entry":
            _pms_opts_ai = _get_pms_opts(include_unknown=True)
            _cb_opts_ai  = ['Yes','No','Unknown']
            _hs_snap_ai  = st.session_state.get('hs_parsed')
            # Merge baseline PODs with any POD values that appear in the HubSpot file
            _hs_pods_extra = []
            if _hs_snap_ai is not None and not _hs_snap_ai.empty and '_pod' in _hs_snap_ai.columns:
                _hs_pods_extra = [
                    p for p in _hs_snap_ai['_pod'].dropna().astype(str).str.strip().unique()
                    if p and p.lower() not in ('nan', 'none', '')
                ]
            _pod_opts_ai = [""] + sorted(set(lista_pods) | set(_hs_pods_extra))

            _hdr_col, _sync_col = st.columns([3, 1])
            with _hdr_col:
                st.write("**New Clients — Manual Entry**")
            with _sync_col:
                _hs_no_file = _hs_snap_ai is None or (_hs_snap_ai is not None and _hs_snap_ai.empty)
                _hs_btn_lbl = ("🔄 Sync from HubSpot" if not _hs_no_file
                               else "📂 Upload HubSpot file")
                _hs_btn_help = (
                    f"HubSpot file loaded ({len(_hs_snap_ai)} rows). "
                    "Click to fill empty POD / Go Live / MRR / PMS / Res Doors / SQFT from HubSpot."
                    if not _hs_no_file
                    else "No HubSpot file loaded yet. Click to upload one and auto-fill fields."
                )
                if st.button(
                    _hs_btn_lbl,
                    key=f"{pfx}_ai_hs_sync_btn",
                    use_container_width=True,
                    help=_hs_btn_help,
                    type="secondary" if _hs_no_file else "primary",
                ):
                    if _hs_no_file:
                        st.session_state[f"{pfx}_ai_hs_show_upload"] = True
                        st.rerun(scope="fragment")
                    else:
                        _tbl_before = st.session_state.get(f"{pfx}_ai_manual_clients", pd.DataFrame())
                        if not _tbl_before.empty:
                            _tbl_after = _enrich_ai_from_hs(_tbl_before, _hs_snap_ai)
                            st.session_state[f"{pfx}_ai_manual_clients"] = _tbl_after
                            # Clear the data_editor's own key so it re-initialises from new data
                            st.session_state.pop(f"{pfx}_ai_manual_editor", None)
                            _fields = ['POD', 'Go Live Date', 'MRR ($)', 'PMS', 'Res Doors', 'SQFT Commercial']
                            _filled = sum(
                                1 for f in _fields
                                if f in _tbl_before.columns and f in _tbl_after.columns
                                and (_tbl_after[f] != _tbl_before[f]).any()
                            )
                            if _filled:
                                st.success(f"✅ Synced — {_filled} field(s) updated.")
                            else:
                                st.info("ℹ️ No new data — all fields already filled or no name matches.")
                            st.rerun(scope="fragment")
                        else:
                            st.warning("Table is empty — add clients first.")

            # If no HubSpot file loaded, show an inline uploader
            if st.session_state.get(f"{pfx}_ai_hs_show_upload"):
                _hs_inline = st.file_uploader(
                    "📂 Upload HubSpot file to sync POD / Go Live / MRR / PMS:",
                    type=["xlsx", "xls"],
                    key=f"{pfx}_ai_hs_inline_upload",
                )
                if _hs_inline is not None:
                    try:
                        _parsed_inline = _parse_hubspot_file(_hs_inline)
                        st.session_state['hs_parsed'] = _parsed_inline
                        _hs_snap_ai = _parsed_inline
                        _tbl_before = st.session_state.get(f"{pfx}_ai_manual_clients", pd.DataFrame())
                        _tbl_after  = _enrich_ai_from_hs(_tbl_before, _hs_snap_ai)
                        st.session_state[f"{pfx}_ai_manual_clients"] = _tbl_after
                        # Clear the data_editor's own key so it re-initialises from new data
                        st.session_state.pop(f"{pfx}_ai_manual_editor", None)
                        _fields = ['POD', 'Go Live Date', 'MRR ($)', 'PMS', 'Res Doors', 'SQFT Commercial']
                        _filled = sum(
                            1 for f in _fields
                            if f in _tbl_before.columns and f in _tbl_after.columns
                            and (_tbl_after[f] != _tbl_before[f]).any()
                        )
                        st.session_state.pop(f"{pfx}_ai_hs_show_upload", None)
                        if _filled:
                            st.success(f"✅ Synced — {_filled} field(s) updated from HubSpot.")
                        else:
                            st.info("ℹ️ No new data — all fields already filled or no name matches.")
                        st.rerun(scope="fragment")
                    except Exception as _hs_err:
                        st.error(f"Could not parse HubSpot file: {_hs_err}")

            if f"{pfx}_ai_manual_clients" not in st.session_state:
                st.session_state[f"{pfx}_ai_manual_clients"] = pd.DataFrame(columns=[
                    "Company Name", "POD", "Go Live Date", "MRR ($)", "PMS",
                    "Res Doors", "Res Properties", "Comm Doors", "Comm Properties",
                    "SQFT Commercial", "Corp Books"
                ])
            st.session_state[f"{pfx}_ai_manual_clients"] = st.data_editor(
                st.session_state[f"{pfx}_ai_manual_clients"],
                num_rows="dynamic",
                use_container_width=True,
                column_config={
                    "Company Name":    st.column_config.TextColumn("Company Name", required=True),
                    "POD":             st.column_config.TextColumn("POD", help="POD / team assignment (auto-filled from HubSpot)"),
                    "Go Live Date":    st.column_config.TextColumn("Go Live Date (YYYY-MM-DD)"),
                    "MRR ($)":         st.column_config.NumberColumn("MRR ($)", min_value=0, format="$%d"),
                    "PMS":             st.column_config.SelectboxColumn("PMS", options=_pms_opts_ai),
                    "Res Doors":       st.column_config.NumberColumn("Res Doors", min_value=0),
                    "Res Properties":  st.column_config.NumberColumn("Res Properties", min_value=0),
                    "Comm Doors":      st.column_config.NumberColumn("Comm Doors", min_value=0),
                    "Comm Properties": st.column_config.NumberColumn("Comm Properties", min_value=0),
                    "SQFT Commercial": st.column_config.NumberColumn("SQFT Commercial", min_value=0),
                    "Corp Books":      st.column_config.SelectboxColumn("Corp Books", options=_cb_opts_ai),
                },
                key=f"{pfx}_ai_manual_editor"
            )

        if st.button("🤖 Run AI Prediction", type="primary", use_container_width=True, key=f"{pfx}_ai_run"):
            _use_manual = (_ai_input == "✏️ Manual Entry")
            _use_queued = not _use_manual and _has_queued and _ai_nc_file is None
            if _use_manual and len(st.session_state.get(f"{pfx}_ai_manual_clients", pd.DataFrame())) == 0:
                st.error("⚠️ Add at least one client to the manual entry table.")
            elif not _use_manual and not _use_queued and _ai_nc_file is None:
                st.error("⚠️ Please upload a New Clients file.")
            else:
                with st.spinner("Training models and generating predictions..."):
                    try:
                        # ── 1. PREPARE TRAINING DATA ──────────────────────────────
                        df_vol = _clean_cols(st.session_state.df_clean.copy())
                        if 'status' in df_vol.columns:
                            df_vol = df_vol[df_vol['status'].astype(str).str.strip().str.lower() == 'client']
                        if 'go_live' in df_vol.columns:
                            df_vol['go_live'] = pd.to_datetime(df_vol['go_live'], errors='coerce')
                            df_vol = df_vol[df_vol['go_live'] < (pd.Timestamp.today() - pd.DateOffset(months=3))]
                        if df_vol.empty:
                            st.error("Training data is empty after filters.")
                            st.stop()

                        for col in ALL_NUM_COLS:
                            if col in df_vol.columns:
                                df_vol[col] = pd.to_numeric(df_vol[col], errors='coerce').fillna(0)
                            else:
                                df_vol[col] = 0.0

                        if 'type' in df_vol.columns and 'subtype' in df_vol.columns:
                            df_vol['task_name'] = df_vol['type'].astype(str) + " - " + df_vol['subtype'].astype(str)
                        else:
                            df_vol['task_name'] = 'General Task'

                        vol_target_candidates = [c for c in df_vol.columns if 'closed_tickets' in c and 'proc' in c]
                        vol_target = vol_target_candidates[0] if vol_target_candidates else None
                        if vol_target is None:
                            st.error("Could not find 'Closed tickets with Proc time' column.")
                            st.stop()
                        df_vol[vol_target] = pd.to_numeric(df_vol[vol_target], errors='coerce').fillna(0)

                        for col in ALL_CAT_COLS:
                            if col not in df_vol.columns:
                                df_vol[col] = 'Unknown'
                            df_vol[col] = df_vol[col].astype(str).replace(['nan','None',''], 'Unknown').fillna('Unknown')

                        client_features = (
                            df_vol[['client_name'] + ALL_NUM_COLS + ALL_CAT_COLS]
                            .drop_duplicates(subset=['client_name'])
                            .set_index('client_name')
                        )

                        pms_enc = OneHotEncoder(handle_unknown='ignore', sparse_output=False, dtype=int)
                        pms_encoded = pms_enc.fit_transform(client_features[['pms']])
                        pms_cols = pms_enc.get_feature_names_out(['pms'])

                        cb_enc = OneHotEncoder(handle_unknown='ignore', sparse_output=False, dtype=int)
                        cb_encoded = cb_enc.fit_transform(client_features[['corp_books']])
                        cb_cols = cb_enc.get_feature_names_out(['corp_books'])

                        df_pms = pd.DataFrame(pms_encoded, columns=pms_cols, index=client_features.index)
                        df_cb  = pd.DataFrame(cb_encoded,  columns=cb_cols,  index=client_features.index)

                        feat_t0 = COLS_T0_NUM                                          # MRR only
                        feat_t1 = COLS_T1_NUM + list(pms_cols)                         # MRR + PMS
                        feat_t2 = COLS_T2_NUM + list(pms_cols)                         # MRR + PMS + res
                        feat_t3 = COLS_T3_NUM + list(pms_cols)                         # MRR + PMS + commercial
                        feat_t4 = COLS_T4_NUM + list(pms_cols) + list(cb_cols)         # Full

                        # ── 2. TRAIN MODELS PER TASK ──────────────────────────────
                        ideal_proc_col = [c for c in df_vol.columns if 'ideal_proc' in c]
                        models = {}
                        unique_tasks = df_vol['task_name'].unique()

                        for task in unique_tasks:
                            df_task = df_vol[df_vol['task_name'] == task]
                            train_data = (
                                client_features.join(df_pms).join(df_cb)
                                .join(df_task.groupby('client_name')[vol_target].sum())
                                .fillna(0)
                            )
                            if len(train_data) <= 3:
                                continue

                            p_aht_col = [c for c in df_task.columns if 'proc_aht' in c or ('final' in c and 'proc' in c)]
                            r_aht_col = [c for c in df_task.columns if 'rev_aht'  in c or ('final' in c and 'rev'  in c)]

                            if ideal_proc_col and not df_task[ideal_proc_col[0]].dropna().empty:
                                proc_role = df_task[ideal_proc_col[0]].mode().iloc[0]
                            else:
                                proc_role = 'Accountant I'

                            models[task] = {
                                'model_t0': RandomForestRegressor(100, random_state=42).fit(train_data[feat_t0], train_data[vol_target]),
                                'model_t1': RandomForestRegressor(100, random_state=42).fit(train_data[feat_t1], train_data[vol_target]),
                                'model_t2': RandomForestRegressor(100, random_state=42).fit(train_data[feat_t2], train_data[vol_target]),
                                'model_t3': RandomForestRegressor(100, random_state=42).fit(train_data[feat_t3], train_data[vol_target]),
                                'model_t4': RandomForestRegressor(100, random_state=42).fit(train_data[feat_t4], train_data[vol_target]),
                                'proc_aht':  df_task[p_aht_col[0]].mean() if p_aht_col else 15.0,
                                'rev_aht':   df_task[r_aht_col[0]].mean() if r_aht_col else 5.0,
                                'proc_role': str(proc_role).strip(),
                                'rev_role':  _reviewer_role(proc_role),
                            }

                        if not models:
                            st.error("No tasks had enough training data (need > 3 clients per task).")
                            st.stop()

                        st.info(f"✅ Trained {len(models)} task models from {len(client_features)} existing clients.")

                        # ── 3. LOAD & PREPARE NEW CLIENTS ─────────────────────────
                        if _use_manual:
                            _m = st.session_state[f"{pfx}_ai_manual_clients"].copy()
                            df_new = pd.DataFrame({
                                'company_name':          _m['Company Name'].fillna('Client'),
                                'pod':                   _m['POD'].fillna('') if 'POD' in _m.columns else '',
                                'go_live_date':          _m['Go Live Date'].fillna('') if 'Go Live Date' in _m.columns else '',
                                'mrr':                   pd.to_numeric(_m['MRR ($)'], errors='coerce').fillna(0),
                                'pms':                   _m['PMS'].fillna('Unknown') if 'PMS' in _m.columns else 'Unknown',
                                'res_doors':             pd.to_numeric(_m.get('Res Doors', 0), errors='coerce').fillna(0),
                                'res_prop':              pd.to_numeric(_m.get('Res Properties', 0), errors='coerce').fillna(0),
                                'commercial_doors':      pd.to_numeric(_m.get('Comm Doors', 0), errors='coerce').fillna(0),
                                'commercial_properties': pd.to_numeric(_m.get('Comm Properties', 0), errors='coerce').fillna(0),
                                'sqft_commercial':       pd.to_numeric(_m.get('SQFT Commercial', 0), errors='coerce').fillna(0),
                                'corp_books':            _m.get('Corp Books', pd.Series(['Unknown']*len(_m))).fillna('Unknown'),
                            })
                            st.info(f"Using {len(df_new)} manually entered clients.")
                        elif _use_queued:
                            df_new = _queued_nc.copy()
                            if 'pod' not in df_new.columns:
                                df_new['pod'] = ''
                            st.info(f"Using {len(df_new)} queued clients.")
                        else:
                            df_new = _clean_cols(pd.read_excel(_ai_nc_file, sheet_name='New Clients'))

                        df_new.rename(columns={
                            'original_cmrr':            'mrr',   # HubSpot export
                            'mrr_':                     'mrr',   # _clean_cols strips MRR ($) → mrr_
                            'mrr__':                    'mrr',   # edge case double-underscore
                            'recent_residential_doors': 'res_doors',
                            'recent_commercial_sqft':   'sqft_commercial',
                        }, inplace=True)

                        default_gl = (pd.Timestamp.today() + pd.DateOffset(days=30)).strftime('%Y-%m-%d')
                        if 'go_live_date' not in df_new.columns:
                            df_new['go_live_date'] = default_gl
                        else:
                            df_new['go_live_date'] = (
                                pd.to_datetime(df_new['go_live_date'], errors='coerce')
                                .dt.strftime('%Y-%m-%d').fillna(default_gl)
                            )
                        if 'pod' not in df_new.columns:
                            df_new['pod'] = ''

                        for col in ALL_NUM_COLS:
                            if col not in df_new.columns: df_new[col] = 0.0
                            df_new[col] = pd.to_numeric(df_new[col], errors='coerce').fillna(0)
                        for col in ALL_CAT_COLS:
                            if col not in df_new.columns: df_new[col] = 'Unknown'
                            df_new[col] = df_new[col].astype(str).replace(['nan','None',''], 'Unknown').fillna('Unknown')

                        new_pms    = pd.DataFrame(pms_enc.transform(df_new[['pms']]),       columns=pms_cols, index=df_new.index)
                        new_cb     = pd.DataFrame(cb_enc.transform(df_new[['corp_books']]),  columns=cb_cols,  index=df_new.index)
                        df_new_enc = df_new.join(new_pms).join(new_cb)

                        # ── 4. PREDICT ─────────────────────────────────────────────
                        hrs_fte_month  = st.session_state.calc_data['dict_hrs_per_fte'][_ai_month]
                        network_days_m = st.session_state.calc_data['dict_workable_days'][_ai_month]
                        vol_scale      = network_days_m / BASELINE_NETWORK_DAYS

                        rows = []
                        for idx, row in df_new.iterrows():
                            client_name = row.get('company_name', f'Client_{idx}')
                            go_live     = row['go_live_date']
                            mrr_val     = row.get('mrr', 0)
                            pod_val     = row.get('pod', '')

                            pms_val  = str(row.get('pms', 'Unknown')).strip()
                            has_pms  = pms_val.lower() not in ['unknown', '', 'nan', 'none']
                            corp_val = str(row.get('corp_books', '')).lower()
                            has_t4   = row.get('sqft_commercial', 0) > 0 or corp_val not in ['unknown', '0', '', 'nan', 'no']
                            has_t3   = row.get('commercial_properties', 0) > 0 or row.get('commercial_doors', 0) > 0
                            has_t2   = row.get('res_doors', 0) > 0 or row.get('res_prop', 0) > 0

                            if   has_t4 and has_pms:   tier, mkey, fcols = "Tier 4 (Full)",         'model_t4', feat_t4
                            elif has_t3 and has_pms:   tier, mkey, fcols = "Tier 3 (+Commercial)",  'model_t3', feat_t3
                            elif has_t2 and has_pms:   tier, mkey, fcols = "Tier 2 (+Residential)", 'model_t2', feat_t2
                            elif has_pms:               tier, mkey, fcols = "Tier 1 (MRR + PMS)",   'model_t1', feat_t1
                            else:                       tier, mkey, fcols = "Tier 0 (MRR only)",    'model_t0', feat_t0

                            feats = df_new_enc.loc[idx, fcols].values.reshape(1, -1)

                            _gl_parsed = pd.to_datetime(go_live, errors='coerce')
                            _mes_start = (today + relativedelta(months=_ai_month)).replace(day=1)
                            if pd.notna(_gl_parsed):
                                _m_diff = (_mes_start.year - _gl_parsed.year) * 12 + (_mes_start.month - _gl_parsed.month)
                                if   _m_diff == 0: _lc = 1.17
                                elif _m_diff == 1: _lc = 0.86
                                elif _m_diff == 2: _lc = 0.99
                                else:              _lc = 1.0
                            else:
                                _lc = 1.17

                            for task, md in models.items():
                                pred_vol = md[mkey].predict(feats)[0] * vol_scale
                                if pred_vol <= 0.1:
                                    continue

                                p_hrs = (pred_vol * md['proc_aht'] * _lc) / 60
                                r_hrs = (pred_vol * md['rev_aht']  * _lc) / 60

                                t_parts  = str(task).split(" - ", 1) if " - " in str(task) else [str(task), ""]
                                v_type, v_subtype = t_parts[0].strip(), t_parts[1].strip()

                                util_p = utilization_map.get(md['proc_role'], util_acc1)
                                util_r = utilization_map.get(md['rev_role'],  util_sr)

                                full_p_hrs = p_hrs * (1 + (1 - util_p) + absenteeism + attrition)
                                full_r_hrs = r_hrs * (1 + (1 - util_r) + absenteeism + attrition)

                                hourly_p = cost_map.get(md['proc_role'], cost_acc1) / hrs_fte_month if hrs_fte_month > 0 else 0
                                hourly_r = cost_map.get(md['rev_role'],  cost_sr)   / hrs_fte_month if hrs_fte_month > 0 else 0

                                rows.append({
                                    '_client':     client_name,
                                    '_pod':        pod_val,
                                    '_go_live':    go_live,
                                    '_mrr':        mrr_val,
                                    '_tier':       tier,
                                    '_task':       task,
                                    '_type':       v_type,
                                    '_subtype':    v_subtype,
                                    '_proc_role':  md['proc_role'],
                                    '_rev_role':   md['rev_role'],
                                    '_pred_vol':   pred_vol,
                                    '_proc_aht':   md['proc_aht'],
                                    '_rev_aht':    md['rev_aht'],
                                    '_prod_p_hrs': p_hrs,
                                    '_prod_r_hrs': r_hrs,
                                    '_full_p_hrs': full_p_hrs,
                                    '_full_r_hrs': full_r_hrs,
                                    '_full_cost_p': full_p_hrs * hourly_p,
                                    '_full_cost_r': full_r_hrs * hourly_r,
                                })

                        if not rows:
                            st.error("All predicted volumes were 0.")
                            st.stop()

                        df_pred = pd.DataFrame(rows)

                        # ── 5. APPLY 35–50% MARGIN RULE ───────────────────────────
                        client_costs = df_pred.groupby('_client').agg(
                            total_full_cost=('_full_cost_p', 'sum'),
                            total_full_cost_r=('_full_cost_r', 'sum'),
                            mrr=('_mrr', 'first')
                        )
                        client_costs['total_cost'] = client_costs['total_full_cost'] + client_costs['total_full_cost_r']

                        for c_name, cdata in client_costs.iterrows():
                            c_mrr, m_cost = cdata['mrr'], cdata['total_cost']
                            if c_mrr <= 0 or m_cost <= 0:
                                continue
                            max_t, min_t = c_mrr * 0.50, c_mrr * 0.35   # cost capped at 35–50% of MRR
                            adj = 1.0
                            if m_cost > max_t: adj = max_t / m_cost
                            elif m_cost < min_t: adj = min_t / m_cost
                            if adj != 1.0:
                                mask = df_pred['_client'] == c_name
                                for col in ['_proc_aht', '_rev_aht']:
                                    df_pred.loc[mask, col] *= adj
                                df_pred.loc[mask, '_prod_p_hrs'] = (df_pred.loc[mask, '_pred_vol'] * df_pred.loc[mask, '_proc_aht']) / 60
                                df_pred.loc[mask, '_prod_r_hrs'] = (df_pred.loc[mask, '_pred_vol'] * df_pred.loc[mask, '_rev_aht'])  / 60
                                util_p_vals = df_pred.loc[mask, '_proc_role'].map(lambda r: utilization_map.get(r, util_acc1))
                                util_r_vals = df_pred.loc[mask, '_rev_role'].map(lambda r: utilization_map.get(r, util_sr))
                                shrink_p = util_p_vals.map(lambda u: 1 + (1 - u) + absenteeism + attrition)
                                shrink_r = util_r_vals.map(lambda u: 1 + (1 - u) + absenteeism + attrition)
                                df_pred.loc[mask, '_full_p_hrs'] = df_pred.loc[mask, '_prod_p_hrs'] * shrink_p.values
                                df_pred.loc[mask, '_full_r_hrs'] = df_pred.loc[mask, '_prod_r_hrs'] * shrink_r.values
                                hp = df_pred.loc[mask, '_proc_role'].map(lambda r: cost_map.get(r, cost_acc1) / hrs_fte_month if hrs_fte_month > 0 else 0)
                                hr = df_pred.loc[mask, '_rev_role'].map(lambda r: cost_map.get(r, cost_sr) / hrs_fte_month if hrs_fte_month > 0 else 0)
                                df_pred.loc[mask, '_full_cost_p'] = df_pred.loc[mask, '_full_p_hrs'] * hp.values
                                df_pred.loc[mask, '_full_cost_r'] = df_pred.loc[mask, '_full_r_hrs'] * hr.values

                        # ── 6. BUILD RESULT TABLES ─────────────────────────────────
                        df_pred['_prod_hrs_total'] = df_pred['_prod_p_hrs'] + df_pred['_prod_r_hrs']
                        df_pred['_full_hrs_total'] = df_pred['_full_p_hrs'] + df_pred['_full_r_hrs']
                        df_pred['_full_cost']      = df_pred['_full_cost_p'] + df_pred['_full_cost_r']
                        df_pred['_proc_fte']       = df_pred['_full_p_hrs'] / hrs_fte_month if hrs_fte_month > 0 else 0
                        df_pred['_rev_fte']        = df_pred['_full_r_hrs'] / hrs_fte_month if hrs_fte_month > 0 else 0
                        df_pred['_total_fte']      = df_pred['_proc_fte'] + df_pred['_rev_fte']

                        summary = df_pred.groupby(['_client', '_pod', '_go_live', '_mrr', '_tier']).agg(
                            Productive_Hours       = ('_prod_hrs_total', 'sum'),
                            Full_Hours_w_Shrinkage = ('_full_hrs_total', 'sum'),
                            FTEs_Required          = ('_total_fte',      'sum'),
                            Full_Cost              = ('_full_cost',       'sum'),
                        ).reset_index()
                        summary['Cost_vs_MRR_%'] = np.where(
                            summary['_mrr'] > 0,
                            (summary['Full_Cost'] / summary['_mrr'] * 100).round(2), 0
                        )
                        summary = summary.rename(columns={
                            '_client': 'Client', '_pod': 'POD', '_go_live': 'Go Live',
                            '_mrr': 'MRR ($)', '_tier': 'Prediction Tier',
                        })
                        for c in ['Productive_Hours','Full_Hours_w_Shrinkage']:
                            summary[c] = summary[c].round(2)
                        summary['FTEs_Required'] = summary['FTEs_Required'].round(3)
                        summary['Full_Cost']     = summary['Full_Cost'].round(2)

                        # ── FTE per role per client ───────────────────────────────
                        _fte_long = pd.concat([
                            df_pred[['_client', '_proc_role', '_proc_fte']].rename(
                                columns={'_proc_role': '_role', '_proc_fte': '_fte'}),
                            df_pred[['_client', '_rev_role', '_rev_fte']].rename(
                                columns={'_rev_role': '_role', '_rev_fte': '_fte'}),
                        ], ignore_index=True)
                        _fte_pivot = (
                            _fte_long.groupby(['_client', '_role'])['_fte']
                            .sum().unstack(fill_value=0.0).reset_index()
                        )
                        _fte_pivot.columns.name = None
                        for _r in roles_permitidos:
                            if _r not in _fte_pivot.columns:
                                _fte_pivot[_r] = 0.0
                        _fte_pivot = _fte_pivot.rename(
                            columns={'_client': 'Client',
                                     **{_r: f'{_r} FTEs' for _r in roles_permitidos}}
                        )[['Client'] + [f'{_r} FTEs' for _r in roles_permitidos]]
                        for _c in [f'{_r} FTEs' for _r in roles_permitidos]:
                            _fte_pivot[_c] = _fte_pivot[_c].round(3)
                        summary = summary.merge(_fte_pivot, on='Client', how='left')

                        detail = df_pred[[
                            '_client', '_pod', '_type', '_subtype', '_proc_role', '_rev_role',
                            '_pred_vol', '_proc_aht', '_rev_aht',
                            '_prod_p_hrs', '_prod_r_hrs', '_prod_hrs_total',
                            '_full_p_hrs', '_full_r_hrs', '_full_hrs_total',
                            '_proc_fte', '_rev_fte', '_total_fte', '_full_cost',
                        ]].copy()
                        detail.columns = [
                            'Client', 'POD', 'Type', 'Subtype', 'Proc Role', 'Rev Role',
                            'Predicted Vol', 'Proc AHT (min)', 'Rev AHT (min)',
                            'Prod. Proc Hrs', 'Prod. Rev Hrs', 'Prod. Total Hrs',
                            'Full Proc Hrs', 'Full Rev Hrs', 'Full Total Hrs',
                            'Proc FTEs', 'Rev FTEs', 'Total FTEs', 'Full Cost ($)',
                        ]
                        for c in ['Predicted Vol','Proc AHT (min)','Rev AHT (min)',
                                  'Prod. Proc Hrs','Prod. Rev Hrs','Prod. Total Hrs',
                                  'Full Proc Hrs','Full Rev Hrs','Full Total Hrs']:
                            detail[c] = detail[c].round(2)
                        for c in ['Proc FTEs','Rev FTEs','Total FTEs']:
                            detail[c] = detail[c].round(3)
                        detail['Full Cost ($)'] = detail['Full Cost ($)'].round(2)

                        st.session_state[f"{pfx}_ai_results"] = {
                            'summary': summary, 'detail': detail,
                            'df_pred': df_pred, 'month_idx': _ai_month,
                            'hrs_fte_month': hrs_fte_month
                        }
                        st.success(f"✅ Predictions for **{summary['Client'].nunique()}** clients across **{len(detail)}** task rows.")

                    except Exception as e:
                        import traceback
                        st.error(f"❌ Error: {e}")
                        st.code(traceback.format_exc())

        # ── DISPLAY RESULTS ────────────────────────────────────────────────
        if f"{pfx}_ai_results" in st.session_state:
            _s4r = st.session_state[f"{pfx}_ai_results"]
            ai_sum = _s4r['summary']
            ai_det = _s4r['detail']
            _mi_r  = _s4r['month_idx']
            _hfm   = _s4r['hrs_fte_month']
            _dfp   = _s4r['df_pred']

            _money = lambda lbl: st.column_config.NumberColumn(lbl, format="$%.2f")
            _pct   = lambda lbl: st.column_config.NumberColumn(lbl, format="%.2f%%")

            t_ai_s, t_ai_d = st.tabs(["📋 Client Summary", "🔍 Task Detail"])
            with t_ai_s:
                st.write(f"### New Client Projections — {meses_proyeccion[_mi_r]}")
                _role_fte_cfg = {f'{_r} FTEs': st.column_config.NumberColumn(f'{_r} FTEs', format="%.3f")
                                 for _r in roles_permitidos}
                st.dataframe(ai_sum, use_container_width=True,
                             column_config={"MRR ($)": _money("MRR ($)"), "Full_Cost": _money("Full Cost ($)"),
                                            "Cost_vs_MRR_%": _pct("Cost vs MRR (%)"),
                                            **_role_fte_cfg})
            with t_ai_d:
                st.write("### Task-Level Breakdown")
                st.dataframe(ai_det, use_container_width=True,
                             column_config={"Full Cost ($)": _money("Full Cost ($)")})

            # Download
            _ai_dl_buf = BytesIO()
            with pd.ExcelWriter(_ai_dl_buf, engine='openpyxl') as _ai_xw:
                ai_sum.to_excel(_ai_xw, sheet_name='AI_Client_Summary', index=False)
                ai_det.to_excel(_ai_xw, sheet_name='AI_Task_Detail', index=False)
            st.download_button(
                "📥 Download AI Prediction",
                _ai_dl_buf.getvalue(),
                file_name=f"AI_NewClients_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"{pfx}_ai_dl"
            )

            # ── Add to Baseline (early section — before Step 2) ────────────────
            if add_to_baseline:
                st.divider()
                st.markdown("**Add new clients to the capacity baseline:**")
                st.caption(
                    "Injects the predicted clients into the raw data, then clears the cascade "
                    "so Step 3 re-runs with them included. Their hours will appear under "
                    "**(+) New Customer Hours** at their Go Live month."
                )
                if st.button("📥 Add to Baseline & Re-run Step 3", type="primary", key=f"{pfx}_ai_add_baseline"):
                    # ── Build property/PMS lookup from manual entry table ──────────
                    _manual_tbl = st.session_state.get(f"{pfx}_ai_manual_clients", pd.DataFrame())
                    _prop_map = {}
                    if not _manual_tbl.empty:
                        for _, _mr in _manual_tbl.iterrows():
                            _cn_key = str(_mr.get('Company Name', '')).lower().strip()
                            _prop_map[_cn_key] = {
                                'Res Prop':              float(_mr.get('Res Properties', 0) or 0),
                                'Commercial Properties': float(_mr.get('Comm Properties', 0) or 0),
                                'Res doors':             float(_mr.get('Res Doors', 0) or 0),
                                'Commercial Doors':      float(_mr.get('Comm Doors', 0) or 0),
                                'PMS':                   str(_mr.get('PMS', '') or ''),
                            }

                    # ── Build synthetic df_clean rows from df_pred ─────────────────
                    # Build Sr. Accountant lookup from existing df_clean BEFORE replacing rows
                    _existing_clean  = st.session_state.df_clean
                    _sr_lookup = {}
                    if 'Sr. Accountant' in _existing_clean.columns and 'client_name' in _existing_clean.columns:
                        _sr_lookup = (
                            _existing_clean.dropna(subset=['client_name'])
                            .groupby(_existing_clean['client_name'].str.lower().str.strip())['Sr. Accountant']
                            .first().to_dict()
                        )

                    _synth_rows = []
                    for _, _rp in _dfp.iterrows():
                        _cn    = str(_rp.get('_client', '')).strip()
                        _props = _prop_map.get(_cn.lower(), {})
                        # Carry Sr. Accountant from the existing rows being replaced (if any)
                        _sr_val = _sr_lookup.get(_cn.lower().strip(), '')
                        _pod_synth = str(_rp.get('_pod', '') or '').strip()
                        if _pod_synth.lower() in ('nan', 'none', ''): _pod_synth = ''
                        _synth_rows.append({
                            'client_name':                    _cn,
                            'POD':                            _pod_synth,
                            'Go Live':                        pd.to_datetime(_rp.get('_go_live'), errors='coerce'),
                            'Final Service Date':             pd.NaT,
                            'MRR':                            float(_rp.get('_mrr', 0) or 0),
                            'type':                           str(_rp.get('_type', '') or ''),
                            'subtype':                        str(_rp.get('_subtype', '') or ''),
                            'Closed tickets with Proc time':  float(_rp.get('_pred_vol', 0) or 0),
                            'Closed tickets with rev time':   float(_rp.get('_pred_vol', 0) or 0),
                            '>>> FINAL Capacity Proc AHT':    float(_rp.get('_proc_aht', 0) or 0),
                            '>>> FINAL Capacity Rev AHT':     float(_rp.get('_rev_aht', 0) or 0),
                            'Ideal Proc':                     str(_rp.get('_proc_role', 'Accountant I') or 'Accountant I'),
                            'Ideal Rev':                      str(_rp.get('_rev_role',  'Sr. Accountant') or 'Sr. Accountant'),
                            'Sr. Accountant':                 _sr_val,
                            'status':                         'client',
                            'PMS':                            _props.get('PMS', ''),
                            'Res Prop':                       _props.get('Res Prop', 0),
                            'Commercial Properties':          _props.get('Commercial Properties', 0),
                            'Res doors':                      _props.get('Res doors', 0),
                            'Commercial Doors':               _props.get('Commercial Doors', 0),
                        })
                    _synth_df = pd.DataFrame(_synth_rows)
                    _new_names_lower = set(_synth_df['client_name'].str.lower().str.strip())

                    # ── Merge into df_clean (replace any previous prediction for same clients)
                    _clean_ex = st.session_state.df_clean.copy()
                    _keep     = ~_clean_ex['client_name'].astype(str).str.lower().str.strip().isin(_new_names_lower)
                    st.session_state.df_clean = pd.concat([_clean_ex[_keep], _synth_df], ignore_index=True)

                    # ── Update df_clients_unique for MRR waterfall tracking ─────────
                    _duc = st.session_state.get('df_clients_unique', pd.DataFrame())
                    if not _duc.empty and 'client_name' in _duc.columns:
                        _duc_keep = _duc[~_duc['client_name'].astype(str).str.lower().str.strip().isin(_new_names_lower)]
                        _new_duc  = _synth_df.groupby('client_name', as_index=False).agg(
                            MRR=('MRR', 'first'), POD=('POD', 'first'),
                        )
                        if 'Go Live' in _synth_df.columns:
                            _new_duc['Go Live'] = _synth_df.groupby('client_name')['Go Live'].first().values
                        _new_duc['Final Service Date'] = pd.NaT
                        st.session_state.df_clients_unique = pd.concat([_duc_keep, _new_duc], ignore_index=True)
                    _build_client_master_map()   # re-sync maps after AI prediction adds new clients

                    # ── For replacement clients: purge stale df_resumen_base entries ──
                    # _ob_replace_set contains names of clients the user flagged as "replace
                    # existing data".  Their old Step-1 Base Hours must be removed so the
                    # cascade picks up only the new AI-predicted Post-Auto Hours.
                    _ob_rep_set = st.session_state.get('_ob_replace_set', set())
                    _replaced_lower = (
                        _ob_rep_set & {c.lower().strip() for c in _synth_df['client_name']}
                        if _ob_rep_set else set()
                    )
                    if _replaced_lower and 'calc_data' in st.session_state:
                        # Purge replaced clients from ALL three base keys so the cascade
                        # (which now reads df_resumen_base_ideal or df_resumen_base_real)
                        # doesn't carry stale hours for clients being replaced by AI projection.
                        for _rb_key in ('df_resumen_base', 'df_resumen_base_ideal', 'df_resumen_base_real'):
                            _rb = st.session_state.calc_data.get(_rb_key, pd.DataFrame())
                            if not _rb.empty and 'Client' in _rb.columns:
                                _rb_mask = ~_rb['Client'].astype(str).str.lower().str.strip().isin(_replaced_lower)
                                st.session_state.calc_data[_rb_key] = _rb[_rb_mask].copy()

                    # ── Clear final_dashboards and auto-trigger cascade (Step 3) ──────
                    if 'final_dashboards' in st.session_state:
                        del st.session_state['final_dashboards']
                    # Flag Step 3 to auto-run the cascade on next render
                    st.session_state['_auto_run_cascade'] = True
                    # Done with replace set — clear it so the panel resets
                    st.session_state['_ob_replace_set'] = set()

                    _n_added    = _synth_df['client_name'].nunique()
                    _n_replaced = len(_replaced_lower)
                    _n_new      = _n_added - _n_replaced
                    _msg_parts  = []
                    if _n_replaced:
                        _msg_parts.append(f"**{_n_replaced}** client(s) replaced with AI projection")
                    if _n_new:
                        _msg_parts.append(f"**{_n_new}** new client(s) added")
                    st.success(
                        f"✅ {' · '.join(_msg_parts) if _msg_parts else f'{_n_added} client(s) updated'}. "
                        "Re-running cascade automatically…"
                    )
                    st.rerun()

            # ── Apply to Scenario (Step 4 only) ────────────────────────────────
            if add_to_scenario:
                if st.session_state.get(f'{pfx}_ai_mode', _ai_mode) == "➕ Add to Step 4 Scenario":
                    st.divider()
                    st.markdown("**Apply predicted hours to Step 4 scenario:**")
                    st.caption(
                        f"This will add the predicted productive hours per role to the "
                        f"**M{_mi_r+1}** column of the Hours Adjustments table (confirmed = True). "
                        f"Existing values are added to, not replaced."
                    )
                    if st.button("✅ Apply to Scenario", type="primary", key=f"{pfx}_ai_apply"):
                        # ── 1. Hours per role ──────────────────────────────────────
                        _role_map = {r: 0.0 for r in roles_permitidos}
                        for _, _row_p in _dfp.iterrows():
                            _pr = str(_row_p.get('_proc_role', '')).strip()
                            _rr = str(_row_p.get('_rev_role',  '')).strip()
                            _ph = float(_row_p.get('_prod_p_hrs', 0) or 0)
                            _rh = float(_row_p.get('_prod_r_hrs', 0) or 0)
                            if _pr in _role_map: _role_map[_pr] += _ph
                            if _rr in _role_map: _role_map[_rr] += _rh

                        _mc_apply = _s4v2_mc[_mi_r] if _mi_r < len(_s4v2_mc) else _s4v2_mc[-1]
                        _hrs_df = st.session_state.s4v2_hrs_role_df.copy()
                        if "Confirmed" not in _hrs_df.columns:
                            _hrs_df.insert(0, "Confirmed", False)
                        for _rl, _hrs_v in _role_map.items():
                            if _hrs_v > 0:
                                _mask = _hrs_df["Role"] == _rl
                                if _mask.any():
                                    _hrs_df.loc[_mask, _mc_apply] = float(_hrs_df.loc[_mask, _mc_apply].values[0] or 0) + _hrs_v
                                    _hrs_df.loc[_mask, "Confirmed"] = True
                        st.session_state.s4v2_hrs_role_df = _hrs_df

                        # ── 2. MRR from new clients → New MRR adjustment row ───────
                        _total_new_mrr = float(
                            _dfp.groupby('_client')['_mrr'].first().sum()
                        ) if '_mrr' in _dfp.columns else 0.0
                        if _total_new_mrr > 0:
                            _mrr_df = st.session_state.s4v2_mrr_adj_df.copy()
                            if "Confirmed" not in _mrr_df.columns:
                                _mrr_df.insert(0, "Confirmed", False)
                            _new_mrr_mask = _mrr_df["Adjustment"] == "⊕ New MRR"
                            if _new_mrr_mask.any() and _mc_apply in _mrr_df.columns:
                                _mrr_df.loc[_new_mrr_mask, _mc_apply] = (
                                    float(_mrr_df.loc[_new_mrr_mask, _mc_apply].values[0] or 0)
                                    + _total_new_mrr
                                )
                                _mrr_df.loc[_new_mrr_mask, "Confirmed"] = True
                            st.session_state.s4v2_mrr_adj_df = _mrr_df

                        # Clear data-editor key so it re-reads the updated session state
                        # on the next full-page render (otherwise the editor's cached
                        # edits would silently overwrite the programmatic changes).
                        for _ek in ('s4v2_hrs_ed', 's4v2_mrr_ed'):
                            if _ek in st.session_state:
                                del st.session_state[_ek]
                        st.success(
                            f"✅ Predicted hours added to M{_mi_r+1} in Hours Adjustments"
                            + (f" and ${_total_new_mrr:,.0f} added to New MRR." if _total_new_mrr > 0 else ".")
                            + " Scroll to Step 4 to review."
                        )
                        st.rerun()  # full-page rerun so _s4v2_fragment also updates

    return _ai_frag


# ==========================================
# TAB: NEW CLIENT PREDICTION (Quick Access)

def _norm_name(s):
    """Normalize a person name for loose matching: lowercase, strip, remove accents."""
    import unicodedata
    s = str(s).strip().lower()
    s = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in s if not unicodedata.combining(c))

def _reviewer_role(proc_role):
    lookup = {k.lower(): v for k, v in ROLE_HIERARCHY.items()}
    return lookup.get(str(proc_role).strip().lower(), 'Sr. Accountant')

# ==========================================
# PIPELINE HELPER FUNCTIONS
# ==========================================

_KNOWN_ROLES = set(ROLE_HIERARCHY.keys())

def _detect_file_type(uploaded_file):
    """Auto-detect which pipeline stage a file belongs to."""
    fname = uploaded_file.name.lower()
    # By filename pattern
    if any(x in fname for x in ['hubspot', 'crm-export', 'crm_export', 'capacity-client', 'capacity_client']):
        return 'hubspot'
    if any(x in fname for x in ['ideal pair', 'ideal_pair', 'idealpair']):
        return 'ideal_pairs'
    if any(x in fname for x in ['door count', 'door_count', 'doorcount']):
        return 'door_count'
    # By sheet/column content
    try:
        uploaded_file.seek(0)
        xl = pd.ExcelFile(uploaded_file)
        uploaded_file.seek(0)
        for sheet in xl.sheet_names[:3]:
            try:
                df_peek = xl.parse(sheet, nrows=3)
                cols_lower = ' '.join(str(c).lower() for c in df_peek.columns)
                if 'last billed mrr' in cols_lower or 'original cmrr' in cols_lower:
                    return 'hubspot'
                if 'sub-process' in cols_lower and 'processor' in cols_lower and 'qa' in cols_lower:
                    return 'ideal_pairs'
                if 'residential doors' in cols_lower or 'commercial doors' in cols_lower:
                    return 'door_count'
                if 'closed tickets with proc time' in cols_lower or 'final capacity proc aht' in cols_lower:
                    return 'volume_aht'
            except Exception:
                continue
    except Exception:
        pass
    return 'unknown'


def _load_volume_aht(uploaded_file, log):
    """Load the Volume & AHT source file."""
    uploaded_file.seek(0)
    xl = pd.ExcelFile(uploaded_file)
    sheet = 'Query result' if 'Query result' in xl.sheet_names else xl.sheet_names[0]
    log(f"  Reading sheet '{sheet}' from {uploaded_file.name}")
    uploaded_file.seek(0)
    df = xl.parse(sheet)
    df.columns = df.columns.str.strip()

    # Fix duplicated 'Rev Role' columns (first one is actually Proc Role)
    new_cols = []
    seen = {}
    for c in df.columns:
        if c in seen:
            seen[c] += 1
            new_cols.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0
            new_cols.append(c)
    df.columns = new_cols

    if 'Proc Role' not in df.columns:
        if 'Rev Role' in df.columns:
            df = df.rename(columns={'Rev Role': 'Proc Role'})
        if 'Rev Role.1' in df.columns:
            df = df.rename(columns={'Rev Role.1': 'Rev Role'})

    log(f"  Loaded {len(df)} rows, {df['client_name'].nunique() if 'client_name' in df.columns else '?'} clients")
    return df


def _apply_ideal_pairs(df_vol, uploaded_file, log):
    """Merge Ideal Pairs into the volume dataframe."""
    uploaded_file.seek(0)
    df_ip = pd.read_excel(uploaded_file)
    df_ip.columns = df_ip.columns.str.strip()
    df_ip = df_ip.dropna(subset=['Process', 'Sub-process'])

    # Build lookup: (type_lower, subtype_lower) → (proc_role, rev_role)
    ideal_lookup = {}
    for _, row in df_ip.iterrows():
        proc_r = str(row.get('Processor', '')).strip()
        qa_r   = str(row.get('QA', '')).strip()
        if proc_r.lower() in ('nan', 'none', ''): continue
        if qa_r.lower()   in ('nan', 'none', ''): qa_r = ROLE_HIERARCHY.get(proc_r, 'Sr. Accountant')
        # Normalize roles not in hierarchy → Sr. Accountant
        if proc_r not in _KNOWN_ROLES: proc_r = 'Sr. Accountant'
        if qa_r   not in _KNOWN_ROLES: qa_r   = 'Sr. Accountant'
        key = (str(row['Process']).strip().lower(), str(row['Sub-process']).strip().lower())
        ideal_lookup[key] = (proc_r, qa_r)

    log(f"  Ideal pairs loaded: {len(ideal_lookup)} task mappings")

    matched, fallback = 0, 0

    def _get_pair(row):
        nonlocal matched, fallback
        key = (str(row.get('type', '')).strip().lower(), str(row.get('subtype', '')).strip().lower())
        if key in ideal_lookup:
            matched += 1
            return ideal_lookup[key]
        # Fallback: use real Proc/Rev Role, normalize to hierarchy
        proc_r = str(row.get('Proc Role', 'Accountant I')).strip()
        rev_r  = str(row.get('Rev Role',  'Sr. Accountant')).strip()
        if proc_r not in _KNOWN_ROLES: proc_r = 'Sr. Accountant'
        if rev_r  not in _KNOWN_ROLES: rev_r  = 'Sr. Accountant'
        fallback += 1
        return (proc_r, rev_r)

    pairs = df_vol.apply(_get_pair, axis=1)
    df_vol['Ideal Proc'] = [p[0] for p in pairs]
    df_vol['Ideal Rev']  = [p[1] for p in pairs]
    log(f"  Ideal pairs matched: {matched} rows | Fallback to real roles: {fallback} rows")
    return df_vol


def _apply_door_count(df_vol, uploaded_file, log):
    """Merge door count data into the volume dataframe."""
    uploaded_file.seek(0)
    # Try header rows 0 and 1 — pick the one that has 'Client Name'
    for h in [1, 0, 2]:
        df_dc = pd.read_excel(uploaded_file, header=h)
        uploaded_file.seek(0)
        if 'Client Name' in df_dc.columns or any('client' in str(c).lower() for c in df_dc.columns):
            break

    # Normalize column names
    col_map = {}
    for c in df_dc.columns:
        cl = str(c).lower().strip()
        if 'client' in cl and 'name' in cl:      col_map[c] = 'client_name'
        elif 'res' in cl and 'door' in cl:        col_map[c] = 'Res doors'
        elif 'res' in cl and 'prop' in cl:        col_map[c] = 'Res Prop'
        elif 'commercial' in cl and 'prop' in cl: col_map[c] = 'Commercial Properties'
        elif 'commercial' in cl and 'door' in cl: col_map[c] = 'Commercial Doors'
        elif 'sqft' in cl or 'commercial' in cl and 'sqft' in cl: col_map[c] = 'SQFT Commercial'
        elif 'corp' in cl:                        col_map[c] = 'Corp Books'
        elif 'volume variation' in cl:            col_map[c] = 'Volume Variation %'
        elif 'status' in cl:                      col_map[c] = 'Status_dc'

    df_dc = df_dc.rename(columns=col_map)
    if 'client_name' not in df_dc.columns:
        log("  ⚠️  Could not identify Client Name column in door count file — skipping merge")
        return df_vol

    df_dc['client_name'] = df_dc['client_name'].astype(str).str.strip()
    dc_cols = ['client_name', 'Res doors', 'Res Prop', 'Commercial Properties',
               'Commercial Doors', 'SQFT Commercial', 'Corp Books', 'Volume Variation %']
    dc_cols = [c for c in dc_cols if c in df_dc.columns]
    df_dc = df_dc[dc_cols].drop_duplicates(subset=['client_name']).dropna(subset=['client_name'])
    df_dc = df_dc[df_dc['client_name'].str.lower() != 'nan']

    # Drop pre-existing door count cols from vol to avoid duplicates
    drop_these = [c for c in dc_cols if c in df_vol.columns and c != 'client_name']
    df_vol = df_vol.drop(columns=drop_these, errors='ignore')
    df_vol = df_vol.merge(df_dc, on='client_name', how='left')
    log(f"  Door count merged: {len(df_dc)} client records, {len(dc_cols)-1} fields")
    return df_vol


def _apply_hubspot(df_vol, uploaded_file, log, is_update=False):
    """
    Enrich vol dataframe with HubSpot data (MRR, Go Live, Final Service Date, PMS, POD).
    Also returns df_new_clients (onboarding/prospect clients not yet in vol).
    """
    uploaded_file.seek(0)
    xl = pd.ExcelFile(uploaded_file)
    uploaded_file.seek(0)
    df_hs = xl.parse(xl.sheet_names[0])
    df_hs.columns = df_hs.columns.str.strip()

    # Normalize company name
    if 'Company name' in df_hs.columns:
        df_hs = df_hs.rename(columns={'Company name': 'client_name'})
    elif 'company_name' in df_hs.columns:
        df_hs = df_hs.rename(columns={'company_name': 'client_name'})
    df_hs['client_name'] = df_hs['client_name'].astype(str).str.strip()

    # MRR: Last Billed MRR → fallback Original CMRR
    df_hs['_mrr_last'] = pd.to_numeric(df_hs.get('Last Billed MRR',    pd.Series(dtype=float)), errors='coerce')
    df_hs['_mrr_orig'] = pd.to_numeric(df_hs.get('Original CMRR',       pd.Series(dtype=float)), errors='coerce')
    df_hs['MRR']       = df_hs['_mrr_last'].fillna(df_hs['_mrr_orig'])

    # Go Live: Delivery Confirmed date → fallback Go Live Date
    df_hs['_gl_conf']  = pd.to_datetime(df_hs.get('Delivery Confirmed Go-Live Date', pd.Series(dtype=str)), errors='coerce')
    df_hs['_gl_raw']   = pd.to_datetime(df_hs.get('Go Live Date',                   pd.Series(dtype=str)), errors='coerce')
    df_hs['Go Live']   = df_hs['_gl_conf'].fillna(df_hs['_gl_raw'])

    # Onboarding clients with no Go Live → today + 30 days
    ls_col = next((c for c in df_hs.columns if 'lifecycle' in c.lower()), None)
    if ls_col:
        mask_onboard = df_hs['Go Live'].isna() & df_hs[ls_col].astype(str).str.lower().isin(
            ['customer', 'onboarding', 'subscriber'])
        df_hs.loc[mask_onboard, 'Go Live'] = pd.Timestamp.today() + pd.DateOffset(days=30)
        log(f"  Go Live defaulted to today+30 for {mask_onboard.sum()} onboarding client(s)")

    df_hs['Final Service Date'] = pd.to_datetime(df_hs.get('Final Service Date', pd.Series(dtype=str)), errors='coerce')

    pms_col = next((c for c in df_hs.columns if 'pms' in c.lower() or 'property management' in c.lower()), None)
    df_hs['PMS'] = df_hs[pms_col].astype(str).fillna('Unknown') if pms_col else 'Unknown'

    df_hs['_hs_pod'] = df_hs.get('POD', pd.Series('', index=df_hs.index)).fillna('').astype(str).str.strip()

    # Supplemental door data from HubSpot
    df_hs['_hs_res_doors'] = pd.to_numeric(df_hs.get('Recent - Residential Doors',  pd.Series(dtype=float)), errors='coerce')
    df_hs['_hs_sqft']      = pd.to_numeric(df_hs.get('Recent - Commercial SQFT',    pd.Series(dtype=float)), errors='coerce')

    hs_merge = df_hs[['client_name', 'MRR', 'Go Live', 'Final Service Date',
                       'PMS', '_hs_pod', '_hs_res_doors', '_hs_sqft']].drop_duplicates(subset=['client_name'])

    # Drop stale cols from vol — POD is intentionally excluded so master DB assignments are preserved
    drop_these = ['MRR', 'Go Live', 'Final Service Date', 'PMS']
    df_vol = df_vol.drop(columns=[c for c in drop_these if c in df_vol.columns], errors='ignore')
    df_vol = df_vol.merge(hs_merge, on='client_name', how='left')

    # POD: keep master DB value; only fall back to HubSpot POD when master DB has nothing
    if 'POD' not in df_vol.columns:
        df_vol['POD'] = ''
    df_vol['POD'] = df_vol['POD'].fillna('').astype(str).str.strip()
    _hs_pod_mask = df_vol['POD'].eq('') & df_vol['_hs_pod'].ne('')
    df_vol.loc[_hs_pod_mask, 'POD'] = df_vol.loc[_hs_pod_mask, '_hs_pod']
    df_vol = df_vol.drop(columns=['_hs_pod'], errors='ignore')

    # Fill door count gaps from HubSpot
    if 'Res doors' not in df_vol.columns: df_vol['Res doors'] = np.nan
    if 'SQFT Commercial' not in df_vol.columns: df_vol['SQFT Commercial'] = np.nan
    df_vol['Res doors']      = df_vol['Res doors'].fillna(df_vol['_hs_res_doors'])
    df_vol['SQFT Commercial'] = df_vol['SQFT Commercial'].fillna(df_vol['_hs_sqft'])
    df_vol = df_vol.drop(columns=['_hs_res_doors', '_hs_sqft'], errors='ignore')

    log(f"  HubSpot enrichment applied: MRR, Go Live, FSD, PMS, POD for {len(hs_merge)} clients")

    # --- NEW CLIENTS (not in vol) ---
    existing_clients = set(df_vol['client_name'].str.lower().str.strip())
    df_new = df_hs[~df_hs['client_name'].str.lower().str.strip().isin(existing_clients)].copy()
    if ls_col:
        df_new = df_new[df_new[ls_col].astype(str).str.lower().isin(
            ['customer', 'onboarding', 'subscriber', 'prospect'])]

    # Build new clients in the format expected by AI prediction
    df_new_clients = pd.DataFrame({
        'company_name':     df_new['client_name'].values,
        'go_live_date':     df_new['Go Live'].dt.strftime('%Y-%m-%d').fillna(
                                (pd.Timestamp.today() + pd.DateOffset(days=30)).strftime('%Y-%m-%d')).values,
        'mrr':              df_new['MRR'].fillna(0).values,
        'pms':              df_new['PMS'].values,
        'res_doors':        df_new['_hs_res_doors'].fillna(0).values,
        'sqft_commercial':  df_new['_hs_sqft'].fillna(0).values,
        'corp_books':       'Unknown',
        'res_prop':         0,
        'commercial_properties': 0,
        'commercial_doors': 0,
        'status':           df_new[ls_col].values if ls_col else 'Onboarding',
    })
    log(f"  New clients identified for AI prediction: {len(df_new_clients)}")
    return df_vol, df_new_clients


def _run_pipeline(files_by_type, log):
    """
    Execute the full pipeline in order:
      1. Volume & AHT  2. Ideal Pairs  3. Door Count  4. HubSpot
    Returns (df_vol_merged, df_new_clients) or raises on error.
    """
    vol_file = files_by_type.get('volume_aht')
    ip_file  = files_by_type.get('ideal_pairs')
    dc_file  = files_by_type.get('door_count')
    hs_file  = files_by_type.get('hubspot')

    if vol_file is None:
        raise ValueError("No Volume & AHT file detected. Please upload it.")

    log("── Step 1/4: Loading Volume & AHT ──────────────────")
    df_vol = _load_volume_aht(vol_file, log)

    if ip_file:
        log("── Step 2/4: Applying Ideal Pairs ──────────────────")
        df_vol = _apply_ideal_pairs(df_vol, ip_file, log)
    else:
        log("── Step 2/4: No Ideal Pairs file — using real roles ─")
        # Use Proc/Rev Role columns as Ideal Proc/Rev
        df_vol['Ideal Proc'] = df_vol.get('Proc Role', 'Accountant I')
        df_vol['Ideal Rev']  = df_vol.get('Rev Role',  'Sr. Accountant')

    if dc_file:
        log("── Step 3/4: Merging Door Count ─────────────────────")
        df_vol = _apply_door_count(df_vol, dc_file, log)
    else:
        log("── Step 3/4: No Door Count file — skipping ──────────")

    df_new_clients = pd.DataFrame()
    if hs_file:
        log("── Step 4/4: Enriching with HubSpot ─────────────────")
        df_vol, df_new_clients = _apply_hubspot(df_vol, hs_file, log)
    else:
        log("── Step 4/4: No HubSpot file — MRR/dates not enriched")

    # Ensure canonical column order matching the rest of the app
    must_have = ['client_name', 'type', 'subtype', 'processor', 'Proc Role',
                 'reviewer', 'Rev Role',
                 'Closed tickets with Proc time', 'Closed tickets with rev time',
                 '>>> FINAL Capacity Proc AHT', '>>> FINAL Capacity Rev AHT',
                 'Capacity Processing Hours', 'Capacity reviewing hours', 'Capacity Hours spent',
                 'Ideal Proc', 'Ideal Rev', 'Volume Variation %',
                 'Res doors', 'Res Prop', 'Commercial Properties', 'Commercial Doors',
                 'SQFT Commercial', 'Corp Books', 'PMS', 'MRR', 'Status',
                 'Go Live', 'Final Service Date', 'POD', 'Sr. Accountant']
    for c in must_have:
        if c not in df_vol.columns:
            df_vol[c] = np.nan

    log(f"✅ Pipeline complete — {len(df_vol)} rows, {df_vol['client_name'].nunique()} clients")
    return df_vol, df_new_clients


def _build_client_master_map():
    """
    Builds a unified client→{pod, sr_accountant, mrr, go_live, fsd} lookup table.

    Priority chain (later overrides earlier):
      1. Master DB  — df_clean     → POD, Sr. Accountant, MRR, Go Live, FSD
      2. Unique ref — df_clients_unique → authoritative MRR, Go Live, FSD
      3. HubSpot    — hs_parsed._pod   → POD (HubSpot beats Master DB for POD)
      4. Reconcile  — hs_client_overrides → MRR, Go Live, FSD confirmed by user

    Result stored in st.session_state.client_master_map (DataFrame) AND as fast
    lookup dicts:
      st.session_state._pod_map  = {client_key_lower: pod_name}
      st.session_state._sr_map   = {client_key_lower: sr_name}
      st.session_state._vol_map  = {client_key_lower: {mrr, go_live, fsd}}
    """
    _EMPTY = {'nan', 'none', '', 'n/a', 'na'}

    def _norm_pod(v):
        s = str(v or '').strip()
        return 'No POD' if s.lower() in _EMPTY else s

    def _norm_str(v):
        s = str(v or '').strip()
        return '' if s.lower() in _EMPTY else s

    def _first_valid(series):
        s = series.dropna()
        return s.iloc[0] if not s.empty else None

    _df  = st.session_state.get('df_clean', pd.DataFrame())
    _duc = st.session_state.get('df_clients_unique', pd.DataFrame())
    _hs  = st.session_state.get('hs_parsed')
    _ov  = st.session_state.get('hs_client_overrides', {})

    _COLS = ['client_name', 'client_key', 'pod', 'sr_accountant',
             'mrr', 'go_live', 'fsd', 'source']

    if _df.empty:
        st.session_state.client_master_map = pd.DataFrame(columns=_COLS)
        st.session_state._pod_map = {}
        st.session_state._sr_map  = {}
        st.session_state._vol_map = {}
        return

    # ── LAYER 1: Master DB (df_clean) ─────────────────────────────────────────
    _df2 = _df.copy()
    _df2['_key'] = _df2['client_name'].astype(str).str.lower().str.strip()

    _rows = []
    for _k, _g in _df2.groupby('_key', sort=False):
        _cn  = _g['client_name'].iloc[0]
        _pod = _norm_pod(_first_valid(_g['POD']) if 'POD' in _g.columns else None)
        _sr  = _norm_str(_first_valid(_g['Sr. Accountant']) if 'Sr. Accountant' in _g.columns else None)
        _mrr = float(_first_valid(_g['MRR']) or 0)              if 'MRR'                  in _g.columns else 0.0
        _gl  = _first_valid(_g['Go Live'])                       if 'Go Live'              in _g.columns else pd.NaT
        _fsd = _first_valid(_g['Final Service Date'])            if 'Final Service Date'   in _g.columns else pd.NaT
        _rows.append({'client_name': _cn, 'client_key': _k,
                      'pod': _pod, 'sr_accountant': _sr,
                      'mrr': _mrr, 'go_live': _gl, 'fsd': _fsd,
                      'source': 'master_db'})

    _cmap = pd.DataFrame(_rows)

    # ── LAYER 2: df_clients_unique (authoritative MRR / Go Live / FSD) ────────
    if not _duc.empty and 'client_name' in _duc.columns:
        _duc2 = _duc.copy()
        _duc2['_key'] = _duc2['client_name'].astype(str).str.lower().str.strip()
        for _col, _tgt in [('MRR', 'mrr'), ('Go Live', 'go_live'), ('Final Service Date', 'fsd')]:
            if _col in _duc2.columns:
                _lkp = dict(zip(_duc2['_key'], _duc2[_col]))
                _hit = _cmap['client_key'].isin(_lkp)
                _cmap.loc[_hit, _tgt] = _cmap.loc[_hit, 'client_key'].map(_lkp)
        # Add clients that are ONLY in df_clients_unique (edge case)
        _dup_keys = set(_cmap['client_key'])
        for _, _dr in _duc2.iterrows():
            if _dr['_key'] in _dup_keys:
                continue
            _cmap = pd.concat([_cmap, pd.DataFrame([{
                'client_name': _dr['client_name'], 'client_key': _dr['_key'],
                'pod': 'No POD', 'sr_accountant': '',
                'mrr':   float(_dr.get('MRR', 0) or 0),
                'go_live': _dr.get('Go Live', pd.NaT),
                'fsd':     _dr.get('Final Service Date', pd.NaT),
                'source': 'master_db'
            }])], ignore_index=True)

    # ── LAYER 3: HubSpot POD enrichment (_pod column) ─────────────────────────
    if _hs is not None and not _hs.empty:
        _hs_cn_col = next((c for c in _hs.columns
                           if c.lower() in ('company name', 'company', 'name')), None)
        if _hs_cn_col and '_pod' in _hs.columns:
            _hs2 = _hs.copy()
            _hs2['_key'] = _hs2[_hs_cn_col].astype(str).str.lower().str.strip()
            _hs_pod_lkp = {}
            for _, _hr in _hs2.iterrows():
                _hk  = _hr['_key']
                _hp  = _norm_pod(_hr.get('_pod', ''))
                if _hp != 'No POD':
                    _hs_pod_lkp[_hk] = _hp
            # Update POD for existing clients
            _hit = _cmap['client_key'].isin(_hs_pod_lkp)
            _cmap.loc[_hit, 'pod']    = _cmap.loc[_hit, 'client_key'].map(_hs_pod_lkp)
            _cmap.loc[_hit, 'source'] = _cmap.loc[_hit, 'source'].apply(
                lambda s: 'hubspot' if s == 'master_db' else s)

            # Also pull MRR / Go Live from HubSpot for existing & new clients
            _hs_mrr_col = next((c for c in _hs.columns
                                if 'mrr' in c.lower() or 'amount' in c.lower()), None)
            _hs_gl_col  = next((c for c in _hs.columns
                                if 'go live' in c.lower() or 'go_live' in c.lower()
                                or 'start date' in c.lower()), None)
            _hs_fsd_col = next((c for c in _hs.columns
                                if 'final service' in c.lower()), None)

            _existing_keys = set(_cmap['client_key'])
            for _, _hr in _hs2.iterrows():
                _hk = _hr['_key']
                _m  = _cmap['client_key'] == _hk
                if _m.any():
                    if _hs_mrr_col:
                        _hmrr = pd.to_numeric(_hr.get(_hs_mrr_col, None), errors='coerce')
                        if pd.notna(_hmrr) and _hmrr > 0:
                            _cmap.loc[_m, 'mrr'] = float(_hmrr)
                    if _hs_gl_col:
                        _hgl = pd.to_datetime(_hr.get(_hs_gl_col, None), errors='coerce')
                        if pd.notna(_hgl):
                            _cmap.loc[_m, 'go_live'] = _hgl
                    if _hs_fsd_col:
                        _hfsd = pd.to_datetime(_hr.get(_hs_fsd_col, None), errors='coerce')
                        if pd.notna(_hfsd):
                            _cmap.loc[_m, 'fsd'] = _hfsd
                else:
                    # New client only in HubSpot (not yet in Master DB)
                    _hmrr = float(pd.to_numeric(_hr.get(_hs_mrr_col, 0), errors='coerce') or 0) if _hs_mrr_col else 0.0
                    _hgl  = pd.to_datetime(_hr.get(_hs_gl_col,  None), errors='coerce') if _hs_gl_col  else pd.NaT
                    _hfsd = pd.to_datetime(_hr.get(_hs_fsd_col, None), errors='coerce') if _hs_fsd_col else pd.NaT
                    _hp   = _hs_pod_lkp.get(_hk, 'No POD')
                    _cmap = pd.concat([_cmap, pd.DataFrame([{
                        'client_name': _hr[_hs_cn_col], 'client_key': _hk,
                        'pod': _hp, 'sr_accountant': '',
                        'mrr': _hmrr, 'go_live': _hgl, 'fsd': _hfsd,
                        'source': 'hubspot'
                    }])], ignore_index=True)

    # ── LAYER 4: hs_client_overrides (user-confirmed reconciliation) ───────────
    for _ov_cli, _ov_vals in _ov.items():
        _ov_key = str(_ov_cli).lower().strip()
        _m = _cmap['client_key'] == _ov_key
        if not _m.any():
            continue
        if _ov_vals.get('mrr'):
            try: _cmap.loc[_m, 'mrr'] = float(_ov_vals['mrr'])
            except Exception: pass
        if _ov_vals.get('start_date'):
            try: _cmap.loc[_m, 'go_live'] = pd.to_datetime(_ov_vals['start_date'])
            except Exception: pass
        if _ov_vals.get('fsd'):
            try: _cmap.loc[_m, 'fsd'] = pd.to_datetime(_ov_vals['fsd'])
            except Exception: pass
        if _ov_vals.get('pod'):
            _cmap.loc[_m, 'pod'] = str(_ov_vals['pod']).strip()
        if _ov_vals.get('is_terminating') and not _ov_vals.get('fsd'):
            pass  # FSD already set above if present
        _cmap.loc[_m, 'source'] = 'reconciled'

    # ── Normalize final POD column ─────────────────────────────────────────────
    _cmap['pod'] = _cmap['pod'].apply(_norm_pod)
    _cmap['sr_accountant'] = _cmap['sr_accountant'].apply(_norm_str)

    # ── Store results ──────────────────────────────────────────────────────────
    st.session_state.client_master_map = _cmap
    st.session_state._pod_map = dict(zip(_cmap['client_key'], _cmap['pod']))
    st.session_state._sr_map  = dict(zip(_cmap['client_key'], _cmap['sr_accountant']))
    st.session_state._vol_map = {
        row['client_key']: {
            'mrr':      row['mrr'],
            'go_live':  row['go_live'],
            'fsd':      row['fsd'],
            'source':   row['source'],
        }
        for _, row in _cmap.iterrows()
    }


def _parse_hubspot_file(uploaded_file):
    """Parse a HubSpot export and return a cleaned DataFrame with derived columns."""
    uploaded_file.seek(0)
    xl = pd.ExcelFile(uploaded_file)
    uploaded_file.seek(0)
    df = xl.parse(xl.sheet_names[0])
    df.columns = df.columns.str.strip()

    # Normalize company name column
    _name_candidates = [c for c in df.columns if c.lower() in ['company name', 'company', 'client name', 'client_name']]
    if _name_candidates:
        df = df.rename(columns={_name_candidates[0]: 'client_name'})
    if 'client_name' not in df.columns:
        raise ValueError("HubSpot file must have a 'Company name' or 'Client name' column.")
    df['client_name'] = df['client_name'].astype(str).str.strip()
    df = df[df['client_name'].str.lower() != 'nan'].copy()

    # MRR: Last Billed MRR first, then Original CMRR
    _mrr1 = next((c for c in df.columns if 'last billed mrr' in c.lower()), None)
    _mrr2 = next((c for c in df.columns if 'original cmrr' in c.lower()), None)
    df['_mrr'] = pd.to_numeric(df[_mrr1] if _mrr1 else pd.Series(dtype=float), errors='coerce')
    if _mrr2:
        df['_mrr'] = df['_mrr'].fillna(pd.to_numeric(df[_mrr2], errors='coerce'))
    df['_mrr'] = df['_mrr'].fillna(0.0)

    # Smart Start Date: Go Live Date first, then Delivery Confirmed Go-Live Date
    _gl_col  = next((c for c in df.columns if c.lower() == 'go live date'), None)
    _cgl_col = next((c for c in df.columns if 'delivery confirmed go-live' in c.lower()
                     or ('confirmed' in c.lower() and 'go-live' in c.lower())
                     or ('confirmed' in c.lower() and 'go live' in c.lower())), None)
    _raw_gl  = pd.to_datetime(df[_gl_col]  if _gl_col  else pd.Series(dtype=str), errors='coerce')
    _raw_cgl = pd.to_datetime(df[_cgl_col] if _cgl_col else pd.Series(dtype=str), errors='coerce')
    df['_go_live']        = _raw_gl
    df['_confirmed_gl']   = _raw_cgl
    df['_start_date']     = _raw_gl.fillna(_raw_cgl)   # Go Live preferred, fallback to Confirmed

    # Final Service Date
    _fsd_col = next((c for c in df.columns if 'final service date' in c.lower()), None)
    df['_fsd'] = pd.to_datetime(df[_fsd_col] if _fsd_col else pd.Series(dtype=str), errors='coerce')

    # Lifecycle Stage
    _ls_col = next((c for c in df.columns if 'lifecycle' in c.lower()), None)
    df['_lifecycle'] = df[_ls_col].astype(str).str.strip() if _ls_col else 'Unknown'

    # Retention Status
    _ret_col = next((c for c in df.columns if 'retention' in c.lower()), None)
    df['_retention'] = df[_ret_col].astype(str).str.strip() if _ret_col else ''

    # PMS
    _pms_col = next((c for c in df.columns if 'pms' in c.lower()
                     or 'property management software' in c.lower()
                     or 'property management system' in c.lower()), None)
    df['_pms'] = df[_pms_col].astype(str).str.strip() if _pms_col else 'Unknown'

    # POD — match 'pod' anywhere in the column name, or team/squad, but
    # exclude columns that are clearly about something else (e.g. "Upload", "Episode")
    _pod_col = next(
        (c for c in df.columns
         if ('pod' in c.lower() or 'team' in c.lower() or 'squad' in c.lower())
         and not any(x in c.lower() for x in ('upload', 'episode', 'period', 'template',
                                               'update', 'report', 'export', 'import'))),
        None
    )
    df['_pod'] = df[_pod_col].astype(str).str.strip() if _pod_col else ''

    # Termination flag: Lifecycle ∈ {Pending Termination, On Notice} AND
    #                   Retention  ∈ {Declined Retention, Pending Termination}
    _term_lc  = {'pending termination', 'on notice'}
    _term_ret = {'declined retention', 'pending termination'}
    df['_is_terminating'] = (
        df['_lifecycle'].str.lower().isin(_term_lc) &
        df['_retention'].str.lower().isin(_term_ret)
    )

    return df


def _enrich_ai_from_hs(df_clients, df_hs_parsed):
    """
    Enrich an ai_manual_clients DataFrame with data from the parsed HubSpot file.
    Only fills fields that are empty / zero / Unknown — never overwrites user edits.

    Columns enriched: POD, Go Live Date, MRR ($), PMS, Res Doors, SQFT Commercial.
    Matching is done case-insensitively on Company Name ↔ client_name.
    """
    if df_hs_parsed is None or df_hs_parsed.empty or df_clients.empty:
        return df_clients

    df = df_clients.copy()

    # Build lookup: lower-stripped client_name → HubSpot row
    _hs_lkp = {}
    for _, _hr in df_hs_parsed.iterrows():
        _cn = str(_hr.get('client_name', '')).strip().lower()
        if _cn and _cn not in ('nan', 'none', ''):
            _hs_lkp[_cn] = _hr

    # Detect door / sqft / pod column names in the raw HubSpot frame
    _rd_col   = next((c for c in df_hs_parsed.columns if 'residential doors'  in c.lower()), None)
    _sqft_col = next((c for c in df_hs_parsed.columns
                      if 'commercial sqft' in c.lower() or 'sqft commercial' in c.lower()), None)
    # Fallback raw POD column — in case _parse_hubspot_file missed the mapping
    _raw_pod_col = next(
        (c for c in df_hs_parsed.columns
         if ('pod' in c.lower() or 'team' in c.lower() or 'squad' in c.lower())
         and not any(x in c.lower() for x in ('upload', 'episode', 'period', 'template',
                                               'update', 'report', 'export', 'import'))),
        None
    )

    _EMPTY = {'', 'nan', 'none', 'unknown', 'nat'}

    for _idx, _row in df.iterrows():
        _key = str(_row.get('Company Name', '') or '').strip().lower()
        if _key not in _hs_lkp:
            continue
        _hs = _hs_lkp[_key]

        # ── POD ──────────────────────────────────────────────────────────────
        if 'POD' in df.columns:
            _cur = str(_row.get('POD', '') or '').strip().lower()
            if _cur in _EMPTY:
                # Try parsed _pod first, then fall back to the raw column
                _val = str(_hs.get('_pod', '') or '').strip()
                if _val.lower() in _EMPTY and _raw_pod_col:
                    _val = str(_hs.get(_raw_pod_col, '') or '').strip()
                if _val.lower() not in _EMPTY:
                    df.at[_idx, 'POD'] = _val

        # ── Go Live Date ─────────────────────────────────────────────────────
        if 'Go Live Date' in df.columns:
            _cur = str(_row.get('Go Live Date', '') or '').strip().lower()
            if _cur in _EMPTY:
                _gl = pd.to_datetime(_hs.get('_start_date', pd.NaT), errors='coerce')
                if pd.notna(_gl):
                    df.at[_idx, 'Go Live Date'] = _gl.strftime('%Y-%m-%d')

        # ── MRR ($) ──────────────────────────────────────────────────────────
        if 'MRR ($)' in df.columns:
            try:
                _cur_mrr = float(_row.get('MRR ($)', 0) or 0)
            except (TypeError, ValueError):
                _cur_mrr = 0.0
            if _cur_mrr == 0.0:
                try:
                    _hs_mrr = float(_hs.get('_mrr', 0) or 0)
                except (TypeError, ValueError):
                    _hs_mrr = 0.0
                if _hs_mrr > 0:
                    df.at[_idx, 'MRR ($)'] = _hs_mrr

        # ── PMS ──────────────────────────────────────────────────────────────
        if 'PMS' in df.columns:
            _cur = str(_row.get('PMS', '') or '').strip().lower()
            if _cur in _EMPTY:
                _val = str(_hs.get('_pms', '') or '').strip()
                if _val.lower() not in _EMPTY:
                    df.at[_idx, 'PMS'] = _val

        # ── Res Doors ────────────────────────────────────────────────────────
        if 'Res Doors' in df.columns and _rd_col:
            try:
                _cur_rd = float(_row.get('Res Doors', 0) or 0)
            except (TypeError, ValueError):
                _cur_rd = 0.0
            if _cur_rd == 0.0:
                try:
                    _hs_rd = float(_hs.get(_rd_col, 0) or 0)
                except (TypeError, ValueError):
                    _hs_rd = 0.0
                if _hs_rd > 0:
                    df.at[_idx, 'Res Doors'] = _hs_rd

        # ── SQFT Commercial ──────────────────────────────────────────────────
        if 'SQFT Commercial' in df.columns and _sqft_col:
            try:
                _cur_sq = float(_row.get('SQFT Commercial', 0) or 0)
            except (TypeError, ValueError):
                _cur_sq = 0.0
            if _cur_sq == 0.0:
                try:
                    _hs_sq = float(_hs.get(_sqft_col, 0) or 0)
                except (TypeError, ValueError):
                    _hs_sq = 0.0
                if _hs_sq > 0:
                    df.at[_idx, 'SQFT Commercial'] = _hs_sq

    return df


def _apply_hubspot_update(df_vol_base, df_new_base, uploaded_file, log):
    """
    Apply a fresh HubSpot export on top of an existing scenario:
     - Updates MRR, Go Live, Final Service Date for existing clients
     - Adds brand-new clients to df_new_clients
    """
    uploaded_file.seek(0)
    xl = pd.ExcelFile(uploaded_file)
    uploaded_file.seek(0)
    df_hs = xl.parse(xl.sheet_names[0])
    df_hs.columns = df_hs.columns.str.strip()

    if 'Company name' in df_hs.columns: df_hs = df_hs.rename(columns={'Company name': 'client_name'})
    df_hs['client_name'] = df_hs['client_name'].astype(str).str.strip()

    df_hs['_mrr'] = pd.to_numeric(df_hs.get('Last Billed MRR', pd.Series(dtype=float)), errors='coerce')
    df_hs['_mrr'] = df_hs['_mrr'].fillna(pd.to_numeric(df_hs.get('Original CMRR', pd.Series(dtype=float)), errors='coerce'))
    df_hs['_gl']  = pd.to_datetime(df_hs.get('Delivery Confirmed Go-Live Date', pd.Series(dtype=str)), errors='coerce')
    df_hs['_gl']  = df_hs['_gl'].fillna(pd.to_datetime(df_hs.get('Go Live Date', pd.Series(dtype=str)), errors='coerce'))
    df_hs['_fsd'] = pd.to_datetime(df_hs.get('Final Service Date', pd.Series(dtype=str)), errors='coerce')

    df_vol = df_vol_base.copy()
    updates = {'MRR': 0, 'Go Live': 0, 'Final Service Date': 0, 'New Clients': 0}

    for _, hs_row in df_hs.iterrows():
        cname = hs_row['client_name']
        mask  = df_vol['client_name'].str.strip().str.lower() == cname.lower()
        if mask.any():
            if pd.notna(hs_row['_mrr']):
                df_vol.loc[mask, 'MRR'] = hs_row['_mrr']; updates['MRR'] += 1
            if pd.notna(hs_row['_gl']):
                df_vol.loc[mask, 'Go Live'] = hs_row['_gl']; updates['Go Live'] += 1
            if pd.notna(hs_row['_fsd']):
                df_vol.loc[mask, 'Final Service Date'] = hs_row['_fsd']; updates['Final Service Date'] += 1

    log(f"  Updated existing clients — MRR: {updates['MRR']}, Go Live: {updates['Go Live']}, FSD: {updates['Final Service Date']}")

    # New clients not in vol
    existing = set(df_vol['client_name'].str.lower().str.strip())
    df_new_extra = df_hs[~df_hs['client_name'].str.lower().str.strip().isin(existing)].copy()

    pms_col = next((c for c in df_hs.columns if 'pms' in c.lower() or 'property management' in c.lower()), None)
    ls_col  = next((c for c in df_hs.columns if 'lifecycle' in c.lower()), None)

    if not df_new_extra.empty:
        df_add = pd.DataFrame({
            'company_name':          df_new_extra['client_name'].values,
            'go_live_date':          df_new_extra['_gl'].dt.strftime('%Y-%m-%d').fillna(
                                         (pd.Timestamp.today()+pd.DateOffset(days=30)).strftime('%Y-%m-%d')).values,
            'mrr':                   df_new_extra['_mrr'].fillna(0).values,
            'pms':                   df_new_extra[pms_col].astype(str).values if pms_col else 'Unknown',
            'res_doors':             pd.to_numeric(df_new_extra.get('Recent - Residential Doors', 0), errors='coerce').fillna(0).values,
            'sqft_commercial':       pd.to_numeric(df_new_extra.get('Recent - Commercial SQFT', 0), errors='coerce').fillna(0).values,
            'corp_books':            'Unknown',
            'res_prop':              0,
            'commercial_properties': 0,
            'commercial_doors':      0,
            'status':                df_new_extra[ls_col].values if ls_col else 'Onboarding',
        })
        df_new_clients = pd.concat([df_new_base, df_add], ignore_index=True).drop_duplicates(subset=['company_name'])
        log(f"  New clients added from HubSpot update: {len(df_add)}")
    else:
        df_new_clients = df_new_base.copy()
        log("  No new clients found in HubSpot update")

    return df_vol, df_new_clients


# ── SCENARIO PERSISTENCE ─────────────────────────────────────────────────────

def _save_scenario(name, description, df_vol, df_new_clients, source_files=None):
    safe_name = "".join(c if c.isalnum() or c in (' ', '_', '-') else '_' for c in name).strip().replace(' ', '_')
    ts        = datetime.now().strftime('%Y%m%d_%H%M%S')
    fpath     = os.path.join(SCENARIOS_DIR, f"{safe_name}_{ts}.pkl")
    payload   = {
        'name':          name,
        'description':   description,
        'vol_merged':    df_vol,
        'new_clients':   df_new_clients,
        'source_files':  source_files or [],
        'saved_at':      datetime.now().isoformat(),
    }
    with open(fpath, 'wb') as f:
        pickle.dump(payload, f)
    return fpath


def _list_scenarios():
    paths = sorted(glob.glob(os.path.join(SCENARIOS_DIR, "*.pkl")), reverse=True)
    result = []
    for p in paths:
        try:
            with open(p, 'rb') as f:
                meta = pickle.load(f)
            result.append({
                'path':        p,
                'name':        meta.get('name', os.path.basename(p)),
                'description': meta.get('description', ''),
                'saved_at':    meta.get('saved_at', ''),
                'clients':     len(meta.get('vol_merged', pd.DataFrame())['client_name'].unique())
                               if 'client_name' in meta.get('vol_merged', pd.DataFrame()).columns else 0,
                'new_clients': len(meta.get('new_clients', pd.DataFrame())),
            })
        except Exception:
            pass
    return result


def _load_scenario(path):
    with open(path, 'rb') as f:
        return pickle.load(f)


def _save_step4_scenario(name, description, payload):
    """Save a full Step 4 scenario (results + all inputs) to disk."""
    safe = "".join(c if c.isalnum() or c in (' ', '_', '-') else '_' for c in name).strip().replace(' ', '_')
    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    fpath = os.path.join(SCENARIOS_DIR, f"S4_{safe}_{ts}.pkl")
    payload['name']     = name
    payload['description'] = description
    payload['saved_at'] = datetime.now().isoformat()
    payload['type']     = 'step4'
    with open(fpath, 'wb') as f:
        pickle.dump(payload, f)
    return fpath


def _list_step4_scenarios():
    paths = sorted(glob.glob(os.path.join(SCENARIOS_DIR, "S4_*.pkl")), reverse=True)
    result = []
    for p in paths:
        try:
            with open(p, 'rb') as f:
                meta = pickle.load(f)
            result.append({
                'path':        p,
                'name':        meta.get('name', os.path.basename(p)),
                'description': meta.get('description', ''),
                'saved_at':    meta.get('saved_at', ''),
            })
        except Exception:
            pass
    return result


# ── GLOBAL PARAMS CONFIG SAVE / LOAD ─────────────────────────────────────────

def _save_params_config(name, params: dict) -> str:
    """Persist a named set of global parameters to disk."""
    safe  = "".join(c if c.isalnum() or c in (' ', '_', '-') else '_' for c in name).strip().replace(' ', '_')
    ts    = datetime.now().strftime('%Y%m%d_%H%M%S')
    fpath = os.path.join(PARAMS_CONFIG_DIR, f"PARAMS_{safe}_{ts}.json")
    payload = {
        'name':     name,
        'saved_at': datetime.now().isoformat(),
        'params':   params,
    }
    with open(fpath, 'w') as f:
        json.dump(payload, f, indent=2)
    return fpath


def _list_params_configs() -> list:
    """Return saved parameter configs sorted newest-first."""
    paths  = sorted(glob.glob(os.path.join(PARAMS_CONFIG_DIR, "PARAMS_*.json")), reverse=True)
    result = []
    for p in paths:
        try:
            with open(p) as f:
                meta = json.load(f)
            result.append({
                'path':     p,
                'name':     meta.get('name', os.path.basename(p)),
                'saved_at': meta.get('saved_at', ''),
                'params':   meta.get('params', {}),
            })
        except Exception:
            pass
    return result


def _load_params_config(path: str) -> dict:
    with open(path) as f:
        return json.load(f).get('params', {})


# ── HC REPORT PROCESSING ─────────────────────────────────────────────────────

_HC_ROLE_MAP = {
    'accountant i':       'Accountant I',
    'accountant ii':      'Accountant II',
    'general accountant': 'General Accountant',
    'sr. accountant':     'Sr. Accountant',
    'sr accountant':      'Sr. Accountant',
}

@st.cache_data(show_spinner=False)
def _process_hc_report(file_bytes: bytes):
    """
    Parse HC Weekly Report.
    Returns dict: {by_role, by_pod_role (DataFrame), total, detail (DataFrame)}
    Active Pod employees only; job titles mapped to 4 capacity planning roles.
    Decorated with @st.cache_data so the same file is not reprocessed on every rerun.
    """
    df = pd.read_excel(BytesIO(file_bytes), sheet_name='Weekly Report')
    df.columns = df.columns.str.strip()

    active = df[df['Worker Status'].astype(str).str.lower() == 'active'].copy()
    pod_mask = active['Department unit'].astype(str).str.strip().str.lower().str.startswith('pod')
    active_pods = active[pod_mask].copy()

    active_pods['Capacity Role'] = (
        active_pods['Job title'].astype(str).str.lower().str.strip()
        .map(_HC_ROLE_MAP).fillna('Other')
    )
    active_pods['POD'] = active_pods['Department unit'].astype(str).str.strip().str.title()

    _jt_lower = active_pods['Job title'].astype(str).str.lower().str.strip()
    _mgr_mask   = (
        _jt_lower.str.contains('accounting manager', na=False) |
        _jt_lower.str.contains('assistant manager',  na=False)
    )
    n_acct_mgr  = int((_jt_lower.str.contains('accounting manager',  na=False)).sum())
    n_asst_mgr  = int((_jt_lower.str.contains('assistant manager',   na=False)).sum())
    n_mgr_total = int(_mgr_mask.sum())

    # Managers per POD (Accounting + Assistant Managers combined)
    _mgr_df = active_pods[_mgr_mask]
    mgr_by_pod = (
        _mgr_df.groupby('POD')['Full name'].count().to_dict()
        if not _mgr_df.empty else {}
    )

    _productive_roles = {'Accountant I', 'Accountant II', 'General Accountant', 'Sr. Accountant'}
    by_role     = {k: v for k, v in active_pods.groupby('Capacity Role')['Full name'].count().to_dict().items()
                   if k in _productive_roles}
    by_pod_role = (active_pods[active_pods['Capacity Role'].isin(_productive_roles)]
                   .groupby(['POD', 'Capacity Role'])['Full name']
                   .count().reset_index().rename(columns={'Full name': 'HC'}))
    total       = int(active_pods['Capacity Role'].ne('Other').sum())  # count only mapped roles

    # ── Sr. Accountant → direct reports mapping (via Manager email — encoding-safe) ──
    _sr_jt_lower  = active_pods['Job title'].astype(str).str.lower().str.strip()
    _sr_staff     = active_pods[_sr_jt_lower.isin({'sr. accountant', 'sr accountant'})]
    # Use email as the reliable key; names are kept for display only
    _sr_email_map = {
        str(r['Work Email']).strip().lower(): str(r['Full name']).strip()
        for _, r in _sr_staff.iterrows()
        if str(r.get('Work Email', '')).strip()
    }  # email → full name
    _sr_email_set = set(_sr_email_map.keys())
    # Search ALL active workers for direct reports (not just pods)
    _all_active_cap = active.copy()
    _all_active_cap['Capacity Role'] = (
        _all_active_cap['Job title'].astype(str).str.lower().str.strip()
        .map(_HC_ROLE_MAP).fillna('Other')
    )
    _all_active_cap['_mgr_email_norm'] = (
        _all_active_cap['Manager email'].astype(str).str.strip().str.lower()
    )
    _dr_mask = _all_active_cap['_mgr_email_norm'].isin(_sr_email_set)
    _dr_df   = _all_active_cap[_dr_mask].copy()
    by_sr       = {}   # exact Sr. Full name key
    by_sr_norm  = {}   # normalized Sr. name key → same data (for fuzzy external lookup)
    by_sr_email = {}   # Sr. email key → same data
    for _email, _sn in _sr_email_map.items():
        _rpts     = _dr_df[_dr_df['_mgr_email_norm'] == _email]
        _dr_roles = _rpts['Capacity Role'].value_counts().to_dict()
        _dr_total = int((_rpts['Capacity Role'] != 'Other').sum())
        # Managers among this Sr.'s direct reports (rare, but counted for display)
        _rpts_jt  = _rpts['Job title'].astype(str).str.lower().str.strip()
        _sr_mgrs  = int((
            _rpts_jt.str.contains('accounting manager', na=False) |
            _rpts_jt.str.contains('assistant manager',  na=False)
        ).sum())
        # Sr. themselves counts as 1 under 'Sr. Accountant' — total = DRs + 1
        _sr_roles = dict(_dr_roles)
        _sr_roles['Sr. Accountant'] = 1          # always exactly 1 (themselves)
        _sr_total = _dr_total + 1
        _sr_data  = {
            'total':    _sr_total,
            'dr_total': _dr_total,               # direct reports only (no Sr.)
            'by_role':  _sr_roles,
            'managers': _sr_mgrs,
            'email':    _email,
        }
        by_sr[_sn]                  = _sr_data
        by_sr_norm[_norm_name(_sn)] = _sr_data
        by_sr_email[_email]         = _sr_data

    # ── Attrited (non-active) employees — kept so the Employee Level tab
    # can surface them with a "<role> Att" suffix, indicating they are no
    # longer productively available. We only include people who had a POD
    # and a Capacity-mapped role at some point; otherwise the list balloons
    # with support / corporate headcount unrelated to capacity planning.
    attrited_raw = df[~df['Worker Status'].astype(str).str.lower().isin(['active', 'ready to start'])].copy()
    if not attrited_raw.empty:
        _att_pod_mask = attrited_raw['Department unit'].astype(str).str.strip().str.lower().str.startswith('pod')
        attrited_pods = attrited_raw[_att_pod_mask].copy()
        attrited_pods['Capacity Role'] = (
            attrited_pods['Job title'].astype(str).str.lower().str.strip()
            .map(_HC_ROLE_MAP).fillna('Other')
        )
        attrited_pods['POD'] = attrited_pods['Department unit'].astype(str).str.strip().str.title()
        attrited_pods = attrited_pods[attrited_pods['Capacity Role'].isin(
            ['Accountant I', 'Accountant II', 'General Accountant', 'Sr. Accountant']
        )]
        attrited_detail = attrited_pods[['Full name', 'Work Email', 'Job title',
                                         'Capacity Role', 'POD', 'Worker Status']].copy()
    else:
        attrited_detail = pd.DataFrame(columns=[
            'Full name', 'Work Email', 'Job title', 'Capacity Role', 'POD', 'Worker Status'
        ])

    return {
        'by_role':        by_role,
        'by_pod_role':    by_pod_role,
        'by_sr':          by_sr,
        'by_sr_norm':     by_sr_norm,
        'by_sr_email':    by_sr_email,
        'total':          total,
        'acct_managers':  n_acct_mgr,
        'asst_managers':  n_asst_mgr,
        'mgr_total':      n_mgr_total,
        'mgr_by_pod':     mgr_by_pod,
        'detail':         active_pods[['Full name', 'Work Email', 'Job title', 'Capacity Role', 'POD']],
        'attrited_detail': attrited_detail,
    }


# --- PAGE CONFIGURATION ---
st.set_page_config(page_title="Capacity & FTE Projections", layout="wide", initial_sidebar_state="collapsed")

# ── Global layout CSS ─────────────────────────────────────────────────────────
st.markdown("""
<style>
/* Remove Streamlit's default max-width so the app uses the full browser window */
.main .block-container {
    max-width: 100% !important;
    padding-left: 1.5rem !important;
    padding-right: 1.5rem !important;
    padding-top: 1rem !important;
}
/* Horizontal-scrollable tab bars everywhere (main + nested) */
div[data-testid="stTabs"] > div:first-child {
    overflow-x: auto !important;
    flex-wrap: nowrap !important;
    scrollbar-width: auto;
    scrollbar-color: #888 #2a2a2a;
    -webkit-overflow-scrolling: touch;
    gap: 0 !important;
    padding-bottom: 4px !important;
}
div[data-testid="stTabs"] > div:first-child::-webkit-scrollbar {
    height: 5px;
}
div[data-testid="stTabs"] > div:first-child::-webkit-scrollbar-thumb {
    background: #888;
    border-radius: 4px;
}
div[data-testid="stTabs"] > div:first-child > button {
    white-space: nowrap !important;
    flex-shrink: 0 !important;
    min-width: fit-content !important;
    min-height: 42px !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
    font-size: 0.88rem !important;
    cursor: pointer !important;
}
/* Dataframes fill available width */
[data-testid="stDataFrame"] > div { width: 100% !important; }
/* Tighten expander padding to save vertical space */
[data-testid="stExpander"] > details > summary {
    padding-top: 0.4rem !important;
    padding-bottom: 0.4rem !important;
}

/* ── Sidebar scroll fix ─────────────────────────────────────────────────
   Force the sidebar to be fully scrollable with the mouse wheel no matter
   how tall the content is. Without this, Streamlit sometimes traps scroll
   events on buttons/dividers and only arrow keys can move the view. */
section[data-testid="stSidebar"] {
    overflow-y: auto !important;
    overflow-x: hidden !important;
}
section[data-testid="stSidebar"] > div:first-child {
    height: 100vh !important;
    max-height: 100vh !important;
    overflow-y: auto !important;
    overflow-x: hidden !important;
    padding-bottom: 3rem !important;   /* breathing room at the bottom */
    -webkit-overflow-scrolling: touch;
    scrollbar-width: thin;
    scrollbar-color: #888 transparent;
}
section[data-testid="stSidebar"] > div:first-child::-webkit-scrollbar {
    width: 8px;
}
section[data-testid="stSidebar"] > div:first-child::-webkit-scrollbar-thumb {
    background: #888;
    border-radius: 4px;
}
/* Make sidebar buttons slightly tighter so more fit on short screens */
section[data-testid="stSidebar"] .stButton > button {
    padding: 0.35rem 0.6rem !important;
    min-height: 2rem !important;
    font-size: 0.86rem !important;
}
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {
    margin-bottom: 0.25rem !important;
}
</style>
""", unsafe_allow_html=True)

st.title("📊 Capacity Planning & FTE Projections")
st.markdown("Operational projection integrating learning curves, automations, financial savings, and cascade adjustments.")

# ── Sidebar — quick-jump navigation ───────────────────────────────────────────
with st.sidebar:
    st.markdown("## Navigation")
    st.caption("Click to jump to a section")

    _nav_sections = [
        ("📂 Data Load & Filters",         0),
        ("📊 Volume & AHT",                1),
        ("🧠 AI Prediction",               2),
        ("🔄 Recon",                        3),
        ("📊 Actual Hours",                4),
    ]
    st.markdown("**Main Tabs**")
    for _nav_label, _nav_idx in _nav_sections:
        if st.button(_nav_label, key=f"_nav_main_{_nav_idx}", use_container_width=True):
            st.session_state['_nav_tab_idx'] = _nav_idx
            st.rerun()

    st.divider()
    st.markdown("**Within Data Load**")

    # Step anchors — use JS scroll to the matching header element
    import streamlit.components.v1 as _stc_nav

    _step_anchors = [
        ("⚙️ Step 1 — Base Math",     "STEP 1"),
        ("🔧 Step 2 — Simulator",      "STEP 2"),
        ("📊 Step 3 — Dashboards",     "STEP 3"),
        ("📐 Step 4 — Scenario",       "STEP 4"),
    ]
    for _sa_label, _sa_text in _step_anchors:
        if st.button(_sa_label, key=f"_nav_{_sa_text.replace(' ','_')}", use_container_width=True):
            # Switch to tab1 (Data Load), then scroll to the step header via JS
            st.session_state['_nav_tab_idx'] = 0
            st.session_state['_nav_scroll_to'] = _sa_text
            st.rerun()

    # Inject scroll JS if a scroll target is queued
    if '_nav_scroll_to' in st.session_state:
        _scroll_text = st.session_state.pop('_nav_scroll_to')
        _stc_nav.html(f"""<script>
        setTimeout(function(){{
            var els = window.parent.document.querySelectorAll('h3, h4, .stMarkdown');
            for (var i = 0; i < els.length; i++) {{
                if (els[i].innerText && els[i].innerText.includes('{_scroll_text}')) {{
                    els[i].scrollIntoView({{behavior:'smooth', block:'start'}});
                    break;
                }}
            }}
        }}, 400);
        </script>""", height=0)

    st.divider()
    st.markdown("**Quick Actions**")
    if st.button("🔄 Clear Cache & Reload", key="_nav_clear_cache", use_container_width=True):
        for _ck in ['df_clean', 'df_clients_unique', 'calc_data', 'final_dashboards',
                    '_cascade_export_buf_ideal', '_cascade_export_buf_real']:
            st.session_state.pop(_ck, None)
        st.rerun()

    # Show data status
    st.divider()
    st.markdown("**Status**")
    _sb_data_ok  = "df_clean"       in st.session_state
    _sb_calc_ok  = "calc_data"      in st.session_state
    _sb_dash_ok  = "final_dashboards" in st.session_state
    st.markdown(f"Data loaded: {'✅' if _sb_data_ok else '❌'}")
    st.markdown(f"Step 1 done: {'✅' if _sb_calc_ok else '❌'}")
    st.markdown(f"Step 3 done: {'✅' if _sb_dash_ok else '❌'}")
    _sb_ideal_ok = bool(st.session_state.get('_cascade_export_buf_ideal'))
    _sb_real_ok  = bool(st.session_state.get('_cascade_export_buf_real'))
    st.markdown(f"Export Ideal: {'✅' if _sb_ideal_ok else '⬜'}")
    st.markdown(f"Export Real:  {'✅' if _sb_real_ok  else '⬜'}")

# --- GENERATE MONTHS: previous month (base) + 5 forward ---
# Starting from last month means the first column is always the reference/base month
# (e.g. if today is March, columns are: Feb, Mar, Apr, May, Jun, Jul)
_actual_today   = datetime.today()
_prev_month_dt  = _actual_today.replace(day=1) - relativedelta(months=1)

# Read user-selected base month/year from session state (set in Master Database expander)
_sel_base_month = st.session_state.get('base_month_sel', _prev_month_dt.month)
_sel_base_year  = st.session_state.get('base_year_sel',  _prev_month_dt.year)

# Derive `today` so that offset -1 always points to the user-chosen base month
_base_date = datetime(_sel_base_year, _sel_base_month, 15)  # mid-month avoids edge cases
today = _base_date + relativedelta(months=1)

_month_offsets   = list(range(-1, 5))   # -1=base month, 0=month after base, 1..4=forward
meses_proyeccion = [(today + relativedelta(months=off)).strftime("%B %Y") for off in _month_offsets]
roles_permitidos = ["Accountant I", "Accountant II", "General Accountant", "Sr. Accountant"]

AFFECTS_OPTIONS = [
    "Vol Proc",
    "Vol Rev",
    "AHT Proc",
    "AHT Rev",
    "Vol Proc + Vol Rev",
    "AHT Proc + AHT Rev",
    "All (Vol + AHT)",
]

# --- INITIALIZE SESSION STATE (INTERACTIVE TABLES) ---
if "automations_df" not in st.session_state:
    st.session_state.automations_df = pd.DataFrame(columns=[
        "Confirmed", "Initiative Name", "POD", "PMS", "Client", "Task (Type - Subtype)",
        "Affects",
        "M1 (%)", "M2 (%)", "M3 (%)", "M4 (%)", "M5 (%)", "M6 (%)"
    ])

meses_hrs_cols = [f"M{i+1} (Hrs)" for i in range(6)]
meses_fte_cols = [f"M{i+1} (FTEs)" for i in range(6)]

meses_pct_cols = [f"M{i+1} (%)" for i in range(6)]   # for door-count variation

if "historical_df" not in st.session_state:
    st.session_state.historical_df = pd.DataFrame(columns=["Confirmed", "POD", "Client", "Required Role"] + meses_hrs_cols)
elif "Confirmed" not in st.session_state.historical_df.columns:
    st.session_state.historical_df.insert(0, "Confirmed", False)
if "POD" not in st.session_state.historical_df.columns:
    st.session_state.historical_df.insert(1, "POD", "")

if "doorcount_df" not in st.session_state:
    st.session_state.doorcount_df = pd.DataFrame(
        columns=["Confirmed", "Client", "POD"] + meses_pct_cols
    )

if "reductions_df" not in st.session_state:
    st.session_state.reductions_df = pd.DataFrame(columns=["Confirmed", "POD", "Client", "Required Role"] + meses_hrs_cols)
else:
    if "Confirmed" not in st.session_state.reductions_df.columns:
        st.session_state.reductions_df.insert(0, "Confirmed", False)
    if "POD" not in st.session_state.reductions_df.columns:
        st.session_state.reductions_df.insert(1, "POD", "")

if "s4v2_hist_df" not in st.session_state:
    st.session_state.s4v2_hist_df = pd.DataFrame(
        columns=["Confirmed", "POD", "Client", "Required Role"] + meses_hrs_cols)
if "s4v2_red_df" not in st.session_state:
    st.session_state.s4v2_red_df = pd.DataFrame(
        columns=["Confirmed", "POD", "Client", "Required Role"] + meses_hrs_cols)
if "s4v2_doorcount_df" not in st.session_state:
    st.session_state.s4v2_doorcount_df = pd.DataFrame(
        columns=["Confirmed", "Client", "POD"] + meses_pct_cols)

if "hs_sync_choice" not in st.session_state:
    st.session_state.hs_sync_choice = None          # None | "yes" | "skip"
if "hs_parsed" not in st.session_state:
    st.session_state.hs_parsed = None               # parsed HubSpot DataFrame
if "hs_recon_df" not in st.session_state:
    st.session_state.hs_recon_df = None             # reconciliation editable df
if "hs_client_overrides" not in st.session_state:
    st.session_state.hs_client_overrides = {}       # {client_name: {mrr, start_date, fsd, is_terminating}}
if "hs_onboarding_clients" not in st.session_state:
    st.session_state.hs_onboarding_clients = None   # df of new onboarding clients for AI prediction
if "_ob_replace_set" not in st.session_state:
    st.session_state["_ob_replace_set"] = set()     # clients checked for AI-prediction replacement
if "s2_efficiency_choice" not in st.session_state:
    st.session_state.s2_efficiency_choice = None  # None | "yes" | "skip"

if "va_client_name" not in st.session_state:
    st.session_state["va_client_name"] = ""
if "va_mrr" not in st.session_state:
    st.session_state["va_mrr"] = 0.0
if "va_pms" not in st.session_state:
    st.session_state["va_pms"] = ""

if "ai_manual_clients" not in st.session_state:
    st.session_state.ai_manual_clients = pd.DataFrame(columns=[
        "Company Name", "POD", "Go Live Date", "MRR ($)", "PMS",
        "Res Doors", "Res Properties", "Comm Doors", "Comm Properties",
        "SQFT Commercial", "Corp Books"
    ])
elif "POD" not in st.session_state.ai_manual_clients.columns:
    st.session_state.ai_manual_clients.insert(1, "POD", "")

meses_mrr_cols = [f"M{i+1} ($)" for i in range(6)]

if "ramp_df" not in st.session_state:
    st.session_state.ramp_df = pd.DataFrame(
        columns=["Confirmed", "POD", "Client", "Required Role"] + meses_fte_cols
    )
else:
    if "Confirmed" not in st.session_state.ramp_df.columns:
        st.session_state.ramp_df.insert(0, "Confirmed", False)
    if "POD" not in st.session_state.ramp_df.columns:
        st.session_state.ramp_df.insert(1, "POD", "")
    # Migrate old Hrs columns to FTEs columns
    for _old, _new in zip(meses_hrs_cols, meses_fte_cols):
        if _old in st.session_state.ramp_df.columns and _new not in st.session_state.ramp_df.columns:
            st.session_state.ramp_df = st.session_state.ramp_df.rename(columns={_old: _new})

if "manual_mrr_df" not in st.session_state:
    st.session_state.manual_mrr_df = pd.DataFrame(
        columns=["Confirmed", "POD", "Client"] + meses_mrr_cols
    )
else:
    if "Confirmed" not in st.session_state.manual_mrr_df.columns:
        st.session_state.manual_mrr_df.insert(0, "Confirmed", False)
    if "POD" not in st.session_state.manual_mrr_df.columns:
        st.session_state.manual_mrr_df.insert(1, "POD", "")

_rev_hc_mrr_cols = [f"M{i+1} ($)" for i in range(6)]
_rev_hc_fte_cols = [f"M{i+1} (FTEs)" for i in range(6)]
if "rev_hc_mrr_df" not in st.session_state:
    st.session_state.rev_hc_mrr_df = pd.DataFrame([[0.0] * 6], columns=_rev_hc_mrr_cols)
if "rev_hc_hc_df" not in st.session_state:
    st.session_state.rev_hc_hc_df = pd.DataFrame(
        {"Role": roles_permitidos, **{c: [0.0] * len(roles_permitidos) for c in _rev_hc_fte_cols}}
    )

# ── Step 4 v2 scenario adjuster session state ─────────────────────────────
_s4v2_mc = [f"M{i+1}" for i in range(6)]
if "s4v2_hc_adj_df" not in st.session_state:
    _hc_init = []
    for _rl in roles_permitidos:
        _hc_init.append({"Confirmed": False, "Direction": "↑ Ramp Up",   "Role": _rl, **{c: 0.0 for c in _s4v2_mc}})
        _hc_init.append({"Confirmed": False, "Direction": "↓ Ramp Down", "Role": _rl, **{c: 0.0 for c in _s4v2_mc}})
    st.session_state.s4v2_hc_adj_df = pd.DataFrame(_hc_init)
elif "Confirmed" not in st.session_state.s4v2_hc_adj_df.columns:
    st.session_state.s4v2_hc_adj_df.insert(0, "Confirmed", False)
if "s4v2_mrr_adj_df" not in st.session_state:
    st.session_state.s4v2_mrr_adj_df = pd.DataFrame([
        {"Confirmed": False, "Adjustment": "⊕ New MRR",   **{c: 0.0 for c in _s4v2_mc}},
        {"Confirmed": False, "Adjustment": "⊖ Churn MRR", **{c: 0.0 for c in _s4v2_mc}},
    ])
elif "Confirmed" not in st.session_state.s4v2_mrr_adj_df.columns:
    st.session_state.s4v2_mrr_adj_df.insert(0, "Confirmed", False)
if "s4v2_hrs_role_df" not in st.session_state:
    _hrs_init = [{"Confirmed": False, "Role": _rl, **{c: 0.0 for c in _s4v2_mc}} for _rl in roles_permitidos]
    st.session_state.s4v2_hrs_role_df = pd.DataFrame(_hrs_init)
elif "Confirmed" not in st.session_state.s4v2_hrs_role_df.columns:
    st.session_state.s4v2_hrs_role_df.insert(0, "Confirmed", False)
_s4v2_auto_cols = ["Confirmed", "Initiative Name", "POD", "PMS", "Client", "Task (Type - Subtype)", "Affects",
                   "M1 (%)", "M2 (%)", "M3 (%)", "M4 (%)", "M5 (%)", "M6 (%)"]
if "s4v2_auto_df" not in st.session_state:
    st.session_state.s4v2_auto_df = pd.DataFrame(columns=_s4v2_auto_cols)

if "s4v2_ai_hrs" not in st.session_state:
    # Stores AI-predicted productive hours per role for each projected month
    st.session_state.s4v2_ai_hrs = {_rl: [0.0]*6 for _rl in roles_permitidos}

# ── Defaults so variables are always defined (overwritten by widgets in tab1 ⚙️ expander) ──
cost_acc1 = 1140.0; cost_acc2 = 1369.0; cost_gen = 1900.0; cost_sr = 2536.0
cost_map  = {'Accountant I': cost_acc1, 'Accountant II': cost_acc2,
             'General Accountant': cost_gen, 'Sr. Accountant': cost_sr}
absenteeism = 0.10; attrition = 0.03
util_acc1 = 0.85;  util_gen  = 0.80;  util_sr = 0.50
utilization_map = {'Accountant I': util_acc1, 'Accountant II': util_acc1,
                   'General Accountant': util_gen, 'Sr. Accountant': util_sr}
calc_mode  = "Fixed days per month"; fixed_days = 22
holidays_per_month = {mes: 0 for mes in meses_proyeccion}

# ── START OVER ───────────────────────────────────────────────────────────────
if st.sidebar.button("🔄 Start Over", use_container_width=True, key="start_over_btn"):
    st.session_state.clear()
    st.rerun()

# ── SIDEBAR COLLAPSE HINT ────────────────────────────────────────────────────
st.sidebar.markdown(
    "<small>💡 Press **`<`** on the left edge to collapse, **`>`** to expand for navigation.</small>",
    unsafe_allow_html=True,
)

# ── SCENARIO MANAGER IN SIDEBAR ──────────────────────────────────────────────
st.sidebar.divider()
with st.sidebar.expander("📦 Scenario Manager", expanded=False):
    _scenarios = _list_scenarios()

    if _scenarios:
        _sc_options = {f"{s['name']}  ({s['saved_at'][:10]})  — {s['clients']} clients": s for s in _scenarios}
        _selected_sc_label = st.selectbox("Load scenario into app:", options=list(_sc_options.keys()), key="sc_select")
        if st.button("⬆️ Load Selected Scenario", key="sc_load_btn"):
            _sc = _load_scenario(_sc_options[_selected_sc_label]['path'])
            st.session_state.pipeline_vol_merged  = _sc['vol_merged']
            st.session_state.pipeline_new_clients = _sc['new_clients']
            st.session_state.active_scenario      = _sc
            # Clear downstream cache so Step 1 re-reads from scenario
            for _k in ['df_clean', 'df_clients_unique', 'calc_data', 'final_dashboards']:
                st.session_state.pop(_k, None)
            st.success(f"Loaded: {_sc['name']}")
            st.rerun()

        if st.button("🗑️ Delete Selected Scenario", key="sc_del_btn"):
            os.remove(_sc_options[_selected_sc_label]['path'])
            st.success("Deleted.")
            st.rerun()
    else:
        st.info("No saved scenarios yet.")

    if "active_scenario" in st.session_state:
        st.caption(f"Active: **{st.session_state.active_scenario['name']}**")

# ── Sidebar scroll forwarder (JS) ────────────────────────────────────────────
# CSS alone does not fix the "wheel gets trapped on buttons/dividers" issue —
# we attach a capture-phase wheel listener on the sidebar that manually scrolls
# the inner container, bypassing any child element that stops propagation.
import streamlit.components.v1 as _stc_sb_scroll
_stc_sb_scroll.html("""
<script>
(function(){
    var doc = window.parent.document;
    if (doc.__capSidebarScrollInit) return;

    function findScrollableAncestor(el, stopAt) {
        while (el && el !== stopAt) {
            var cs = window.getComputedStyle(el);
            if ((cs.overflowY === 'auto' || cs.overflowY === 'scroll')
                && el.scrollHeight > el.clientHeight + 1) {
                return el;
            }
            el = el.parentElement;
        }
        return null;
    }

    function attach() {
        var sb = doc.querySelector('section[data-testid="stSidebar"]');
        if (!sb) { setTimeout(attach, 300); return; }
        doc.__capSidebarScrollInit = true;

        // Pick the inner scroll container
        var container = sb.querySelector('[data-testid="stSidebarContent"]')
                     || sb.querySelector('[data-testid="stSidebarUserContent"]')
                     || sb.firstElementChild;
        if (container) {
            container.style.overflowY = 'auto';
            container.style.overscrollBehavior = 'contain';
        }

        sb.addEventListener('wheel', function(ev){
            var c = doc.querySelector('section[data-testid="stSidebar"] [data-testid="stSidebarContent"]')
                 || doc.querySelector('section[data-testid="stSidebar"] [data-testid="stSidebarUserContent"]')
                 || (doc.querySelector('section[data-testid="stSidebar"]') || {}).firstElementChild;
            // If the wheel originated inside a nested scrollable (e.g. a table),
            // let the browser handle it normally.
            var nested = findScrollableAncestor(ev.target, c);
            if (nested && nested !== c) return;
            if (c && c.scrollHeight > c.clientHeight + 1) {
                c.scrollTop += ev.deltaY;
                ev.preventDefault();
                ev.stopPropagation();
            }
        }, { passive: false, capture: true });
    }
    attach();
})();
</script>
""", height=0)

# ── Main content wheel forwarder (JS) ────────────────────────────────────────
# Streamlit's `st.dataframe` (and Plotly, AgGrid, etc.) trap wheel events so
# the mouse wheel only scrolls the inner table — not the page — when the
# cursor hovers over a waterfall. When the inner table has no more room to
# scroll in the wheel direction, OR when the table content fits without its
# own scrollbar (common for collapsed groups), we need to forward the wheel
# to the main page scroll container. This listener does exactly that.
import streamlit.components.v1 as _stc_main_scroll
_stc_main_scroll.html("""
<script>
(function(){
    var doc = window.parent.document;
    var win = window.parent;
    if (doc.__capMainScrollInit) return;
    doc.__capMainScrollInit = true;

    function isScrollable(el){
        if (!el || el === doc.body || el === doc.documentElement) return false;
        var cs = win.getComputedStyle(el);
        var oy = cs.overflowY;
        return (oy === 'auto' || oy === 'scroll') && el.scrollHeight > el.clientHeight + 1;
    }

    function findInnerScrollable(startEl, stopAt){
        var el = startEl;
        while (el && el !== stopAt && el !== doc.body){
            if (isScrollable(el)) return el;
            el = el.parentElement;
        }
        return null;
    }

    function pageScroller(){
        // Streamlit app main scroll container — try a few selectors
        return doc.querySelector('section.main')
            || doc.querySelector('[data-testid="stAppViewContainer"] section')
            || doc.querySelector('[data-testid="stAppViewContainer"]')
            || doc.scrollingElement
            || doc.documentElement;
    }

    doc.addEventListener('wheel', function(ev){
        // Ignore sidebar (handled by its own listener)
        var sb = doc.querySelector('section[data-testid="stSidebar"]');
        if (sb && sb.contains(ev.target)) return;

        var page = pageScroller();
        // Find innermost scrollable ancestor of the wheel target
        var inner = findInnerScrollable(ev.target, page);
        if (!inner){
            // No nested scroller — browser default handles page scroll
            return;
        }
        // If inner can still scroll in the wheel direction, let it.
        var dy = ev.deltaY;
        var atTop    = inner.scrollTop <= 0;
        var atBottom = (inner.scrollTop + inner.clientHeight) >= (inner.scrollHeight - 1);
        if ((dy < 0 && !atTop) || (dy > 0 && !atBottom)){
            return; // inner consumes it naturally
        }
        // Forward to page
        if (page){
            page.scrollTop += dy;
            ev.preventDefault();
            ev.stopPropagation();
        }
    }, { passive: false, capture: true });

    // ── MutationObserver: auto-refresh scroll whenever content changes ────
    function refreshScroll(){
        win.dispatchEvent(new Event('resize'));
        var p = pageScroller();
        if (p) { var t = p.scrollTop; p.scrollTop = t + 1; p.scrollTop = t; }
    }
    var _capObserver = new MutationObserver(function(){ refreshScroll(); });
    var _capTarget = doc.querySelector('[data-testid="stAppViewBlockContainer"]')
                  || doc.querySelector('.main .block-container')
                  || doc.querySelector('section.main');
    if (_capTarget) _capObserver.observe(_capTarget, { childList: true, subtree: false });

    // ── Floating "Fix Scroll" button ─────────────────────────────────────
    if (!doc.getElementById('cap-fix-scroll-btn')) {
        var _btn = doc.createElement('button');
        _btn.id = 'cap-fix-scroll-btn';
        _btn.title = 'Fix scroll bar';
        _btn.textContent = '⇕';
        _btn.style.cssText = [
            'position:fixed',
            'bottom:18px',
            'right:18px',
            'z-index:99999',
            'width:36px',
            'height:36px',
            'border-radius:50%',
            'border:1.5px solid rgba(255,255,255,0.18)',
            'background:rgba(30,33,48,0.88)',
            'color:#cdd3f0',
            'font-size:16px',
            'cursor:pointer',
            'box-shadow:0 2px 10px rgba(0,0,0,0.45)',
            'display:flex',
            'align-items:center',
            'justify-content:center',
            'transition:background 0.2s',
            'line-height:1',
        ].join(';');
        _btn.onmouseover = function(){ _btn.style.background = 'rgba(255,75,75,0.85)'; };
        _btn.onmouseout  = function(){ _btn.style.background = 'rgba(30,33,48,0.88)'; };
        _btn.onclick = function(){
            refreshScroll();
            _btn.style.background = 'rgba(50,200,120,0.85)';
            setTimeout(function(){ _btn.style.background = 'rgba(30,33,48,0.88)'; }, 600);
        };
        doc.body.appendChild(_btn);
    }
})();
</script>
""", height=0)


# ==========================================
# INITIAL CONFIGURATION TABS
# ==========================================
_tab_names = [
    "📂 Data Load & Filters",
    "📊 New Client per Volume & AHT",
    "🧠 New Client Prediction",
    "🔄 Recon",
    "📊 Actual Hours Distribution",
]
_tabs              = st.tabs(_tab_names)
tab1               = _tabs[0]
tab_vol_aht        = _tabs[1]
tab_predict        = _tabs[2]
tab_recon          = _tabs[3]
tab_actual_hours   = _tabs[4]
tab_params   = None   # removed — parameters now live inside tab1
tab_pipe     = None
tab3         = None
_data_loaded = "df_clean" in st.session_state

# ── Programmatic tab navigation via JS ───────────────────────────────────────
if '_nav_tab_idx' in st.session_state:
    _nav_idx = st.session_state.pop('_nav_tab_idx')
    import streamlit.components.v1 as _stc
    _stc.html(f"""<script>
    setTimeout(function(){{
        var tabs=window.parent.document.querySelectorAll('[data-baseweb="tab"]');
        if(tabs.length>{_nav_idx})tabs[{_nav_idx}].click();
    }},250);
    </script>""", height=0)

lista_clientes = ["All"]
lista_tareas   = ["All"]
lista_pods     = []
lista_pms      = []
# Filter state — always initialised so module-level code that references these
# (e.g. cascade lines that run in the same render) never hits a NameError.
selected_pods          = []
selected_srs           = []
selected_clients_final = []

with tab1:
    # ── Apply pending params config load BEFORE any widgets are instantiated ──
    if '_pending_params_load' in st.session_state:
        for _pk, _pv in st.session_state.pop('_pending_params_load').items():
            st.session_state[_pk] = _pv

    # ── Restore filter state from session state (survives reruns & pipeline flow) ──
    # These globals may not be (re)assigned if uploaded_file is falsy, so always
    # pull the last-known values from session state as a safe fallback.
    selected_pods          = st.session_state.get('_filt_pods',    [])
    selected_srs           = st.session_state.get('_filt_srs',     [])
    selected_clients_final = st.session_state.get('_filt_clients', [])

    # ── Global Parameters (collapsed by default) ──────────────────────────────
    with st.expander("⚙️ Global Parameters", expanded=False):
        gp_col1, gp_col2 = st.columns(2)

        with gp_col1:
            st.subheader("💰 Monthly Cost per Role ($)")
            cost_acc1 = st.number_input("Accountant I",       value=1140.0, key="gp_cost_acc1")
            cost_acc2 = st.number_input("Accountant II",      value=1369.0, key="gp_cost_acc2")
            cost_gen  = st.number_input("General Accountant", value=1900.0, key="gp_cost_gen")
            cost_sr   = st.number_input("Sr. Accountant",     value=2536.0, key="gp_cost_sr")
            cost_map  = {
                'Accountant I':       cost_acc1,
                'Accountant II':      cost_acc2,
                'General Accountant': cost_gen,
                'Sr. Accountant':     cost_sr,
            }

            st.subheader("📉 Shrinkage")
            absenteeism = st.number_input("Absenteeism (%) – Max 10%", min_value=0.0, max_value=10.0, value=10.0, key="gp_abs") / 100
            attrition   = st.number_input("Attrition (%)",             min_value=0.0, value=3.0,  key="gp_att") / 100

        with gp_col2:
            st.subheader("🎯 Utilization by Role (%)")
            util_acc1 = st.number_input("Accountant I & II",  value=85.0, key="gp_util_acc1") / 100
            util_gen  = st.number_input("General Accountant", value=80.0, key="gp_util_gen")  / 100
            util_sr   = st.number_input("Sr. Accountant",     value=50.0, key="gp_util_sr")   / 100
            utilization_map = {
                'Accountant I':       util_acc1,
                'Accountant II':      util_acc1,
                'General Accountant': util_gen,
                'Sr. Accountant':     util_sr,
            }

            st.subheader("📅 Calendar & Working Days")
            calc_mode = st.radio(
                "Calculation method:",
                ("Actual network days of the month", "Fixed days per month"),
                index=1,
                key="gp_calc_mode"
            )
            fixed_days = 22
            if calc_mode == "Fixed days per month":
                fixed_days = st.number_input("Fixed days per month", value=22, key="gp_fixed_days")

            with st.expander("🏖️ Configure Holidays per Month"):
                holidays_per_month = {}
                for i, mes in enumerate(meses_proyeccion):
                    holidays_per_month[mes] = st.number_input(
                        f"Holidays in {mes}", value=0, min_value=0, key=f"hol_{i}"
                    )

        # ── Config Save / Load ────────────────────────────────────────────────
        st.divider()
        _pc_col1, _pc_col2 = st.columns([2, 1])

        # ── Load saved config ──────────────────────────────────────────────
        with _pc_col1:
            _saved_pcs = _list_params_configs()
            if _saved_pcs:
                _pc_options = {
                    f"{c['name']}  ({c['saved_at'][:10]})": c
                    for c in _saved_pcs
                }
                _sel_pc_label = st.selectbox(
                    "📂 Load saved config:",
                    options=list(_pc_options.keys()),
                    key="pc_select"
                )
                _lc1, _lc2 = st.columns(2)
                with _lc1:
                    if st.button("⬆️ Load Config", key="pc_load_btn", use_container_width=True):
                        _pc_vals = _pc_options[_sel_pc_label]['params']
                        # Stage values in a pending key — applied BEFORE widgets instantiate on next rerun
                        _pending = {
                            'gp_cost_acc1':  _pc_vals.get('cost_acc1', 1140.0),
                            'gp_cost_acc2':  _pc_vals.get('cost_acc2', 1369.0),
                            'gp_cost_gen':   _pc_vals.get('cost_gen',  1900.0),
                            'gp_cost_sr':    _pc_vals.get('cost_sr',   2536.0),
                            'gp_abs':        _pc_vals.get('absenteeism_pct', 10.0),
                            'gp_att':        _pc_vals.get('attrition_pct',   3.0),
                            'gp_util_acc1':  _pc_vals.get('util_acc1_pct', 85.0),
                            'gp_util_gen':   _pc_vals.get('util_gen_pct',  80.0),
                            'gp_util_sr':    _pc_vals.get('util_sr_pct',   50.0),
                            'gp_calc_mode':  _pc_vals.get('calc_mode', "Actual network days of the month"),
                            'gp_fixed_days': _pc_vals.get('fixed_days', 22),
                            **{f'hol_{i}': _pc_vals.get(f'hol_{i}', 0) for i in range(len(meses_proyeccion))},
                        }
                        st.session_state['_pending_params_load'] = _pending
                        st.rerun()
                with _lc2:
                    if st.button("🗑️ Delete Config", key="pc_del_btn", use_container_width=True):
                        os.remove(_pc_options[_sel_pc_label]['path'])
                        st.success("Deleted.")
                        st.rerun()
            else:
                st.caption("No saved configs yet — save one below.")

        # ── Save current config ────────────────────────────────────────────
        with _pc_col2:
            _pc_name = st.text_input("Config name:", placeholder="e.g. Conservative", key="pc_save_name")
            if st.button("💾 Save Current Config", type="primary", key="pc_save_btn", use_container_width=True):
                if _pc_name.strip():
                    _pc_payload = {
                        'cost_acc1':        cost_acc1,
                        'cost_acc2':        cost_acc2,
                        'cost_gen':         cost_gen,
                        'cost_sr':          cost_sr,
                        'absenteeism_pct':  round(absenteeism * 100, 4),
                        'attrition_pct':    round(attrition   * 100, 4),
                        'util_acc1_pct':    round(util_acc1   * 100, 4),
                        'util_gen_pct':     round(util_gen    * 100, 4),
                        'util_sr_pct':      round(util_sr     * 100, 4),
                        'calc_mode':        calc_mode,
                        'fixed_days':       fixed_days,
                        **{f'hol_{i}': holidays_per_month.get(mes, 0)
                           for i, mes in enumerate(meses_proyeccion)},
                    }
                    _saved_path = _save_params_config(_pc_name.strip(), _pc_payload)
                    st.success(f"✅ Saved **{_pc_name.strip()}**")
                    st.session_state.pop('pc_save_name', None)
                    st.rerun()
                else:
                    st.warning("Please enter a config name before saving.")

    with st.expander("📂 Master Database & Filters", expanded=(not st.session_state.get('_s0_collapsed', False) and "calc_data" not in st.session_state)):

        # ── Base month / year selector ─────────────────────────────────────────
        _month_names = ["January","February","March","April","May","June",
                        "July","August","September","October","November","December"]
        _cy = _actual_today.year
        _base_year_opts = sorted(set(
            [_cy - 1, _cy, _cy + 1] +
            st.session_state.get('base_extra_years', [])
        ))

        _bm_col, _by_col, _badd_col = st.columns([2, 2, 1])
        with _bm_col:
            _bm_idx = _sel_base_month - 1          # 0-based index
            _new_bm = st.selectbox(
                "📅 Base month",
                options=list(range(1, 13)),
                format_func=lambda m: _month_names[m - 1],
                index=_bm_idx,
                key="base_month_sel",
            )
        with _by_col:
            _by_idx = _base_year_opts.index(_sel_base_year) if _sel_base_year in _base_year_opts else 0
            _new_by = st.selectbox(
                "📅 Base year",
                options=_base_year_opts,
                index=_by_idx,
                key="base_year_sel",
            )
        with _badd_col:
            _extra_yr = st.number_input("➕ Add year", min_value=2000, max_value=2100,
                                        value=_cy, step=1, key="base_extra_yr_input",
                                        label_visibility="visible")
            if st.button("Add", key="base_add_yr_btn", use_container_width=True):
                _extras = st.session_state.get('base_extra_years', [])
                if _extra_yr not in _extras:
                    st.session_state['base_extra_years'] = _extras + [_extra_yr]
                    st.rerun()

        # Compute and display network days for the selected base month
        _bm_start = datetime(_new_by, _new_bm, 1)
        _bm_end   = _bm_start + relativedelta(months=1) - relativedelta(days=1)
        _bm_nd    = int(np.busday_count(_bm_start.strftime('%Y-%m-%d'),
                                        (_bm_end + relativedelta(days=1)).strftime('%Y-%m-%d')))
        st.info(f"📆 **{_month_names[_new_bm-1]} {_new_by}** — **{_bm_nd} network days** (Mon–Fri, no holidays)")

        st.divider()

        # Allow using data from the pipeline (no file upload needed)
        if "pipeline_vol_merged" in st.session_state:
            _pipe_active = st.session_state.pipeline_vol_merged
            st.success(
                f"✅ Pipeline data active — {len(_pipe_active)} rows, "
                f"{_pipe_active['client_name'].nunique()} clients. "
                "You can also upload a different Excel file below to override."
            )

        _tab1_left, _tab1_right = st.columns(2)
        with _tab1_left:
            uploaded_file = st.file_uploader(
                "📂 Capacity data (Excel — vol_merged or vol sheet)",
                type=["xlsx", "xls"],
                key="main_data_upload"
            )
        with _tab1_right:
            st.markdown("**👥 HC Weekly Report**")
            _hc_upload = st.file_uploader(
                "Upload HC Weekly Report to show Actual HC in the Capacity Overview",
                type=["xlsx", "xls"],
                key="hc_file_upload"
            )
            if _hc_upload:
                try:
                    st.session_state.hc_data = _process_hc_report(_hc_upload.read())
                    st.session_state['_hc_version'] = st.session_state.get('_hc_version', 0) + 1
                    _hc_loaded = st.session_state.hc_data
                    _n_mgrs = _hc_loaded.get('acct_managers', 0) + _hc_loaded.get('asst_managers', 0)
                    st.success(
                        f"HC Loaded: **{_n_mgrs} Managers** (Accounting Managers & Assistant Managers)"
                        f" + **{_hc_loaded['total']}** Active accounting staff  \n"
                        f"AccI: {_hc_loaded['by_role'].get('Accountant I',0)} · "
                        f"AccII: {_hc_loaded['by_role'].get('Accountant II',0)} · "
                        f"GenAcc: {_hc_loaded['by_role'].get('General Accountant',0)} · "
                        f"Sr.: {_hc_loaded['by_role'].get('Sr. Accountant',0)} · "
                        f"Asm: {_hc_loaded.get('asst_managers',0)} · "
                        f"AM: {_hc_loaded.get('acct_managers',0)}"
                    )
                    # ── Temporary Sr. / Direct-Reports checker ────────────────────
                    with st.expander("🔍 [DEBUG] Sr. Accountant HC Checker", expanded=False):
                        _by_sr_dbg  = _hc_loaded.get('by_sr', {})
                        _sr_count   = len(_by_sr_dbg)
                        _dr_grand   = sum(v.get('dr_total', v.get('total', 1) - 1) for v in _by_sr_dbg.values())
                        _tot_grand  = sum(v.get('total', 0) for v in _by_sr_dbg.values())
                        st.markdown(
                            f"**Sr. Accountants found:** {_sr_count}  \n"
                            f"**Total direct reports (excl. Sr.):** {_dr_grand}  \n"
                            f"**Grand total (DRs + Srs):** {_tot_grand}"
                        )
                        if _by_sr_dbg:
                            _dbg_rows = [
                                {
                                    'Sr. Accountant': _sn,
                                    'Email': v.get('email', '—'),
                                    'DRs only': v.get('dr_total', v.get('total', 1) - 1),
                                    'Total (incl. Sr.)': v.get('total', 0),
                                    'AccI': v['by_role'].get('Accountant I', 0),
                                    'AccII': v['by_role'].get('Accountant II', 0),
                                    'GenAcc': v['by_role'].get('General Accountant', 0),
                                    'Sr. (self)': v['by_role'].get('Sr. Accountant', 0),
                                }
                                for _sn, v in sorted(_by_sr_dbg.items())
                            ]
                            st.dataframe(pd.DataFrame(_dbg_rows), use_container_width=True, hide_index=True)
                        else:
                            st.warning("by_sr is empty — no Sr. Accountants detected. Check if 'Work Email' / 'Manager email' columns exist in the HC file.")
                except Exception as _hc_err:
                    st.error(f"HC error: {_hc_err}")

        if uploaded_file:
            try:
                df_temp = pd.read_excel(uploaded_file, sheet_name="vol_merged")
            except:
                df_temp = pd.read_excel(uploaded_file)

            df_temp.columns = df_temp.columns.str.strip()
            st.success("File loaded into memory successfully!")
            st.info(
                "👇 **Select your filters below** — narrow down by POD, Sr. Accountant, or specific clients.  \n"
                "Leave everything empty to process the **full database**.  \n"
                "⚡ When done, hit **▶ Next → Step 1** to run the calculations."
            )

            all_clients = sorted(df_temp['client_name'].dropna().astype(str).unique().tolist()) if 'client_name' in df_temp.columns else []
            all_pods    = sorted(df_temp['POD'].dropna().astype(str).unique().tolist())          if 'POD'         in df_temp.columns else []
            all_pms     = sorted(
                v for v in df_temp['PMS'].dropna().astype(str).str.strip().unique().tolist()
                if v and v.lower() not in ('nan', 'none', '')
            ) if 'PMS' in df_temp.columns else []
            lista_pods  = all_pods
            lista_pms   = all_pms
            st.session_state['_lista_pms']  = all_pms   # persist for fragment access
            st.session_state['_lista_pods'] = all_pods  # persist for fragment access
            all_srs     = sorted(df_temp['Sr. Accountant'].dropna().astype(str).unique().tolist()) if 'Sr. Accountant' in df_temp.columns else []

            lista_clientes = ["All"] + all_clients

            if 'type' in df_temp.columns and 'subtype' in df_temp.columns:
                df_temp['task_name'] = df_temp['type'].astype(str) + " - " + df_temp['subtype'].astype(str)
                lista_tareas = ["All"] + sorted(df_temp['task_name'].dropna().unique().tolist())

            col1, col2, col3 = st.columns(3)
            with col1:
                selected_pods = st.multiselect(
                    "1. Filter by POD", options=all_pods,
                    default=st.session_state.get('_filt_pods', []),
                    key="_s0_pods_ms",
                )
            with col2:
                selected_srs = st.multiselect(
                    "2. Filter by Sr. Accountant", options=all_srs,
                    default=st.session_state.get('_filt_srs', []),
                    key="_s0_srs_ms",
                )

            # Persist selections immediately so they survive any rerun
            st.session_state['_filt_pods'] = selected_pods
            st.session_state['_filt_srs']  = selected_srs

            default_clients = []
            if selected_pods or selected_srs:
                mask = pd.Series(True, index=df_temp.index)
                if selected_pods: mask &= df_temp['POD'].isin(selected_pods)
                if selected_srs:  mask &= df_temp['Sr. Accountant'].isin(selected_srs)
                default_clients = sorted(df_temp[mask]['client_name'].dropna().astype(str).unique().tolist())

            with col3:
                selected_clients_final = st.multiselect(
                    "3. Clients to Process", options=all_clients,
                    default=default_clients,
                    key="_s0_clients_ms",
                )
                if not selected_clients_final:
                    st.info("ℹ️ If left empty, the entire database will be processed.")

            # Persist client selection as well
            st.session_state['_filt_clients'] = selected_clients_final

            # ── Confirmation of active filter — shown BEFORE onboarding list ─────
            st.divider()
            _cf_parts = []
            if selected_pods:
                _cf_parts.append(f"POD: **{', '.join(selected_pods)}**")
            if selected_srs:
                _cf_parts.append(f"Sr.: **{', '.join(selected_srs)}**")
            if selected_clients_final and not (selected_pods or selected_srs):
                _n_cl = len(selected_clients_final)
                _cf_parts.append(
                    f"Clients: **{selected_clients_final[0]}**"
                    if _n_cl == 1
                    else f"**{_n_cl} clients** selected"
                )
            if _cf_parts:
                st.success("📊 Will process: " + " · ".join(_cf_parts))
            else:
                st.info("📊 Will process: **Overall capacity** (no filters — full database)")

            # ── 🆕 New / Onboarding Clients Alert ────────────────────────────
            # Shown AFTER the filter confirmation so the user knows which scope
            # is active before deciding to add / replace these clients.
            _today_ts   = pd.Timestamp.today().normalize()
            _ci_m       = {c.strip().lower(): c for c in df_temp.columns}
            _st_col_r   = _ci_m.get('status')
            _gl_col_r   = _ci_m.get('go live')
            _pod_col_r  = _ci_m.get('pod')
            _pms_col_r  = _ci_m.get('pms')
            _mrr_col_r  = _ci_m.get('mrr')

            _ob_grp_agg = {}
            for _oc in [_st_col_r, _gl_col_r, _pod_col_r, _pms_col_r, _mrr_col_r]:
                if _oc and _oc not in _ob_grp_agg:
                    _ob_grp_agg[_oc] = 'first'
            for _hc in ['Capacity Processing Hours', 'Capacity reviewing hours']:
                if _hc in df_temp.columns:
                    _ob_grp_agg[_hc] = 'sum'

            _ob_grp = (
                df_temp.groupby('client_name', as_index=False).agg(_ob_grp_agg)
                if _ob_grp_agg else
                df_temp[['client_name']].drop_duplicates()
            )

            _ob_mask = pd.Series(False, index=_ob_grp.index)
            # Match: lifecycle/status = Onboarding (or similar)
            if _st_col_r:
                _ob_mask |= _ob_grp[_st_col_r].astype(str).str.lower().str.strip().isin(
                    ['onboarding', 'new client', 'subscriber']
                )
            # Match: Go Live within 20 days of today (includes recently started + upcoming)
            if _gl_col_r:
                _gl_ser  = pd.to_datetime(_ob_grp[_gl_col_r], errors='coerce')
                _gl_diff = (_gl_ser - _today_ts).dt.days   # positive = future, negative = past
                # Within 20 days of today in either direction (recently started or starting soon)
                _ob_mask |= _gl_ser.notna() & (_gl_diff >= -20) & (_gl_diff <= 20)

            # ── Apply active filter so only relevant clients appear ────────────
            # If a POD/Sr./Client filter is selected, restrict the onboarding list
            # to clients that fall within that scope.
            _ob_scope_mask = pd.Series(True, index=_ob_grp.index)
            if selected_clients_final and 'client_name' in _ob_grp.columns:
                _ob_scope_mask &= _ob_grp['client_name'].isin(selected_clients_final)
            else:
                if selected_pods and _pod_col_r:
                    _ob_scope_mask &= _ob_grp[_pod_col_r].astype(str).str.strip().isin(selected_pods)
                if selected_srs:
                    _sr_col_r = _ci_m.get('sr. accountant') or _ci_m.get('sr accountant') or _ci_m.get('senior accountant')
                    if _sr_col_r:
                        _ob_scope_mask &= _ob_grp[_sr_col_r].astype(str).str.strip().isin(selected_srs)

            _ob_sel = _ob_grp[_ob_mask & _ob_scope_mask].copy()

            if not _ob_sel.empty:
                _ob_rows = []
                for _, _r in _ob_sel.iterrows():
                    _cn   = _r['client_name']
                    _proc = float(_r['Capacity Processing Hours']) if 'Capacity Processing Hours' in _r.index and pd.notna(_r['Capacity Processing Hours']) else 0.0
                    _rev  = float(_r['Capacity reviewing hours'])  if 'Capacity reviewing hours'  in _r.index and pd.notna(_r['Capacity reviewing hours'])  else 0.0
                    _tot  = _proc + _rev
                    _st_v = str(_r[_st_col_r] or '') if _st_col_r and pd.notna(_r.get(_st_col_r)) else ''
                    _gl_v = _r[_gl_col_r] if _gl_col_r else pd.NaT
                    _gl_s = pd.to_datetime(_gl_v, errors='coerce')
                    _gl_str = _gl_s.strftime('%Y-%m-%d') if pd.notna(_gl_s) else '—'
                    _pod_v = str(_r[_pod_col_r] or '') if _pod_col_r and pd.notna(_r.get(_pod_col_r)) else ''
                    _pms_v = str(_r[_pms_col_r] or '') if _pms_col_r and pd.notna(_r.get(_pms_col_r)) else ''
                    _mrr_v = float(_r[_mrr_col_r] or 0) if _mrr_col_r and pd.notna(_r.get(_mrr_col_r)) else 0.0
                    _ob_rows.append({
                        'Client':    _cn,
                        'Status':    _st_v,
                        'Go Live':   _gl_str,
                        'POD':       _pod_v,
                        'PMS':       _pms_v,
                        'MRR ($)':   round(_mrr_v, 0),
                        'Proc Hrs':  round(_proc, 2),
                        'Rev Hrs':   round(_rev, 2),
                        'Total Hrs': round(_tot, 2),
                        '_gl_raw':   _gl_v,
                        '_pod':      _pod_v,
                        '_pms':      _pms_v,
                        '_mrr':      _mrr_v,
                    })

                _ob_disp_df  = pd.DataFrame(_ob_rows)
                _has_data_m  = _ob_disp_df['Total Hrs'] > 0
                _n_ob_total  = len(_ob_disp_df)
                _n_ob_w_data = int(_has_data_m.sum())

                st.markdown(f"#### 🆕 New / Onboarding Clients — {_n_ob_total} found")
                with st.container():
                    _disp_cols = ['Client', 'Status', 'Go Live', 'POD', 'PMS', 'MRR ($)',
                                  'Proc Hrs', 'Rev Hrs', 'Total Hrs']
                    st.dataframe(_ob_disp_df[_disp_cols], use_container_width=True, hide_index=True)

                    if _n_ob_w_data > 0:
                        st.warning(
                            f"⚠️ **{_n_ob_w_data} of these client(s)** already have hours in the input file. "
                            "Click below to queue them for AI prediction — their existing hours will be replaced."
                        )

                    if st.button(
                        f"🤖 Queue All {_n_ob_total} Client(s) for AI Prediction",
                        key="btn_ob_queue_all", type="primary"
                    ):
                        st.session_state['_ob_replace_set'] = set(_ob_disp_df[_has_data_m]['Client'].tolist())
                        _existing_ai2 = st.session_state.get('ai_manual_clients', pd.DataFrame())
                        _new_ai_rows2 = []
                        for _, _ob_r in _ob_disp_df.iterrows():
                            _gl_ai_s = pd.to_datetime(_ob_r['_gl_raw'], errors='coerce')
                            _new_ai_rows2.append({
                                'Company Name':    _ob_r['Client'],
                                'POD':             _ob_r['_pod'],
                                'Go Live Date':    _gl_ai_s.strftime('%Y-%m-%d') if pd.notna(_gl_ai_s) else '',
                                'MRR ($)':         float(_ob_r['_mrr'] or 0),
                                'PMS':             _ob_r['_pms'],
                                'Res Doors':       0,
                                'Res Properties':  0,
                                'Comm Doors':      0,
                                'Comm Properties': 0,
                                'SQFT Commercial': 0,
                                'Corp Books':      '',
                            })
                        _new_ai_df2 = pd.DataFrame(_new_ai_rows2)
                        _new_ai_df2 = _enrich_ai_from_hs(_new_ai_df2, st.session_state.get('hs_parsed'))
                        if not _existing_ai2.empty and 'Company Name' in _existing_ai2.columns:
                            _ex2_keys = set(_existing_ai2['Company Name'].astype(str).str.lower().str.strip())
                            _truly_new2 = [r for _, r in _new_ai_df2.iterrows()
                                           if str(r['Company Name']).lower().strip() not in _ex2_keys]
                            if _truly_new2:
                                st.session_state.ai_manual_clients = pd.concat(
                                    [_existing_ai2, pd.DataFrame(_truly_new2)], ignore_index=True)
                            _q_count = len(_truly_new2)
                        else:
                            st.session_state.ai_manual_clients = _new_ai_df2
                            _q_count = len(_new_ai_df2)
                        st.success(
                            f"✅ {_q_count} client(s) queued for AI Prediction.  \n"
                            "👇 Run AI Prediction below, then click **Add to Baseline** to apply."
                        )
                        st.rerun()

                    # Clients WITHOUT hours — informational note
                    _no_data_list = _ob_disp_df[~_has_data_m]['Client'].tolist()
                    if _no_data_list:
                        st.info(
                            f"💡 **{len(_no_data_list)} client(s) have no hours yet** — AI prediction recommended:  \n"
                            + ", ".join(f"**{c}**" for c in _no_data_list[:8])
                            + (f", and {len(_no_data_list) - 8} more" if len(_no_data_list) > 8 else "")
                        )

            # ── Next → Step 1 ────────────────────────────────────────────────
            st.divider()
            if st.button("▶ Next → Step 1", type="primary", use_container_width=True, key="btn_next_step1"):
                st.session_state['_auto_run_baseline'] = True
                st.session_state['_show_step1'] = True
                st.session_state['_s0_collapsed'] = True
                st.session_state['_s1_collapsed'] = True
                st.rerun()


with tab1:
    # ==========================================
    # STEP 1: BASE MATH ENGINE
    # ==========================================
    if st.session_state.get('_show_step1') or 'calc_data' in st.session_state:
      with st.expander("🚀 Step 1: Calculate Baseline", expanded=(not st.session_state.get('_s1_collapsed', False) and "calc_data" not in st.session_state)):
        st.markdown(
            "Calculates base operational hours and shrinkage without any adjustments.  \n"
            "**Learning curve** (+17% / −14% / −1% AHT) applies **only to new clients** "
            "— specifically clients whose Go Live falls within the 3-month window of each projected month. "
            "Existing clients (Go Live ≥ 3 months ago) are unaffected (multiplier = 1.0×)."
        )

        _auto_baseline = st.session_state.pop('_auto_run_baseline', False)
        if st.button("⚙️ Generate Baseline", type="primary", use_container_width=True) or _auto_baseline:
            _has_pipeline = "pipeline_vol_merged" in st.session_state
            if not uploaded_file and not _has_pipeline:
                st.error("⚠️ Please upload an Excel file or run the Data Pipeline first.")
            else:
                with st.spinner("Calculating operational baseline (without automations)..."):
                    try:
                        if uploaded_file:
                            try:
                                df = pd.read_excel(uploaded_file, sheet_name="vol_merged")
                            except:
                                df = pd.read_excel(uploaded_file)
                            df.columns = df.columns.str.strip()
                        else:
                            # Use pipeline / scenario data
                            df = st.session_state.pipeline_vol_merged.copy()
                            df.columns = df.columns.str.strip()

                        # ── Apply active filters to df ────────────────────────────────
                        # Priority: client list > Sr. Accountant > POD.
                        # All three are applied so that selecting only a POD (without
                        # explicit clients) still restricts the baseline to that POD.
                        _s1_filt_clients = st.session_state.get('_filt_clients', selected_clients_final)
                        _s1_filt_srs     = st.session_state.get('_filt_srs',     selected_srs)
                        _s1_filt_pods    = st.session_state.get('_filt_pods',     selected_pods)

                        _s1_mask = pd.Series(True, index=df.index)
                        if _s1_filt_clients and 'client_name' in df.columns:
                            _s1_mask &= df['client_name'].isin(_s1_filt_clients)
                        else:
                            # No explicit client list — apply POD / Sr. directly
                            if _s1_filt_pods and 'POD' in df.columns:
                                _s1_mask &= df['POD'].isin(_s1_filt_pods)
                            if _s1_filt_srs and 'Sr. Accountant' in df.columns:
                                _s1_mask &= df['Sr. Accountant'].isin(_s1_filt_srs)

                        if not _s1_mask.all():
                            df = df[_s1_mask].copy()

                        # DATE CLEANING
                        def clean_date_column(col_name):
                            if col_name not in df.columns:
                                return pd.Series(pd.NaT, index=df.index)
                            s = df[col_name].astype(str).str.strip().str.lower()
                            garbage = ['0', '0.0', 'nan', 'nat', 'none', 'null', '', ' ', '00:00:00', '00:00:00.000000']
                            s = s.replace(garbage, np.nan)
                            s = pd.to_datetime(s, errors='coerce').dt.normalize()
                            s = s.where(s.dt.year >= 2000, pd.NaT)
                            return s

                        df['Go Live']            = clean_date_column('Go Live')
                        df['Final Service Date'] = clean_date_column('Final Service Date')

                        if 'MRR' in df.columns:
                            df['MRR'] = df['MRR'].astype(str).str.replace(r'[$, ]', '', regex=True)
                            df['MRR'] = pd.to_numeric(df['MRR'], errors='coerce').fillna(0.0)
                        else:
                            df['MRR'] = 0.0

                        cols_to_clean = [
                            'Capacity Processing Hours', 'Capacity reviewing hours',
                            'Closed tickets with Proc time', 'Closed tickets with rev time',
                            '>>> FINAL Capacity Proc AHT', '>>> FINAL Capacity Rev AHT'
                        ]
                        for col in cols_to_clean:
                            if col in df.columns:
                                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                            else:
                                df[col] = 0

                        # MRR BASE VALIDATION
                        df_clients_unique = df.groupby('client_name', as_index=False).agg({
                            'MRR':                'max',
                            'Go Live':            lambda x: x.dropna().iloc[0] if not x.dropna().empty else pd.NaT,
                            'Final Service Date': lambda x: x.dropna().iloc[0] if not x.dropna().empty else pd.NaT,
                        })

                        st.session_state.df_clean         = df.copy()
                        st.session_state.df_clients_unique = df_clients_unique.copy()
                        _build_client_master_map()   # build POD / Sr. / Vol maps from Step 1 data

                        dict_hrs_per_fte   = {}
                        dict_workable_days = {}

                        # ── Pre-compute month params + vectorised apct/lc (Step 1) ───
                        _b_gl  = df['Go Live'].values
                        _b_fsd = df['Final Service Date'].values
                        _b_ptix = pd.to_numeric(df['Closed tickets with Proc time'], errors='coerce').fillna(0).values
                        _b_rtix = pd.to_numeric(df['Closed tickets with rev time'],  errors='coerce').fillna(0).values
                        _b_paht = pd.to_numeric(df['>>> FINAL Capacity Proc AHT'],   errors='coerce').fillna(0).values
                        _b_raht = pd.to_numeric(df['>>> FINAL Capacity Rev AHT'],    errors='coerce').fillna(0).values

                        # ── IDEAL + REAL role arrays — vectorised (no iterrows) ──────────
                        _INVALID_ROLES = {'nan', 'None', ''}
                        _ip_pref = (df['Ideal Proc'].astype(str).str.strip()
                                    if 'Ideal Proc' in df.columns
                                    else pd.Series([''] * len(df), index=df.index))
                        _ip_fb   = (df['Proc Role'].astype(str).str.strip()
                                    if 'Proc Role'  in df.columns
                                    else pd.Series([''] * len(df), index=df.index))
                        _ir_pref = (df['Ideal Rev'].astype(str).str.strip()
                                    if 'Ideal Rev'  in df.columns
                                    else pd.Series([''] * len(df), index=df.index))
                        _ir_fb   = (df['Rev Role'].astype(str).str.strip()
                                    if 'Rev Role'   in df.columns
                                    else pd.Series([''] * len(df), index=df.index))

                        # Ideal: prefer Ideal Proc/Rev, fall back to Proc/Rev Role
                        _b_ip = np.where(
                            _ip_pref.isin(_INVALID_ROLES),
                            np.where(_ip_fb.isin(_INVALID_ROLES), 'Accountant I',   _ip_fb),
                            _ip_pref
                        )
                        _b_ir = np.where(
                            _ir_pref.isin(_INVALID_ROLES),
                            np.where(_ir_fb.isin(_INVALID_ROLES), 'Sr. Accountant', _ir_fb),
                            _ir_pref
                        )
                        _b_up = np.array([utilization_map.get(ip, util_acc1) for ip in _b_ip])
                        _b_ur = np.array([utilization_map.get(ir, util_sr)   for ir in _b_ir])

                        # Real: use Proc/Rev Role directly (no ideal fallback)
                        _b_ip_real = np.where(_ip_fb.isin(_INVALID_ROLES), 'Accountant I',   _ip_fb)
                        _b_ir_real = np.where(_ir_fb.isin(_INVALID_ROLES), 'Sr. Accountant', _ir_fb)
                        _b_up_real = np.array([utilization_map.get(ip, util_acc1) for ip in _b_ip_real])
                        _b_ur_real = np.array([utilization_map.get(ir, util_sr)   for ir in _b_ir_real])
                        _b_month_frames_real = []
                        _b_pod_s = df.get('POD', pd.Series(['No POD']*len(df))).fillna('No POD').astype(str).str.strip()
                        _b_pod   = _b_pod_s.where(~_b_pod_s.str.lower().isin({'nan','none',''}), 'No POD').values
                        _b_cli = df.get('client_name', pd.Series(['Unknown']*len(df))).fillna('Unknown').astype(str).values

                        _b_month_frames = []

                        # Base month = previous calendar month (e.g. Feb when today is Mar).
                        # Its network-day count is the reference for demand fluctuation.
                        _bm_start     = pd.Timestamp((today.replace(day=1) - relativedelta(months=1)).date())
                        _bm_end       = pd.Timestamp((_bm_start + relativedelta(months=1) - relativedelta(days=1)).date())
                        base_net_days = max(1, int(np.busday_count(
                            _bm_start.strftime('%Y-%m-%d'),
                            (_bm_end + pd.Timedelta(days=1)).strftime('%Y-%m-%d')
                        )))

                        for i, mes_str in enumerate(meses_proyeccion):
                            mes_date = today + relativedelta(months=_month_offsets[i])
                            start_m  = pd.Timestamp(mes_date.replace(day=1).date())
                            end_m    = pd.Timestamp((start_m + relativedelta(months=1) - relativedelta(days=1)).date())
                            # Actual network days (Mon-Fri) for this month — always from
                            # busday_count regardless of the fixed-days setting.
                            # Used on the demand side: customer hours reflect real workload.
                            actual_net_days = int(np.busday_count(start_m.strftime('%Y-%m-%d'),
                                                                  (end_m + pd.Timedelta(days=1)).strftime('%Y-%m-%d')))
                            # Capacity days: user-supplied fixed value OR actual network days.
                            # This drives FTE availability — fewer days = more FTEs needed.
                            net_days = fixed_days if calc_mode == "Fixed days per month" else actual_net_days
                            holidays      = holidays_per_month[mes_str]
                            workable_days = max(1, net_days - holidays)
                            dict_hrs_per_fte[i]   = workable_days * 7.5
                            dict_workable_days[i] = workable_days

                            # Vectorised active_pct (pd fillna/clip avoids NPY_ITER_REFS_OK with object arrays)
                            _gl_s  = pd.to_datetime(pd.Series(_b_gl),  errors='coerce').fillna(start_m).clip(lower=start_m, upper=end_m)
                            _fsd_s = pd.to_datetime(pd.Series(_b_fsd), errors='coerce').fillna(end_m).clip(lower=start_m, upper=end_m)
                            _dias  = np.busday_count(
                                _gl_s.values.astype('datetime64[D]'),
                                (_fsd_s + pd.Timedelta(days=1)).values.astype('datetime64[D]')
                            )
                            # active_pct: always relative to real network days of this month
                            _ap = np.clip(np.maximum(_dias.astype(float), 0) / actual_net_days, 0.0, 1.0) if actual_net_days > 0 else np.zeros(len(df))

                            # Day-scale: how this month's network days compare to the base month.
                            #   Fixed days  → scale = 1.0  (requirement stays constant every month)
                            #   Network days → scale = current / base  (hours flex with day count)
                            day_scale = 1.0 if calc_mode == "Fixed days per month" \
                                        else (actual_net_days / base_net_days)

                            # Vectorised learning_mult
                            _hgl   = pd.notna(pd.Series(_b_gl)).values
                            _gl_ts = pd.DatetimeIndex(pd.to_datetime(pd.Series(_b_gl), errors='coerce').fillna(start_m))
                            _md    = np.where(_hgl, (start_m.year-_gl_ts.year.values)*12+(start_m.month-_gl_ts.month.values), 999).astype(int)
                            _lc   = np.select(
                                [~_hgl|(_ap==0), (_md==0)&_hgl&(_ap>0), (_md==1)&_hgl&(_ap>0), (_md==2)&_hgl&(_ap>0)],
                                [1.0, 1.17, 0.86, 0.99], default=1.0
                            )

                            _bp = (_b_ptix * _ap * _b_paht * _lc * day_scale) / 60
                            _br = (_b_rtix * _ap * _b_raht * _lc * day_scale) / 60
                            _tp = _bp * (2 - _b_up + absenteeism + attrition)
                            _tr = _br * (2 - _b_ur + absenteeism + attrition)

                            col_name = f"M{i+1} ({mes_str}) - Base Hours"
                            col_prod = f"M{i+1} ({mes_str}) - Productive Hours"
                            _b_month_frames.append(pd.DataFrame({
                                'POD':           np.concatenate([_b_pod, _b_pod]),
                                'Client':        np.concatenate([_b_cli, _b_cli]),
                                'Required Role': np.concatenate([_b_ip,  _b_ir]),
                                col_name:        np.concatenate([_tp,    _tr]),
                                col_prod:        np.concatenate([_bp,    _br])
                            }))
                            # Real-role base: same productive hours, but shrinkage uses real role utilisation
                            _tp_real = _bp * (2 - _b_up_real + absenteeism + attrition)
                            _tr_real = _br * (2 - _b_ur_real + absenteeism + attrition)
                            _b_month_frames_real.append(pd.DataFrame({
                                'POD':           np.concatenate([_b_pod,    _b_pod]),
                                'Client':        np.concatenate([_b_cli,    _b_cli]),
                                'Required Role': np.concatenate([_b_ip_real, _b_ir_real]),
                                col_name:        np.concatenate([_tp_real,  _tr_real]),
                                col_prod:        np.concatenate([_bp,       _br])
                            }))

                        df_sum_raw      = pd.concat(_b_month_frames, ignore_index=True)
                        cols_meses      = [c for c in df_sum_raw.columns if " - Base Hours" in c or " - Productive Hours" in c]
                        df_resumen_base = df_sum_raw.groupby(['POD', 'Client', 'Required Role'])[cols_meses].sum().reset_index()

                        # Real-role base
                        df_sum_raw_real      = pd.concat(_b_month_frames_real, ignore_index=True)
                        cols_meses_real      = [c for c in df_sum_raw_real.columns if " - Base Hours" in c or " - Productive Hours" in c]
                        df_resumen_base_real = df_sum_raw_real.groupby(['POD', 'Client', 'Required Role'])[cols_meses_real].sum().reset_index()

                        st.session_state.calc_data = {
                            'df_resumen_base':       df_resumen_base,        # backward-compat alias (ideal)
                            'df_resumen_base_ideal': df_resumen_base,
                            'df_resumen_base_real':  df_resumen_base_real,
                            'dict_hrs_per_fte':      dict_hrs_per_fte,
                            'dict_workable_days':    dict_workable_days,
                            'clientes_validos':      sorted(df_resumen_base['Client'].unique().tolist())
                        }

                        if "final_dashboards" in st.session_state:
                            del st.session_state.final_dashboards

                        # Reset sync state that depends on the client list (recon + choices).
                        # NOTE: hs_parsed is intentionally kept — the raw HubSpot file is an
                        # independent data source and does NOT change when filters or the
                        # baseline are regenerated. Clearing it forced users to re-upload the
                        # HubSpot file on every baseline run, which broke "Sync from HubSpot".
                        st.session_state.hs_sync_choice = None
                        st.session_state.hs_recon_df = None
                        st.session_state.hs_client_overrides = {}
                        st.session_state.hs_onboarding_clients = None
                        st.session_state.s2_efficiency_choice = None

                        st.success("✅ Baseline calculated.")

                    except Exception as e:
                        st.error(f"❌ Error in baseline calculation: {e}")
                        import traceback
                        st.write(traceback.format_exc())

# ==========================================
# HUBSPOT SYNC (optional, after Step 1)
# ==========================================
# ── Re-enter the Data Load & Filters tab for all remaining sections
# (HubSpot sync, Step 2 Efficiency, Step 3 Dashboards, Step 4 Scenario).
# Without this, these blocks render at the page level and leak into the
# other sibling tabs (Vol & AHT, Prediction, Recon, Actual Hours).
# We use manual __enter__ to keep zero-indent code unchanged, and close
# the context right before the next sibling `with tab_predict:` block.
tab1.__enter__()

if "calc_data" in st.session_state:

    @st.fragment
    def _hs_sync_fragment():
        if st.session_state.hs_sync_choice == "skip":
            return   # collapsed — nothing rendered

        st.divider()
        st.subheader("🔄 HubSpot Sync")

        # ── Choice gate ──────────────────────────────────────────────────
        if st.session_state.hs_sync_choice is None:
            st.info(
                "**Would you like to sync with a HubSpot export?**\n\n"
                "Upload a HubSpot CRM export to review client statuses, flag terminations, "
                "reconcile data, and queue new onboarding clients for AI prediction."
            )
            _hsc1, _hsc2, _ = st.columns([1.5, 2, 4])
            with _hsc1:
                if st.button("🔄 Sync HubSpot", type="primary", use_container_width=True, key="hs_yes_btn"):
                    st.session_state.hs_sync_choice = "yes"
                    st.rerun(scope="fragment")
            with _hsc2:
                if st.button("⏭️ Skip HubSpot Sync", use_container_width=True, key="hs_skip_btn"):
                    st.session_state.hs_sync_choice = "skip"
                    st.rerun(scope="fragment")
            return

        # ── "yes" branch ─────────────────────────────────────────────────
        _hs_file = st.file_uploader(
            "📂 Upload HubSpot CRM export (.xlsx / .xls)",
            type=["xlsx", "xls"], key="hs_upload"
        )

        if _hs_file is not None and st.session_state.hs_parsed is None:
            try:
                _df_hs = _parse_hubspot_file(_hs_file)
                st.session_state.hs_parsed = _df_hs
                st.session_state.hs_recon_df = None   # force rebuild
                _build_client_master_map()   # re-sync POD/Sr./Vol maps with HubSpot data
                st.rerun(scope="fragment")
            except Exception as _e:
                st.error(f"Could not parse HubSpot file: {_e}")
                return

        if st.session_state.hs_parsed is None:
            st.info("Upload a HubSpot export above to continue.")
            if st.button("⏭️ Skip for now", key="hs_skip2_btn"):
                st.session_state.hs_sync_choice = "skip"
                st.rerun(scope="fragment")
            return

        _df_hs = st.session_state.hs_parsed

        # ── POD Column Mapping ───────────────────────────────────────────────
        # Auto-detect; let user confirm / override in case the column has an unusual name
        _hs_raw_cols = [c for c in _df_hs.columns if not c.startswith('_')]
        _pod_auto = next(
            (c for c in _hs_raw_cols
             if ('pod' in c.lower() or 'team' in c.lower() or 'squad' in c.lower())
             and not any(x in c.lower() for x in (
                 'upload', 'episode', 'period', 'template',
                 'update', 'report', 'export', 'import'))),
            None
        )
        _pod_override = st.session_state.get('hs_pod_col_override')
        _pod_col_options = ['— None (no POD column) —'] + _hs_raw_cols
        _pod_col_default = (
            _pod_override if _pod_override in _hs_raw_cols
            else (_pod_auto if _pod_auto in _hs_raw_cols else '— None (no POD column) —')
        )
        _pod_sel = st.selectbox(
            "🏷️ Which column in the HubSpot file contains the **POD / Team** assignment?",
            options=_pod_col_options,
            index=_pod_col_options.index(_pod_col_default),
            key="hs_pod_col_sel",
            help="Select the column that holds the POD or team name for each client. "
                 "This fills the POD field in AI Prediction and all sync sections."
        )
        if _pod_sel and _pod_sel != '— None (no POD column) —':
            if st.session_state.get('hs_pod_col_override') != _pod_sel:
                st.session_state['hs_pod_col_override'] = _pod_sel
                # Re-map _pod in the stored parsed frame
                _df_hs = _df_hs.copy()
                _df_hs['_pod'] = _df_hs[_pod_sel].astype(str).str.strip().replace({'nan': '', 'None': ''})
                st.session_state.hs_parsed = _df_hs
                # Re-enrich the AI prediction tables so POD values appear immediately
                for _ai_tbl_key in ['ai_manual_clients', 'hs_pre_ai_manual_clients']:
                    _tbl = st.session_state.get(_ai_tbl_key, pd.DataFrame())
                    if not _tbl.empty:
                        st.session_state[_ai_tbl_key] = _enrich_ai_from_hs(_tbl, _df_hs)
                # Clear all data_editor caches so tables re-initialise with new data
                for _pfx_k in list(st.session_state.keys()):
                    if _pfx_k.endswith('_ai_manual_editor'):
                        st.session_state.pop(_pfx_k, None)
                st.session_state.pop('hs_recon_df', None)  # force recon rebuild with new POD
                st.rerun(scope="fragment")
            else:
                # Ensure _pod is always mapped even without re-run
                if _df_hs['_pod'].eq('').all() or _df_hs['_pod'].isna().all():
                    _df_hs = _df_hs.copy()
                    _df_hs['_pod'] = _df_hs[_pod_sel].astype(str).str.strip().replace({'nan': '', 'None': ''})
                    st.session_state.hs_parsed = _df_hs
                    # Re-enrich AI tables too
                    for _ai_tbl_key in ['ai_manual_clients', 'hs_pre_ai_manual_clients']:
                        _tbl = st.session_state.get(_ai_tbl_key, pd.DataFrame())
                        if not _tbl.empty:
                            st.session_state[_ai_tbl_key] = _enrich_ai_from_hs(_tbl, _df_hs)
                    for _pfx_k in list(st.session_state.keys()):
                        if _pfx_k.endswith('_ai_manual_editor'):
                            st.session_state.pop(_pfx_k, None)
        else:
            st.session_state['hs_pod_col_override'] = None

        _pod_filled = _df_hs['_pod'].ne('').sum()
        if _pod_filled > 0:
            st.caption(f"✅ POD mapped from **'{_pod_sel}'** — {_pod_filled} clients have a POD assigned.")
        elif _pod_sel != '— None (no POD column) —':
            st.warning(f"⚠️ Column **'{_pod_sel}'** found but all POD values are empty. "
                       "Check that this is the right column.")

        # ── POD filter for summary / onboarding sections ─────────────────
        # The reconciliation table always shows all clients; the three summary
        # sections (Lifecycle Stage, At-Risk, New Onboarding) respect the POD
        # filter chosen in the sidebar (Data Load & Filters).
        _hs_filt_pods = st.session_state.get('_filt_pods', [])
        if _hs_filt_pods and _pod_filled > 0:
            _df_hs_view = _df_hs[
                _df_hs['_pod'].astype(str).str.strip().isin(_hs_filt_pods)
            ].copy()
            st.caption(f"🔍 Filtered to POD: **{', '.join(_hs_filt_pods)}** — {len(_df_hs_view)} client(s) shown.")
        else:
            _df_hs_view = _df_hs

        # ── A. Lifecycle Stage Summary ────────────────────────────────────
        with st.expander("📊 Lifecycle Stage Summary", expanded=True):
            _lc_summary = (
                _df_hs_view.groupby('_lifecycle')
                .agg(Clients=('client_name', 'count'))
                .reset_index()
                .rename(columns={'_lifecycle': 'Lifecycle Stage'})
                .sort_values('Clients', ascending=False)
            )
            _lc_cols = st.columns(min(len(_lc_summary), 5))
            for _ci, (_lc_col_w, (_lc_ri, _lc_row)) in enumerate(zip(_lc_cols, _lc_summary.iterrows())):
                with _lc_col_w:
                    st.metric(_lc_row['Lifecycle Stage'], int(_lc_row['Clients']))
            if len(_lc_summary) > 5:
                st.dataframe(_lc_summary, use_container_width=True, hide_index=True)

        # ── B. Termination / At-Risk Clients ─────────────────────────────
        _df_term = _df_hs_view[_df_hs_view['_is_terminating']].copy()
        if not _df_term.empty:
            with st.expander(f"⚠️ Terminating / At-Risk Clients ({len(_df_term)})", expanded=True):
                st.warning(
                    f"**{len(_df_term)} client(s)** flagged as terminating "
                    "(Lifecycle: Pending Termination / On Notice **AND** "
                    "Retention: Declined Retention / Pending Termination)."
                )
                _term_cols_sel = ['client_name', '_pod', '_lifecycle', '_retention', '_fsd']
                _term_disp = _df_term[[c for c in _term_cols_sel if c in _df_term.columns]].copy()
                _rename_map = {
                    'client_name': 'Client', '_pod': 'POD',
                    '_lifecycle': 'Lifecycle Stage',
                    '_retention': 'Retention Status', '_fsd': 'Final Service Date'
                }
                _term_disp = _term_disp.rename(columns={k: v for k, v in _rename_map.items() if k in _term_disp.columns})
                if 'Final Service Date' in _term_disp.columns:
                    _term_disp['Final Service Date'] = pd.to_datetime(
                        _term_disp['Final Service Date'], errors='coerce'
                    ).dt.strftime('%Y-%m-%d').fillna('—')
                if 'POD' in _term_disp.columns:
                    _term_disp['POD'] = _term_disp['POD'].replace({'nan': '—', '': '—', 'None': '—'})
                st.dataframe(_term_disp, use_container_width=True, hide_index=True)

        # ── C. Reconciliation Table ───────────────────────────────────────
        # Normalize helper: lower + collapse all internal whitespace
        def _norm_name(n):
            return ' '.join(str(n).lower().split())

        _baseline_clients_raw = st.session_state.calc_data.get('clientes_validos', [])
        # Build normalized → original name map for volume clients
        _vol_norm_map = {_norm_name(c): c for c in _baseline_clients_raw}
        _baseline_clients = set(_vol_norm_map.keys())   # normalized keys

        # Filter HubSpot to active clients only:
        # - lifecycle must not start with 'churn'
        # - lifecycle must not be blank / none / nan / —
        # - must have a POD assigned
        _hs_lc_norm = _df_hs['_lifecycle'].astype(str).str.lower().str.strip()
        _hs_active_mask = (
            ~_hs_lc_norm.str.startswith('churn') &
            ~_hs_lc_norm.isin({'', 'none', 'nan', '—'}) &
            _df_hs['_pod'].astype(str).str.strip().ne('') &
            ~_df_hs['_pod'].astype(str).str.strip().isin({'nan', 'None', '—'})
        )
        _df_hs_active = _df_hs[_hs_active_mask].copy()
        _df_hs_active['_name_norm'] = _df_hs_active['client_name'].apply(_norm_name)

        # Build normalized name set for active HubSpot clients
        _hs_norm_set = set(_df_hs_active['_name_norm'].values)

        # Get current client data from session state
        _df_curr = st.session_state.get('df_clients_unique', pd.DataFrame())
        _curr_map = {}
        if not _df_curr.empty and 'client_name' in _df_curr.columns:
            for _, _cr in _df_curr.iterrows():
                _curr_map[_norm_name(_cr['client_name'])] = _cr

        # Build reconciliation rows
        if st.session_state.get('hs_recon_df') is None:
            _recon_rows = []
            _recon_seen = set()   # track normalized names already added

            # ── Pass 1: all active HubSpot clients ───────────────────────
            for _, _hr in _df_hs_active.iterrows():
                _hname_norm = _hr['_name_norm']
                _recon_seen.add(_hname_norm)
                _in_vol = _hname_norm in _baseline_clients
                _status = 'In Both' if _in_vol else 'Only in HubSpot'

                _cur     = _curr_map.get(_hname_norm, {})
                _cur_mrr = float(_cur.get('MRR', 0)) if hasattr(_cur, 'get') else 0.0
                _cur_gl  = _cur.get('Go Live', pd.NaT) if hasattr(_cur, 'get') else pd.NaT
                _cur_fsd = _cur.get('Final Service Date', pd.NaT) if hasattr(_cur, 'get') else pd.NaT

                _cmap_now = st.session_state.get('client_master_map', pd.DataFrame())
                _cur_pod  = ''
                if not _cmap_now.empty:
                    _pm = _cmap_now[_cmap_now['client_key'] == _hname_norm]
                    _cur_pod = str(_pm['pod'].iloc[0]).strip() if not _pm.empty else ''
                if not _cur_pod or _cur_pod.lower() in ('nan', 'none', ''):
                    _cur_pod = str(_cur.get('POD', '') or '').strip() if hasattr(_cur, 'get') else ''

                _hs_pod  = str(_hr.get('_pod', '') or '').strip()
                _new_mrr = float(_hr['_mrr']) if pd.notna(_hr['_mrr']) else _cur_mrr
                _new_gl  = _hr['_start_date'].strftime('%Y-%m-%d') if pd.notna(_hr['_start_date']) else (
                               _cur_gl.strftime('%Y-%m-%d') if pd.notna(_cur_gl) else '')
                _new_fsd = _hr['_fsd'].strftime('%Y-%m-%d') if pd.notna(_hr['_fsd']) else ''
                _cur_fsd_str = str(_cur_fsd.date()) if pd.notna(_cur_fsd) else ''
                _pod_diff = bool(_hs_pod) and _hs_pod.lower() != _cur_pod.lower()
                _mrr_diff = abs(_new_mrr - _cur_mrr) > 0.01
                _gl_diff  = (pd.notna(_hr['_start_date']) and pd.notna(_cur_gl) and
                             _hr['_start_date'].date() != pd.Timestamp(_cur_gl).date())
                _fsd_diff = bool(_new_fsd) and _new_fsd != _cur_fsd_str  # only diff if FSD actually changed
                _has_diff = (_status != 'In Both') or _mrr_diff or _gl_diff or _fsd_diff or _pod_diff

                _recon_rows.append({
                    'Apply':              _has_diff,
                    'Client':             _hr['client_name'],
                    'Current POD':        _cur_pod or '—',
                    'New POD':            _hs_pod or _cur_pod or '',
                    'Status':             _status,
                    'Lifecycle':          _hr['_lifecycle'],
                    'Terminating':        bool(_hr['_is_terminating']),
                    'Current MRR ($)':    _cur_mrr,
                    'New MRR ($)':        _new_mrr,
                    'Current Start Date': str(_cur_gl.date()) if pd.notna(_cur_gl) else '',
                    'New Start Date':     _new_gl,
                    'Final Svc Date':     _new_fsd,
                })

            # ── Pass 2: volume clients not in active HubSpot ─────────────
            for _bc in _baseline_clients_raw:
                _bc_norm = _norm_name(_bc)
                if _bc_norm not in _recon_seen:
                    _bc_cur = _curr_map.get(_bc_norm, {})
                    _bc_pod = str(_bc_cur.get('POD', '') or '').strip() if hasattr(_bc_cur, 'get') else ''
                    _bc_mrr = float(_bc_cur.get('MRR', 0)) if hasattr(_bc_cur, 'get') else 0.0
                    _bc_gl  = _bc_cur.get('Go Live', pd.NaT) if hasattr(_bc_cur, 'get') else pd.NaT
                    _bc_fsd = _bc_cur.get('Final Service Date', pd.NaT) if hasattr(_bc_cur, 'get') else pd.NaT
                    _recon_rows.append({
                        'Apply':              False,
                        'Client':             _bc,
                        'Current POD':        _bc_pod or '—',
                        'New POD':            _bc_pod or '',
                        'Status':             'Only in Volume',
                        'Lifecycle':          '—',
                        'Terminating':        False,
                        'Current MRR ($)':    _bc_mrr,
                        'New MRR ($)':        _bc_mrr,
                        'Current Start Date': str(_bc_gl.date()) if pd.notna(_bc_gl) else '',
                        'New Start Date':     '',
                        'Final Svc Date':     str(_bc_fsd.date()) if pd.notna(_bc_fsd) else '',
                    })

            st.session_state.hs_recon_df = pd.DataFrame(_recon_rows) if _recon_rows else pd.DataFrame()

        with st.expander("📋 Client Reconciliation", expanded=True):
            _recon_df_now = st.session_state.get('hs_recon_df')
            if _recon_df_now is None or _recon_df_now.empty:
                st.info("No matched clients to reconcile.")
            else:
                st.caption(
                    "☑️ Check **Apply** to apply the proposed change. "
                    "Edit **New MRR**, **New Start Date**, or **Final Svc Date** cells manually if needed. "
                    "Rows marked **Not in HubSpot** may indicate churned clients."
                )
                # Rebuild recon table if columns changed
                if 'Current POD' not in _recon_df_now.columns or 'New POD' not in _recon_df_now.columns:
                    st.session_state['hs_recon_df'] = None
                    st.rerun(scope="fragment")
                _recon_pod_opts = [''] + sorted(set(lista_pods) | set(
                    _recon_df_now['Current POD'].replace('—', '').dropna().astype(str).unique().tolist()
                ))
                _recon_clients_opts = st.session_state.hs_recon_df['Client'].tolist()
                st.session_state.hs_recon_df = st.data_editor(
                    st.session_state.hs_recon_df,
                    use_container_width=True,
                    hide_index=True,
                    disabled=['Client', 'Current POD', 'Status', 'Lifecycle', 'Terminating',
                               'Current MRR ($)', 'Current Start Date'],
                    column_config={
                        'Apply':           st.column_config.CheckboxColumn('Apply', default=False),
                        'Client':          st.column_config.TextColumn('Client'),
                        'Current POD':     st.column_config.TextColumn('Current POD'),
                        'New POD':         st.column_config.SelectboxColumn('New POD', options=_recon_pod_opts, default=''),
                        'Status':          st.column_config.TextColumn('Status'),
                        'Lifecycle':       st.column_config.TextColumn('Lifecycle'),
                        'Terminating':     st.column_config.CheckboxColumn('Terminating', disabled=True),
                        'Current MRR ($)': st.column_config.NumberColumn('Current MRR ($)', format='$%.2f'),
                        'New MRR ($)':     st.column_config.NumberColumn('New MRR ($)', format='$%.2f'),
                        'Current Start Date': st.column_config.TextColumn('Current Start Date'),
                        'New Start Date':  st.column_config.TextColumn('New Start Date'),
                        'Final Svc Date':  st.column_config.TextColumn('Final Svc Date'),
                    },
                    key="hs_recon_editor",
                )

                _apply_rows = st.session_state.hs_recon_df[st.session_state.hs_recon_df['Apply'] == True]
                if st.button(
                    f"✅ Apply {len(_apply_rows)} Selected Changes",
                    type="primary", key="hs_apply_btn",
                    disabled=len(_apply_rows) == 0
                ):
                    _overrides = {}
                    for _, _ar in _apply_rows.iterrows():
                        _new_pod_val = str(_ar.get('New POD', '') or '').strip()
                        _cur_pod_val = str(_ar.get('Current POD', '') or '').replace('—', '').strip()
                        _overrides[_ar['Client']] = {
                            'mrr':            _ar['New MRR ($)'],
                            'start_date':     _ar['New Start Date'],
                            'fsd':            _ar['Final Svc Date'],
                            'is_terminating': bool(_ar['Terminating']),
                            'lifecycle':      _ar['Lifecycle'],
                            'pod':            _new_pod_val if _new_pod_val else _cur_pod_val,
                        }
                    st.session_state.hs_client_overrides = _overrides
                    # Also update df_clients_unique so it reflects in downstream calculations
                    if 'df_clients_unique' in st.session_state:
                        _dcu = st.session_state.df_clients_unique.copy()
                        for _cname, _ov in _overrides.items():
                            _mask = _dcu['client_name'].str.lower().str.strip() == _cname.lower().strip()
                            if _mask.any():
                                # Existing client — update in place
                                if _ov['mrr'] is not None:
                                    _dcu.loc[_mask, 'MRR'] = _ov['mrr']
                                if _ov['start_date']:
                                    try:
                                        _dcu.loc[_mask, 'Go Live'] = pd.to_datetime(_ov['start_date'])
                                    except Exception:
                                        pass
                                if _ov['fsd']:
                                    try:
                                        _dcu.loc[_mask, 'Final Service Date'] = pd.to_datetime(_ov['fsd'])
                                    except Exception:
                                        pass
                                if _ov.get('pod'):
                                    _dcu.loc[_mask, 'POD'] = str(_ov['pod']).strip()
                            else:
                                # "New in HubSpot" client — add a new row to baseline
                                _new_row = {c: None for c in _dcu.columns}
                                _new_row['client_name'] = _cname
                                _new_row['MRR']         = _ov.get('mrr', 0.0)
                                _new_row['POD']         = str(_ov.get('pod', '') or '').strip()
                                try:
                                    _new_row['Go Live'] = pd.to_datetime(_ov.get('start_date')) if _ov.get('start_date') else pd.NaT
                                except Exception:
                                    _new_row['Go Live'] = pd.NaT
                                try:
                                    _new_row['Final Service Date'] = pd.to_datetime(_ov.get('fsd')) if _ov.get('fsd') else pd.NaT
                                except Exception:
                                    _new_row['Final Service Date'] = pd.NaT
                                _dcu = pd.concat([_dcu, pd.DataFrame([_new_row])], ignore_index=True)
                        st.session_state.df_clients_unique = _dcu
                    _build_client_master_map()   # re-sync maps with reconciliation overrides
                    # ── Patch df_vol_export immediately so the download reflects
                    # the reconciled values without needing a cascade re-run ──────
                    _vol_exp_now = st.session_state.get('df_vol_export', pd.DataFrame())
                    if not _vol_exp_now.empty and 'client_name' in _vol_exp_now.columns:
                        _ve = _vol_exp_now.copy()
                        _ve_key = _ve['client_name'].astype(str).str.lower().str.strip()
                        for _ov_cn, _ov_v in _overrides.items():
                            _ov_k = _ov_cn.lower().strip()
                            _ve_m = _ve_key == _ov_k
                            if not _ve_m.any():
                                continue
                            if _ov_v.get('mrr') and 'MRR' in _ve.columns:
                                _ve.loc[_ve_m, 'MRR'] = float(_ov_v['mrr'])
                            if _ov_v.get('start_date') and 'Go Live' in _ve.columns:
                                try: _ve.loc[_ve_m, 'Go Live'] = pd.to_datetime(_ov_v['start_date'])
                                except Exception: pass
                            if _ov_v.get('fsd') and 'Final Service Date' in _ve.columns:
                                try: _ve.loc[_ve_m, 'Final Service Date'] = pd.to_datetime(_ov_v['fsd'])
                                except Exception: pass
                            if _ov_v.get('pod') and 'POD' in _ve.columns:
                                _ve.loc[_ve_m, 'POD'] = str(_ov_v['pod']).strip()
                        st.session_state['df_vol_export'] = _ve
                    # Also force reconciliation rebuild so it reflects the applied values
                    st.session_state['hs_recon_df'] = None
                    st.success(
                        f"✅ Applied changes for {len(_overrides)} client(s). "
                        "Volume export updated — download reflects the latest reconciliation."
                    )
                    if _overrides:
                        _term_count = sum(1 for v in _overrides.values() if v['is_terminating'])
                        if _term_count:
                            st.warning(
                                f"⚠️ {_term_count} terminating client(s) flagged — "
                                "their capacity will be marked at-risk in the results."
                            )

        # ── D. Onboarding New Clients ─────────────────────────────────────
        # Any HubSpot client not in baseline — exclude Churn and blank lifecycle
        _ob_lc_blank = {'—', '', 'none', 'nan'}
        _ob_lc_norm  = _df_hs_view['_lifecycle'].astype(str).str.lower().str.strip()
        _df_onboard = _df_hs_view[
            ~_df_hs_view['client_name'].str.lower().str.strip().isin(_baseline_clients) &
            ~_ob_lc_norm.str.startswith('churn') &
            ~_ob_lc_norm.isin(_ob_lc_blank) &
            _df_hs_view['_lifecycle'].astype(str).str.strip().ne('')
        ].copy()

        if not _df_onboard.empty:
            with st.expander(f"🆕 New Clients — Not in Baseline ({len(_df_onboard)})", expanded=True):
                st.info(
                    f"**{len(_df_onboard)} client(s)** found in HubSpot but not present in the current baseline. "
                    "You can run the AI Prediction for them in Step 4."
                )
                _ob_disp = _df_onboard[[
                    'client_name', '_pod', '_lifecycle', '_pms', '_mrr', '_start_date'
                ]].copy()
                _ob_disp.columns = ['Client', 'POD', 'Lifecycle', 'PMS', 'MRR ($)', 'Start Date']
                _ob_disp['Start Date'] = pd.to_datetime(_ob_disp['Start Date'], errors='coerce').dt.strftime('%Y-%m-%d').fillna('—')
                _ob_disp['POD'] = _ob_disp['POD'].replace({'nan': '—', '': '—', 'None': '—'}).fillna('—')
                # Display as read-only table — avoids data_editor React #185 crash
                # inside fragments when data shape changes between renders
                st.dataframe(
                    _ob_disp,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        'MRR ($)': st.column_config.NumberColumn('MRR ($)', format='$%.0f'),
                    },
                )
                _ob_all_clients = _ob_disp['Client'].tolist()
                _hs_ob_ms_key = f"hs_ob_ms_{abs(hash(tuple(sorted(_ob_all_clients))))}"
                _sel_clients = st.multiselect(
                    "Select clients to queue:",
                    options=_ob_all_clients,
                    default=_ob_all_clients,
                    key=_hs_ob_ms_key,
                )
                _sel_ob = _ob_disp[_ob_disp['Client'].isin(_sel_clients)]
                if st.button(
                    f"🤖 Queue {len(_sel_ob)} Clients for AI Prediction (Step 4)",
                    type="primary", key="hs_queue_ai_btn",
                    disabled=len(_sel_ob) == 0
                ):
                    # Pre-populate ai_manual_clients with these new clients
                    _new_ai_rows = []
                    for _, _nr in _sel_ob.iterrows():
                        _new_ai_rows.append({
                            'Company Name':      _nr['Client'],
                            'POD':               _nr['POD'] if _nr['POD'] not in ('', 'nan', 'None', '—', None) else '',
                            'Go Live Date':      _nr['Start Date'] if _nr['Start Date'] != '—' else '',
                            'MRR ($)':           float(_nr['MRR ($)']) if pd.notna(_nr['MRR ($)']) else 0.0,
                            'PMS':               _nr['PMS'] if _nr['PMS'] not in ('nan', 'Unknown', None) else '',
                            'Res Doors':         0,
                            'Res Properties':    0,
                            'Comm Doors':        0,
                            'Comm Properties':   0,
                            'SQFT Commercial':   0,
                            'Corp Books':        '',
                        })
                    _new_ai_df = pd.DataFrame(_new_ai_rows)
                    # Enrich from HubSpot before queuing (fills POD, Go Live, MRR, PMS, Res Doors, SQFT)
                    _new_ai_df = _enrich_ai_from_hs(_new_ai_df, st.session_state.get('hs_parsed'))
                    # Upsert: match on Company Name + POD; update existing rows, append truly new ones
                    _existing_ai = st.session_state.get('ai_manual_clients', pd.DataFrame())
                    _added   = 0
                    _updated = 0
                    if not _existing_ai.empty and 'Company Name' in _existing_ai.columns:
                        _ex = _existing_ai.copy()
                        _ex['_key'] = (
                            _ex['Company Name'].astype(str).str.lower().str.strip() + '||' +
                            _ex.get('POD', pd.Series([''] * len(_ex))).astype(str).str.lower().str.strip()
                        )
                        _truly_new = []
                        for _, _nr in _new_ai_df.iterrows():
                            _k = (str(_nr['Company Name']).lower().strip() + '||' +
                                  str(_nr.get('POD', '')).lower().strip())
                            if _k in _ex['_key'].values:
                                # Update existing row in place
                                _ex.loc[_ex['_key'] == _k, list(_nr.index)] = _nr.values
                                _updated += 1
                            else:
                                _truly_new.append(_nr)
                                _added += 1
                        _ex = _ex.drop(columns=['_key'])
                        if _truly_new:
                            _ex = pd.concat([_ex, pd.DataFrame(_truly_new)], ignore_index=True)
                        st.session_state.ai_manual_clients = _ex
                    else:
                        st.session_state.ai_manual_clients = _new_ai_df.copy()
                        _added = len(_new_ai_df)
                    st.session_state.hs_onboarding_clients = _new_ai_df
                    # ── Also sync hs_pre_ai_manual_clients so the early AI section refreshes ──
                    # (top-level page code only runs on full rerun; fragment fires don't trigger it)
                    _hs_pre_key_q = 'hs_pre_ai_manual_clients'
                    _hs_pre_ex_q  = st.session_state.get(_hs_pre_key_q, pd.DataFrame())
                    _hs_snap_q    = st.session_state.get('hs_parsed')
                    if _hs_pre_ex_q.empty:
                        _seeded_q = _new_ai_df.copy()
                        _seeded_q = _enrich_ai_from_hs(_seeded_q, _hs_snap_q)
                        st.session_state[_hs_pre_key_q] = _seeded_q
                    else:
                        _pre_keys_q = set(
                            _hs_pre_ex_q['Company Name'].astype(str).str.lower().str.strip() + '||' +
                            _hs_pre_ex_q.get('POD', pd.Series(['']*len(_hs_pre_ex_q))).astype(str).str.lower().str.strip()
                        )
                        _new_pre_rows_q = []
                        _update_rows_q  = []
                        for _, _qr_q in _new_ai_df.iterrows():
                            _qk_q = (str(_qr_q['Company Name']).lower().strip() + '||' +
                                     str(_qr_q.get('POD', '')).lower().strip())
                            if _qk_q in _pre_keys_q:
                                _update_rows_q.append((_qk_q, _qr_q))
                            else:
                                _new_pre_rows_q.append(_qr_q)
                        # Apply updates in-place
                        if _update_rows_q:
                            _pre_key_ser = (
                                _hs_pre_ex_q['Company Name'].astype(str).str.lower().str.strip() + '||' +
                                _hs_pre_ex_q.get('POD', pd.Series(['']*len(_hs_pre_ex_q))).astype(str).str.lower().str.strip()
                            )
                            for _uk_q, _urow_q in _update_rows_q:
                                _umask_q = _pre_key_ser == _uk_q
                                _hs_pre_ex_q.loc[_umask_q, list(_urow_q.index)] = _urow_q.values
                        # Append truly new rows
                        if _new_pre_rows_q:
                            _new_pre_df_q = _enrich_ai_from_hs(pd.DataFrame(_new_pre_rows_q), _hs_snap_q)
                            _hs_pre_ex_q = pd.concat([_hs_pre_ex_q, _new_pre_df_q], ignore_index=True)
                        st.session_state[_hs_pre_key_q] = _hs_pre_ex_q
                    # Clear the editor key so data_editor re-initialises with the new data
                    st.session_state.pop('hs_pre_ai_manual_editor', None)
                    if _added > 0 and _updated == 0:
                        st.success(
                            f"✅ {_added} client(s) added to the AI Prediction table in Step 4. "
                            "Scroll down to Step 4 → New Clients AI Prediction to run predictions."
                        )
                    elif _updated > 0 and _added == 0:
                        st.success(
                            f"✅ {_updated} client(s) updated in the AI Prediction table in Step 4. "
                            "Scroll down to Step 4 → New Clients AI Prediction to run predictions."
                        )
                    elif _added > 0 or _updated > 0:
                        st.success(
                            f"✅ {_added} client(s) added, {_updated} updated in the AI Prediction table in Step 4. "
                            "Scroll down to Step 4 → New Clients AI Prediction to run predictions."
                        )
                    else:
                        st.warning(
                            "⚠️ No clients were queued — all selected clients are already in the AI Prediction table "
                            "with the same name and POD. Check Step 4 → New Clients AI Prediction."
                        )

        if st.button("🔄 Re-upload HubSpot file", key="hs_reupload_btn"):
            st.session_state.hs_parsed = None
            st.session_state.hs_recon_df = None
            st.rerun(scope="fragment")

    with st.expander("🔄 HubSpot Sync", expanded=(st.session_state.get("hs_sync_choice") is None and "calc_data" in st.session_state)):
        _hs_sync_fragment()

    # ── Client Master Map viewer ──────────────────────────────────────────────
    _cmap_view = st.session_state.get('client_master_map', pd.DataFrame())
    if not _cmap_view.empty:
        with st.expander(f"🗺️ Client Master Map ({len(_cmap_view)} clients)", expanded=False):
            st.caption(
                "Single source of truth: **Client → POD → Sr. Accountant → MRR → Go Live → FSD**. "
                "Built from Master DB (Layer 1) → HubSpot (Layer 2) → Reconciliation (Layer 3). "
                "This map is applied automatically before every cascade run."
            )
            _mv_tab_pod, _mv_tab_sr, _mv_tab_vol, _mv_tab_all = st.tabs(
                ["🏷️ POD Map", "👤 Sr. Map", "📊 Volume Map", "📋 Full Table"]
            )

            with _mv_tab_pod:
                _pod_grp = (
                    _cmap_view.groupby('pod')['client_name']
                    .apply(lambda x: ', '.join(sorted(x.tolist())))
                    .reset_index()
                    .rename(columns={'pod': 'POD', 'client_name': 'Clients'})
                    .sort_values('POD')
                )
                st.caption(f"**{_cmap_view['pod'].nunique()}** PODs · **{len(_cmap_view)}** clients")
                st.dataframe(_pod_grp, use_container_width=True, hide_index=True,
                             column_config={"POD": st.column_config.TextColumn("POD", width="small"),
                                            "Clients": st.column_config.TextColumn("Clients")})

            with _mv_tab_sr:
                _sr_view = _cmap_view[['client_name', 'sr_accountant', 'pod']].copy()
                _sr_view.columns = ['Client', 'Sr. Accountant', 'POD']
                _sr_view = _sr_view.sort_values(['Sr. Accountant', 'POD', 'Client'])
                _sr_no   = _sr_view['Sr. Accountant'].eq('')
                st.caption(
                    f"**{_sr_view['Sr. Accountant'].replace('', pd.NA).notna().sum()}** clients have a Sr. Accountant · "
                    f"**{_sr_no.sum()}** unassigned"
                )
                if _sr_no.any():
                    st.warning(f"⚠️ {_sr_no.sum()} client(s) have no Sr. Accountant assigned.")
                st.dataframe(_sr_view, use_container_width=True, hide_index=True)

            with _mv_tab_vol:
                _vol_view = _cmap_view[['client_name', 'pod', 'mrr', 'go_live', 'fsd', 'source']].copy()
                _vol_view.columns = ['Client', 'POD', 'MRR ($)', 'Go Live', 'FSD', 'Source']
                _vol_view['Go Live'] = pd.to_datetime(_vol_view['Go Live'], errors='coerce').dt.strftime('%Y-%m-%d').fillna('—')
                _vol_view['FSD']     = pd.to_datetime(_vol_view['FSD'],     errors='coerce').dt.strftime('%Y-%m-%d').fillna('—')
                _vol_view = _vol_view.sort_values(['POD', 'Client'])
                _no_gl  = (_vol_view['Go Live'] == '—').sum()
                _has_fsd = (_vol_view['FSD'] != '—').sum()
                st.caption(f"**{_no_gl}** clients missing Go Live · **{_has_fsd}** clients with a Final Service Date")
                st.dataframe(_vol_view, use_container_width=True, hide_index=True,
                             column_config={"MRR ($)": st.column_config.NumberColumn("MRR ($)", format="$%.0f")})

            with _mv_tab_all:
                _all_view = _cmap_view[['client_name', 'pod', 'sr_accountant', 'mrr', 'go_live', 'fsd', 'source']].copy()
                _all_view.columns = ['Client', 'POD', 'Sr. Accountant', 'MRR ($)', 'Go Live', 'FSD', 'Source']
                _all_view['Go Live'] = pd.to_datetime(_all_view['Go Live'], errors='coerce').dt.strftime('%Y-%m-%d').fillna('—')
                _all_view['FSD']     = pd.to_datetime(_all_view['FSD'],     errors='coerce').dt.strftime('%Y-%m-%d').fillna('—')
                _all_view = _all_view.sort_values(['POD', 'Sr. Accountant', 'Client'])
                st.dataframe(_all_view, use_container_width=True, hide_index=True,
                             column_config={"MRR ($)": st.column_config.NumberColumn("MRR ($)", format="$%.0f")})

# ── AI Prediction for queued onboarding clients (early access, before Step 2) ──
_hs_queued = st.session_state.get('ai_manual_clients', pd.DataFrame())
if "calc_data" in st.session_state and not _hs_queued.empty:
    # Keep hs_pre manual table in sync with the HubSpot queue (upsert, don't overwrite user edits)
    _hs_pre_key = 'hs_pre_ai_manual_clients'
    _hs_pre_existing = st.session_state.get(_hs_pre_key, pd.DataFrame())
    _hs_parsed_snap = st.session_state.get('hs_parsed')
    if _hs_pre_existing.empty:
        # First time — seed from HubSpot queue, then enrich with HubSpot data
        _seeded = _hs_queued.copy()
        _seeded = _enrich_ai_from_hs(_seeded, _hs_parsed_snap)
        st.session_state[_hs_pre_key] = _seeded
    else:
        # Upsert: add any newly queued clients not yet in the pre-table
        _pre_keys = set(
            (_hs_pre_existing['Company Name'].astype(str).str.lower().str.strip() + '||' +
             _hs_pre_existing.get('POD', pd.Series(['']*len(_hs_pre_existing))).astype(str).str.lower().str.strip())
        )
        _new_pre_rows = []
        for _, _qr in _hs_queued.iterrows():
            _qk = str(_qr['Company Name']).lower().strip() + '||' + str(_qr.get('POD', '')).lower().strip()
            if _qk not in _pre_keys:
                _new_pre_rows.append(_qr)
        if _new_pre_rows:
            _new_pre_df = _enrich_ai_from_hs(pd.DataFrame(_new_pre_rows), _hs_parsed_snap)
            st.session_state[_hs_pre_key] = pd.concat(
                [_hs_pre_existing, _new_pre_df], ignore_index=True
            )
    # Default widget states for first render
    if 'hs_pre_ai_input_mode' not in st.session_state:
        st.session_state['hs_pre_ai_input_mode'] = "✏️ Manual Entry"

    _n_queued = len(_hs_queued)
    _ai_already_run = 'hs_pre_ai_results' in st.session_state
    with st.expander(
        f"🤖 AI Prediction — {_n_queued} New Onboarding Client(s)",
        expanded=(not _ai_already_run)
    ):
        _make_ai_prediction_fragment(pfx="hs_pre", add_to_scenario=False, add_to_baseline=True)()

# ==========================================
# STEP 2: SIMULATOR (AUTOMATIONS & ADJUSTMENTS)
# ==========================================
if "calc_data" in st.session_state:
    with st.expander("✍️ Step 2: Efficiency Initiatives", expanded=(not st.session_state.get('_s2_collapsed', False) and "final_dashboards" not in st.session_state)):

        # ── Prompt: ask the user before showing the efficiency tables ────────────
        if st.session_state.s2_efficiency_choice is None:
            st.info(
                "**Would you like to configure efficiency initiatives?**\n\n"
                "You can add Automations, extra hours, or reductions — or skip directly to Step 3 & 4."
            )
            _eff_col1, _eff_col2, _ = st.columns([1.5, 2, 4])
            with _eff_col1:
                if st.button("⚙️ Add Efficiency", type="primary", use_container_width=True):
                    st.session_state.s2_efficiency_choice = "yes"
                    st.session_state['_s2_proceed'] = False   # reset gate
                    st.rerun()
            with _eff_col2:
                if st.button("⏭️ Skip to Step 3 & 4", use_container_width=True):
                    st.session_state.s2_efficiency_choice = "skip"
                    st.session_state['_s2_collapsed'] = True
                    st.rerun()
            st.stop()

        # ── Allow the user to change their mind ─────────────────────────────────
        if st.session_state.s2_efficiency_choice == "skip":
            st.success("Efficiency step skipped — no automations or adjustments will be applied.")
            if st.button("⚙️ Go back and add efficiency initiatives", key="s2_back_btn"):
                st.session_state.s2_efficiency_choice = "yes"
                st.session_state['_s2_proceed'] = False   # reset gate
                st.rerun()
        else:
            # Show the full efficiency tables
            st.markdown("1️⃣ Apply Automations **$\\rightarrow$** 2️⃣ Add Historical Hours **$\\rightarrow$** 3️⃣ Apply Reductions.")
            if st.button("⏭️ Skip efficiency / remove all adjustments", key="s2_skip_btn"):
                st.session_state.s2_efficiency_choice = "skip"
                st.rerun()

        # ── Fragment functions — each editor reruns independently, no full-page flash ──

        @st.fragment
        def _auto_tab_fragment():
            st.markdown(
                "**Choose Client, Task, and the metric(s) this automation reduces (%).** "
                "Use the **Affects** dropdown to select which factors are impacted."
            )
            _auto_tmpl = pd.DataFrame([
                {"Initiative Name": "AP Automation", "Client": "All",
                 "Task (Type - Subtype)": "AP - Invoice Processing",
                 "Affects": "Vol Proc + Vol Rev",
                 "M1 (%)": 10, "M2 (%)": 15, "M3 (%)": 20, "M4 (%)": 20, "M5 (%)": 20, "M6 (%)": 20},
                {"Initiative Name": "AI Coding", "Client": "Acme Corp",
                 "Task (Type - Subtype)": "All",
                 "Affects": "AHT Proc",
                 "M1 (%)": 5, "M2 (%)": 5, "M3 (%)": 10, "M4 (%)": 10, "M5 (%)": 10, "M6 (%)": 10},
            ])
            _tmpl_buf = BytesIO()
            _auto_tmpl.to_excel(_tmpl_buf, index=False)
            st.download_button("📄 Download Automations Template", _tmpl_buf.getvalue(),
                               file_name="Automations_Template.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="dl_auto_tmpl")
            _auto_upload = st.file_uploader("📂 Load automations from Excel (optional)",
                                            type=["xlsx","xls"], key="auto_file_load")
            if _auto_upload:
                try:
                    _auto_loaded = pd.read_excel(_auto_upload)
                    _need_cols   = ["Initiative Name","Client","Task (Type - Subtype)","Affects",
                                    "M1 (%)","M2 (%)","M3 (%)","M4 (%)","M5 (%)","M6 (%)"]
                    for _nc in _need_cols:
                        if _nc not in _auto_loaded.columns: _auto_loaded[_nc] = None
                    st.session_state.automations_df = _auto_loaded[_need_cols].copy()
                    st.success(f"Loaded {len(_auto_loaded)} automation rows.")
                except Exception as _e:
                    st.error(f"Could not load file: {_e}")
            if "Confirmed" not in st.session_state.automations_df.columns:
                st.session_state.automations_df.insert(0, "Confirmed", False)
            if "POD" not in st.session_state.automations_df.columns:
                st.session_state.automations_df.insert(2, "POD", "")
            if "PMS" not in st.session_state.automations_df.columns:
                st.session_state.automations_df.insert(3, "PMS", "")
            _auto_pods_opts = ["", "All"] + lista_pods
            _auto_pms_opts  = _get_pms_opts(include_all=True)
            st.caption("☑️ Check **Confirmed** to include a row. Leave **POD / PMS / Client / Task** blank or 'All' to apply broadly.")
            st.session_state.automations_df = st.data_editor(
                st.session_state.automations_df,
                num_rows="dynamic",
                use_container_width=True,
                key="auto_ed",
                column_config={
                    "Confirmed": st.column_config.CheckboxColumn("✅", default=False, help="Only confirmed rows are applied in Step 3"),
                    "POD": st.column_config.SelectboxColumn("POD", options=_auto_pods_opts, default="", help="Leave blank or 'All' to apply to every POD"),
                    "PMS": st.column_config.SelectboxColumn("PMS", options=_auto_pms_opts, default="", help="Leave blank or 'All' to apply to every PMS"),
                    "Client": st.column_config.SelectboxColumn("Client", options=lista_clientes, default="All"),
                    "Task (Type - Subtype)": st.column_config.SelectboxColumn("Task", options=lista_tareas, default="All"),
                    "Affects": st.column_config.SelectboxColumn("Affects", options=AFFECTS_OPTIONS, default="All (Vol + AHT)",
                        help="Vol Proc = Processing Volume | Vol Rev = Review Volume | AHT Proc = Processing Handle Time | AHT Rev = Review Handle Time"),
                    "M1 (%)": st.column_config.NumberColumn("M1 (%)", min_value=0.0, max_value=100.0, default=0.0, format="%.1f%%"),
                    "M2 (%)": st.column_config.NumberColumn("M2 (%)", min_value=0.0, max_value=100.0, default=0.0, format="%.1f%%"),
                    "M3 (%)": st.column_config.NumberColumn("M3 (%)", min_value=0.0, max_value=100.0, default=0.0, format="%.1f%%"),
                    "M4 (%)": st.column_config.NumberColumn("M4 (%)", min_value=0.0, max_value=100.0, default=0.0, format="%.1f%%"),
                    "M5 (%)": st.column_config.NumberColumn("M5 (%)", min_value=0.0, max_value=100.0, default=0.0, format="%.1f%%"),
                    "M6 (%)": st.column_config.NumberColumn("M6 (%)", min_value=0.0, max_value=100.0, default=0.0, format="%.1f%%"),
                }
            )

            # ── Enforce cumulative roll-over: each month ≥ previous month ────────
            _pct_cols = ["M1 (%)", "M2 (%)", "M3 (%)", "M4 (%)", "M5 (%)", "M6 (%)"]
            if not st.session_state.automations_df.empty:
                _existing = [c for c in _pct_cols if c in st.session_state.automations_df.columns]
                if _existing:
                    _num_before = st.session_state.automations_df[_existing].apply(pd.to_numeric, errors='coerce').fillna(0.0)
                    _num_after  = _num_before.cummax(axis=1)
                    if not _num_before.equals(_num_after):
                        st.session_state.automations_df[_existing] = _num_after
                        st.rerun(scope="fragment")

        @st.fragment
        def _hist_tab_fragment():
            st.markdown("**Pure extra hours to add per month (e.g. Historical Accounting). Include the POD for new clients.**")
            _avail_pods_h = [""] + lista_pods
            _avail_cli_h  = st.session_state.calc_data.get('clientes_validos', [])
            _hist_tmpl = pd.DataFrame([
                {"POD": "POD A", "Client": "Acme Corp", "Required Role": "Accountant I",
                 "M1 (Hrs)": 20, "M2 (Hrs)": 20, "M3 (Hrs)": 20,
                 "M4 (Hrs)": 0,  "M5 (Hrs)": 0,  "M6 (Hrs)": 0},
            ])
            _ht_buf = BytesIO()
            _hist_tmpl.to_excel(_ht_buf, index=False)
            _ht_dl_col, _ht_ul_col = st.columns([1, 1])
            with _ht_dl_col:
                st.download_button("📄 Download Add Hours Template", _ht_buf.getvalue(),
                                   file_name="Add_Hours_Template.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key="dl_hist_tmpl")
            with _ht_ul_col:
                _hist_upload = st.file_uploader("📂 Upload Add Hours file", type=["xlsx"],
                                                key="fu_hist_tmpl", label_visibility="collapsed")
                st.caption("📂 Upload Add Hours file (.xlsx)")
                if _hist_upload is not None:
                    try:
                        _hist_up_df = pd.read_excel(_hist_upload)
                        if 'Role' in _hist_up_df.columns and 'Required Role' not in _hist_up_df.columns:
                            _hist_up_df = _hist_up_df.rename(columns={'Role': 'Required Role'})
                        _hist_expected = ['Confirmed', 'POD', 'Client', 'Required Role',
                                          'M1 (Hrs)', 'M2 (Hrs)', 'M3 (Hrs)', 'M4 (Hrs)', 'M5 (Hrs)', 'M6 (Hrs)']
                        for _hc2 in _hist_expected:
                            if _hc2 not in _hist_up_df.columns:
                                _hist_up_df[_hc2] = True if _hc2 == 'Confirmed' else (0.0 if '(Hrs)' in _hc2 else '')
                        _hist_up_df = _hist_up_df[_hist_expected]
                        # All rows from upload are confirmed by default
                        _hist_up_df['Confirmed'] = True
                        for _hc2 in ['M1 (Hrs)', 'M2 (Hrs)', 'M3 (Hrs)', 'M4 (Hrs)', 'M5 (Hrs)', 'M6 (Hrs)']:
                            _hist_up_df[_hc2] = pd.to_numeric(_hist_up_df[_hc2], errors='coerce').fillna(0.0)
                        # Normalize client names case-insensitively
                        _cli_norm_h = {c.lower().strip(): c for c in _avail_cli_h if c}
                        _hist_up_df['Client'] = _hist_up_df['Client'].fillna('').astype(str).str.strip().apply(
                            lambda v: _cli_norm_h.get(v.lower(), v) if v else ''
                        )
                        # Normalize POD
                        _pod_norm_h = {p.lower().strip(): p for p in lista_pods if p}
                        _hist_up_df['POD'] = _hist_up_df['POD'].fillna('').astype(str).str.strip().apply(
                            lambda v: _pod_norm_h.get(v.lower(), v) if v else ''
                        )
                        st.session_state.historical_df = _hist_up_df
                        st.success(f"✅ Loaded {len(_hist_up_df)} row(s) — all marked Confirmed.")
                    except Exception as _he:
                        st.error(f"❌ Could not read file: {_he}")
            if "Confirmed" not in st.session_state.historical_df.columns:
                st.session_state.historical_df.insert(0, "Confirmed", False)
            if "POD" not in st.session_state.historical_df.columns:
                st.session_state.historical_df.insert(1, "POD", "")
            st.caption("☑️ **Confirmed** rows are included in the cascade. All uploaded rows are confirmed by default.")
            st.session_state.historical_df = st.data_editor(
                st.session_state.historical_df,
                num_rows="dynamic",
                use_container_width=True,
                key="hist_ed",
                column_config={
                    "Confirmed": st.column_config.CheckboxColumn("✅", default=True, help="Only confirmed rows are applied in Step 3"),
                    "POD": st.column_config.SelectboxColumn("POD", options=_avail_pods_h, default="", help="Required for new clients not yet in the cascade"),
                    "Client": st.column_config.SelectboxColumn("Client", options=_avail_cli_h, required=True),
                    "Required Role": st.column_config.SelectboxColumn("Required Role", options=roles_permitidos, required=True),
                }
            )

        @st.fragment
        def _red_tab_fragment():
            st.markdown(
                "**Manual hours to subtract.** Targeting is hierarchical — specify as much or as little as needed:\n\n"
                "- **POD only** → prorates across all clients & roles in that POD\n"
                "- **POD + Role** → prorates across all clients in that POD for that role\n"
                "- **POD + Client** → prorates across all roles for that client\n"
                "- **POD + Client + Role** → applies to that exact combination\n"
                "- **Client + Role** → applies to that client/role (no POD filter)"
            )
            _red_tmpl = pd.DataFrame([
                {"Confirmed": True, "POD": "POD A", "Client": "", "Required Role": "",
                 "M1 (Hrs)": 10, "M2 (Hrs)": 10, "M3 (Hrs)": 0, "M4 (Hrs)": 0, "M5 (Hrs)": 0, "M6 (Hrs)": 0},
                {"Confirmed": True, "POD": "POD A", "Client": "", "Required Role": "Accountant I",
                 "M1 (Hrs)": 5,  "M2 (Hrs)": 5,  "M3 (Hrs)": 0, "M4 (Hrs)": 0, "M5 (Hrs)": 0, "M6 (Hrs)": 0},
                {"Confirmed": True, "POD": "",      "Client": "Acme Corp", "Required Role": "Accountant I",
                 "M1 (Hrs)": 8,  "M2 (Hrs)": 0,  "M3 (Hrs)": 0, "M4 (Hrs)": 0, "M5 (Hrs)": 0, "M6 (Hrs)": 0},
            ])
            _rt_buf = BytesIO()
            _red_tmpl.to_excel(_rt_buf, index=False)
            _rt_dl_col, _rt_ul_col = st.columns([1, 1])
            with _rt_dl_col:
                st.download_button("📄 Download Remove Hours Template", _rt_buf.getvalue(),
                                   file_name="Remove_Hours_Template.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key="dl_red_tmpl")
            with _rt_ul_col:
                _red_upload = st.file_uploader("📂 Upload Remove Hours file", type=["xlsx"],
                                               key="fu_red_tmpl", label_visibility="collapsed")
                st.caption("📂 Upload Remove Hours file (.xlsx)")
                if _red_upload is not None:
                    try:
                        _red_up_df = pd.read_excel(_red_upload)
                        # Normalise column names — accept "Role" as alias for "Required Role"
                        if 'Role' in _red_up_df.columns and 'Required Role' not in _red_up_df.columns:
                            _red_up_df = _red_up_df.rename(columns={'Role': 'Required Role'})
                        _red_expected = ['Confirmed', 'POD', 'Client', 'Required Role',
                                         'M1 (Hrs)', 'M2 (Hrs)', 'M3 (Hrs)', 'M4 (Hrs)', 'M5 (Hrs)', 'M6 (Hrs)']
                        for _rc in _red_expected:
                            if _rc not in _red_up_df.columns:
                                _red_up_df[_rc] = False if _rc == 'Confirmed' else (0.0 if '(Hrs)' in _rc else '')
                        _red_up_df = _red_up_df[_red_expected]
                        # All rows from upload are confirmed by default
                        _red_up_df['Confirmed'] = True
                        for _hc in ['M1 (Hrs)', 'M2 (Hrs)', 'M3 (Hrs)', 'M4 (Hrs)', 'M5 (Hrs)', 'M6 (Hrs)']:
                            _red_up_df[_hc] = pd.to_numeric(_red_up_df[_hc], errors='coerce').fillna(0.0)
                        # Normalize POD/Client values to match exact dropdown options (case-insensitive)
                        _avail_pods = st.session_state.get('_lista_pods', lista_pods)
                        _pod_norm_map = {p.lower().strip(): p for p in _avail_pods if p}
                        _red_up_df['POD'] = _red_up_df['POD'].fillna('').astype(str).str.strip().apply(
                            lambda v: _pod_norm_map.get(v.lower(), v) if v else ''
                        )
                        _avail_clients = st.session_state.calc_data.get('clientes_validos', [])
                        _cli_norm_map = {c.lower().strip(): c for c in _avail_clients if c}
                        _red_up_df['Client'] = _red_up_df['Client'].fillna('').astype(str).str.strip().apply(
                            lambda v: _cli_norm_map.get(v.lower(), v) if v else ''
                        )
                        st.session_state.reductions_df = _red_up_df
                        st.success(f"✅ Loaded {len(_red_up_df)} row(s) — all marked Confirmed.")
                    except Exception as _re:
                        st.error(f"❌ Could not read file: {_re}")
            if "Confirmed" not in st.session_state.reductions_df.columns:
                st.session_state.reductions_df.insert(0, "Confirmed", False)
            if "POD" not in st.session_state.reductions_df.columns:
                st.session_state.reductions_df.insert(1, "POD", "")
            _red_clients_opts = [""] + st.session_state.calc_data.get('clientes_validos', [])
            _red_pods_opts    = [""] + lista_pods
            st.caption("☑️ Check **Confirmed** on each row to include it in the cascade. Unconfirmed rows are ignored.")
            st.session_state.reductions_df = st.data_editor(
                st.session_state.reductions_df,
                num_rows="dynamic",
                use_container_width=True,
                key="red_ed",
                column_config={
                    "Confirmed": st.column_config.CheckboxColumn("✅", default=False, help="Only confirmed rows are applied in Step 3"),
                    "POD": st.column_config.SelectboxColumn("POD", options=_red_pods_opts, default="", help="Leave blank to target by Client only"),
                    "Client": st.column_config.SelectboxColumn("Client", options=_red_clients_opts, default="", help="Leave blank to target entire POD"),
                    "Required Role": st.column_config.SelectboxColumn("Role", options=[""] + roles_permitidos, default="", help="Leave blank to prorate across all roles"),
                    "M1 (Hrs)": st.column_config.NumberColumn("M1 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M2 (Hrs)": st.column_config.NumberColumn("M2 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M3 (Hrs)": st.column_config.NumberColumn("M3 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M4 (Hrs)": st.column_config.NumberColumn("M4 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M5 (Hrs)": st.column_config.NumberColumn("M5 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M6 (Hrs)": st.column_config.NumberColumn("M6 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                }
            )

        @st.fragment
        def _doorcount_tab_fragment():
            st.markdown(
                "**Door / property count variation per client.** "
                "Applies a percentage change to that client's capacity hours for the selected months — "
                "positive % = increase (more doors), negative % = decrease. "
                "Only ✅ Confirmed rows are applied in the cascade."
            )
            _dc_avail_cli  = [""] + st.session_state.calc_data.get('clientes_validos', [])
            _dc_avail_pods = [""] + lista_pods

            # ── Template download ─────────────────────────────────────────────
            _dc_tmpl = pd.DataFrame([
                {"Client": "Acme Corp", "POD": "POD A",
                 "M1 (%)": 5.0, "M2 (%)": 5.0, "M3 (%)": 0.0,
                 "M4 (%)": 0.0, "M5 (%)": 0.0, "M6 (%)": 0.0},
                {"Client": "Beta LLC",  "POD": "",
                 "M1 (%)": -10.0, "M2 (%)": -10.0, "M3 (%)": 0.0,
                 "M4 (%)": 0.0,  "M5 (%)": 0.0,    "M6 (%)": 0.0},
            ])
            _dc_buf = BytesIO()
            _dc_tmpl.to_excel(_dc_buf, index=False)
            _dc_dl_col, _dc_ul_col = st.columns([1, 1])
            with _dc_dl_col:
                st.download_button(
                    "📄 Download Door Count Template",
                    _dc_buf.getvalue(),
                    file_name="DoorCount_Variation_Template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_dc_tmpl",
                )
            with _dc_ul_col:
                _dc_upload = st.file_uploader("📂 Upload Door Count file", type=["xlsx"],
                                              key="fu_dc_tmpl", label_visibility="collapsed")
                st.caption("📂 Upload Door Count file (.xlsx)")
                if _dc_upload is not None:
                    try:
                        _dc_up_df = pd.read_excel(_dc_upload)
                        _dc_expected = ['Confirmed', 'Client', 'POD'] + meses_pct_cols
                        for _dcc in _dc_expected:
                            if _dcc not in _dc_up_df.columns:
                                _dc_up_df[_dcc] = True if _dcc == 'Confirmed' else (0.0 if '(%)' in _dcc else '')
                        _dc_up_df = _dc_up_df[_dc_expected]
                        _dc_up_df['Confirmed'] = True   # all uploaded rows confirmed
                        for _dcc in meses_pct_cols:
                            _dc_up_df[_dcc] = pd.to_numeric(_dc_up_df[_dcc], errors='coerce').fillna(0.0)
                        _dc_cli_norm = {c.lower().strip(): c for c in _dc_avail_cli if c}
                        _dc_up_df['Client'] = _dc_up_df['Client'].fillna('').astype(str).str.strip().apply(
                            lambda v: _dc_cli_norm.get(v.lower(), v) if v else ''
                        )
                        _dc_pod_norm = {p.lower().strip(): p for p in lista_pods if p}
                        _dc_up_df['POD'] = _dc_up_df['POD'].fillna('').astype(str).str.strip().apply(
                            lambda v: _dc_pod_norm.get(v.lower(), v) if v else ''
                        )
                        st.session_state.doorcount_df = _dc_up_df
                        st.success(f"✅ Loaded {len(_dc_up_df)} row(s) — all marked Confirmed.")
                    except Exception as _dce:
                        st.error(f"❌ Could not read file: {_dce}")

            if "Confirmed" not in st.session_state.doorcount_df.columns:
                st.session_state.doorcount_df.insert(0, "Confirmed", True)
            if "POD" not in st.session_state.doorcount_df.columns:
                st.session_state.doorcount_df.insert(2, "POD", "")

            st.caption("☑️ **Confirmed** rows are applied in the cascade. Positive % = hours increase, negative % = decrease.")
            st.session_state.doorcount_df = st.data_editor(
                st.session_state.doorcount_df,
                num_rows="dynamic",
                use_container_width=True,
                key="dc_ed",
                column_config={
                    "Confirmed": st.column_config.CheckboxColumn("✅", default=True, help="Only confirmed rows are applied in Step 3"),
                    "Client": st.column_config.SelectboxColumn("Client", options=_dc_avail_cli, required=True),
                    "POD": st.column_config.SelectboxColumn("POD", options=_dc_avail_pods, default=""),
                    "M1 (%)": st.column_config.NumberColumn("M1 (%)", format="%.1f%%", help="% change for month 1"),
                    "M2 (%)": st.column_config.NumberColumn("M2 (%)", format="%.1f%%"),
                    "M3 (%)": st.column_config.NumberColumn("M3 (%)", format="%.1f%%"),
                    "M4 (%)": st.column_config.NumberColumn("M4 (%)", format="%.1f%%"),
                    "M5 (%)": st.column_config.NumberColumn("M5 (%)", format="%.1f%%"),
                    "M6 (%)": st.column_config.NumberColumn("M6 (%)", format="%.1f%%"),
                },
            )

        if st.session_state.s2_efficiency_choice == "yes":
            t_auto, t_hist, t_red, t_door = st.tabs([
                "⚙️ Automations", "➕ Add Hours", "➖ Reduce Hours", "🚪 Door Count Variation"
            ])

            with t_auto:
                _auto_tab_fragment()

            with t_hist:
                _hist_tab_fragment()

            with t_red:
                _red_tab_fragment()

            with t_door:
                _doorcount_tab_fragment()

            # ── Proceed gate: hide Step 3 & 4 while the user is actively
            # configuring Step 2. A "Proceed" button opens the gate; having
            # already-run cascade results also keeps the gate open.
            _s2_proc = st.session_state.get('_s2_proceed', False)
            if not _s2_proc and 'final_dashboards' not in st.session_state:
                st.divider()
                _s2p_c1, _s2p_c2 = st.columns([3, 1])
                with _s2p_c1:
                    st.info("👆 Configure efficiency initiatives above. Click **Proceed to Step 3** when ready.")
                with _s2p_c2:
                    if st.button("▶ Proceed to Step 3", type="primary", key="s2_proceed_btn", use_container_width=True):
                        st.session_state['_s2_proceed'] = True
                        st.rerun()
                st.stop()   # hides Step 3 & 4 until user clicks Proceed

        # ==========================================
    # STEP 3: RECALCULATION & DASHBOARDS
    # ==========================================
    # Expander state: open by default until cascade runs once, then collapsed.
    # We track the desired state in session state so reruns (e.g. from radio
    # change) don't force-close the expander while the user is still interacting
    # with the radio or cascade button.
    if 'final_dashboards' not in st.session_state:
        # Haven't run yet → always open
        st.session_state['_s3_exp_open'] = True
    # If session key missing (first ever load), default to open
    _s3_exp_open = st.session_state.get('_s3_exp_open', True)
    with st.expander("📊 Step 3: Generate Results", expanded=_s3_exp_open):

        # ── Pre-flight: show confirmed row counts ─────────────────────────────
        def _confirmed_count(df_s):
            if df_s.empty or "Confirmed" not in df_s.columns: return 0
            return int(df_s["Confirmed"].eq(True).sum())

        _n_auto = _confirmed_count(st.session_state.automations_df)
        _n_hist = _confirmed_count(st.session_state.historical_df)
        _n_red  = _confirmed_count(st.session_state.reductions_df)
        _total  = _n_auto + _n_hist + _n_red

        _unconf_total = (
            (len(st.session_state.automations_df) - _n_auto) +
            (len(st.session_state.historical_df)  - _n_hist) +
            (len(st.session_state.reductions_df)  - _n_red)
        )

        _parts = []
        if _n_auto: _parts.append(f"**{_n_auto}** automation(s)")
        if _n_hist: _parts.append(f"**{_n_hist}** add-hours row(s)")
        if _n_red:  _parts.append(f"**{_n_red}** reduction(s)")

        if not _parts:
            st.info("ℹ️ No confirmed inputs — cascade will run on baseline only.")
        else:
            st.success(f"✅ Will apply: {', '.join(_parts)}.")

        if _unconf_total > 0:
            st.warning(f"⚠️ {_unconf_total} unconfirmed row(s) will be **skipped**. Check ✅ on any row you want included.")

        # ── Role pair selector ────────────────────────────────────────────────────
        _s3_role_mode = st.radio(
            "📋 Role pair for calculations",
            ["🎯 Ideal (Ideal Proc / Ideal Rev)", "👥 Real (Proc Role / Rev Role)"],
            horizontal=True,
            key="s3_role_mode_radio",
            help=(
                "**Ideal** uses the standardised role assigned to each task type (Ideal Proc / Ideal Rev).\n\n"
                "**Real** uses the actual role of the employee who performed the work (Proc Role / Rev Role)."
            ),
        )
        _s3_use_real = _s3_role_mode.startswith("👥")

        _auto_cascade = st.session_state.pop('_auto_run_cascade', False)
        if st.button("🔄 Apply Cascade & Generate Dashboards", type="primary", use_container_width=True) or _auto_cascade:
            with st.spinner("Calculating financial savings, automations, and final FTEs…"):

                df               = st.session_state.df_clean.copy()
                df_clients_unique = st.session_state.df_clients_unique.copy()

                # ── Apply master map: overwrite POD / Sr. / Go Live / FSD from
                #    the unified lookup built across Master DB + HubSpot + reconciliation
                _build_client_master_map()   # ensure map is fresh before cascade
                _cmap_cas = st.session_state.get('client_master_map', pd.DataFrame())
                if not _cmap_cas.empty and 'client_name' in df.columns:
                    _df_key = df['client_name'].astype(str).str.lower().str.strip()
                    _pod_lkp = dict(zip(_cmap_cas['client_key'], _cmap_cas['pod']))
                    _sr_lkp  = dict(zip(_cmap_cas['client_key'], _cmap_cas['sr_accountant']))
                    _gl_lkp  = dict(zip(_cmap_cas['client_key'], _cmap_cas['go_live']))
                    _fsd_lkp = dict(zip(_cmap_cas['client_key'], _cmap_cas['fsd']))
                    _mrr_lkp = dict(zip(_cmap_cas['client_key'], _cmap_cas['mrr']))
                    # POD: fill from map (HubSpot/reconcile beats Master DB blank)
                    df['POD'] = _df_key.map(_pod_lkp).fillna(df.get('POD', 'No POD'))
                    df['POD'] = df['POD'].fillna('No POD').astype(str).str.strip()
                    df['POD'] = df['POD'].where(~df['POD'].str.lower().isin({'nan','none',''}), 'No POD')
                    # Sr. Accountant: fill gaps from map
                    if 'Sr. Accountant' not in df.columns:
                        df['Sr. Accountant'] = ''
                    _sr_from_map = _df_key.map(_sr_lkp)
                    _sr_empty = df['Sr. Accountant'].astype(str).str.strip().str.lower().isin({'','nan','none'})
                    df.loc[_sr_empty, 'Sr. Accountant'] = _sr_from_map[_sr_empty]
                    # Go Live / FSD: apply reconciliation dates (overrides Master DB)
                    _gl_from_map  = _df_key.map(_gl_lkp)
                    _fsd_from_map = _df_key.map(_fsd_lkp)
                    _mrr_from_map = _df_key.map(_mrr_lkp)
                    _gl_has_map  = _gl_from_map.notna()
                    _fsd_has_map = _fsd_from_map.notna()
                    df.loc[_gl_has_map,  'Go Live']              = pd.to_datetime(_gl_from_map[_gl_has_map],   errors='coerce')
                    df.loc[_fsd_has_map, 'Final Service Date']   = pd.to_datetime(_fsd_from_map[_fsd_has_map], errors='coerce')
                    if 'MRR' in df.columns:
                        _mrr_has_map = _mrr_from_map.notna() & (_mrr_from_map > 0)
                        df.loc[_mrr_has_map, 'MRR'] = _mrr_from_map[_mrr_has_map]
                    # Persist the fully-updated df so the Volume Input export captures
                    # all master-map changes (POD, Sr., Go Live, FSD, MRR, AI clients)
                    st.session_state['df_vol_export'] = df.copy()
                    # Also sync df_clients_unique with master map dates/MRR
                    if not df_clients_unique.empty and 'client_name' in df_clients_unique.columns:
                        _duc_key = df_clients_unique['client_name'].astype(str).str.lower().str.strip()
                        for _col_tgt, _lkp_d in [('Go Live', _gl_lkp), ('Final Service Date', _fsd_lkp), ('MRR', _mrr_lkp)]:
                            if _col_tgt in df_clients_unique.columns:
                                _mapped = _duc_key.map(_lkp_d)
                                _has    = _mapped.notna()
                                if _col_tgt == 'MRR':
                                    _has = _has & (_mapped > 0)
                                df_clients_unique.loc[_has, _col_tgt] = _mapped[_has]
                dict_hrs_per_fte  = st.session_state.calc_data['dict_hrs_per_fte']
                dict_workable_days = st.session_state.calc_data['dict_workable_days']

                # Resolve role mode BEFORE the main loop.
                # Use _s3_use_real (set directly from the radio widget return value
                # at render time) — this is authoritative and avoids any session-state
                # read ambiguity inside the button callback.
                _s3_use_real_cas = _s3_use_real

                summary_data_auto = []
                monthly_prod_hrs  = {i: 0.0 for i in range(6)}
                monthly_util_hrs  = {i: 0.0 for i in range(6)}
                monthly_abs_hrs   = {i: 0.0 for i in range(6)}
                monthly_att_hrs   = {i: 0.0 for i in range(6)}
                # Per-POD productive hours: {pod_name: {month_idx: prod_hrs}}
                _pod_prod_hrs_m   = {}

                # 1. RECALCULATE WITH AUTOMATIONS
                # ── Base month for demand-side day scaling (same reference as baseline) ─
                _bm_s2      = pd.Timestamp((today.replace(day=1) - relativedelta(months=1)).date())
                _bm_e2      = pd.Timestamp((_bm_s2 + relativedelta(months=1) - relativedelta(days=1)).date())
                _base_nd_s2 = max(1, int(np.busday_count(
                    _bm_s2.strftime('%Y-%m-%d'),
                    (_bm_e2 + pd.Timedelta(days=1)).strftime('%Y-%m-%d')
                )))

                # ── Pre-compute month params ──────────────────────────────────────
                _month_params = []
                for _mi, _ms in enumerate(meses_proyeccion):
                    _md  = today + relativedelta(months=_month_offsets[_mi])
                    _sm  = pd.Timestamp(_md.replace(day=1).date())
                    _em  = pd.Timestamp((_sm + relativedelta(months=1) - relativedelta(days=1)).date())
                    # Actual network days (demand side — always real Mon-Fri count)
                    _actual_nd = int(np.busday_count(_sm.strftime('%Y-%m-%d'),
                                                     (_em + pd.Timedelta(days=1)).strftime('%Y-%m-%d')))
                    # Capacity days (FTE denominator — may be fixed by user)
                    _nd  = fixed_days if calc_mode == "Fixed days per month" else _actual_nd
                    # Day-scale: fixed mode keeps hours constant; network mode scales vs base month
                    _dscale = 1.0 if calc_mode == "Fixed days per month" else (_actual_nd / _base_nd_s2)
                    _month_params.append((_mi, _ms, _sm, _em, _nd, _actual_nd, _dscale))

                # ── Vectorised active_pct and learning_mult for all rows × months ─
                _gl_raw  = df['Go Live'].values
                _fsd_raw = df['Final Service Date'].values
                _n_rows  = len(df)
                _apct_m  = {}   # month_idx → np.array(n_rows)
                _lc_m    = {}   # month_idx → np.array(n_rows)

                for _mi, _ms, _sm, _em, _nd, _actual_nd, _dscale in _month_params:
                    # pd fillna/clip avoids NPY_ITER_REFS_OK error with object datetime arrays
                    _gl_s  = pd.to_datetime(pd.Series(_gl_raw),  errors='coerce').fillna(_sm).clip(lower=_sm, upper=_em)
                    _fsd_s = pd.to_datetime(pd.Series(_fsd_raw), errors='coerce').fillna(_em).clip(lower=_sm, upper=_em)
                    _dias  = np.busday_count(
                        _gl_s.values.astype('datetime64[D]'),
                        (_fsd_s + pd.Timedelta(days=1)).values.astype('datetime64[D]')
                    )
                    # Use actual network days (not fixed_days) so active_pct matches baseline
                    _apct_m[_mi] = np.clip(np.maximum(_dias.astype(float), 0) / _actual_nd, 0.0, 1.0) if _actual_nd > 0 else np.zeros(_n_rows)
                    # Explicitly zero out clients whose FSD fell BEFORE this month's start.
                    # The clip(lower=_sm) above would otherwise assign ~1/21 of a month to churned clients.
                    _fsd_raw_s = pd.to_datetime(pd.Series(_fsd_raw), errors='coerce')
                    _fully_churned = _fsd_raw_s.notna() & (_fsd_raw_s < _sm)
                    if _fully_churned.any():
                        _apct_m[_mi] = np.where(_fully_churned.values, 0.0, _apct_m[_mi])

                    _has_gl = pd.notna(pd.Series(_gl_raw)).values
                    _gl_ts  = pd.DatetimeIndex(pd.to_datetime(pd.Series(_gl_raw), errors='coerce').fillna(_sm))
                    _mdiff  = np.where(_has_gl,
                                       (_sm.year  - _gl_ts.year.values)  * 12 +
                                       (_sm.month - _gl_ts.month.values),
                                       999).astype(int)
                    _ap     = _apct_m[_mi]
                    _lc_m[_mi] = np.select(
                        [~_has_gl | (_ap == 0),
                         (_mdiff == 0) & _has_gl & (_ap > 0),
                         (_mdiff == 1) & _has_gl & (_ap > 0),
                         (_mdiff == 2) & _has_gl & (_ap > 0)],
                        [1.0, 1.17, 0.86, 0.99],
                        default=1.0
                    )

                # ── Pre-compute automation efficiencies per (client, task, pod, pms) ─
                _autos_raw = st.session_state.automations_df
                _autos_src = _autos_raw[_autos_raw.get("Confirmed", pd.Series(True, index=_autos_raw.index)) == True].copy() if not _autos_raw.empty else _autos_raw
                _auto_cache = {}   # (client, task, pod, pms) → list[6] of (vp, vr, ap, ar)

                def _auto_matches_all(v):
                    return pd.isna(v) or str(v).strip() in ('', 'All')

                if not _autos_src.empty:
                    _uniq_pairs = (
                        df.assign(_task=df['type'].astype(str) + ' - ' + df['subtype'].astype(str))
                          .assign(_pod=df.get('POD', pd.Series('', index=df.index)).fillna('').astype(str).str.strip())
                          .assign(_pms=df.get('PMS', pd.Series('', index=df.index)).fillna('').astype(str).str.strip())
                          [['client_name', '_task', '_pod', '_pms']].drop_duplicates()
                    )
                    for _, _up in _uniq_pairs.iterrows():
                        _ck   = str(_up['client_name']).strip()
                        _tk   = str(_up['_task']).strip()
                        _pk   = str(_up['_pod']).strip()
                        _pmsk = str(_up['_pms']).strip()
                        _mc   = (_autos_src["Client"].apply(_auto_matches_all)) | (_autos_src["Client"] == _ck)
                        _mt   = (_autos_src["Task (Type - Subtype)"].apply(_auto_matches_all)) | (_autos_src["Task (Type - Subtype)"] == _tk)
                        _mpod = (_autos_src["POD"].apply(_auto_matches_all)) | (_autos_src["POD"] == _pk)   if "POD" in _autos_src.columns else pd.Series(True, index=_autos_src.index)
                        _mpms = (_autos_src["PMS"].apply(_auto_matches_all)) | (_autos_src["PMS"] == _pmsk) if "PMS" in _autos_src.columns else pd.Series(True, index=_autos_src.index)
                        _ap_rows = _autos_src[_mc & _mt & _mpod & _mpms]
                        _effs = []
                        for _mi2 in range(6):
                            _evp = _evr = _eap = _ear = 0.0
                            if not _ap_rows.empty:
                                _mc2 = f"M{_mi2+1} (%)"
                                for _, _au in _ap_rows.iterrows():
                                    _v = pd.to_numeric(_au.get(_mc2, 0), errors='coerce')
                                    if pd.isna(_v): _v = 0.0
                                    _v /= 100.0
                                    _af = str(_au.get("Affects", ""))
                                    _ia = _af == "All (Vol + AHT)"
                                    if "Vol Proc" in _af or _ia: _evp += _v
                                    if "Vol Rev"  in _af or _ia: _evr += _v
                                    if "AHT Proc" in _af or _ia: _eap += _v
                                    if "AHT Rev"  in _af or _ia: _ear += _v
                            _effs.append((min(1.0,_evp), min(1.0,_evr), min(1.0,_eap), min(1.0,_ear)))
                        _auto_cache[(_ck, _tk, _pk, _pmsk)] = _effs

                # ── Main loop: rows (outer) × months (inner) ─────────────────────
                for _ri, (index, row) in enumerate(df.iterrows()):
                    _ck    = str(row.get('client_name', '')).strip()
                    _tk    = str(row.get('type', '')) + ' - ' + str(row.get('subtype', ''))
                    _pk    = str(row.get('POD', '')).strip()
                    _pmsk  = str(row.get('PMS', '')).strip()
                    _effs  = _auto_cache.get((_ck, _tk, _pk, _pmsk), [(0.0,0.0,0.0,0.0)]*6)

                    if _s3_use_real_cas:
                        _ideal_p = str(row.get('Proc Role', '')).strip()
                        _ideal_r = str(row.get('Rev Role',  '')).strip()
                        # Fall back to Ideal Proc/Rev when real roles are blank/invalid
                        if _ideal_p in ['nan', 'None', '']:
                            _ideal_p = str(row.get('Ideal Proc', row.get('Proc Role', 'Accountant I'))).strip()
                        if _ideal_r in ['nan', 'None', '']:
                            _ideal_r = str(row.get('Ideal Rev', row.get('Rev Role', 'Sr. Accountant'))).strip()
                        if _ideal_p in ['nan', 'None', '']: _ideal_p = 'Accountant I'
                        if _ideal_r in ['nan', 'None', '']: _ideal_r = 'Sr. Accountant'
                    else:
                        _ideal_p = str(row.get('Ideal Proc', row.get('Proc Role', 'Accountant I'))).strip()
                        _ideal_r = str(row.get('Ideal Rev',  row.get('Rev Role',  'Sr. Accountant'))).strip()
                        if _ideal_p in ['nan','None','']: _ideal_p = str(row.get('Proc Role','Accountant I')).strip()
                        if _ideal_r in ['nan','None','']: _ideal_r = str(row.get('Rev Role','Sr. Accountant')).strip()

                    _util_p = utilization_map.get(_ideal_p, util_acc1)
                    _util_r = utilization_map.get(_ideal_r, util_sr)
                    _ptix   = float(row.get('Closed tickets with Proc time', 0) or 0)
                    _rtix   = float(row.get('Closed tickets with rev time',  0) or 0)
                    _paht   = float(row.get('>>> FINAL Capacity Proc AHT',   0) or 0)
                    _raht   = float(row.get('>>> FINAL Capacity Rev AHT',    0) or 0)
                    _pod_raw = row.get('POD', '')
                    _pod    = str(_pod_raw if pd.notna(_pod_raw) else '').strip()
                    if _pod.lower() in ('nan', 'none', ''): _pod = 'No POD'
                    _cli    = str(row.get('client_name', 'Unknown')).strip()

                    for _mi, _ms, _sm, _em, _nd, _actual_nd, _dscale in _month_params:
                        _apct = float(_apct_m[_mi][_ri])
                        _lc   = float(_lc_m[_mi][_ri])
                        _evp, _evr, _eap, _ear = _effs[_mi]

                        # Apply day_scale so post-auto hours match the baseline when no automations exist
                        _bp = (_ptix * _apct * (1 - _evp) * _paht * (1 - _eap) * _lc * _dscale) / 60
                        _br = (_rtix * _apct * (1 - _evr) * _raht * (1 - _ear) * _lc * _dscale) / 60

                        _tp = _bp + _bp*(1-_util_p) + _bp*absenteeism + _bp*attrition
                        _tr = _br + _br*(1-_util_r) + _br*absenteeism + _br*attrition

                        summary_data_auto.append({'POD': _pod, 'Client': _cli, 'Required Role': _ideal_p, f"M{_mi+1} ({_ms}) - Post-Auto Hours": _tp})
                        summary_data_auto.append({'POD': _pod, 'Client': _cli, 'Required Role': _ideal_r, f"M{_mi+1} ({_ms}) - Post-Auto Hours": _tr})

                        monthly_prod_hrs[_mi] += (_bp + _br)
                        monthly_util_hrs[_mi] += _bp*(1-_util_p) + _br*(1-_util_r)
                        monthly_abs_hrs[_mi]  += (_bp + _br) * absenteeism
                        monthly_att_hrs[_mi]  += (_bp + _br) * attrition
                        # Accumulate per-POD productive hours for POD-level waterfall
                        if _pod not in _pod_prod_hrs_m:
                            _pod_prod_hrs_m[_pod] = {j: 0.0 for j in range(6)}
                        _pod_prod_hrs_m[_pod][_mi] += (_bp + _br)

                df_sum_auto    = pd.DataFrame(summary_data_auto)
                cols_auto      = [c for c in df_sum_auto.columns if " - Post-Auto Hours" in c]
                df_resumen_auto = df_sum_auto.groupby(['POD', 'Client', 'Required Role'])[cols_auto].sum().reset_index()

                # Store per-POD productive hours so the POD waterfall can use them
                # (includes both baseline and AI clients; covers AI-only PODs where
                #  df_resumen_base has no 'Productive Hours' entries)
                st.session_state.calc_data['pod_prod_hrs'] = _pod_prod_hrs_m

                # 2. MERGE BASE AND AUTOMATED
                # Normalize POD in both frames before merging so NaN/'nan' → 'No POD' consistently
                def _norm_pod_col(df_arg):
                    if 'POD' in df_arg.columns:
                        df_arg = df_arg.copy()
                        df_arg['POD'] = (
                            df_arg['POD'].fillna('No POD').astype(str).str.strip()
                            .where(lambda s: ~s.str.lower().isin({'nan','none',''}), 'No POD')
                        )
                    return df_arg

                # Re-use the same flag (already set from radio widget at render time)
                _s3_use_real_cas = _s3_use_real
                _rb_base_key = 'df_resumen_base_real' if _s3_use_real_cas else 'df_resumen_base_ideal'
                _rb_normed  = _norm_pod_col(st.session_state.calc_data.get(_rb_base_key,
                                            st.session_state.calc_data['df_resumen_base']))
                _ra_normed  = _norm_pod_col(df_resumen_auto)

                # ── Re-assign POD in df_resumen_base using master map ──────────────
                # Step 1 builds df_resumen_base from Master DB which may have blank/No POD.
                # The cascade (Step 3) now enriches df with HubSpot PODs, so df_resumen_auto
                # has the correct PODs. Without this fix, the outer merge creates split rows:
                #   "No POD" row (base hours, c_final=0 → filtered) vs
                #   "POD 1"  row (auto hours, no base) — losing all baseline hours per POD.
                _pod_lkp_rb = st.session_state.get('_pod_map', {})
                if _pod_lkp_rb and 'Client' in _rb_normed.columns:
                    _rb_cli_key  = _rb_normed['Client'].astype(str).str.lower().str.strip()
                    _rb_map_pod  = _rb_cli_key.map(_pod_lkp_rb)
                    # Only override where master map has a real POD assignment
                    _rb_override = _rb_map_pod.notna() & (_rb_map_pod != 'No POD')
                    if _rb_override.any():
                        _rb_normed = _rb_normed.copy()
                        _rb_normed.loc[_rb_override, 'POD'] = _rb_map_pod[_rb_override]
                        # Re-aggregate: rows that share (POD, Client, Role) after reassignment
                        # must be summed so hours aren't duplicated
                        _rb_num_cols = [c for c in _rb_normed.columns
                                        if c not in ('POD', 'Client', 'Required Role')]
                        _rb_normed = (_rb_normed
                                      .groupby(['POD', 'Client', 'Required Role'], as_index=False)
                                      [_rb_num_cols].sum())
                        # Persist updated POD assignments back so future cascade runs start clean
                        st.session_state.calc_data['df_resumen_base'] = _rb_normed.copy()
                        st.session_state.calc_data[_rb_base_key]      = _rb_normed.copy()

                df_resumen = pd.merge(
                    _rb_normed,
                    _ra_normed,
                    on=['POD', 'Client', 'Required Role'],
                    how='outer'
                ).fillna(0.0)

                df_resumen['Monthly_Cost'] = df_resumen['Required Role'].map(cost_map).fillna(0.0)

                # Map Sr. Accountant from master map (single source of truth)
                _sr_lkp_res = st.session_state.get('_sr_map', {})
                if not _sr_lkp_res:
                    # Fallback: build from df if master map not available
                    if 'Sr. Accountant' in df.columns:
                        _sr_src = df.dropna(subset=['client_name']).copy()
                        _sr_src['_key_lower'] = _sr_src['client_name'].astype(str).str.lower().str.strip()
                        _sr_lkp_res = _sr_src.groupby('_key_lower')['Sr. Accountant'].first().to_dict()
                df_resumen['Sr. Accountant'] = (
                    df_resumen['Client'].astype(str).str.lower().str.strip()
                    .map(_sr_lkp_res).fillna('')
                )

                for i, mes_str in enumerate(meses_proyeccion):
                    df_resumen[f"M{i+1} ({mes_str}) - Adjustments (+) Hrs"] = 0.0
                    df_resumen[f"M{i+1} ({mes_str}) - Adjustments (-) Hrs"] = 0.0

                # 3. APPLY (+) HISTORICAL ADJUSTMENTS
                _hist_raw = st.session_state.historical_df
                _hist_confirmed = _hist_raw[_hist_raw.get("Confirmed", pd.Series(True, index=_hist_raw.index)) == True] if not _hist_raw.empty else _hist_raw
                for _, r in _hist_confirmed.iterrows():
                    c_name, rol = r.get("Client"), r.get("Required Role")
                    if pd.notna(c_name) and pd.notna(rol):
                        mask = (df_resumen["Client"] == c_name) & (df_resumen["Required Role"] == rol)
                        if not mask.any():
                            _h_pod = str(r.get("POD", "") or "").strip()
                            if not _h_pod or _h_pod.lower() in ('nan', 'none', ''):
                                _h_pod = "Manual/Historical"
                            new_row = {"POD": _h_pod, "Client": c_name, "Required Role": rol, "Monthly_Cost": cost_map.get(rol, 0)}
                            for col in df_resumen.columns:
                                if col not in new_row: new_row[col] = 0.0
                            df_resumen = pd.concat([df_resumen, pd.DataFrame([new_row])], ignore_index=True)
                            mask = (df_resumen["Client"] == c_name) & (df_resumen["Required Role"] == rol)

                        for i, mes_str in enumerate(meses_proyeccion):
                            val = pd.to_numeric(r.get(f"M{i+1} (Hrs)", 0), errors='coerce')
                            if pd.notna(val) and val > 0:
                                df_resumen.loc[mask, f"M{i+1} ({mes_str}) - Adjustments (+) Hrs"] += val

                # 3b. APPLY DOOR COUNT VARIATION
                # Each confirmed row specifies a % change per month for a client.
                # The percentage scales that client's Post-Auto Hours for the chosen months.
                # Positive = more doors (more hours), negative = fewer doors (fewer hours).
                _dc_raw = st.session_state.get('doorcount_df', pd.DataFrame())
                if not _dc_raw.empty:
                    _dc_conf = _dc_raw[_dc_raw.get('Confirmed', pd.Series(True, index=_dc_raw.index)) == True]
                    for _, _dcr in _dc_conf.iterrows():
                        _dc_cli = str(_dcr.get('Client', '') or '').strip()
                        if not _dc_cli or _dc_cli.lower() in ('nan', 'none', ''):
                            continue
                        _dc_mask = df_resumen['Client'] == _dc_cli
                        for _dci, _dcms in enumerate(meses_proyeccion):
                            _dc_pct = pd.to_numeric(_dcr.get(f"M{_dci+1} (%)", 0), errors='coerce') or 0.0
                            if _dc_pct == 0.0:
                                continue
                            _dc_post_col = f"M{_dci+1} ({_dcms}) - Post-Auto Hours"
                            _dc_adj_col  = (
                                f"M{_dci+1} ({_dcms}) - Adjustments (+) Hrs"
                                if _dc_pct > 0
                                else f"M{_dci+1} ({_dcms}) - Adjustments (-) Hrs"
                            )
                            if _dc_post_col in df_resumen.columns and _dc_mask.any():
                                _dc_delta = df_resumen.loc[_dc_mask, _dc_post_col] * (abs(_dc_pct) / 100.0)
                                df_resumen.loc[_dc_mask, _dc_adj_col] += _dc_delta.values

                # 4. APPLY (-) REDUCTION ADJUSTMENTS  (POD-hierarchy)
                _red_raw = st.session_state.reductions_df
                _red_confirmed = _red_raw[_red_raw.get("Confirmed", pd.Series(True, index=_red_raw.index)) == True] if not _red_raw.empty else _red_raw

                def _blank(v):
                    return pd.isna(v) or str(v).strip() in ('', 'nan', 'None', 'All')

                for _, r in _red_confirmed.iterrows():
                    _pod_v  = r.get("POD", "")
                    _cli_v  = r.get("Client", "")
                    _rol_v  = r.get("Required Role", "")
                    has_pod = not _blank(_pod_v)
                    has_cli = not _blank(_cli_v)
                    has_rol = not _blank(_rol_v)

                    for i, mes_str in enumerate(meses_proyeccion):
                        val = pd.to_numeric(r.get(f"M{i+1} (Hrs)", 0), errors='coerce')
                        if pd.isna(val) or val <= 0:
                            continue
                        col_post = f"M{i+1} ({mes_str}) - Post-Auto Hours"
                        col_minus = f"M{i+1} ({mes_str}) - Adjustments (-) Hrs"

                        # Build base mask
                        _m_pod = (df_resumen['POD'].astype(str).str.strip() == str(_pod_v).strip()) if has_pod else pd.Series(True, index=df_resumen.index)
                        _m_cli = (df_resumen['Client'] == str(_cli_v).strip()) if has_cli else pd.Series(True, index=df_resumen.index)
                        _m_rol = (df_resumen['Required Role'] == str(_rol_v).strip()) if has_rol else pd.Series(True, index=df_resumen.index)
                        mask = _m_pod & _m_cli & _m_rol

                        if has_cli and has_rol:
                            # Exact match — apply directly
                            df_resumen.loc[mask, col_minus] += val
                        else:
                            # Prorate by each row's share of total hours in the target group
                            tot = df_resumen.loc[mask, col_post].sum()
                            if tot > 0:
                                df_resumen.loc[mask, col_minus] += val * (df_resumen.loc[mask, col_post] / tot)

                # 5. BUILD FINAL CASCADE & SAVINGS ($)
                columnas_ordenadas = ['POD', 'Sr. Accountant', 'Client', 'Required Role']

                # Automation savings are only real if the user actually configured automations.
                # Without automations, small differences between Step-1 Base Hours and cascade
                # Post-Auto Hours (caused by master-map date updates changing active_pct) would
                # produce phantom "savings". Gate on: efficiency was enabled AND the automations
                # table has at least one confirmed entry.
                _autos_df_check = st.session_state.get('s2_automations', pd.DataFrame())
                _has_real_autos = (
                    st.session_state.get('s2_efficiency_choice') == 'yes'
                    and not _autos_df_check.empty
                    and (
                        _autos_df_check.get('Confirmed', pd.Series(dtype=bool)).any()
                        if 'Confirmed' in _autos_df_check.columns
                        else len(_autos_df_check) > 0
                    )
                )

                for i, mes_str in enumerate(meses_proyeccion):
                    c_base      = f"M{i+1} ({mes_str}) - Base Hours"
                    c_post      = f"M{i+1} ({mes_str}) - Post-Auto Hours"
                    c_save_hrs  = f"M{i+1} ({mes_str}) - Auto Saving (Hrs)"
                    c_save_usd  = f"M{i+1} ({mes_str}) - Auto Saving ($)"
                    c_plus      = f"M{i+1} ({mes_str}) - Adjustments (+) Hrs"
                    c_minus     = f"M{i+1} ({mes_str}) - Adjustments (-) Hrs"
                    c_final     = f"M{i+1} ({mes_str}) - Final Hours"
                    c_fte       = f"M{i+1} ({mes_str}) - Final FTEs"

                    # Only compute real automation savings when automations are actually configured;
                    # otherwise keep it zero to avoid phantom values from date-update drift.
                    if _has_real_autos:
                        df_resumen[c_save_hrs] = np.maximum(0, df_resumen[c_base] - df_resumen[c_post])
                    else:
                        df_resumen[c_save_hrs] = 0.0

                    hrs_fte = dict_hrs_per_fte[i]
                    df_resumen[c_save_usd] = (
                        (df_resumen[c_save_hrs] / hrs_fte * df_resumen['Monthly_Cost']).round(2)
                        if hrs_fte > 0 else 0.0
                    )

                    df_resumen[c_minus] = np.minimum(df_resumen[c_minus], df_resumen[c_post] + df_resumen[c_plus])
                    df_resumen[c_final] = df_resumen[c_post] + df_resumen[c_plus] - df_resumen[c_minus]
                    df_resumen[c_fte]   = (df_resumen[c_final] / hrs_fte).round(4) if hrs_fte > 0 else 0.0

                    for col in [c_base, c_save_hrs, c_post, c_plus, c_minus, c_final]:
                        df_resumen[col] = df_resumen[col].round(2)

                    columnas_ordenadas.extend([c_base, c_save_hrs, c_save_usd, c_post, c_plus, c_minus, c_final, c_fte])

                df_resumen = df_resumen[columnas_ordenadas]
                cols_check  = [c for c in df_resumen.columns if "Final Hours" in c]
                df_resumen  = df_resumen.loc[(df_resumen[cols_check] > 0.01).any(axis=1)]

                # 5b. BASELINE AUDIT TABLE — all per-row calculated fields
                # Build bidirectional lookup from HC report:
                #   _n2email : full_name_lower → email
                #   _e2name  : email_lower     → full_name   (for display when processor IS an email)
                _hc_snap = st.session_state.get('hc_data')
                _n2email = {}
                _e2name  = {}
                if _hc_snap:
                    _hc_det_snap = _hc_snap.get('detail', pd.DataFrame())
                    if not _hc_det_snap.empty:
                        for _, _hr in _hc_det_snap.iterrows():
                            _fn = str(_hr.get('Full name', '')).strip()
                            _em = str(_hr.get('Work Email', '')).strip()
                            if _fn and _em and _em.lower() not in ('nan', 'none', ''):
                                _n2email[_fn.lower()]  = _em
                                _e2name[_em.lower()]   = _fn

                _audit_rows = []
                _hrs_fte_m1 = dict_hrs_per_fte.get(0, 1)
                _wdays_m1   = dict_workable_days.get(0, 21)
                _m1_start   = pd.Timestamp((today + relativedelta(months=0)).replace(day=1).date())
                _m1_end     = pd.Timestamp((_m1_start + relativedelta(months=1) - relativedelta(days=1)).date())

                # Case-insensitive column lookup — master DB may use any casing
                _ci_col = {c.strip().lower(): c for c in df.columns}

                def _get_ci(row, col_lower, default=''):
                    actual = _ci_col.get(col_lower)
                    if actual is None:
                        return default
                    v = row.get(actual, default)
                    return default if (v is None or (isinstance(v, float) and pd.isna(v))) else v

                for _idx, _row in df.iterrows():
                    _cli   = str(_row.get('client_name', 'Unknown')).strip()
                    _pod_r = _row.get('POD', '')
                    _pod   = str(_pod_r if pd.notna(_pod_r) else '').strip()
                    if _pod.lower() in ('nan', 'none', ''): _pod = 'No POD'
                    _sr    = str(_row.get('Sr. Accountant', '')).strip()
                    _mrr   = float(_row.get('MRR', 0) or 0)
                    _gl    = _row['Go Live']
                    _fsd   = _row['Final Service Date']

                    if _s3_use_real_cas:
                        _ip = str(_row.get('Proc Role', 'Accountant I')).strip()
                        _ir = str(_row.get('Rev Role',  'Sr. Accountant')).strip()
                        if _ip in ['nan','None','']: _ip = 'Accountant I'
                        if _ir in ['nan','None','']: _ir = 'Sr. Accountant'
                    else:
                        _ip = str(_row.get('Ideal Proc', _row.get('Proc Role', 'Accountant I'))).strip()
                        _ir = str(_row.get('Ideal Rev',  _row.get('Rev Role',  'Sr. Accountant'))).strip()
                        if _ip in ['nan','None','']: _ip = str(_row.get('Proc Role','Accountant I')).strip()
                        if _ir in ['nan','None','']: _ir = str(_row.get('Rev Role','Sr. Accountant')).strip()

                    _ptix = float(_row.get('Closed tickets with Proc time', 0) or 0)
                    _rtix = float(_row.get('Closed tickets with rev time',  0) or 0)
                    _paht = float(_row.get('>>> FINAL Capacity Proc AHT', 0) or 0)
                    _raht = float(_row.get('>>> FINAL Capacity Rev AHT',  0) or 0)

                    _as = _m1_start if pd.isna(_gl) else max(_m1_start, _gl)
                    _ae = _m1_end   if pd.isna(_fsd) else min(_m1_end, _fsd)
                    if _as <= _ae:
                        _da = np.busday_count(_as.strftime('%Y-%m-%d'), (_ae + pd.Timedelta(days=1)).strftime('%Y-%m-%d'))
                        _apct = max(0.0, min(1.0, _da / _wdays_m1)) if _wdays_m1 > 0 else 0.0
                    else:
                        _apct = 0.0

                    _lc = 1.0
                    if pd.notna(_gl) and _apct > 0:
                        _md = (_m1_start.year - _gl.year) * 12 + (_m1_start.month - _gl.month)
                        if   _md == 0: _lc = 1.17
                        elif _md == 1: _lc = 0.86
                        elif _md == 2: _lc = 0.99

                    _bp = (_ptix * _apct * _paht * _lc) / 60
                    _br = (_rtix * _apct * _raht * _lc) / 60
                    _up = utilization_map.get(_ip, util_acc1)
                    _ur = utilization_map.get(_ir, util_sr)
                    _tp = _bp + (_bp * (1 - _up)) + (_bp * absenteeism) + (_bp * attrition)
                    _tr = _br + (_br * (1 - _ur)) + (_br * absenteeism) + (_br * attrition)

                    # Resolve processor / reviewer names and emails.
                    # The Master DB 'processor' / 'reviewer' columns may contain either:
                    #   (a) an email address directly  → maria.tamayo@proper.ai
                    #   (b) a full name                → Maria Tamayo
                    # Detect by presence of '@' and handle both cases.
                    _proc_raw  = str(_get_ci(_row, 'processor')).strip()
                    _rev_raw   = str(_get_ci(_row, 'reviewer')).strip()
                    if _proc_raw.lower() in ('nan', 'none'): _proc_raw = ''
                    if _rev_raw.lower()  in ('nan', 'none'): _rev_raw  = ''

                    if '@' in _proc_raw:                      # field already is an email
                        _proc_email = _proc_raw.lower()
                        _proc_name  = _e2name.get(_proc_email, _proc_raw)
                    else:                                     # field is a name
                        _proc_name  = _proc_raw
                        _proc_email = str(_get_ci(_row, 'processor email')).strip()
                        if not _proc_email or _proc_email.lower() in ('nan', 'none'):
                            _proc_email = _n2email.get(_proc_name.lower(), '')

                    if '@' in _rev_raw:
                        _rev_email = _rev_raw.lower()
                        _rev_name  = _e2name.get(_rev_email, _rev_raw)
                    else:
                        _rev_name  = _rev_raw
                        _rev_email = str(_get_ci(_row, 'reviewer email')).strip()
                        if not _rev_email or _rev_email.lower() in ('nan', 'none'):
                            _rev_email = _n2email.get(_rev_name.lower(), '')

                    _audit_rows.append({
                        'POD':                          _pod,
                        'Sr. Accountant':               _sr,
                        'Client':                       _cli,
                        'Process':                      str(_row.get('type',    '')).strip(),
                        'Sub-process':                  str(_row.get('subtype', '')).strip(),
                        'Processor':                    _proc_name,
                        'Processor Email':              _proc_email,
                        'Processor Role':               _ip,
                        'Reviewer':                     _rev_name,
                        'Reviewer Email':               _rev_email,
                        'Reviewer Role':                _ir,
                        'Closed Tix (Proc)':            round(_ptix, 0),
                        'AHT Proc (min)':               round(_paht, 2),
                        'Closed Tix (Rev)':             round(_rtix, 0),
                        'AHT Rev (min)':                round(_raht, 2),
                        'MRR ($)':                      round(_mrr, 2),
                        'Res Doors':                    _row.get('Res doors', ''),
                        'Res Prop':                     _row.get('Res Prop', ''),
                        'Comm Doors':                   _row.get('Commercial Doors', ''),
                        'Comm Properties':              _row.get('Commercial Properties', ''),
                        'SQFT':                         _row.get('SQFT Commercial', ''),
                        'Corp Books':                   _row.get('Corp Books', ''),
                        'PMS':                          str(_row.get('PMS', '')).strip(),
                        'Go Live':                      str(_gl)[:10] if pd.notna(_gl) else '',
                        'Final Service Date':           str(_fsd)[:10] if pd.notna(_fsd) else '',
                        'Active % (M1)':                round(_apct * 100, 1),
                        'Learning Curve (M1)':          round(_lc, 2),
                        'Prod Hrs Proc (M1)':           round(_bp, 2),
                        'Prod Hrs Rev (M1)':            round(_br, 2),
                        'Total Hrs Proc w/ Shrinkage':  round(_tp, 2),
                        'Total Hrs Rev w/ Shrinkage':   round(_tr, 2),
                        'Total Hrs (M1)':               round(_tp + _tr, 2),
                        'FTEs (M1)':                    round((_tp + _tr) / _hrs_fte_m1, 4) if _hrs_fte_m1 > 0 else 0,
                        'Util Rate Proc':               round(_up * 100, 1),
                        'Util Rate Rev':                round(_ur * 100, 1),
                        'Absenteeism Rate':             round(absenteeism * 100, 1),
                        'Attrition Rate':               round(attrition * 100, 1),
                        'Working Days (M1)':            _wdays_m1,
                    })

                df_baseline_audit = pd.DataFrame(_audit_rows)

                # 6. SUMMARY BY POD
                todas_cols     = [c for c in df_resumen.columns if "M" in c]
                df_pod_roles   = df_resumen.groupby(['POD', 'Required Role'])[todas_cols].sum().reset_index()
                df_pod_totales = df_resumen.groupby(['POD'])[todas_cols].sum().reset_index()
                df_pod_totales['Required Role'] = '>>> POD TOTAL'
                df_pod_final   = (
                    pd.concat([df_pod_roles, df_pod_totales], ignore_index=True)
                    .sort_values(by=['POD', 'Required Role'])
                    .reset_index(drop=True)
                )

                # 7. EXECUTIVE GENERAL DASHBOARD
                resumen_ejecutivo        = []
                _pod_churn_store         = {}   # {mes_str: {pod_name: churn_hrs (required)}}
                _pod_churn_prod_store    = {}   # {mes_str: {pod_name: churn_hrs (productive)}}
                _pod_new_hrs_store       = {}   # {mes_str: {pod_name: new_client_hrs (required)}}
                _pod_new_prod_hrs_store  = {}   # {mes_str: {pod_name: new_client_hrs (productive)}}
                _pod_new_mrr_store       = {}   # {mes_str: {pod_name: new_mrr}}
                _pod_churn_mrr_store     = {}   # {mes_str: {pod_name: churn_mrr}}

                # Build client→POD mapping once (for per-POD MRR look-ups)
                _cli_pod_map = {}
                if 'Client' in df_resumen.columns and 'POD' in df_resumen.columns:
                    _tmp_cpod = df_resumen.dropna(subset=['Client']).groupby('Client')['POD'].first()
                    _cli_pod_map = {str(k).strip().lower(): str(v) for k, v in _tmp_cpod.items()}

                for i, mes_str in enumerate(meses_proyeccion):
                    c_base      = f"M{i+1} ({mes_str}) - Base Hours"
                    c_prod      = f"M{i+1} ({mes_str}) - Productive Hours"
                    c_post      = f"M{i+1} ({mes_str}) - Post-Auto Hours"
                    c_save_hrs  = f"M{i+1} ({mes_str}) - Auto Saving (Hrs)"
                    c_save_usd  = f"M{i+1} ({mes_str}) - Auto Saving ($)"
                    c_plus      = f"M{i+1} ({mes_str}) - Adjustments (+) Hrs"
                    c_minus     = f"M{i+1} ({mes_str}) - Adjustments (-) Hrs"
                    c_final     = f"M{i+1} ({mes_str}) - Final Hours"
                    c_fte       = f"M{i+1} ({mes_str}) - Final FTEs"

                    tot_base      = df_resumen[c_base].sum()     if c_base     in df_resumen.columns else 0
                    tot_save_hrs  = df_resumen[c_save_hrs].sum() if c_save_hrs in df_resumen.columns else 0
                    tot_save_usd  = df_resumen[c_save_usd].sum() if c_save_usd in df_resumen.columns else 0
                    tot_plus      = df_resumen[c_plus].sum()     if c_plus     in df_resumen.columns else 0
                    tot_minus     = df_resumen[c_minus].sum()    if c_minus    in df_resumen.columns else 0
                    tot_final     = df_resumen[c_final].sum()    if c_final    in df_resumen.columns else 0
                    tot_fte       = df_resumen[c_fte].sum()      if c_fte      in df_resumen.columns else 0

                    fte_acc1 = df_resumen[df_resumen['Required Role'] == 'Accountant I'][c_fte].sum()
                    fte_acc2 = df_resumen[df_resumen['Required Role'] == 'Accountant II'][c_fte].sum()
                    fte_gen  = df_resumen[df_resumen['Required Role'] == 'General Accountant'][c_fte].sum()
                    fte_sr   = df_resumen[df_resumen['Required Role'] == 'Sr. Accountant'][c_fte].sum()

                    mes_date = today + relativedelta(months=_month_offsets[i])
                    start_m  = pd.Timestamp(mes_date.replace(day=1).date())
                    end_m    = pd.Timestamp((start_m + relativedelta(months=1) - relativedelta(days=1)).date())

                    mask_active_mrr = (
                        (df_clients_unique['Go Live'] <= end_m)   | df_clients_unique['Go Live'].isna()
                    ) & (
                        (df_clients_unique['Final Service Date'] >= start_m) | df_clients_unique['Final Service Date'].isna()
                    )
                    total_mrr = df_clients_unique.loc[mask_active_mrr, 'MRR'].sum()

                    mask_churn  = (df_clients_unique['Final Service Date'] >= start_m) & (df_clients_unique['Final Service Date'] <= end_m)
                    churn_count = mask_churn.sum()
                    churn_mrr   = df_clients_unique.loc[mask_churn, 'MRR'].sum()

                    # Hours from churning clients (base hours for clients leaving this month)
                    _churn_names = set(
                        df_clients_unique.loc[mask_churn, 'client_name']
                        .astype(str).str.strip().str.lower()
                    )
                    _churn_row_mask = (
                        df_resumen['Client'].astype(str).str.strip().str.lower().isin(_churn_names)
                    ) if _churn_names else pd.Series(False, index=df_resumen.index)
                    # Use c_final so AI-predicted churning clients (c_base=0) are counted correctly
                    churn_hrs_val = (
                        df_resumen.loc[_churn_row_mask, c_final].sum()
                        if c_final in df_resumen.columns and _churn_names else 0.0
                    )
                    churn_prod_hrs_val = (
                        df_resumen.loc[_churn_row_mask, c_post].sum()
                        if c_post in df_resumen.columns and _churn_names else 0.0
                    )

                    # Per-POD churn hours for the Capacity Overview POD tabs
                    _pod_churn_mes      = {}
                    _pod_churn_prod_mes = {}
                    if _churn_names and 'POD' in df_resumen.columns:
                        for _pn in df_resumen['POD'].dropna().astype(str).unique():
                            _pmask = _churn_row_mask & (df_resumen['POD'].astype(str) == _pn)
                            _pod_churn_mes[_pn]      = float(df_resumen.loc[_pmask, c_final].sum()) if c_final in df_resumen.columns else 0.0
                            _pod_churn_prod_mes[_pn] = float(df_resumen.loc[_pmask, c_post].sum())  if c_post  in df_resumen.columns else 0.0
                    _pod_churn_store[mes_str]      = _pod_churn_mes
                    _pod_churn_prod_store[mes_str] = _pod_churn_prod_mes

                    mask_new     = (df_clients_unique['Go Live'] >= start_m) & (df_clients_unique['Go Live'] <= end_m)
                    new_count    = mask_new.sum()
                    new_mrr      = df_clients_unique.loc[mask_new, 'MRR'].sum()

                    # Hours from new clients going live this month
                    # Use c_final (not c_base) so AI-predicted clients are included.
                    # AI clients only exist in df_resumen_auto → c_base = 0 after outer merge,
                    # but c_final = c_post + adjustments is correctly non-zero.
                    _new_names = set(
                        df_clients_unique.loc[mask_new, 'client_name']
                        .astype(str).str.strip().str.lower()
                    )
                    _new_row_mask = (
                        df_resumen['Client'].astype(str).str.strip().str.lower().isin(_new_names)
                    ) if _new_names else pd.Series(False, index=df_resumen.index)
                    new_hrs_val = float(
                        df_resumen.loc[_new_row_mask, c_final].sum()
                    ) if c_final in df_resumen.columns else 0.0
                    new_prod_hrs_val = float(
                        df_resumen.loc[_new_row_mask, c_post].sum()
                    ) if c_post in df_resumen.columns else 0.0

                    # Per-POD new hours, new MRR, churn MRR
                    _all_pods = (df_resumen['POD'].dropna().astype(str).unique().tolist()
                                 if 'POD' in df_resumen.columns else [])
                    _pod_new_mes, _pod_new_prod_mes, _pod_new_mrr_mes, _pod_churn_mrr_mes = {}, {}, {}, {}
                    for _pn in _all_pods:
                        _pod_mask = df_resumen['POD'].astype(str) == _pn
                        _pod_new_mes[_pn] = float(
                            df_resumen.loc[_new_row_mask & _pod_mask, c_final].sum()
                        ) if c_final in df_resumen.columns else 0.0
                        _pod_new_prod_mes[_pn] = float(
                            df_resumen.loc[_new_row_mask & _pod_mask, c_post].sum()
                        ) if c_post in df_resumen.columns else 0.0
                        _pn_new_cli   = {k for k, v in _cli_pod_map.items() if v == _pn} & _new_names
                        _pn_churn_cli = {k for k, v in _cli_pod_map.items() if v == _pn} & _churn_names
                        _pod_new_mrr_mes[_pn] = float(
                            df_clients_unique.loc[
                                mask_new & df_clients_unique['client_name']
                                .astype(str).str.strip().str.lower().isin(_pn_new_cli), 'MRR'
                            ].sum()
                        )
                        _pod_churn_mrr_mes[_pn] = float(
                            df_clients_unique.loc[
                                mask_churn & df_clients_unique['client_name']
                                .astype(str).str.strip().str.lower().isin(_pn_churn_cli), 'MRR'
                            ].sum()
                        )
                    _pod_new_hrs_store[mes_str]      = _pod_new_mes
                    _pod_new_prod_hrs_store[mes_str] = _pod_new_prod_mes
                    _pod_new_mrr_store[mes_str]      = _pod_new_mrr_mes
                    _pod_churn_mrr_store[mes_str]    = _pod_churn_mrr_mes

                    prod_hrs        = monthly_prod_hrs[i]
                    u_hrs, a_hrs, att_hrs = monthly_util_hrs[i], monthly_abs_hrs[i], monthly_att_hrs[i]
                    shrinkage_total = u_hrs + a_hrs + att_hrs

                    resumen_ejecutivo.append({
                        "Projected Month":                   mes_str,
                        "Total MRR ($)":                     round(total_mrr, 2),
                        "Working Days (Used)":               dict_workable_days[i],
                        "1. Productive Hours (Pure Base)":   round(prod_hrs, 2),
                        "2. Total Shrinkage (Hrs)":          round(shrinkage_total, 2),
                        "3. Total Hours (Pre-Auto)":         round(tot_base, 2),
                        "4. Automation Saving (Hrs)":        round(tot_save_hrs, 2),
                        "4.1 Cost Saving ($)":               round(tot_save_usd, 2),
                        "5. Manual Adjustments (+) Hrs":     round(tot_plus, 2),
                        "6. Manual Adjustments (-) Hrs":     round(tot_minus, 2),
                        "7. Total Required Hours (Final)":   round(tot_final, 2),
                        "Total FTEs":                        round(tot_fte, 2),
                        "FTEs Accountant I":                 round(fte_acc1, 2),
                        "FTEs Accountant II":                round(fte_acc2, 2),
                        "FTEs General Acc.":                 round(fte_gen, 2),
                        "FTEs Sr. Accountant":               round(fte_sr, 2),
                        "New MRR ($)":                       round(new_mrr, 2),
                        "New Clients (Go Live)":             new_count,
                        "New Clients Hours":                 round(new_hrs_val, 2),
                        "New Clients Prod Hours":            round(new_prod_hrs_val, 2),
                        "Lost MRR (Churn) ($)":              round(churn_mrr, 2),
                        "Clients Ending (#)":                churn_count,
                        "Confirmed Churn (Hrs)":             round(churn_hrs_val, 2),
                        "Confirmed Churn Prod Hrs":          round(churn_prod_hrs_val, 2),
                    })

                df_resumen_general = pd.DataFrame(resumen_ejecutivo)

                st.session_state.final_dashboards = {
                    'general':    df_resumen_general,
                    'pod':        df_pod_final,
                    'cliente':    df_resumen,
                    'baseline':   df_baseline_audit,
                    'pod_churn':          _pod_churn_store,         # {mes_str: {pod_name: churn_hrs (required)}}
                    'pod_churn_prod':     _pod_churn_prod_store,    # {mes_str: {pod_name: churn_hrs (productive)}}
                    'pod_new_hrs':        _pod_new_hrs_store,       # {mes_str: {pod_name: new_client_hrs (required)}}
                    'pod_new_prod_hrs':   _pod_new_prod_hrs_store,  # {mes_str: {pod_name: new_client_hrs (productive)}}
                    'pod_new_mrr':        _pod_new_mrr_store,       # {mes_str: {pod_name: new_mrr}}
                    'pod_churn_mrr': _pod_churn_mrr_store, # {mes_str: {pod_name: churn_mrr}}
                }
                # ── Tag the role mode used for this cascade run ───────────────────────
                _mode_tag = 'real' if _s3_use_real_cas else 'ideal'
                st.session_state['_cascade_role_mode'] = _mode_tag
                st.session_state[f'final_dashboards_{_mode_tag}'] = st.session_state.final_dashboards

                # ── Build Client MRR per month table ──────────────────────────────────
                _duc_mrr = st.session_state.get('df_clients_unique', pd.DataFrame()).copy()
                if not _duc_mrr.empty and 'client_name' in _duc_mrr.columns:
                    # Join POD from df_resumen (POD→Client map, deduplicated).
                    # Drop any existing POD column first to avoid pandas creating
                    # POD_x / POD_y conflicts that make _crow.get('POD') return ''.
                    if 'POD' in df_resumen.columns and 'Client' in df_resumen.columns:
                        _duc_mrr = _duc_mrr.drop(columns=[c for c in ['POD'] if c in _duc_mrr.columns])
                        _cli_pod_map_mrr = (
                            df_resumen[['POD', 'Client']].drop_duplicates()
                            .rename(columns={'Client': 'client_name'})
                        )
                        _duc_mrr = _duc_mrr.merge(_cli_pod_map_mrr, on='client_name', how='left')
                    _cli_mrr_rows = []
                    for _, _crow in _duc_mrr.iterrows():
                        _cm_name = str(_crow.get('client_name', '')).strip()
                        _cm_mrr  = float(_crow.get('MRR', 0) or 0)
                        _cm_gl   = pd.to_datetime(_crow.get('Go Live'), errors='coerce')
                        _cm_fsd  = pd.to_datetime(_crow.get('Final Service Date'), errors='coerce')
                        _cm_pod  = str(_crow.get('POD', '')).strip()
                        _mrr_row = {'Client': _cm_name, 'POD': _cm_pod, 'MRR (Base)': _cm_mrr}
                        for _moi, _msi in enumerate(meses_proyeccion):
                            _mdi  = today + relativedelta(months=_month_offsets[_moi])
                            _smi  = pd.Timestamp(_mdi.replace(day=1).date())
                            _emi  = pd.Timestamp((_smi + relativedelta(months=1) - relativedelta(days=1)).date())
                            _active = (
                                (pd.isna(_cm_gl)  or _cm_gl  <= _emi) and
                                (pd.isna(_cm_fsd) or _cm_fsd >= _smi)
                            )
                            _mrr_row[_msi] = round(_cm_mrr, 2) if _active else 0.0
                        _cli_mrr_rows.append(_mrr_row)
                    df_client_mrr = pd.DataFrame(_cli_mrr_rows)
                    _cmrr_sort = [c for c in ['POD', 'Client'] if c in df_client_mrr.columns]
                    df_client_mrr = df_client_mrr.sort_values(_cmrr_sort).reset_index(drop=True)
                else:
                    df_client_mrr = pd.DataFrame()
                st.session_state.final_dashboards['client_mrr'] = df_client_mrr

                # ── Save filter context for view-level rendering ──────────────────────
                # Always read from session state so this works even in pipeline flow
                # where the uploaded-file block (and its local variables) never ran.
                _cas_pods    = st.session_state.get('_filt_pods',    selected_pods)
                _cas_srs     = st.session_state.get('_filt_srs',     selected_srs)
                _cas_clients = st.session_state.get('_filt_clients', selected_clients_final)

                st.session_state['_dash_sel_pods']    = list(_cas_pods)
                st.session_state['_dash_sel_srs']     = list(_cas_srs)
                st.session_state['_dash_sel_clients'] = list(_cas_clients)

                # Compute dashboard view level from applied filters
                _n_sc = len(_cas_clients)
                # Sr. filter takes priority — client list auto-populates from Sr. selection
                if _cas_srs and not _cas_pods:
                    _dlvl = 'sr'
                elif len(_cas_pods) == 1 and not _cas_srs:
                    _dlvl = 'pod'
                elif len(_cas_pods) > 1 and not _cas_srs:
                    _dlvl = 'multi_pod'
                elif _n_sc == 1 and not _cas_srs and not _cas_pods:
                    _dlvl = 'client'
                elif _n_sc > 1 and not _cas_srs and not _cas_pods:
                    _aff_pods = df[df['client_name'].isin(_cas_clients)]['POD'].dropna().unique().tolist() if 'POD' in df.columns else []
                    _dlvl = 'multi_pod' if len(_aff_pods) > 1 else ('pod' if len(_aff_pods) == 1 else 'overall')
                else:
                    _dlvl = 'overall'
                st.session_state['_dash_level'] = _dlvl
                # Bump version so waterfall cache knows to rebuild
                st.session_state['_fd_version'] = st.session_state.get('_fd_version', 0) + 1
                # Collapse Step 3 after cascade completes so dashboards come into view
                st.session_state['_s3_exp_open'] = False

    # ==========================================
    # RENDER DASHBOARDS IF AVAILABLE
    # ==========================================
    if "final_dashboards" in st.session_state:
        st.success("✅ Dashboards and Final Reports Generated Successfully!")

        # ── View mode selector ────────────────────────────────────────────────────
        _view_mode = st.radio(
            "Select view:",
            ["⚡ Quick Overview", "📊 Complete View", "🎮 Complete + Playground"],
            index=1,          # default to Complete View
            horizontal=True,
            key="view_mode_radio",
        )

        # ── View level context (populated when cascade runs) ──────────────────────
        _dash_level       = st.session_state.get('_dash_level', 'overall')
        _dash_sel_pods    = st.session_state.get('_dash_sel_pods', [])
        _dash_sel_srs     = st.session_state.get('_dash_sel_srs', [])
        _dash_sel_clients = st.session_state.get('_dash_sel_clients', [])
        # Hide Actual HC only when clients are the sole filter (no POD / Sr. context available)
        _hide_actual_hc   = bool(_dash_sel_clients) and not _dash_sel_pods and not _dash_sel_srs

        _level_label_map = {
            'overall':   '🌎 Overall — all data',
            'pod':       '🚀 POD View',
            'sr':        '👤 Sr. Accountant View',
            'client':    '👤 Single Client View',
            'multi_pod': '🌎 Multi-POD View',
        }
        if _dash_level != 'overall':
            _flbl = []
            if _dash_sel_pods:    _flbl.append(f"POD: {', '.join(_dash_sel_pods)}")
            if _dash_sel_srs:     _flbl.append(f"Sr: {', '.join(_dash_sel_srs)}")
            if _dash_sel_clients: _flbl.append(f"{len(_dash_sel_clients)} client(s) selected")
            st.info(f"{_level_label_map.get(_dash_level, '')}  •  {' | '.join(_flbl)}")
        if _hide_actual_hc:
            st.warning("⚠️ Client filter active — Actual HC indicators are hidden (HC is tracked at team level, not per client).")

        # ── Tab jump dropdown ─────────────────────────────────────────────────
        _s3_tab_options = [
            "📋 Capacity Overview",
            "🌎 General Waterfall Summary",
            "🚀 Summary by POD (Cascade)",
            "🏢 POD × Sr. Accountant",
            "📊 Client & Role Summary (Cascade)",
            "💰 Client MRR by Month",
            "🔬 Baseline Audit",
            "👤 Employee Level",
        ]
        _s3_tab_prev = st.session_state.get('_s3_tab_jump_prev')
        _s3_tab_sel  = st.selectbox(
            "Go to tab:",
            _s3_tab_options,
            key="s3_tab_jump",
            label_visibility="collapsed",
        )
        if _s3_tab_sel != _s3_tab_prev:
            st.session_state['_s3_tab_jump_prev'] = _s3_tab_sel
            import streamlit.components.v1 as _stc_s3
            _stc_s3.html(f"""<script>
(function() {{
    var target = {json.dumps(_s3_tab_sel)};
    function tryClick(n) {{
        var btns = window.parent.document.querySelectorAll('[data-testid="stTabs"] button[role="tab"]');
        for (var i = 0; i < btns.length; i++) {{
            if (btns[i].innerText.trim() === target) {{ btns[i].click(); return; }}
        }}
        if (n < 8) setTimeout(function() {{ tryClick(n + 1); }}, 150);
    }}
    tryClick(0);
}})();
</script>""", height=0)

        # (Tab scrolling handled by global CSS at page top)
        t_overview, t_gral, t_pod, t_pod_sr, t_cli, t_cli_mrr, t_baseline, t_employee = st.tabs([
            "📋 Capacity Overview",
            "🌎 General Waterfall Summary",
            "🚀 Summary by POD (Cascade)",
            "🏢 POD × Sr. Accountant",
            "📊 Client & Role Summary (Cascade)",
            "💰 Client MRR by Month",
            "🔬 Baseline Audit",
            "👤 Employee Level",
        ])

        # Column configs for monetary/FTE columns in the general dashboard
        _money_col  = lambda label: st.column_config.NumberColumn(label, format="$%.2f")
        _number_col = lambda label: st.column_config.NumberColumn(label, format="%.2f")

        general_col_cfg = {
            "Total MRR ($)":            _money_col("Total MRR ($)"),
            "4.1 Cost Saving ($)":      _money_col("4.1 Cost Saving ($)"),
            "New MRR ($)":              _money_col("New MRR ($)"),
            "Lost MRR (Churn) ($)":     _money_col("Lost MRR (Churn) ($)"),
        }

        # Build column config for client-level table ($ saving columns)
        cliente_col_cfg = {}
        for i, mes_str in enumerate(meses_proyeccion):
            col_usd = f"M{i+1} ({mes_str}) - Auto Saving ($)"
            cliente_col_cfg[col_usd] = _money_col(col_usd)

        # ── CAPACITY OVERVIEW (WATERFALL VIEW) ───────────────────────────────────
        with t_overview:
            # ── Quick Overview (compact KPI table shown when Quick mode is selected) ─
            if _view_mode == "⚡ Quick Overview":
                # Sr. level uses the Sr. waterfall; all others use the overall waterfall
                _qv_src = (
                    st.session_state.get('_wf_sr_export')
                    if _dash_level == 'sr'
                    else st.session_state.get('_wf_overall_export')
                )
                if _qv_src is not None:
                    st.markdown("### ⚡ Quick Overview")
                    st.caption("Key metrics across projected months. Switch to **Complete View** for the full waterfall breakdown.")

                    # Rows to display — full set when HC available, reduced when client filter active
                    _QV_ALL = [
                        "━ Required Hours",
                        "━ MRR ($)",
                        "━ Required HC (FTEs)",
                        "━ Actual HC (Report)",
                        "━ HC Δ (Actual − Required)",
                        "  Expected Margin ($)",
                        "  Expected Margin (%)",
                        "  Revenue / HC ($)",
                    ]
                    _QV_NO_HC = [
                        "━ Required Hours",
                        "━ MRR ($)",
                        "━ Required HC (FTEs)",
                        "  Capacity Margin ($)",
                        "  Capacity Margin (%)",
                    ]
                    # _qv_src always contains the full waterfall — filter rows here for display
                    _qv_rows_target = _QV_NO_HC if _hide_actual_hc else _QV_ALL
                    _qv_rows_exist  = [r for r in _qv_rows_target if r in _qv_src.index]
                    _df_qv = _qv_src.loc[_qv_rows_exist]

                    # Clean display labels (remove leading dashes/spaces)
                    _df_qv = _df_qv.copy()
                    _df_qv.index = (
                        _df_qv.index
                        .str.replace(r'^[━·\s(+)\-]+\s*', '', regex=True)
                        .str.strip()
                    )

                    st.dataframe(
                        _df_qv,
                        use_container_width=True,
                        height=len(_df_qv) * 35 + 38,
                    )
                else:
                    st.info("⏳ Quick Overview builds on first load — switch to **Complete View** once, then return here.")

            # ── Full waterfall expander ──────────────────────────────────────────────
            _show_full = _view_mode in ["📊 Complete View", "🎮 Complete + Playground"]
            with st.expander("📋 Capacity Overview — Waterfall", expanded=_show_full):
                st.caption(
                    "Baseline month + 5 projections. "
                    "Upload HC Weekly Report in the sidebar to see Actual HC and over/under. "
                    "Run AI Prediction to include New Customers hours."
                )

                _exec   = st.session_state.final_dashboards['general']
                _pod_df = st.session_state.final_dashboards['pod']
                _cli_df = st.session_state.final_dashboards['cliente']
                _hc     = st.session_state.get('hc_data', None)
                _ai     = st.session_state.get('ai_results', None)
                _df_raw = st.session_state.get('df_clean', pd.DataFrame())
                _duc    = st.session_state.get('df_clients_unique', pd.DataFrame())

                # ── Aggregate metrics from raw data ─────────────────────────────────
                def _safe_num(series):
                    return pd.to_numeric(series, errors='coerce').fillna(0)

                if not _df_raw.empty:
                    _client_snap = _df_raw.groupby('client_name', as_index=False).agg({
                        c: 'first' for c in ['Res Prop','Commercial Properties','Res doors','Commercial Doors','SQFT Commercial']
                        if c in _df_raw.columns
                    })
                    _res_prop_count  = int(_safe_num(_client_snap.get('Res Prop', 0)).sum())
                    _comm_prop_count = int(_safe_num(_client_snap.get('Commercial Properties', 0)).sum())
                    _prop_count      = _res_prop_count + _comm_prop_count
                    _res_door_count  = int(_safe_num(_client_snap.get('Res doors', 0)).sum())
                    _comm_door_count = int(_safe_num(_client_snap.get('Commercial Doors', 0)).sum())
                    _door_count      = _res_door_count + _comm_door_count
                    _sqft_count      = int(_safe_num(_client_snap.get('SQFT Commercial', 0)).sum())
                else:
                    _res_prop_count = _comm_prop_count = _prop_count = 0
                    _res_door_count = _comm_door_count = _door_count = _sqft_count = 0

                # (AI results kept in _ai for other uses, but not shown in the waterfall rows)
                _ai_base_wdays = st.session_state.get('calc_data', {}).get('dict_workable_days', {}).get(0, 21)

                # ── Scope HC report to cascade-selected PODs (multi_pod case) ──────
                # When the cascade ran for a subset of PODs, the Overall waterfall must
                # use only those PODs' employees from the HC report — not the full roster.
                # When no POD filter is active (_dash_sel_pods empty) OR the Overall
                # cascade was run, _hc['total'] / _hc['by_role'] are used as-is.
                _hc_wf = _hc  # default: use full HC data
                if _hc and _dash_sel_pods:
                    _hbpr = _hc.get('by_pod_role', pd.DataFrame())
                    if not _hbpr.empty:
                        def _nhp_ov(s): return str(s).lower().replace(' ', '').strip()
                        _sel_pods_norm = {_nhp_ov(p) for p in _dash_sel_pods}
                        _hbpr_mask = _hbpr['POD'].apply(_nhp_ov).isin(_sel_pods_norm)
                        _hbpr_filt = _hbpr[_hbpr_mask]
                        _sc_by_role = _hbpr_filt.groupby('Capacity Role')['HC'].sum().to_dict()
                        _sc_total   = sum(v for k, v in _sc_by_role.items() if k != 'Other')
                        _sc_mgr_pd  = {
                            k: v for k, v in _hc.get('mgr_by_pod', {}).items()
                            if _nhp_ov(str(k)) in _sel_pods_norm
                        }
                        _sc_mgr_tot = sum(_sc_mgr_pd.values())
                        # Build a lightweight "scoped HC" dict with the same keys
                        _hc_wf = dict(_hc)   # shallow copy
                        _hc_wf['by_role']   = _sc_by_role
                        _hc_wf['total']     = _sc_total
                        _hc_wf['mgr_total'] = _sc_mgr_tot

                # ── Helper to format by metric type ─────────────────────────────────
                def _fmt(val, kind):
                    if val is None: return "—"
                    try:
                        v = float(val)
                    except (TypeError, ValueError):
                        return str(val)
                    if kind == '$':   return f"${v:,.2f}"
                    if kind == '%':   return f"{v:.2f}%"
                    if kind == 'fte': return f"{v:.2f}"
                    if kind == 'dec': return f"{v:.2f}"
                    return f"{v:,.0f}"

                # ── Build the waterfall rows ─────────────────────────────────────────
                def _build_overall_wf():
                    rows   = {}
                    months = []
                    _wdays = st.session_state.get('calc_data', {}).get('dict_workable_days', {})

                    for i, mes_str in enumerate(meses_proyeccion):
                        if i >= len(_exec): break
                        r = _exec.iloc[i]
                        months.append(mes_str)

                        wdays        = _wdays.get(i, 21)
                        current_hrs  = r.get("3. Total Hours (Pre-Auto)", 0)
                        auto_hrs     = r.get("4. Automation Saving (Hrs)", 0)
                        churn_hrs    = r.get("Confirmed Churn (Hrs)", 0)
                        new_cli_hrs  = r.get("New Clients Hours", 0)
                        new_mrr_val  = r.get("New MRR ($)", 0)
                        churn_mrr_val = r.get("Lost MRR (Churn) ($)", 0)
                        adj_plus     = r.get("5. Manual Adjustments (+) Hrs", 0)
                        adj_minus    = r.get("6. Manual Adjustments (-) Hrs", 0)
                        final_hrs    = r.get("7. Total Required Hours (Final)", 0)
                        prod_hrs_b   = r.get("1. Productive Hours (Pure Base)", 0)
                        fte_total    = r.get("Total FTEs", 0)
                        fte_acc1     = r.get("FTEs Accountant I", 0)
                        fte_acc2     = r.get("FTEs Accountant II", 0)
                        fte_gen      = r.get("FTEs General Acc.", 0)
                        fte_sr       = r.get("FTEs Sr. Accountant", 0)
                        mrr          = r.get("Total MRR ($)", 0)
                        workdays_val = r.get("Working Days (Used)", wdays)
                        holidays_val = holidays_per_month.get(mes_str, 0)

                        cap_prod = (prod_hrs_b / current_hrs * 100) if current_hrs > 0 else 0

                        hc_total = _hc_wf['total']              if _hc_wf else None
                        hc_acc1  = _hc_wf['by_role'].get('Accountant I', 0)      if _hc_wf else None
                        hc_acc2  = _hc_wf['by_role'].get('Accountant II', 0)     if _hc_wf else None
                        hc_gen   = _hc_wf['by_role'].get('General Accountant', 0) if _hc_wf else None
                        hc_sr    = _hc_wf['by_role'].get('Sr. Accountant', 0)    if _hc_wf else None
                        hc_other = _hc_wf['by_role'].get('Other', 0)             if _hc_wf else None
                        hc_mgr   = _hc_wf.get('mgr_total', 0)                    if _hc_wf else None

                        act_hc_prod = (
                            (float(hc_acc1 or 0) * util_acc1 +
                             float(hc_acc2 or 0) * util_acc1 +
                             float(hc_gen  or 0) * util_gen  +
                             float(hc_sr   or 0) * util_sr)
                            / float(hc_total) * 100
                        ) if hc_total and float(hc_total) > 0 else None

                        d_total  = round(hc_total - fte_total, 2) if hc_total is not None else None
                        d_acc1   = round(hc_acc1  - fte_acc1,  2) if hc_acc1  is not None else None
                        d_acc2   = round(hc_acc2  - fte_acc2,  2) if hc_acc2  is not None else None
                        d_gen    = round(hc_gen   - fte_gen,   2) if hc_gen   is not None else None
                        d_sr     = round(hc_sr    - fte_sr,    2) if hc_sr    is not None else None

                        # Revenue / HC uses accounting staff only (Acc I–Sr.), excluding managers
                        hc_staff = (
                            float(hc_acc1 or 0) + float(hc_acc2 or 0) +
                            float(hc_gen  or 0) + float(hc_sr   or 0)
                        ) if _hc else None
                        rev_per_hc = mrr / hc_staff if (hc_staff and hc_staff > 0) else None

                        cap_cost = (
                            float(fte_acc1 or 0) * cost_acc1 +
                            float(fte_acc2 or 0) * cost_acc2 +
                            float(fte_gen  or 0) * cost_gen  +
                            float(fte_sr   or 0) * cost_sr
                        )
                        cap_margin     = float(mrr or 0) - cap_cost
                        cap_margin_pct = (cap_margin / float(mrr) * 100) if mrr and float(mrr) != 0 else None
                        if hc_total is not None:
                            exp_cost = (
                                float(hc_acc1 or 0) * cost_acc1 +
                                float(hc_acc2 or 0) * cost_acc2 +
                                float(hc_gen  or 0) * cost_gen  +
                                float(hc_sr   or 0) * cost_sr
                            )
                            exp_margin     = float(mrr or 0) - exp_cost
                            exp_margin_pct = (exp_margin / float(mrr) * 100) if mrr and float(mrr) != 0 else None
                        else:
                            exp_cost = exp_margin = exp_margin_pct = None

                        shrinkage_hrs = current_hrs - prod_hrs_b

                        col = mes_str
                        rows.setdefault("━ Required Hours",               {})[col] = _fmt(final_hrs, 'n')
                        rows.setdefault("  Current Customer Hours",       {})[col] = _fmt(prod_hrs_b, 'n')
                        rows.setdefault("  Shrinkage (Hrs)",              {})[col] = _fmt(shrinkage_hrs if shrinkage_hrs > 0 else None, 'n')
                        rows.setdefault("  (+) New Customer Hours",       {})[col] = _fmt(new_cli_hrs if new_cli_hrs else None, 'n')
                        rows.setdefault("  (-) Confirmed Churn (Hrs)",    {})[col] = _fmt(churn_hrs if churn_hrs else None, 'n')
                        rows.setdefault("  (-) Automations",              {})[col] = _fmt(auto_hrs if auto_hrs else None, 'n')
                        rows.setdefault("  (+) Manual Adjustments",       {})[col] = _fmt(adj_plus - adj_minus if (adj_plus - adj_minus) != 0 else None, 'n')
                        rows.setdefault("(/) Capacity Productivity",      {})[col] = _fmt(cap_prod, '%')
                        rows.setdefault("(/) Shrinkage (%)",              {})[col] = _fmt(100 - cap_prod if current_hrs > 0 else None, '%')
                        rows.setdefault("(/) Actual HC Productivity",     {})[col] = _fmt(act_hc_prod, '%')
                        rows.setdefault("━ Required HC (FTEs)",           {})[col] = _fmt(fte_total, 'fte')
                        rows.setdefault("  · Accountant I",               {})[col] = _fmt(fte_acc1, 'fte')
                        rows.setdefault("  · Accountant II",              {})[col] = _fmt(fte_acc2, 'fte')
                        rows.setdefault("  · General Accountant",         {})[col] = _fmt(fte_gen, 'fte')
                        rows.setdefault("  · Sr. Accountant",             {})[col] = _fmt(fte_sr, 'fte')
                        rows.setdefault("━ Actual HC (Report)",           {})[col] = _fmt(hc_total, 'fte')
                        rows.setdefault("  · Accountant I (actual)",      {})[col] = _fmt(hc_acc1, 'fte')
                        rows.setdefault("  · Accountant II (actual)",     {})[col] = _fmt(hc_acc2, 'fte')
                        rows.setdefault("  · General Acc. (actual)",      {})[col] = _fmt(hc_gen, 'fte')
                        rows.setdefault("  · Sr. Accountant (actual)",    {})[col] = _fmt(hc_sr, 'fte')
                        rows.setdefault("  · Managers (actual)",          {})[col] = _fmt(hc_mgr, 'fte')
                        rows.setdefault("━ HC Δ (Actual − Required)",     {})[col] = _fmt(d_total, 'dec')
                        rows.setdefault("  · Δ Accountant I",             {})[col] = _fmt(d_acc1, 'dec')
                        rows.setdefault("  · Δ Accountant II",            {})[col] = _fmt(d_acc2, 'dec')
                        rows.setdefault("  · Δ General Accountant",       {})[col] = _fmt(d_gen, 'dec')
                        rows.setdefault("  · Δ Sr. Accountant",           {})[col] = _fmt(d_sr, 'dec')
                        rows.setdefault("━ MRR ($)",                      {})[col] = _fmt(mrr, '$')
                        rows.setdefault("  (+) New MRR ($)",              {})[col] = _fmt(new_mrr_val if new_mrr_val else None, '$')
                        rows.setdefault("  (-) Churn MRR ($)",            {})[col] = _fmt(churn_mrr_val if churn_mrr_val else None, '$')
                        rows.setdefault("  Revenue / HC ($)",             {})[col] = _fmt(rev_per_hc, '$')
                        rows.setdefault("━ Cost & Margin",               {})[col] = _fmt(exp_margin_pct, '%')
                        rows.setdefault("  Capacity Cost ($)",            {})[col] = _fmt(cap_cost, '$')
                        rows.setdefault("  Capacity Margin ($)",          {})[col] = _fmt(cap_margin, '$')
                        rows.setdefault("  Capacity Margin (%)",          {})[col] = _fmt(cap_margin_pct, '%')
                        rows.setdefault("  Expected Cost ($)",            {})[col] = _fmt(exp_cost, '$')
                        rows.setdefault("  Expected Margin ($)",          {})[col] = _fmt(exp_margin, '$')
                        rows.setdefault("  Expected Margin (%)",          {})[col] = _fmt(exp_margin_pct, '%')
                        # Per-month learning-curve-aware AHT and split ticket counts
                        # Filter to clients active in this specific month:
                        #   Go Live <= month end  AND  (FSD >= month start OR FSD missing)
                        _aht_start_m = pd.Timestamp((today + relativedelta(months=_month_offsets[i])).replace(day=1).date())
                        _aht_end_m   = pd.Timestamp((_aht_start_m + relativedelta(months=1) - relativedelta(days=1)).date())
                        if not _df_raw.empty:
                            _aht_gl_all  = pd.to_datetime(_df_raw.get('Go Live',              pd.Series(dtype='datetime64[ns]')), errors='coerce')
                            _aht_fsd_all = pd.to_datetime(_df_raw.get('Final Service Date',   pd.Series(dtype='datetime64[ns]')), errors='coerce')
                            _aht_active  = (
                                (_aht_gl_all.isna()  | (_aht_gl_all  <= _aht_end_m)) &
                                (_aht_fsd_all.isna() | (_aht_fsd_all >= _aht_start_m))
                            )
                            _df_m = _df_raw[_aht_active].copy()
                        else:
                            _df_m = _df_raw
                        if not _df_m.empty and 'Go Live' in _df_m.columns:
                            _aht_gl   = pd.to_datetime(_df_m['Go Live'], errors='coerce')
                            _aht_hagl = _aht_gl.notna()
                            _aht_gl_f = _aht_gl.fillna(_aht_start_m)
                            _aht_md   = np.where(_aht_hagl,
                                (_aht_start_m.year  - _aht_gl_f.dt.year)  * 12 +
                                (_aht_start_m.month - _aht_gl_f.dt.month), 999)
                            _aht_lc = np.select(
                                [~_aht_hagl, (_aht_md==0), (_aht_md==1), (_aht_md==2)],
                                [1.0, 1.17, 0.86, 0.99], default=1.0)
                        else:
                            _aht_lc = np.ones(len(_df_m)) if not _df_m.empty else np.array([1.0])
                        _aht_ptix = _safe_num(_df_m.get('Closed tickets with Proc time', 0)) if not _df_m.empty else pd.Series([0.0])
                        _aht_rtix = _safe_num(_df_m.get('Closed tickets with rev time',  0)) if not _df_m.empty else pd.Series([0.0])
                        _aht_paht = _safe_num(_df_m.get('>>> FINAL Capacity Proc AHT',   0)) if not _df_m.empty else pd.Series([0.0])
                        _aht_raht = _safe_num(_df_m.get('>>> FINAL Capacity Rev AHT',    0)) if not _df_m.empty else pd.Series([0.0])
                        _aht_proc_tix = int(_aht_ptix.sum())
                        _aht_rev_tix  = int(_aht_rtix.sum())
                        _aht_tot_tix  = _aht_proc_tix + _aht_rev_tix
                        _avg_aht = (
                            (_aht_ptix * _aht_paht * _aht_lc).sum() +
                            (_aht_rtix * _aht_raht * _aht_lc).sum()
                        ) / _aht_tot_tix if _aht_tot_tix > 0 else 0.0
                        # Active client count for this month (Go Live ≤ month end
                        # AND (FSD ≥ month start OR FSD missing)).
                        _m_cli_count = 0
                        if not _duc.empty and 'client_name' in _duc.columns:
                            _duc_gl_o  = pd.to_datetime(_duc.get('Go Live',            pd.Series(dtype='datetime64[ns]')), errors='coerce')
                            _duc_fsd_o = pd.to_datetime(_duc.get('Final Service Date', pd.Series(dtype='datetime64[ns]')), errors='coerce')
                            _m_cli_mask = (
                                (_duc_gl_o.isna()  | (_duc_gl_o  <= _aht_end_m)) &
                                (_duc_fsd_o.isna() | (_duc_fsd_o >= _aht_start_m))
                            )
                            _m_cli_count = int(_m_cli_mask.sum())
                        # ── Per-month properties/doors/sqft (active clients only) ──
                        if not _df_m.empty:
                            _m_snap = _df_m.groupby('client_name', as_index=False).agg({
                                c: 'first' for c in ['Res Prop','Commercial Properties','Res doors','Commercial Doors','SQFT Commercial']
                                if c in _df_m.columns
                            })
                            _res_prop_count  = int(_safe_num(_m_snap.get('Res Prop', pd.Series(dtype=float))).sum())
                            _comm_prop_count = int(_safe_num(_m_snap.get('Commercial Properties', pd.Series(dtype=float))).sum())
                            _prop_count      = _res_prop_count + _comm_prop_count
                            _res_door_count  = int(_safe_num(_m_snap.get('Res doors', pd.Series(dtype=float))).sum())
                            _comm_door_count = int(_safe_num(_m_snap.get('Commercial Doors', pd.Series(dtype=float))).sum())
                            _door_count      = _res_door_count + _comm_door_count
                            _sqft_count      = int(_safe_num(_m_snap.get('SQFT Commercial', pd.Series(dtype=float))).sum())
                        else:
                            _res_prop_count = _comm_prop_count = _prop_count = 0
                            _res_door_count = _comm_door_count = _door_count = _sqft_count = 0
                        rows.setdefault("━ Property Count",               {})[col] = _fmt(_prop_count, 'n')
                        rows.setdefault("  Res Properties",               {})[col] = _fmt(_res_prop_count if _res_prop_count else None, 'n')
                        rows.setdefault("  Comm Properties",              {})[col] = _fmt(_comm_prop_count if _comm_prop_count else None, 'n')
                        rows.setdefault("  Client Count",                 {})[col] = _fmt(_m_cli_count, 'n')
                        rows.setdefault("  Res Doors",                    {})[col] = _fmt(_res_door_count if _res_door_count else None, 'n')
                        rows.setdefault("  Comm Doors",                   {})[col] = _fmt(_comm_door_count if _comm_door_count else None, 'n')
                        rows.setdefault("  SQFT (Comm)",                  {})[col] = _fmt(_sqft_count if _sqft_count else None, 'n')
                        rows.setdefault("  Tickets to Process",           {})[col] = _fmt(_aht_proc_tix, 'n')
                        rows.setdefault("  Tickets to Review",            {})[col] = _fmt(_aht_rev_tix,  'n')
                        rows.setdefault("  AHT (min)",                    {})[col] = _fmt(_avg_aht, 'dec')
                        rows.setdefault("━ Working Days",                  {})[col] = _fmt(workdays_val, 'n')
                        rows.setdefault("  Holidays",                     {})[col] = _fmt(holidays_val, 'n')

                    df_wf = pd.DataFrame(rows, index=months).T
                    df_wf.index.name = ""
                    return df_wf

                # ── Overall view — rebuild only when cascade or HC data changes ────────
                _wf_cache_key = (
                    st.session_state.get('_fd_version', 0),
                    st.session_state.get('_hc_version', 0),
                    tuple(sorted(_dash_sel_pods)),   # scoped HC depends on selected pods
                )
                if st.session_state.get('_wf_cache_key') != _wf_cache_key:
                    _df_wf_overall_full = _build_overall_wf()
                    st.session_state['_wf_overall_export'] = _df_wf_overall_full
                    st.session_state['_wf_cache_key'] = _wf_cache_key
                else:
                    _df_wf_overall_full = st.session_state['_wf_overall_export']

                # Apply Actual HC row filter for display when client filters are active
                _actual_hc_rows_to_hide = {
                    "━ Actual HC (Report)", "  · Accountant I (actual)", "  · Accountant II (actual)",
                    "  · General Acc. (actual)", "  · Sr. Accountant (actual)", "  · Managers (actual)",
                    "━ HC Δ (Actual − Required)", "  · Δ Accountant I", "  · Δ Accountant II",
                    "  · Δ General Accountant", "  · Δ Sr. Accountant",
                    "(/) Actual HC Productivity", "  Expected Cost ($)", "  Expected Margin ($)",
                    "  Expected Margin (%)", "  Revenue / HC ($)",
                }
                _df_wf_overall = (
                    _df_wf_overall_full[~_df_wf_overall_full.index.isin(_actual_hc_rows_to_hide)]
                    if _hide_actual_hc else _df_wf_overall_full
                )

                # Collapsible row groups (pivot-style)
                _ov_groups_full = {
                    "━ Required Hours":           ["  Current Customer Hours", "  Shrinkage (Hrs)", "  (+) New Customer Hours", "  (-) Confirmed Churn (Hrs)", "  (-) Automations", "  (+) Manual Adjustments"],
                    "━ Required HC (FTEs)":       ["  · Accountant I", "  · Accountant II", "  · General Accountant", "  · Sr. Accountant"],
                    "━ Actual HC (Report)":       ["  · Accountant I (actual)", "  · Accountant II (actual)", "  · General Acc. (actual)", "  · Sr. Accountant (actual)", "  · Managers (actual)"],
                    "━ HC Δ (Actual − Required)": ["  · Δ Accountant I", "  · Δ Accountant II", "  · Δ General Accountant", "  · Δ Sr. Accountant"],
                    "━ MRR ($)":                  ["  (+) New MRR ($)", "  (-) Churn MRR ($)", "  Revenue / HC ($)"],
                    "━ Cost & Margin":            ["  Capacity Cost ($)", "  Capacity Margin ($)", "  Capacity Margin (%)",
                                                   "  Expected Cost ($)", "  Expected Margin ($)", "  Expected Margin (%)"],
                    "━ Property Count":           ["  Res Properties", "  Comm Properties", "  Client Count", "  Res Doors", "  Comm Doors", "  SQFT (Comm)"],
                    "━ Working Days":             ["  Holidays"],
                }
                _ov_groups_nohc = {
                    "━ Required Hours":           ["  Current Customer Hours", "  Shrinkage (Hrs)", "  (+) New Customer Hours", "  (-) Confirmed Churn (Hrs)", "  (-) Automations", "  (+) Manual Adjustments"],
                    "━ Required HC (FTEs)":       ["  · Accountant I", "  · Accountant II", "  · General Accountant", "  · Sr. Accountant"],
                    "━ MRR ($)":                  ["  (+) New MRR ($)", "  (-) Churn MRR ($)"],
                    "━ Cost & Margin":            ["  Capacity Cost ($)", "  Capacity Margin ($)", "  Capacity Margin (%)"],
                    "━ Property Count":           ["  Res Properties", "  Comm Properties", "  Client Count", "  Res Doors", "  Comm Doors", "  SQFT (Comm)"],
                    "━ Working Days":             ["  Holidays"],
                }
                _ov_groups = _ov_groups_nohc if _hide_actual_hc else _ov_groups_full
                for _gh in _ov_groups:
                    if f"_ov_exp_{_gh}" not in st.session_state:
                        st.session_state[f"_ov_exp_{_gh}"] = False  # collapsed by default

                @st.fragment
                def _ov_table_fragment():
                    _short = lambda s: s.replace("━ ", "").replace(" (FTEs)", "").replace(" ($)", "").replace(" (Report)", "")
                    # Expand / Collapse all
                    _ca_col, _ea_col, *_grp_cols = st.columns([1, 1] + [1] * len(_ov_groups))
                    if _ca_col.button("▶ Collapse All", key="_ov_collapse_all", use_container_width=True):
                        for _gh in _ov_groups:
                            st.session_state[f"_ov_exp_{_gh}"] = False
                        st.rerun(scope="fragment")
                    if _ea_col.button("► Expand All", key="_ov_expand_all", use_container_width=True):
                        for _gh in _ov_groups:
                            st.session_state[f"_ov_exp_{_gh}"] = True
                        st.rerun(scope="fragment")
                    for _ci, (_gh, _drs) in enumerate(_ov_groups.items()):
                        _ek = f"_ov_exp_{_gh}"
                        _icon = "▼" if st.session_state.get(_ek, False) else "▶"
                        if _grp_cols[_ci].button(f"{_icon} {_short(_gh)}", key=f"_ovbtn_{_ci}", use_container_width=True):
                            st.session_state[_ek] = not st.session_state.get(_ek, False)
                            st.rerun(scope="fragment")

                    _all_detail = {_r for _drs in _ov_groups.values() for _r in _drs}
                    _visible = []
                    for _r in _df_wf_overall.index:
                        if _r in _all_detail:
                            for _gh, _drs in _ov_groups.items():
                                if _r in _drs and st.session_state.get(f"_ov_exp_{_gh}", False):
                                    _visible.append(_r)
                                    break
                        else:
                            _visible.append(_r)
                    # Height fitted to content — no large minimum so empty space
                    # doesn't trap mouse-wheel scroll events.
                    _h = min(800, len(_visible) * 35 + 42)
                    st.dataframe(_df_wf_overall.loc[_visible], use_container_width=True, height=_h)

                if _dash_level in ('overall', 'client', 'multi_pod'):
                    st.markdown("#### Overall")
                    _ov_table_fragment()

                if _dash_level in ('overall', 'pod'):
                    st.divider()
                    st.markdown("#### By POD")
                    if 'POD' in _pod_df.columns:
                        _pod_df = _pod_df.copy()
                        _pod_df['POD'] = (
                            _pod_df['POD'].fillna('').astype(str).str.strip()
                            .where(lambda s: ~s.str.lower().isin({'nan', 'none', ''}), 'No POD')
                        )
                    _pod_names = sorted(
                        _pod_df['POD'].unique().tolist() if 'POD' in _pod_df.columns else []
                    )

                    # ── FTE Summary Table (all PODs × 6 months) ──────────────────
                    if _pod_names and not _pod_df.empty:
                        st.markdown("##### 📊 FTE Summary by POD")
                        def _norm_hc_pod_sum(s):
                            return str(s).lower().replace(' ', '').strip()
                        _fte_sum_rows = []
                        for _pn in _pod_names:
                            _pm_tot = _pod_df[
                                (_pod_df['POD'] == _pn) &
                                (_pod_df['Required Role'] == '>>> POD TOTAL')
                            ]
                            _prow = {'POD': _pn}
                            # Actual HC from HC report
                            _act_hc_sum = 0
                            if _hc:
                                _hbp_s = _hc['by_pod_role']
                                _hbp_s_norm = _hbp_s['POD'].apply(_norm_hc_pod_sum)
                                _pm_hc_s = _hbp_s[_hbp_s_norm == _norm_hc_pod_sum(_pn)]
                                _act_hc_sum = int(_pm_hc_s['HC'].sum()) if not _pm_hc_s.empty else 0
                            _prow['Actual HC'] = _act_hc_sum if _act_hc_sum > 0 else '—'
                            for _si, _smes in enumerate(meses_proyeccion):
                                _sc_fte = f"M{_si+1} ({_smes}) - Final FTEs"
                                _spfte = float(_pm_tot[_sc_fte].sum()) if (not _pm_tot.empty and _sc_fte in _pm_tot.columns) else 0.0
                                _prow[f"M{_si+1} ({_smes})"] = round(_spfte, 2) if _spfte > 0 else '—'
                            _fte_sum_rows.append(_prow)
                        if _fte_sum_rows:
                            _df_fte_sum = pd.DataFrame(_fte_sum_rows).set_index('POD')
                            st.dataframe(
                                _df_fte_sum,
                                use_container_width=True,
                                height=min(600, len(_fte_sum_rows) * 35 + 60),
                            )
                        st.divider()

                    if _pod_names:
                        _pod_tab_labels = _pod_names
                        _pod_tabs = st.tabs(_pod_tab_labels)
                        _wf_pod_all = {}   # accumulate {pod_name: df_pod_wf} for export
                        for _pt, _pod_name in zip(_pod_tabs, _pod_names):
                            with _pt:
                                _pod_rows = {}
                                _pod_hc   = {}
                                if _hc:
                                    _hbp = _hc['by_pod_role']
                                    # Normalize both sides: lowercase + strip spaces for matching
                                    def _norm_hc_pod(s):
                                        return str(s).lower().replace(' ', '').strip()
                                    _hbp_norm   = _hbp['POD'].apply(_norm_hc_pod)
                                    _pod_nm_norm = _norm_hc_pod(_pod_name)
                                    _mask_p = _hbp_norm == _pod_nm_norm
                                    for _, _hr in _hbp[_mask_p].iterrows():
                                        _pod_hc[_hr['Capacity Role']] = int(_hr['HC'])

                                # Pre-compute POD clients for MRR and property lookups
                                _duc = st.session_state.get('df_clients_unique', pd.DataFrame())
                                if 'POD' in _cli_df.columns and 'Client' in _cli_df.columns:
                                    _cli_pod_norm_p = (
                                        _cli_df['POD'].fillna('').astype(str).str.strip()
                                        .where(lambda s: ~s.str.lower().isin({'nan', 'none', ''}), 'No POD')
                                    )
                                    _pod_clients_lower = set(
                                        _cli_df[_cli_pod_norm_p == _pod_name]['Client']
                                        .dropna().astype(str).str.strip().str.lower().unique()
                                    )
                                else:
                                    _pod_clients_lower = set()

                                # Pre-compute POD-level property / ticket / AHT from raw data
                                # Normalize df_clean POD values the same way as df_resumen
                                # (NaN → 'No POD', string 'nan'/'none'/'' → 'No POD') before filtering
                                if not _df_raw.empty and 'POD' in _df_raw.columns:
                                    _raw_pod_norm = _df_raw['POD'].fillna('').astype(str).str.strip()
                                    _raw_pod_norm = _raw_pod_norm.where(
                                        ~_raw_pod_norm.str.lower().isin({'nan', 'none', ''}), 'No POD'
                                    )
                                    _pdf_raw = _df_raw[_raw_pod_norm == _pod_name]
                                else:
                                    _pdf_raw = pd.DataFrame()
                                if not _pdf_raw.empty:
                                    _p_client_snap = _pdf_raw.groupby('client_name', as_index=False).agg({
                                        c: 'first' for c in ['Res Prop','Commercial Properties','Res doors','Commercial Doors','SQFT Commercial']
                                        if c in _pdf_raw.columns
                                    })
                                    _p_res_prop_count  = int(_safe_num(_p_client_snap.get('Res Prop', 0)).sum())
                                    _p_comm_prop_count = int(_safe_num(_p_client_snap.get('Commercial Properties', 0)).sum())
                                    _p_prop_count      = _p_res_prop_count + _p_comm_prop_count
                                    _p_res_door_count  = int(_safe_num(_p_client_snap.get('Res doors', 0)).sum())
                                    _p_comm_door_count = int(_safe_num(_p_client_snap.get('Commercial Doors', 0)).sum())
                                    _p_door_count      = _p_res_door_count + _p_comm_door_count
                                    _p_sqft_count      = int(_safe_num(_p_client_snap.get('SQFT Commercial', 0)).sum())
                                else:
                                    _p_res_prop_count = _p_comm_prop_count = _p_prop_count = 0
                                    _p_res_door_count = _p_comm_door_count = _p_door_count = _p_sqft_count = 0

                                _pm_all  = _pod_df[(_pod_df['POD'] == _pod_name) & (_pod_df['Required Role'] == '>>> POD TOTAL')]
                                _proles_all = _pod_df[(_pod_df['POD'] == _pod_name) & (_pod_df['Required Role'] != '>>> POD TOTAL')]

                                for i, mes_str in enumerate(meses_proyeccion):
                                    if i >= len(_exec): break
                                    _wdays = st.session_state.get('calc_data',{}).get('dict_workable_days',{})
                                    wdays  = _wdays.get(i, 21)
                                    c_base_col = f"M{i+1} ({mes_str}) - Base Hours"
                                    c_fte_col  = f"M{i+1} ({mes_str}) - Final FTEs"
                                    c_save_col = f"M{i+1} ({mes_str}) - Auto Saving (Hrs)"
                                    c_fin_col  = f"M{i+1} ({mes_str}) - Final Hours"
                                    c_plus_col = f"M{i+1} ({mes_str}) - Adjustments (+) Hrs"
                                    c_minus_col= f"M{i+1} ({mes_str}) - Adjustments (-) Hrs"

                                    _pm     = _pm_all
                                    _proles = _proles_all

                                    p_base  = float(_pm[c_base_col].sum()) if c_base_col  in _pm.columns else 0
                                    p_auto  = float(_pm[c_save_col].sum()) if c_save_col  in _pm.columns else 0
                                    p_fin   = float(_pm[c_fin_col].sum())  if c_fin_col   in _pm.columns else 0
                                    p_fte   = float(_pm[c_fte_col].sum())  if c_fte_col   in _pm.columns else 0
                                    p_plus  = float(_pm[c_plus_col].sum()) if c_plus_col  in _pm.columns else 0
                                    p_minus = float(_pm[c_minus_col].sum())if c_minus_col in _pm.columns else 0

                                    def _prole_fte(role_name):
                                        _r = _proles[_proles['Required Role'] == role_name]
                                        return float(_r[c_fte_col].sum()) if c_fte_col in _r.columns else 0

                                    # Total HC only counts the 4 productive roles
                                    # (Accountant I/II, General Acc., Sr. Acc.) — managers
                                    # live in 'Other' and are surfaced in their own row.
                                    hc_p_tot  = sum(
                                        v for k, v in _pod_hc.items() if k != 'Other'
                                    ) or None
                                    hc_p_acc1 = _pod_hc.get('Accountant I')
                                    hc_p_acc2 = _pod_hc.get('Accountant II')
                                    hc_p_gen  = _pod_hc.get('General Accountant')
                                    hc_p_sr   = _pod_hc.get('Sr. Accountant')
                                    # Manager count for this POD (normalized lookup)
                                    _p_mgr_by_pod = _hc.get('mgr_by_pod', {}) if _hc else {}
                                    _p_mgr_norm   = {
                                        str(k).lower().replace(' ', '').strip(): int(v)
                                        for k, v in _p_mgr_by_pod.items()
                                    }
                                    hc_p_mgr  = _p_mgr_norm.get(
                                        str(_pod_name).lower().replace(' ', '').strip(), 0
                                    )
                                    d_pod     = round(hc_p_tot - p_fte, 2) if hc_p_tot is not None else None
                                    d_p_acc1  = round(hc_p_acc1 - _prole_fte('Accountant I'),       2) if hc_p_acc1 is not None else None
                                    d_p_acc2  = round(hc_p_acc2 - _prole_fte('Accountant II'),      2) if hc_p_acc2 is not None else None
                                    d_p_gen   = round(hc_p_gen  - _prole_fte('General Accountant'), 2) if hc_p_gen  is not None else None
                                    d_p_sr    = round(hc_p_sr   - _prole_fte('Sr. Accountant'),     2) if hc_p_sr   is not None else None

                                    # POD MRR: sum active clients belonging to this POD
                                    _pod_mrr = 0.0
                                    if not _duc.empty and _pod_clients_lower and 'MRR' in _duc.columns:
                                        mes_date = today + relativedelta(months=_month_offsets[i])
                                        _start_m = pd.Timestamp(mes_date.replace(day=1).date())
                                        _end_m   = pd.Timestamp((_start_m + relativedelta(months=1) - relativedelta(days=1)).date())
                                        _m_pod_mrr = (
                                            (_duc['Go Live'].isna() | (_duc['Go Live'] <= _end_m)) &
                                            (_duc['Final Service Date'].isna() | (_duc['Final Service Date'] >= _start_m)) &
                                            (_duc['client_name'].astype(str).str.strip().str.lower().isin(_pod_clients_lower))
                                        )
                                        _pod_mrr = float(_duc.loc[_m_pod_mrr, 'MRR'].sum())

                                    _rev_per_hc = _pod_mrr / hc_p_tot if (hc_p_tot and hc_p_tot > 0) else None

                                    _p_cap_cost = (
                                        float(_prole_fte('Accountant I')       or 0) * cost_acc1 +
                                        float(_prole_fte('Accountant II')      or 0) * cost_acc2 +
                                        float(_prole_fte('General Accountant') or 0) * cost_gen  +
                                        float(_prole_fte('Sr. Accountant')     or 0) * cost_sr
                                    )
                                    _p_cap_margin     = float(_pod_mrr or 0) - _p_cap_cost
                                    _p_cap_margin_pct = (_p_cap_margin / float(_pod_mrr) * 100) if _pod_mrr and float(_pod_mrr) != 0 else None
                                    if hc_p_tot is not None:
                                        _p_exp_cost = (
                                            float(hc_p_acc1 or 0) * cost_acc1 +
                                            float(hc_p_acc2 or 0) * cost_acc2 +
                                            float(hc_p_gen  or 0) * cost_gen  +
                                            float(hc_p_sr   or 0) * cost_sr
                                        )
                                        _p_exp_margin     = float(_pod_mrr or 0) - _p_exp_cost
                                        _p_exp_margin_pct = (_p_exp_margin / float(_pod_mrr) * 100) if _pod_mrr and float(_pod_mrr) != 0 else None
                                    else:
                                        _p_exp_cost = _p_exp_margin = _p_exp_margin_pct = None

                                    # New/churn hours and MRR from pre-computed stores
                                    _p_churn_hrs  = st.session_state.final_dashboards.get('pod_churn',   {}).get(mes_str, {}).get(_pod_name, 0.0)
                                    _p_new_hrs    = st.session_state.final_dashboards.get('pod_new_hrs', {}).get(mes_str, {}).get(_pod_name, 0.0)
                                    _p_new_mrr    = st.session_state.final_dashboards.get('pod_new_mrr',   {}).get(mes_str, {}).get(_pod_name, 0.0)
                                    _p_churn_mrr  = st.session_state.final_dashboards.get('pod_churn_mrr', {}).get(mes_str, {}).get(_pod_name, 0.0)

                                    # Productive hours for this POD — use cascade-accumulated
                                    # per-POD data (covers AI-only PODs not in df_resumen_base)
                                    _pod_prod_store = st.session_state.get('calc_data', {}).get('pod_prod_hrs', {})
                                    if _pod_prod_store and _pod_name in _pod_prod_store:
                                        p_prod_hrs = float(_pod_prod_store[_pod_name].get(i, 0.0))
                                    else:
                                        # Fallback: read from df_resumen_base (baseline-only clients)
                                        _c_prod_pod = f"M{i+1} ({mes_str}) - Productive Hours"
                                        _rb_pod = st.session_state.get('calc_data', {}).get('df_resumen_base', pd.DataFrame())
                                        if not _rb_pod.empty and 'POD' in _rb_pod.columns and _c_prod_pod in _rb_pod.columns:
                                            _rb_pod_norm = (
                                                _rb_pod['POD'].fillna('').astype(str).str.strip()
                                                .where(lambda s: ~s.str.lower().isin({'nan', 'none', ''}), 'No POD')
                                            )
                                            p_prod_hrs = float(_rb_pod[_rb_pod_norm == _pod_name][_c_prod_pod].sum())
                                        else:
                                            p_prod_hrs = p_fin  # last resort: no shrinkage shown

                                    # Use p_fin (Final Hours) as the total so shrinkage is correct
                                    # for both baseline clients (p_fin ≈ p_base) and AI-only PODs
                                    # (p_base = 0, but p_fin = correct post-auto total)
                                    _p_shrinkage_hrs = p_fin - p_prod_hrs

                                    col = mes_str
                                    _pod_rows.setdefault("━ Required Hours",               {})[col] = _fmt(p_fin, 'n')
                                    _pod_rows.setdefault("  Current Customer Hours",       {})[col] = _fmt(p_prod_hrs, 'n')
                                    _pod_rows.setdefault("  Shrinkage (Hrs)",              {})[col] = _fmt(_p_shrinkage_hrs if _p_shrinkage_hrs > 0 else None, 'n')
                                    _pod_rows.setdefault("  (+) New Customer Hours",       {})[col] = _fmt(_p_new_hrs if _p_new_hrs > 0 else None, 'n')
                                    _pod_rows.setdefault("  (-) Confirmed Churn (Hrs)",    {})[col] = _fmt(_p_churn_hrs if _p_churn_hrs > 0 else None, 'n')
                                    _pod_rows.setdefault("  (-) Automations",              {})[col] = _fmt(p_auto if p_auto else None, 'n')
                                    _pod_rows.setdefault("  (+) Manual Adjustments",       {})[col] = _fmt(p_plus - p_minus if (p_plus - p_minus) != 0 else None, 'n')
                                    _p_cap_prod = (p_prod_hrs / p_fin * 100) if p_fin and p_fin > 0 else 0
                                    _p_act_hc_prod = (
                                        (float(hc_p_acc1 or 0) * util_acc1 +
                                         float(hc_p_acc2 or 0) * util_acc1 +
                                         float(hc_p_gen  or 0) * util_gen  +
                                         float(hc_p_sr   or 0) * util_sr)
                                        / float(hc_p_tot) * 100
                                    ) if hc_p_tot and float(hc_p_tot) > 0 else None
                                    _pod_rows.setdefault("(/) Capacity Productivity",      {})[col] = _fmt(_p_cap_prod, '%')
                                    _pod_rows.setdefault("(/) Shrinkage (%)",              {})[col] = _fmt(100 - _p_cap_prod if p_fin > 0 else None, '%')
                                    _pod_rows.setdefault("(/) Actual HC Productivity",     {})[col] = _fmt(_p_act_hc_prod, '%')
                                    _pod_rows.setdefault("━ Required HC (FTEs)",           {})[col] = _fmt(p_fte, 'fte')
                                    _pod_rows.setdefault("  · Accountant I",               {})[col] = _fmt(_prole_fte('Accountant I'), 'fte')
                                    _pod_rows.setdefault("  · Accountant II",              {})[col] = _fmt(_prole_fte('Accountant II'), 'fte')
                                    _pod_rows.setdefault("  · General Accountant",         {})[col] = _fmt(_prole_fte('General Accountant'), 'fte')
                                    _pod_rows.setdefault("  · Sr. Accountant",             {})[col] = _fmt(_prole_fte('Sr. Accountant'), 'fte')
                                    _pod_rows.setdefault("━ Actual HC (Report)",           {})[col] = _fmt(hc_p_tot, 'fte')
                                    _pod_rows.setdefault("  · Accountant I (actual)",      {})[col] = _fmt(hc_p_acc1, 'fte')
                                    _pod_rows.setdefault("  · Accountant II (actual)",     {})[col] = _fmt(hc_p_acc2, 'fte')
                                    _pod_rows.setdefault("  · General Acc. (actual)",      {})[col] = _fmt(hc_p_gen, 'fte')
                                    _pod_rows.setdefault("  · Sr. Accountant (actual)",    {})[col] = _fmt(hc_p_sr, 'fte')
                                    _pod_rows.setdefault("  · Managers (actual)",          {})[col] = _fmt(hc_p_mgr if _hc else None, 'fte')
                                    _pod_rows.setdefault("━ HC Δ (Actual − Required)",     {})[col] = _fmt(d_pod, 'dec')
                                    _pod_rows.setdefault("  · Δ Accountant I",             {})[col] = _fmt(d_p_acc1, 'dec')
                                    _pod_rows.setdefault("  · Δ Accountant II",            {})[col] = _fmt(d_p_acc2, 'dec')
                                    _pod_rows.setdefault("  · Δ General Accountant",       {})[col] = _fmt(d_p_gen, 'dec')
                                    _pod_rows.setdefault("  · Δ Sr. Accountant",           {})[col] = _fmt(d_p_sr, 'dec')
                                    _pod_rows.setdefault("━ MRR ($)",                      {})[col] = _fmt(_pod_mrr, '$')
                                    _pod_rows.setdefault("  (+) New MRR ($)",              {})[col] = _fmt(_p_new_mrr if _p_new_mrr > 0 else None, '$')
                                    _pod_rows.setdefault("  (-) Churn MRR ($)",            {})[col] = _fmt(_p_churn_mrr if _p_churn_mrr > 0 else None, '$')
                                    _pod_rows.setdefault("  Revenue / HC ($)",             {})[col] = _fmt(_rev_per_hc, '$')
                                    _pod_rows.setdefault("━ Cost & Margin",               {})[col] = _fmt(_p_exp_margin_pct, '%')
                                    _pod_rows.setdefault("  Capacity Cost ($)",            {})[col] = _fmt(_p_cap_cost, '$')
                                    _pod_rows.setdefault("  Capacity Margin ($)",          {})[col] = _fmt(_p_cap_margin, '$')
                                    _pod_rows.setdefault("  Capacity Margin (%)",          {})[col] = _fmt(_p_cap_margin_pct, '%')
                                    _pod_rows.setdefault("  Expected Cost ($)",            {})[col] = _fmt(_p_exp_cost, '$')
                                    _pod_rows.setdefault("  Expected Margin ($)",          {})[col] = _fmt(_p_exp_margin, '$')
                                    _pod_rows.setdefault("  Expected Margin (%)",          {})[col] = _fmt(_p_exp_margin_pct, '%')
                                    # Per-month learning-curve-aware AHT and split ticket counts
                                    # Filter to clients active in this month (Go Live ≤ end, FSD ≥ start or missing)
                                    _paht_sm  = pd.Timestamp((today + relativedelta(months=_month_offsets[i])).replace(day=1).date())
                                    _paht_em  = pd.Timestamp((_paht_sm + relativedelta(months=1) - relativedelta(days=1)).date())
                                    if not _pdf_raw.empty:
                                        _paht_gl_all  = pd.to_datetime(_pdf_raw.get('Go Live',            pd.Series(dtype='datetime64[ns]')), errors='coerce')
                                        _paht_fsd_all = pd.to_datetime(_pdf_raw.get('Final Service Date', pd.Series(dtype='datetime64[ns]')), errors='coerce')
                                        _paht_active  = (
                                            (_paht_gl_all.isna()  | (_paht_gl_all  <= _paht_em)) &
                                            (_paht_fsd_all.isna() | (_paht_fsd_all >= _paht_sm))
                                        )
                                        _pdf_m = _pdf_raw[_paht_active].copy()
                                    else:
                                        _pdf_m = _pdf_raw
                                    if not _pdf_m.empty and 'Go Live' in _pdf_m.columns:
                                        _paht_gl  = pd.to_datetime(_pdf_m['Go Live'], errors='coerce')
                                        _paht_hgl = _paht_gl.notna()
                                        _paht_glf = _paht_gl.fillna(_paht_sm)
                                        _paht_md  = np.where(_paht_hgl,
                                            (_paht_sm.year  - _paht_glf.dt.year)  * 12 +
                                            (_paht_sm.month - _paht_glf.dt.month), 999)
                                        _paht_lc = np.select(
                                            [~_paht_hgl, (_paht_md==0), (_paht_md==1), (_paht_md==2)],
                                            [1.0, 1.17, 0.86, 0.99], default=1.0)
                                    else:
                                        _paht_lc = np.ones(len(_pdf_m)) if not _pdf_m.empty else np.array([1.0])
                                    _paht_ptix = _safe_num(_pdf_m.get('Closed tickets with Proc time', 0)) if not _pdf_m.empty else pd.Series([0.0])
                                    _paht_rtix = _safe_num(_pdf_m.get('Closed tickets with rev time',  0)) if not _pdf_m.empty else pd.Series([0.0])
                                    _paht_pa   = _safe_num(_pdf_m.get('>>> FINAL Capacity Proc AHT',   0)) if not _pdf_m.empty else pd.Series([0.0])
                                    _paht_ra   = _safe_num(_pdf_m.get('>>> FINAL Capacity Rev AHT',    0)) if not _pdf_m.empty else pd.Series([0.0])
                                    _p_proc_tix = int(_paht_ptix.sum())
                                    _p_rev_tix  = int(_paht_rtix.sum())
                                    _p_tot_tix2 = _p_proc_tix + _p_rev_tix
                                    _p_avg_aht  = (
                                        (_paht_ptix * _paht_pa * _paht_lc).sum() +
                                        (_paht_rtix * _paht_ra * _paht_lc).sum()
                                    ) / _p_tot_tix2 if _p_tot_tix2 > 0 else 0.0
                                    # Active client count for this POD this month
                                    _p_cli_count = 0
                                    if not _duc.empty and _pod_clients_lower and 'client_name' in _duc.columns:
                                        _pcl_gl  = pd.to_datetime(_duc.get('Go Live',            pd.Series(dtype='datetime64[ns]')), errors='coerce')
                                        _pcl_fsd = pd.to_datetime(_duc.get('Final Service Date', pd.Series(dtype='datetime64[ns]')), errors='coerce')
                                        _pcl_mask = (
                                            (_pcl_gl.isna()  | (_pcl_gl  <= _paht_em)) &
                                            (_pcl_fsd.isna() | (_pcl_fsd >= _paht_sm)) &
                                            (_duc['client_name'].astype(str).str.strip().str.lower().isin(_pod_clients_lower))
                                        )
                                        _p_cli_count = int(_pcl_mask.sum())
                                    # ── Per-month properties/doors/sqft (active clients only) ──
                                    if not _pdf_m.empty:
                                        _pm_snap = _pdf_m.groupby('client_name', as_index=False).agg({
                                            c: 'first' for c in ['Res Prop','Commercial Properties','Res doors','Commercial Doors','SQFT Commercial']
                                            if c in _pdf_m.columns
                                        })
                                        _p_res_prop_count  = int(_safe_num(_pm_snap.get('Res Prop', pd.Series(dtype=float))).sum())
                                        _p_comm_prop_count = int(_safe_num(_pm_snap.get('Commercial Properties', pd.Series(dtype=float))).sum())
                                        _p_prop_count      = _p_res_prop_count + _p_comm_prop_count
                                        _p_res_door_count  = int(_safe_num(_pm_snap.get('Res doors', pd.Series(dtype=float))).sum())
                                        _p_comm_door_count = int(_safe_num(_pm_snap.get('Commercial Doors', pd.Series(dtype=float))).sum())
                                        _p_door_count      = _p_res_door_count + _p_comm_door_count
                                        _p_sqft_count      = int(_safe_num(_pm_snap.get('SQFT Commercial', pd.Series(dtype=float))).sum())
                                    else:
                                        _p_res_prop_count = _p_comm_prop_count = _p_prop_count = 0
                                        _p_res_door_count = _p_comm_door_count = _p_door_count = _p_sqft_count = 0
                                    _pod_rows.setdefault("━ Property Count",               {})[col] = _fmt(_p_prop_count, 'n')
                                    _pod_rows.setdefault("  Res Properties",               {})[col] = _fmt(_p_res_prop_count if _p_res_prop_count else None, 'n')
                                    _pod_rows.setdefault("  Comm Properties",              {})[col] = _fmt(_p_comm_prop_count if _p_comm_prop_count else None, 'n')
                                    _pod_rows.setdefault("  Client Count",                 {})[col] = _fmt(_p_cli_count, 'n')
                                    _pod_rows.setdefault("  Res Doors",                    {})[col] = _fmt(_p_res_door_count if _p_res_door_count else None, 'n')
                                    _pod_rows.setdefault("  Comm Doors",                   {})[col] = _fmt(_p_comm_door_count if _p_comm_door_count else None, 'n')
                                    _pod_rows.setdefault("  SQFT (Comm)",                  {})[col] = _fmt(_p_sqft_count if _p_sqft_count else None, 'n')
                                    _pod_rows.setdefault("  Tickets to Process",           {})[col] = _fmt(_p_proc_tix, 'n')
                                    _pod_rows.setdefault("  Tickets to Review",            {})[col] = _fmt(_p_rev_tix,  'n')
                                    _pod_rows.setdefault("  AHT (min)",                    {})[col] = _fmt(_p_avg_aht, 'dec')
                                    _pod_rows.setdefault("━ Working Days",                 {})[col] = _fmt(wdays, 'n')
                                    _pod_rows.setdefault("  Holidays",                     {})[col] = _fmt(holidays_per_month.get(mes_str, 0), 'n')

                                df_pod_wf = pd.DataFrame(_pod_rows, index=[m for m in meses_proyeccion if m in list(_pod_rows.get("━ Required Hours",{}).keys())]).T
                                df_pod_wf.index.name = ""
                                _wf_pod_all[_pod_name] = df_pod_wf   # cache for export

                                # Collapsible groups for POD table (mirrors Overall)
                                _pod_grp_defs = {
                                    "━ Required Hours":           ["  Current Customer Hours", "  Shrinkage (Hrs)", "  (+) New Customer Hours", "  (-) Confirmed Churn (Hrs)", "  (-) Automations", "  (+) Manual Adjustments"],
                                    "━ Required HC (FTEs)":       ["  · Accountant I", "  · Accountant II", "  · General Accountant", "  · Sr. Accountant"],
                                    "━ Actual HC (Report)":       ["  · Accountant I (actual)", "  · Accountant II (actual)", "  · General Acc. (actual)", "  · Sr. Accountant (actual)", "  · Managers (actual)"],
                                    "━ HC Δ (Actual − Required)": ["  · Δ Accountant I", "  · Δ Accountant II", "  · Δ General Accountant", "  · Δ Sr. Accountant"],
                                    "━ MRR ($)":                  ["  (+) New MRR ($)", "  (-) Churn MRR ($)", "  Revenue / HC ($)"],
                                    "━ Cost & Margin":            ["  Capacity Cost ($)", "  Capacity Margin ($)", "  Capacity Margin (%)",
                                                                   "  Expected Cost ($)", "  Expected Margin ($)", "  Expected Margin (%)"],
                                    "━ Property Count":           ["  Res Properties", "  Comm Properties", "  Client Count", "  Res Doors", "  Comm Doors", "  SQFT (Comm)"],
                                    "━ Working Days":             ["  Holidays"],
                                }
                                for _gh in _pod_grp_defs:
                                    if f"_pod_exp_{_pod_name}_{_gh}" not in st.session_state:
                                        st.session_state[f"_pod_exp_{_pod_name}_{_gh}"] = False  # collapsed by default

                                @st.fragment
                                def _pod_table_fragment(df_pod_wf=df_pod_wf, pod_name=_pod_name, pod_grp_defs=_pod_grp_defs):
                                    _short = lambda s: s.replace("━ ", "").replace(" (FTEs)", "").replace(" ($)", "").replace(" (Report)", "")
                                    _ca_col, _ea_col, *_grp_cols = st.columns([1, 1] + [1] * len(pod_grp_defs))
                                    if _ca_col.button("▶ Collapse All", key=f"_pod_collapse_{pod_name}", use_container_width=True):
                                        for _gh in pod_grp_defs:
                                            st.session_state[f"_pod_exp_{pod_name}_{_gh}"] = False
                                        st.rerun(scope="fragment")
                                    if _ea_col.button("► Expand All", key=f"_pod_expand_{pod_name}", use_container_width=True):
                                        for _gh in pod_grp_defs:
                                            st.session_state[f"_pod_exp_{pod_name}_{_gh}"] = True
                                        st.rerun(scope="fragment")
                                    for _ci, (_gh, _drs) in enumerate(pod_grp_defs.items()):
                                        _ek = f"_pod_exp_{pod_name}_{_gh}"
                                        _icon = "▼" if st.session_state.get(_ek, False) else "▶"
                                        if _grp_cols[_ci].button(f"{_icon} {_short(_gh)}", key=f"_podbtn_{pod_name}_{_ci}", use_container_width=True):
                                            st.session_state[_ek] = not st.session_state.get(_ek, False)
                                            st.rerun(scope="fragment")
                                    _all_detail = {_r for _drs in pod_grp_defs.values() for _r in _drs}
                                    _visible = []
                                    for _r in df_pod_wf.index:
                                        if _r in _all_detail:
                                            for _gh, _drs in pod_grp_defs.items():
                                                if _r in _drs and st.session_state.get(f"_pod_exp_{pod_name}_{_gh}", False):
                                                    _visible.append(_r)
                                                    break
                                        else:
                                            _visible.append(_r)
                                    _h = min(700, len(_visible) * 35 + 42)
                                    st.dataframe(df_pod_wf.loc[_visible], use_container_width=True, height=_h)

                                _pod_table_fragment()
                    else:
                        st.info("No POD data available.")
                    # Persist all pod waterfalls for Excel export
                    st.session_state['_wf_pod_all_export'] = _wf_pod_all

                if _dash_level in ('overall', 'pod', 'sr'):
                    st.divider()
                    st.markdown("#### Sr. Accountant Level")
                    _has_sr_col = 'Sr. Accountant' in _cli_df.columns

                    if not _has_sr_col or _cli_df['Sr. Accountant'].eq('').all():
                        st.info("Sr. Accountant column not found in the data. Make sure your capacity file includes a 'Sr. Accountant' column.")
                    else:
                        _all_srs = sorted(
                            _cli_df['Sr. Accountant'].dropna()
                            .astype(str).replace('', pd.NA).dropna().unique().tolist()
                        )
                        if not _all_srs:
                            st.info("No Sr. Accountant values found in the current dataset.")
                        else:
                            _sel_sr = st.selectbox(
                                "Select Sr. Accountant",
                                options=_all_srs,
                                key="ov_sr_selectbox"
                            )

                            _sr_sel_df = _cli_df[_cli_df['Sr. Accountant'].astype(str) == _sel_sr]

                            # ── Waterfall for selected Sr. — full fields matching POD view ─────────
                            _sr_rows = {}

                            # HC for this Sr. from the HC report (direct reports)
                            # Sr. Accountant column may store email OR name — try all lookup keys
                            _sr_hc_data = {}
                            if _hc:
                                _sr_hc_data = (
                                    _hc.get('by_sr_email', {}).get(str(_sel_sr).strip().lower())   # email match (primary)
                                    or _hc.get('by_sr',      {}).get(_sel_sr)                       # exact name match
                                    or _hc.get('by_sr_norm', {}).get(_norm_name(_sel_sr))           # normalized name match
                                    or {}
                                )
                            hc_sr_tot  = _sr_hc_data.get('total', None) or None
                            _sr_by_role = _sr_hc_data.get('by_role', {})
                            hc_sr_mgr   = int(_sr_hc_data.get('managers', 0) or 0) if _sr_hc_data else None

                            # ── Direct reports label + debug table ────────────────────────
                            if _sr_hc_data:
                                _dr_cnt = _sr_hc_data.get('dr_total', hc_sr_tot - 1 if hc_sr_tot else 0)
                                _lbl_parts = [f"**{_dr_cnt} direct report{'s' if _dr_cnt != 1 else ''}**"]
                                for _rl, _cnt in [
                                    ('Accountant I',       _sr_by_role.get('Accountant I', 0)),
                                    ('Accountant II',      _sr_by_role.get('Accountant II', 0)),
                                    ('General Accountant', _sr_by_role.get('General Accountant', 0)),
                                ]:
                                    if _cnt: _lbl_parts.append(f"{_rl}: {int(_cnt)}")
                                st.caption("  ·  ".join(_lbl_parts))
                            elif _hc:
                                st.caption("⚠️ No HC data matched for this Sr. — check debug table below.")

                            # Temporary debug: show all Srs and their DR counts from HC file
                            with st.expander("🔍 [DEBUG] All Sr. Accountants — Direct Reports from HC file", expanded=False):
                                _by_sr_dbg = (_hc or {}).get('by_sr_email', {})
                                if _by_sr_dbg:
                                    _dbg_rows = sorted([
                                        {
                                            'Email (HC key)': _em,
                                            'Name (HC file)': v.get('email', _em),  # email stored in data
                                            'DRs only': v.get('dr_total', 0),
                                            'Total (incl. Sr.)': v.get('total', 0),
                                            'AccI': v['by_role'].get('Accountant I', 0),
                                            'AccII': v['by_role'].get('Accountant II', 0),
                                            'GenAcc': v['by_role'].get('General Accountant', 0),
                                            'Sr. (self)': v['by_role'].get('Sr. Accountant', 0),
                                        }
                                        for _em, v in _by_sr_dbg.items()
                                    ], key=lambda x: x['Email (HC key)'])
                                    st.caption(f"HC file: {len(_dbg_rows)} Sr. Accountants · "
                                               f"Total DRs: {sum(r['DRs only'] for r in _dbg_rows)} · "
                                               f"Selected Sr. key: `{str(_sel_sr).strip().lower()}`")
                                    st.dataframe(pd.DataFrame(_dbg_rows), use_container_width=True, hide_index=True)
                                else:
                                    st.warning("No by_sr_email data found. Reload the HC file.")
                                    if _hc:
                                        st.code(f"by_sr keys (first 5): {list((_hc.get('by_sr') or {}).keys())[:5]}")
                            # At Sr. level: total = direct reports + 1 (the Sr. themselves)
                            # Sr. Accountant row is always 1 (themselves); DRs are AccI/AccII/GenAcc
                            hc_sr_acc1 = float(_sr_by_role.get('Accountant I',       0)) if hc_sr_tot is not None else None
                            hc_sr_acc2 = float(_sr_by_role.get('Accountant II',      0)) if hc_sr_tot is not None else None
                            hc_sr_gen  = float(_sr_by_role.get('General Accountant', 0)) if hc_sr_tot is not None else None
                            hc_sr_sr   = 1.0                                               if hc_sr_tot is not None else None

                            # Property / ticket / AHT from raw data filtered to Sr.'s clients
                            _sr_clients_set = set(
                                _sr_sel_df['Client'].dropna().astype(str).str.strip().unique()
                            ) if 'Client' in _sr_sel_df.columns else set()
                            _sr_clients_lower = {c.lower() for c in _sr_clients_set}
                            _sr_raw = (
                                _df_raw[_df_raw['client_name'].astype(str).str.strip().str.lower().isin(_sr_clients_lower)]
                                if not _df_raw.empty and 'client_name' in _df_raw.columns else pd.DataFrame()
                            )
                            if not _sr_raw.empty:
                                _sr_snap = _sr_raw.groupby('client_name', as_index=False).agg({
                                    c: 'first' for c in ['Res Prop','Commercial Properties','Res doors','Commercial Doors','SQFT Commercial']
                                    if c in _sr_raw.columns
                                })
                                _sr_res_prop_count  = int(_safe_num(_sr_snap.get('Res Prop', 0)).sum())
                                _sr_comm_prop_count = int(_safe_num(_sr_snap.get('Commercial Properties', 0)).sum())
                                _srp_count          = _sr_res_prop_count + _sr_comm_prop_count
                                _sr_res_door_count  = int(_safe_num(_sr_snap.get('Res doors', 0)).sum())
                                _sr_comm_door_count = int(_safe_num(_sr_snap.get('Commercial Doors', 0)).sum())
                                _srd_count          = _sr_res_door_count + _sr_comm_door_count
                                _srs_count          = int(_safe_num(_sr_snap.get('SQFT Commercial', 0)).sum())
                            else:
                                _sr_res_prop_count = _sr_comm_prop_count = _srp_count = 0
                                _sr_res_door_count = _sr_comm_door_count = _srd_count = _srs_count = 0

                            _duc = st.session_state.get('df_clients_unique', pd.DataFrame())

                            for i, mes_str in enumerate(meses_proyeccion):
                                if i >= len(_exec): break
                                _wdays_sr = st.session_state.get('calc_data', {}).get('dict_workable_days', {})
                                wdays_sr  = _wdays_sr.get(i, 21)
                                c_fte_col  = f"M{i+1} ({mes_str}) - Final FTEs"
                                c_base_col = f"M{i+1} ({mes_str}) - Base Hours"
                                c_fin_col  = f"M{i+1} ({mes_str}) - Final Hours"
                                c_save_col = f"M{i+1} ({mes_str}) - Auto Saving (Hrs)"
                                c_plus_col = f"M{i+1} ({mes_str}) - Adjustments (+) Hrs"
                                c_minus_col= f"M{i+1} ({mes_str}) - Adjustments (-) Hrs"

                                def _srv(col): return float(_sr_sel_df[col].sum()) if col in _sr_sel_df.columns else 0.0
                                def _srole_fte(role):
                                    _s = _sr_sel_df[_sr_sel_df['Required Role'] == role] if 'Required Role' in _sr_sel_df.columns else pd.DataFrame()
                                    return float(_s[c_fte_col].sum()) if c_fte_col in _s.columns else 0.0

                                # MRR + new/churn for this Sr.'s clients
                                _sr_mrr = 0.0
                                _sr_new_hrs = 0.0
                                _sr_churn_hrs = 0.0
                                _sr_new_mrr = 0.0
                                _sr_churn_mrr = 0.0
                                if not _duc.empty and 'client_name' in _duc.columns and 'MRR' in _duc.columns:
                                    _mes_d = today + relativedelta(months=_month_offsets[i])
                                    _s_m   = pd.Timestamp(_mes_d.replace(day=1).date())
                                    _e_m   = pd.Timestamp((_s_m + relativedelta(months=1) - relativedelta(days=1)).date())
                                    _mask_active = (
                                        (_duc['Go Live'].isna()  | (_duc['Go Live'] <= _e_m)) &
                                        (_duc['Final Service Date'].isna() | (_duc['Final Service Date'] >= _s_m)) &
                                        (_duc['client_name'].isin(_sr_clients_set))
                                    )
                                    _sr_mrr = float(_duc.loc[_mask_active, 'MRR'].sum())
                                    # New clients going live this month
                                    _mask_new = (
                                        (_duc['Go Live'] >= _s_m) & (_duc['Go Live'] <= _e_m) &
                                        (_duc['client_name'].isin(_sr_clients_set))
                                    )
                                    _new_cli_names = set(_duc.loc[_mask_new, 'client_name'].astype(str).str.strip().str.lower())
                                    _new_rows = _sr_sel_df[_sr_sel_df['Client'].astype(str).str.strip().str.lower().isin(_new_cli_names)]
                                    _sr_new_hrs = 0.0  # reassigned below from c_base_col
                                    _sr_new_mrr = float(_duc.loc[_mask_new, 'MRR'].sum())
                                    # Churning clients this month
                                    _mask_churn = (
                                        (_duc['Final Service Date'] >= _s_m) & (_duc['Final Service Date'] <= _e_m) &
                                        (_duc['client_name'].isin(_sr_clients_set))
                                    )
                                    _churn_cli_names = set(_duc.loc[_mask_churn, 'client_name'].astype(str).str.strip().str.lower())
                                    _churn_rows = _sr_sel_df[_sr_sel_df['Client'].astype(str).str.strip().str.lower().isin(_churn_cli_names)]
                                    _sr_churn_hrs = 0.0  # reassigned below from c_base_col
                                    _sr_churn_mrr = float(_duc.loc[_mask_churn, 'MRR'].sum())

                                d_sr_tot  = round(hc_sr_tot  - _srv(c_fte_col),          2) if hc_sr_tot  is not None else None
                                d_sr_acc1 = round(hc_sr_acc1 - _srole_fte('Accountant I'),       2) if hc_sr_acc1 is not None else None
                                d_sr_acc2 = round(hc_sr_acc2 - _srole_fte('Accountant II'),      2) if hc_sr_acc2 is not None else None
                                d_sr_gen  = round(hc_sr_gen  - _srole_fte('General Accountant'), 2) if hc_sr_gen  is not None else None
                                d_sr_sr   = round(hc_sr_sr   - _srole_fte('Sr. Accountant'),     2) if hc_sr_sr   is not None else None

                                _rev_sr   = _sr_mrr / hc_sr_tot if (hc_sr_tot and hc_sr_tot > 0) else None
                                _sr_cap_cost = (
                                    float(_srole_fte('Accountant I')       or 0) * cost_acc1 +
                                    float(_srole_fte('Accountant II')      or 0) * cost_acc2 +
                                    float(_srole_fte('General Accountant') or 0) * cost_gen  +
                                    float(_srole_fte('Sr. Accountant')     or 0) * cost_sr
                                )
                                _sr_cap_margin     = float(_sr_mrr or 0) - _sr_cap_cost
                                _sr_cap_margin_pct = (_sr_cap_margin / float(_sr_mrr) * 100) if _sr_mrr and float(_sr_mrr) != 0 else None
                                if hc_sr_tot is not None:
                                    _sr_exp_cost = (
                                        float(hc_sr_acc1 or 0) * cost_acc1 +
                                        float(hc_sr_acc2 or 0) * cost_acc2 +
                                        float(hc_sr_gen  or 0) * cost_gen  +
                                        1.0                    * cost_sr
                                    )
                                    _sr_exp_margin     = float(_sr_mrr or 0) - _sr_exp_cost
                                    _sr_exp_margin_pct = (_sr_exp_margin / float(_sr_mrr) * 100) if _sr_mrr and float(_sr_mrr) != 0 else None
                                else:
                                    _sr_exp_cost = _sr_exp_margin = _sr_exp_margin_pct = None
                                # Productive hours for this Sr.'s clients from df_resumen_base
                                _c_prod_sr = f"M{i+1} ({mes_str}) - Productive Hours"
                                _rb_sr = st.session_state.get('calc_data', {}).get('df_resumen_base', pd.DataFrame())
                                if not _rb_sr.empty and 'Client' in _rb_sr.columns and _c_prod_sr in _rb_sr.columns:
                                    _sr_prod_hrs = float(
                                        _rb_sr[_rb_sr['Client'].astype(str).str.strip().str.lower().isin(_sr_clients_lower)][_c_prod_sr].sum()
                                    )
                                else:
                                    _sr_prod_hrs = _srv(c_base_col)

                                _sr_new_hrs   = float(_new_rows[c_base_col].sum())   if c_base_col   in _new_rows.columns   else 0.0
                                _sr_churn_hrs = float(_churn_rows[c_base_col].sum()) if c_base_col in _churn_rows.columns else 0.0
                                _sr_shrinkage_hrs = _srv(c_base_col) - _sr_prod_hrs

                                col = mes_str
                                _sr_rows.setdefault("━ Required Hours",               {})[col] = _fmt(_srv(c_fin_col), 'n')
                                _sr_rows.setdefault("  Current Customer Hours",       {})[col] = _fmt(_sr_prod_hrs, 'n')
                                _sr_rows.setdefault("  Shrinkage (Hrs)",              {})[col] = _fmt(_sr_shrinkage_hrs if _sr_shrinkage_hrs > 0 else None, 'n')
                                _sr_rows.setdefault("  (+) New Customer Hours",       {})[col] = _fmt(_sr_new_hrs   if _sr_new_hrs   > 0 else None, 'n')
                                _sr_rows.setdefault("  (-) Confirmed Churn (Hrs)",    {})[col] = _fmt(_sr_churn_hrs if _sr_churn_hrs > 0 else None, 'n')
                                _sr_rows.setdefault("  (-) Automations",              {})[col] = _fmt(_srv(c_save_col) if _srv(c_save_col) else None, 'n')
                                _sr_rows.setdefault("  (+) Manual Adjustments",       {})[col] = _fmt(_srv(c_plus_col) - _srv(c_minus_col) if (_srv(c_plus_col) - _srv(c_minus_col)) != 0 else None, 'n')
                                _sr_cap_prod = (_sr_prod_hrs / _srv(c_fin_col) * 100) if _srv(c_fin_col) > 0 else 0
                                _sr_act_hc_prod = (
                                    (float(hc_sr_acc1 or 0) * util_acc1 +
                                     float(hc_sr_acc2 or 0) * util_acc1 +
                                     float(hc_sr_gen  or 0) * util_gen  +
                                     1.0                    * util_sr)
                                    / float(hc_sr_tot) * 100
                                ) if hc_sr_tot and float(hc_sr_tot) > 0 else None
                                _sr_rows.setdefault("(/) Capacity Productivity",      {})[col] = _fmt(_sr_cap_prod, '%')
                                _sr_rows.setdefault("(/) Shrinkage (%)",              {})[col] = _fmt(100 - _sr_cap_prod if _srv(c_fin_col) > 0 else None, '%')
                                _sr_rows.setdefault("(/) Actual HC Productivity",     {})[col] = _fmt(_sr_act_hc_prod, '%')
                                _sr_rows.setdefault("━ Required HC (FTEs)",           {})[col] = _fmt(_srv(c_fte_col), 'fte')
                                _sr_rows.setdefault("  · Accountant I",               {})[col] = _fmt(_srole_fte('Accountant I'), 'fte')
                                _sr_rows.setdefault("  · Accountant II",              {})[col] = _fmt(_srole_fte('Accountant II'), 'fte')
                                _sr_rows.setdefault("  · General Accountant",         {})[col] = _fmt(_srole_fte('General Accountant'), 'fte')
                                _sr_rows.setdefault("  · Sr. Accountant",             {})[col] = _fmt(_srole_fte('Sr. Accountant'), 'fte')
                                _sr_rows.setdefault("━ Actual HC (Report)",           {})[col] = _fmt(hc_sr_tot, 'fte')
                                _sr_rows.setdefault("  · Accountant I (actual)",      {})[col] = _fmt(hc_sr_acc1, 'fte')
                                _sr_rows.setdefault("  · Accountant II (actual)",     {})[col] = _fmt(hc_sr_acc2, 'fte')
                                _sr_rows.setdefault("  · General Acc. (actual)",      {})[col] = _fmt(hc_sr_gen, 'fte')
                                _sr_rows.setdefault("  · Sr. Accountant (actual)",    {})[col] = _fmt(hc_sr_sr, 'fte')
                                _sr_rows.setdefault("  · Managers (actual)",          {})[col] = _fmt(hc_sr_mgr, 'fte')
                                _sr_rows.setdefault("━ HC Δ (Actual − Required)",     {})[col] = _fmt(d_sr_tot, 'dec')
                                _sr_rows.setdefault("  · Δ Accountant I",             {})[col] = _fmt(d_sr_acc1, 'dec')
                                _sr_rows.setdefault("  · Δ Accountant II",            {})[col] = _fmt(d_sr_acc2, 'dec')
                                _sr_rows.setdefault("  · Δ General Accountant",       {})[col] = _fmt(d_sr_gen, 'dec')
                                _sr_rows.setdefault("  · Δ Sr. Accountant",           {})[col] = _fmt(d_sr_sr, 'dec')
                                _sr_rows.setdefault("━ MRR ($)",                      {})[col] = _fmt(_sr_mrr, '$')
                                _sr_rows.setdefault("  (+) New MRR ($)",              {})[col] = _fmt(_sr_new_mrr   if _sr_new_mrr   > 0 else None, '$')
                                _sr_rows.setdefault("  (-) Churn MRR ($)",            {})[col] = _fmt(_sr_churn_mrr if _sr_churn_mrr > 0 else None, '$')
                                _sr_rows.setdefault("  Revenue / HC ($)",             {})[col] = _fmt(_rev_sr, '$')
                                _sr_rows.setdefault("━ Cost & Margin",               {})[col] = _fmt(_sr_exp_margin_pct, '%')
                                _sr_rows.setdefault("  Capacity Cost ($)",            {})[col] = _fmt(_sr_cap_cost, '$')
                                _sr_rows.setdefault("  Capacity Margin ($)",          {})[col] = _fmt(_sr_cap_margin, '$')
                                _sr_rows.setdefault("  Capacity Margin (%)",          {})[col] = _fmt(_sr_cap_margin_pct, '%')
                                _sr_rows.setdefault("  Expected Cost ($)",            {})[col] = _fmt(_sr_exp_cost, '$')
                                _sr_rows.setdefault("  Expected Margin ($)",          {})[col] = _fmt(_sr_exp_margin, '$')
                                _sr_rows.setdefault("  Expected Margin (%)",          {})[col] = _fmt(_sr_exp_margin_pct, '%')
                                # Per-month learning-curve-aware AHT and split ticket counts
                                # Filter to clients active in this month (Go Live ≤ end, FSD ≥ start or missing)
                                _saht_sm  = pd.Timestamp((today + relativedelta(months=_month_offsets[i])).replace(day=1).date())
                                _saht_em  = pd.Timestamp((_saht_sm + relativedelta(months=1) - relativedelta(days=1)).date())
                                if not _sr_raw.empty:
                                    _saht_gl_all  = pd.to_datetime(_sr_raw.get('Go Live',            pd.Series(dtype='datetime64[ns]')), errors='coerce')
                                    _saht_fsd_all = pd.to_datetime(_sr_raw.get('Final Service Date', pd.Series(dtype='datetime64[ns]')), errors='coerce')
                                    _saht_active  = (
                                        (_saht_gl_all.isna()  | (_saht_gl_all  <= _saht_em)) &
                                        (_saht_fsd_all.isna() | (_saht_fsd_all >= _saht_sm))
                                    )
                                    _sr_m = _sr_raw[_saht_active].copy()
                                else:
                                    _sr_m = _sr_raw
                                if not _sr_m.empty and 'Go Live' in _sr_m.columns:
                                    _saht_gl  = pd.to_datetime(_sr_m['Go Live'], errors='coerce')
                                    _saht_hgl = _saht_gl.notna()
                                    _saht_glf = _saht_gl.fillna(_saht_sm)
                                    _saht_md  = np.where(_saht_hgl,
                                        (_saht_sm.year  - _saht_glf.dt.year)  * 12 +
                                        (_saht_sm.month - _saht_glf.dt.month), 999)
                                    _saht_lc = np.select(
                                        [~_saht_hgl, (_saht_md==0), (_saht_md==1), (_saht_md==2)],
                                        [1.0, 1.17, 0.86, 0.99], default=1.0)
                                else:
                                    _saht_lc = np.ones(len(_sr_m)) if not _sr_m.empty else np.array([1.0])
                                _saht_ptix = _safe_num(_sr_m.get('Closed tickets with Proc time', 0)) if not _sr_m.empty else pd.Series([0.0])
                                _saht_rtix = _safe_num(_sr_m.get('Closed tickets with rev time',  0)) if not _sr_m.empty else pd.Series([0.0])
                                _saht_pa   = _safe_num(_sr_m.get('>>> FINAL Capacity Proc AHT',   0)) if not _sr_m.empty else pd.Series([0.0])
                                _saht_ra   = _safe_num(_sr_m.get('>>> FINAL Capacity Rev AHT',    0)) if not _sr_m.empty else pd.Series([0.0])
                                _sr_proc_tix = int(_saht_ptix.sum())
                                _sr_rev_tix  = int(_saht_rtix.sum())
                                _sr_tot_tix2 = _sr_proc_tix + _sr_rev_tix
                                _sr_avg_aht  = (
                                    (_saht_ptix * _saht_pa * _saht_lc).sum() +
                                    (_saht_rtix * _saht_ra * _saht_lc).sum()
                                ) / _sr_tot_tix2 if _sr_tot_tix2 > 0 else 0.0
                                # Active client count for this Sr. this month
                                _sr_cli_count = 0
                                if not _duc.empty and _sr_clients_set and 'client_name' in _duc.columns:
                                    _sc_gl  = pd.to_datetime(_duc.get('Go Live',            pd.Series(dtype='datetime64[ns]')), errors='coerce')
                                    _sc_fsd = pd.to_datetime(_duc.get('Final Service Date', pd.Series(dtype='datetime64[ns]')), errors='coerce')
                                    _sc_mask = (
                                        (_sc_gl.isna()  | (_sc_gl  <= _saht_em)) &
                                        (_sc_fsd.isna() | (_sc_fsd >= _saht_sm)) &
                                        (_duc['client_name'].isin(_sr_clients_set))
                                    )
                                    _sr_cli_count = int(_sc_mask.sum())
                                # ── Per-month properties/doors/sqft (active clients only) ──
                                if not _sr_m.empty:
                                    _sm_snap = _sr_m.groupby('client_name', as_index=False).agg({
                                        c: 'first' for c in ['Res Prop','Commercial Properties','Res doors','Commercial Doors','SQFT Commercial']
                                        if c in _sr_m.columns
                                    })
                                    _sr_res_prop_count  = int(_safe_num(_sm_snap.get('Res Prop', pd.Series(dtype=float))).sum())
                                    _sr_comm_prop_count = int(_safe_num(_sm_snap.get('Commercial Properties', pd.Series(dtype=float))).sum())
                                    _srp_count          = _sr_res_prop_count + _sr_comm_prop_count
                                    _sr_res_door_count  = int(_safe_num(_sm_snap.get('Res doors', pd.Series(dtype=float))).sum())
                                    _sr_comm_door_count = int(_safe_num(_sm_snap.get('Commercial Doors', pd.Series(dtype=float))).sum())
                                    _srd_count          = _sr_res_door_count + _sr_comm_door_count
                                    _srs_count          = int(_safe_num(_sm_snap.get('SQFT Commercial', pd.Series(dtype=float))).sum())
                                else:
                                    _sr_res_prop_count = _sr_comm_prop_count = _srp_count = 0
                                    _sr_res_door_count = _sr_comm_door_count = _srd_count = _srs_count = 0
                                _sr_rows.setdefault("━ Property Count",               {})[col] = _fmt(_srp_count, 'n')
                                _sr_rows.setdefault("  Res Properties",               {})[col] = _fmt(_sr_res_prop_count if _sr_res_prop_count else None, 'n')
                                _sr_rows.setdefault("  Comm Properties",              {})[col] = _fmt(_sr_comm_prop_count if _sr_comm_prop_count else None, 'n')
                                _sr_rows.setdefault("  Client Count",                 {})[col] = _fmt(_sr_cli_count, 'n')
                                _sr_rows.setdefault("  Res Doors",                    {})[col] = _fmt(_sr_res_door_count if _sr_res_door_count else None, 'n')
                                _sr_rows.setdefault("  Comm Doors",                   {})[col] = _fmt(_sr_comm_door_count if _sr_comm_door_count else None, 'n')
                                _sr_rows.setdefault("  SQFT (Comm)",                  {})[col] = _fmt(_srs_count if _srs_count else None, 'n')
                                _sr_rows.setdefault("  Tickets to Process",           {})[col] = _fmt(_sr_proc_tix, 'n')
                                _sr_rows.setdefault("  Tickets to Review",            {})[col] = _fmt(_sr_rev_tix,  'n')
                                _sr_rows.setdefault("  AHT (min)",                    {})[col] = _fmt(_sr_avg_aht, 'dec')
                                _sr_rows.setdefault("━ Working Days",                 {})[col] = _fmt(wdays_sr, 'n')
                                _sr_rows.setdefault("  Holidays",                     {})[col] = _fmt(holidays_per_month.get(mes_str, 0), 'n')

                                # Clients breakdown
                                if 'Client' in _sr_sel_df.columns:
                                    for _cli_n in sorted(_sr_sel_df['Client'].dropna().unique()):
                                        _cli_sub = _sr_sel_df[_sr_sel_df['Client'] == _cli_n]
                                        _sr_rows.setdefault(f"  · {_cli_n}", {})[col] = _fmt(
                                            float(_cli_sub[c_fte_col].sum()) if c_fte_col in _cli_sub.columns else 0, 'fte'
                                        )

                            if _sr_rows:
                                _valid_months = [m for m in meses_proyeccion if m in list(list(_sr_rows.values())[0].keys())]
                                df_sr_wf = pd.DataFrame(_sr_rows, index=_valid_months).T
                                df_sr_wf.index.name = ""
                                # Cache for Quick Overview and for export
                                st.session_state['_wf_sr_export'] = df_sr_wf
                                if '_wf_sr_all_export' not in st.session_state or not isinstance(st.session_state.get('_wf_sr_all_export'), dict):
                                    st.session_state['_wf_sr_all_export'] = {}
                                st.session_state['_wf_sr_all_export'][_sel_sr] = df_sr_wf

                                # Collapsible groups matching the POD / Overall pattern
                                _sr_grp_defs = {
                                    "━ Required Hours":           ["  Current Customer Hours", "  Shrinkage (Hrs)", "  (+) New Customer Hours", "  (-) Confirmed Churn (Hrs)", "  (-) Automations", "  (+) Manual Adjustments"],
                                    "━ Required HC (FTEs)":       ["  · Accountant I", "  · Accountant II", "  · General Accountant", "  · Sr. Accountant"],
                                    "━ Actual HC (Report)":       ["  · Accountant I (actual)", "  · Accountant II (actual)", "  · General Acc. (actual)", "  · Sr. Accountant (actual)", "  · Managers (actual)"],
                                    "━ HC Δ (Actual − Required)": ["  · Δ Accountant I", "  · Δ Accountant II", "  · Δ General Accountant", "  · Δ Sr. Accountant"],
                                    "━ MRR ($)":                  ["  (+) New MRR ($)", "  (-) Churn MRR ($)", "  Revenue / HC ($)"],
                                    "━ Cost & Margin":            ["  Capacity Cost ($)", "  Capacity Margin ($)", "  Capacity Margin (%)",
                                                                   "  Expected Cost ($)", "  Expected Margin ($)", "  Expected Margin (%)"],
                                    "━ Property Count":           ["  Res Properties", "  Comm Properties", "  Client Count", "  Res Doors", "  Comm Doors", "  SQFT (Comm)"],
                                    "━ Working Days":             ["  Holidays"],
                                }
                                # Client rows are also collapsible
                                _sr_cli_rows = [k for k in _sr_rows if k.startswith("  · ") and not any(k in v for v in _sr_grp_defs.values())]
                                if _sr_cli_rows:
                                    _sr_grp_defs["━ Required HC (FTEs)"] = _sr_grp_defs["━ Required HC (FTEs)"]  # keep intact
                                    _sr_grp_defs["━ Clients (FTE breakdown)"] = _sr_cli_rows

                                for _gh in _sr_grp_defs:
                                    if f"_sr_exp_{_sel_sr}_{_gh}" not in st.session_state:
                                        st.session_state[f"_sr_exp_{_sel_sr}_{_gh}"] = False

                                @st.fragment
                                def _sr_table_fragment(df_sr_wf=df_sr_wf, sel_sr=_sel_sr, sr_grp_defs=_sr_grp_defs):
                                    _short = lambda s: s.replace("━ ", "").replace(" (FTEs)", "").replace(" ($)", "").replace(" (Report)", "")
                                    _ca_col, _ea_col, *_grp_cols = st.columns([1, 1] + [1] * len(sr_grp_defs))
                                    if _ca_col.button("▶ Collapse All", key=f"_sr_collapse_{sel_sr}", use_container_width=True):
                                        for _gh in sr_grp_defs:
                                            st.session_state[f"_sr_exp_{sel_sr}_{_gh}"] = False
                                        st.rerun(scope="fragment")
                                    if _ea_col.button("► Expand All", key=f"_sr_expand_{sel_sr}", use_container_width=True):
                                        for _gh in sr_grp_defs:
                                            st.session_state[f"_sr_exp_{sel_sr}_{_gh}"] = True
                                        st.rerun(scope="fragment")
                                    for _ci, (_gh, _drs) in enumerate(sr_grp_defs.items()):
                                        _ek = f"_sr_exp_{sel_sr}_{_gh}"
                                        _icon = "▼" if st.session_state.get(_ek, False) else "▶"
                                        if _grp_cols[_ci].button(f"{_icon} {_short(_gh)}", key=f"_srbtn_{sel_sr}_{_ci}", use_container_width=True):
                                            st.session_state[_ek] = not st.session_state.get(_ek, False)
                                            st.rerun(scope="fragment")
                                    _all_detail = {_r for _drs in sr_grp_defs.values() for _r in _drs}
                                    _visible = []
                                    for _r in df_sr_wf.index:
                                        if _r in _all_detail:
                                            for _gh, _drs in sr_grp_defs.items():
                                                if _r in _drs and st.session_state.get(f"_sr_exp_{sel_sr}_{_gh}", False):
                                                    _visible.append(_r)
                                                    break
                                        else:
                                            _visible.append(_r)
                                    _h = min(800, len(_visible) * 35 + 42)
                                    st.dataframe(df_sr_wf.loc[_visible], use_container_width=True, height=_h)

                                _sr_table_fragment()

                            # ── Detailed data table for audit ─────────────────────────────
                            with st.expander("🔍 Detailed Data — Audit View", expanded=False):
                                st.caption(f"All rows for Sr. Accountant: **{_sel_sr}** — {len(_sr_sel_df)} records")
                                # Show all columns, format numeric ones
                                _detail_show = _sr_sel_df.copy()
                                # Round float columns for readability
                                for _dc in _detail_show.select_dtypes(include='number').columns:
                                    _detail_show[_dc] = _detail_show[_dc].round(2)
                                st.dataframe(_detail_show, use_container_width=True, height=min(600, len(_detail_show) * 35 + 38))

                                # Download button for this Sr.'s data
                                _sr_buf = BytesIO()
                                _sr_sel_df.to_excel(_sr_buf, index=False)
                                st.download_button(
                                    label=f"📥 Download {_sel_sr} data",
                                    data=_sr_buf.getvalue(),
                                    file_name=f"Sr_{_sel_sr.replace('@','_').replace('.','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="dl_sr_detail"
                                )

        with t_gral:
            st.write("### Executive Dashboard (Cascade Impact & ROI)")
            st.dataframe(
                st.session_state.final_dashboards['general'],
                use_container_width=True,
                column_config=general_col_cfg
            )

        with t_pod:
            with st.expander("🚀 Summary by POD", expanded=True):
                pod_col_cfg = {}
                for i, mes_str in enumerate(meses_proyeccion):
                    pod_col_cfg[f"M{i+1} ({mes_str}) - Auto Saving ($)"] = _money_col(f"M{i+1} ({mes_str}) - Auto Saving ($)")
                st.dataframe(
                    st.session_state.final_dashboards['pod'],
                    use_container_width=True,
                    column_config=pod_col_cfg
                )

        with t_pod_sr:
            with st.expander("👤 POD × Sr. Accountant", expanded=True):
                st.caption("Select a POD to see each Sr. Accountant's capacity waterfall within that POD.")
                _psr_cli   = st.session_state.final_dashboards['cliente']
                _psr_has   = 'Sr. Accountant' in _psr_cli.columns and 'POD' in _psr_cli.columns

                if not _psr_has:
                    st.info("Sr. Accountant or POD column not found in the cascade results.")
                else:
                    # Normalize POD column in the client dashboard to avoid 'nan' strings
                    _psr_cli = _psr_cli.copy()
                    _psr_cli['POD'] = (
                        _psr_cli['POD'].fillna('').astype(str).str.strip()
                        .where(lambda s: ~s.str.lower().isin({'nan', 'none', ''}), 'No POD')
                    )
                    _psr_pods = sorted(_psr_cli['POD'].unique().tolist())
                    if not _psr_pods:
                        st.info("No POD data available.")
                    else:
                        _sel_psr_pod = st.selectbox("Select POD", options=_psr_pods, key="pod_sr_pod_sel")
                        _psr_pod_df  = _psr_cli[_psr_cli['POD'] == _sel_psr_pod]
                        _psr_srs     = sorted(
                            _psr_pod_df['Sr. Accountant'].dropna().astype(str).replace('', pd.NA).dropna().unique().tolist()
                        )

                        if not _psr_srs:
                            st.info(f"No Sr. Accountant data for {_sel_psr_pod}.")
                        else:
                            # One sub-tab per Sr. within the selected POD
                            _psr_tabs = st.tabs(_psr_srs)
                            for _psr_tab, _psr_sr in zip(_psr_tabs, _psr_srs):
                                with _psr_tab:
                                    _psr_sr_df = _psr_pod_df[_psr_pod_df['Sr. Accountant'].astype(str) == _psr_sr]
                                    _psr_rows  = {}
                                    _duc_psr   = st.session_state.get('df_clients_unique', pd.DataFrame())

                                    for i, mes_str in enumerate(meses_proyeccion):
                                        if i >= len(st.session_state.final_dashboards['general']): break
                                        c_base_col = f"M{i+1} ({mes_str}) - Base Hours"
                                        c_fte_col  = f"M{i+1} ({mes_str}) - Final FTEs"
                                        c_save_col = f"M{i+1} ({mes_str}) - Auto Saving (Hrs)"
                                        c_fin_col  = f"M{i+1} ({mes_str}) - Final Hours"
                                        c_plus_col = f"M{i+1} ({mes_str}) - Adjustments (+) Hrs"
                                        c_minus_col= f"M{i+1} ({mes_str}) - Adjustments (-) Hrs"

                                        def _psrv(col): return float(_psr_sr_df[col].sum()) if col in _psr_sr_df.columns else 0.0
                                        def _psrole(role):
                                            _s = _psr_sr_df[_psr_sr_df['Required Role'] == role] if 'Required Role' in _psr_sr_df.columns else pd.DataFrame()
                                            return float(_s[c_fte_col].sum()) if c_fte_col in _s.columns else 0.0

                                        # MRR for this Sr. in this POD
                                        _psr_clients = _psr_sr_df['Client'].dropna().unique() if 'Client' in _psr_sr_df.columns else []
                                        _psr_mrr = 0.0
                                        if not _duc_psr.empty and 'client_name' in _duc_psr.columns and 'MRR' in _duc_psr.columns:
                                            _pm_s = pd.Timestamp((today + relativedelta(months=_month_offsets[i])).replace(day=1).date())
                                            _pm_e = pd.Timestamp((_pm_s + relativedelta(months=1) - relativedelta(days=1)).date())
                                            _mm   = (
                                                (_duc_psr['Go Live'].isna() | (_duc_psr['Go Live'] <= _pm_e)) &
                                                (_duc_psr['Final Service Date'].isna() | (_duc_psr['Final Service Date'] >= _pm_s)) &
                                                (_duc_psr['client_name'].isin(_psr_clients))
                                            )
                                            _psr_mrr = float(_duc_psr.loc[_mm, 'MRR'].sum())

                                        col = mes_str
                                        _psr_rows.setdefault("Base Hours",              {})[col] = _fmt(_psrv(c_base_col), 'n')
                                        _psr_rows.setdefault("  ─ Auto Savings (Hrs)",  {})[col] = _fmt(_psrv(c_save_col), 'n')
                                        _psr_rows.setdefault("  + Adj (+) Hrs",         {})[col] = _fmt(_psrv(c_plus_col), 'n')
                                        _psr_rows.setdefault("  − Adj (−) Hrs",         {})[col] = _fmt(_psrv(c_minus_col), 'n')
                                        _psr_rows.setdefault("━ Required Hours",        {})[col] = _fmt(_psrv(c_fin_col), 'n')
                                        _psr_rows.setdefault("━ Required FTEs",         {})[col] = _fmt(_psrv(c_fte_col), 'fte')
                                        _psr_rows.setdefault("  · Accountant I",        {})[col] = _fmt(_psrole('Accountant I'), 'fte')
                                        _psr_rows.setdefault("  · Accountant II",       {})[col] = _fmt(_psrole('Accountant II'), 'fte')
                                        _psr_rows.setdefault("  · General Accountant",  {})[col] = _fmt(_psrole('General Accountant'), 'fte')
                                        _psr_rows.setdefault("  · Sr. Accountant",      {})[col] = _fmt(_psrole('Sr. Accountant'), 'fte')
                                        _psr_rows.setdefault("━ MRR ($)",               {})[col] = _fmt(_psr_mrr, '$')

                                        # Clients breakdown
                                        for _cn in sorted(_psr_sr_df['Client'].dropna().unique()):
                                            _cs = _psr_sr_df[_psr_sr_df['Client'] == _cn]
                                            _psr_rows.setdefault(f"  · {_cn}", {})[col] = _fmt(
                                                float(_cs[c_fte_col].sum()) if c_fte_col in _cs.columns else 0, 'fte'
                                            )

                                    if _psr_rows:
                                        _vm = [m for m in meses_proyeccion if m in list(list(_psr_rows.values())[0].keys())]
                                        df_psr_wf = pd.DataFrame(_psr_rows, index=_vm).T
                                        df_psr_wf.index.name = ""
                                        # Cache for export (all Srs. rendered for the selected POD)
                                        if '_wf_sr_all_export' not in st.session_state or not isinstance(st.session_state.get('_wf_sr_all_export'), dict):
                                            st.session_state['_wf_sr_all_export'] = {}
                                        st.session_state['_wf_sr_all_export'][_psr_sr] = df_psr_wf
                                        st.dataframe(df_psr_wf, use_container_width=True)

        with t_cli:
            st.write("### Client & Ideal Role Level")
            _cli_raw = st.session_state.final_dashboards['cliente'].copy()
            if not _cli_raw.empty:
                _cli_grp_cols = [c for c in ['POD', 'Sr. Accountant', 'Client', 'Required Role'] if c in _cli_raw.columns]
                _cli_id_cols  = [c for c in ['POD', 'Sr. Accountant', 'Client'] if c in _cli_raw.columns]
                _cli_num_cols = [c for c in _cli_raw.columns if c not in _cli_grp_cols]
                # Build display df with a CLIENT TOTAL row after each client's roles
                _cli_frames = []
                _group_by = [c for c in ['POD', 'Sr. Accountant', 'Client'] if c in _cli_raw.columns]
                for _cli_key, _cli_sub in _cli_raw.groupby(_group_by, sort=False):
                    _cli_frames.append(_cli_sub)
                    _tot_row = {c: '' for c in _cli_raw.columns}
                    for _ic in _cli_id_cols:
                        _tot_row[_ic] = _cli_sub[_ic].iloc[0] if not _cli_sub.empty else ''
                    if 'Required Role' in _tot_row:
                        _tot_row['Required Role'] = '>>> CLIENT TOTAL'
                    for _nc in _cli_num_cols:
                        try:
                            _tot_row[_nc] = _cli_sub[_nc].sum()
                        except Exception:
                            _tot_row[_nc] = ''
                    _cli_frames.append(pd.DataFrame([_tot_row]))
                _cli_display = pd.concat(_cli_frames, ignore_index=True) if _cli_frames else _cli_raw
                st.dataframe(
                    _cli_display,
                    use_container_width=True,
                    column_config=cliente_col_cfg,
                    height=min(800, len(_cli_display) * 35 + 38),
                )
            else:
                st.dataframe(_cli_raw, use_container_width=True, column_config=cliente_col_cfg)

        # ── CLIENT MRR BY MONTH ──────────────────────────────────────────────────
        with t_cli_mrr:
            st.write("### 💰 Client MRR by Month")
            st.caption(
                "Expected MRR per client per projected month. "
                "MRR drops to **0** in months after the client's Final Service Date, "
                "and is **0** in months before the client's Go Live date."
            )
            _cmrr_df = st.session_state.final_dashboards.get('client_mrr', pd.DataFrame())
            if not _cmrr_df.empty:
                # Summary metrics
                _cmrr_cols = [c for c in _cmrr_df.columns if c in meses_proyeccion]
                _cmrr_m1, _cmrr_m2 = st.columns(2)
                _cmrr_m1.metric("Active Clients (M1)",
                                 int((_cmrr_df[_cmrr_cols[0]] > 0).sum()) if _cmrr_cols else 0)
                _cmrr_m2.metric("Total MRR M1 ($)",
                                 f"${_cmrr_df[_cmrr_cols[0]].sum():,.0f}" if _cmrr_cols else "$0")
                # Format MRR columns as currency
                _cmrr_col_cfg = {c: st.column_config.NumberColumn(c, format="$%.0f") for c in _cmrr_cols}
                _cmrr_col_cfg['MRR (Base)'] = st.column_config.NumberColumn('MRR (Base)', format="$%.0f")
                st.dataframe(_cmrr_df, use_container_width=True, column_config=_cmrr_col_cfg)

                # Per-POD MRR summary (pivot)
                if 'POD' in _cmrr_df.columns and _cmrr_cols:
                    st.divider()
                    st.markdown("**MRR by POD × Month**")
                    _pod_mrr_summary = (
                        _cmrr_df.groupby('POD')[_cmrr_cols].sum().reset_index()
                    )
                    _pod_mrr_summary.loc[len(_pod_mrr_summary)] = (
                        ['**Total**'] + [_cmrr_df[c].sum() for c in _cmrr_cols]
                    )
                    _pod_mrr_col_cfg = {c: st.column_config.NumberColumn(c, format="$%.0f") for c in _cmrr_cols}
                    st.dataframe(_pod_mrr_summary, use_container_width=True, column_config=_pod_mrr_col_cfg)
            else:
                st.info("Run the cascade to generate Client MRR data.")

        with t_baseline:
            st.write("### Baseline Audit — All Calculated Fields")
            st.caption(
                "One row per client × process × role. Shows every input and calculated field used for M1. "
                "Use this to verify volumes, AHTs, active %, learning curve, shrinkage, and resulting hours/FTEs."
            )
            _bl_df = st.session_state.final_dashboards.get('baseline', pd.DataFrame())
            if _bl_df.empty:
                st.info("No baseline data. Click 'Apply Cascade & Generate Dashboards' to compute.")
            else:
                # Normalize POD in baseline to avoid 'nan' strings
                _bl_df = _bl_df.copy()
                if 'POD' in _bl_df.columns:
                    _bl_df['POD'] = (
                        _bl_df['POD'].fillna('').astype(str).str.strip()
                        .where(lambda s: ~s.str.lower().isin({'nan', 'none', ''}), 'No POD')
                    )
                # Filters
                _bl_c1, _bl_c2, _bl_c3 = st.columns(3)
                with _bl_c1:
                    _bl_pods = ["All"] + sorted(_bl_df['POD'].unique().tolist()) if 'POD' in _bl_df.columns else ["All"]
                    _bl_pod  = st.selectbox("Filter by POD", _bl_pods, key="bl_pod_filter")
                with _bl_c2:
                    _bl_srs  = ["All"] + sorted(_bl_df['Sr. Accountant'].dropna().astype(str).replace('', pd.NA).dropna().unique().tolist())
                    _bl_sr   = st.selectbox("Filter by Sr. Accountant", _bl_srs, key="bl_sr_filter")
                with _bl_c3:
                    _bl_clis = ["All"] + sorted(_bl_df['Client'].dropna().astype(str).unique().tolist())
                    _bl_cli  = st.selectbox("Filter by Client", _bl_clis, key="bl_cli_filter")

                _bl_view = _bl_df.copy()
                if _bl_pod != "All":  _bl_view = _bl_view[_bl_view['POD'] == _bl_pod]
                if _bl_sr  != "All":  _bl_view = _bl_view[_bl_view['Sr. Accountant'].astype(str) == _bl_sr]
                if _bl_cli != "All":  _bl_view = _bl_view[_bl_view['Client'].astype(str) == _bl_cli]

                # Always hide email columns (internal use by Employee Hours only) +
                # drop any remaining fully-blank columns
                _bl_hide_cols = {'Processor Email', 'Reviewer Email'}
                _bl_view = _bl_view[[c for c in _bl_view.columns if c not in _bl_hide_cols]]
                _bl_view = _bl_view.loc[:, ~_bl_view.apply(
                    lambda col: col.replace('', pd.NA).isna().all()
                )]

                st.caption(f"Showing **{len(_bl_view)}** of {len(_bl_df)} rows")
                st.dataframe(_bl_view, use_container_width=True, height=min(700, len(_bl_view) * 35 + 38))

                _bl_buf = BytesIO()
                _bl_view.to_excel(_bl_buf, index=False)
                st.download_button(
                    label="📥 Download Baseline Audit (filtered)",
                    data=_bl_buf.getvalue(),
                    file_name=f"Baseline_Audit_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_baseline"
                )

        # ── EMPLOYEE LEVEL ─────────────────────────────────────────────────────
        with t_employee:
            st.write("### 👤 Employee Level Analysis")
            _hc_el      = st.session_state.get('hc_data', None)
            _el_bl      = st.session_state.final_dashboards.get('baseline', pd.DataFrame())
            _el_cli     = st.session_state.final_dashboards.get('cliente', pd.DataFrame())
            _el_gen     = st.session_state.final_dashboards.get('general', pd.DataFrame())
            _el_hrs_fte = st.session_state.get('calc_data', {}).get('dict_hrs_per_fte', {})
            _el_month_cols = [f"M{i+1} ({meses_proyeccion[i]})" for i in range(6)]

            # ── Overall Role Summary — mirrors Capacity Overview HC section ──────
            _el_role_fte_map_ov = {
                'Accountant I':       'FTEs Accountant I',
                'Accountant II':      'FTEs Accountant II',
                'General Accountant': 'FTEs General Acc.',
                'Sr. Accountant':     'FTEs Sr. Accountant',
            }

            # ── Cascade-aware scope: POD / Sr. filter from Step 3 ────────────
            _ov_cascade_pods = st.session_state.get('_dash_sel_pods', [])
            _ov_cascade_srs  = st.session_state.get('_dash_sel_srs', [])
            _ov_scope_label  = "Overall"
            if _ov_cascade_pods:
                _ov_scope_label = f"POD · {', '.join(_ov_cascade_pods)}"
            elif _ov_cascade_srs:
                _ov_scope_label = f"Sr. · {', '.join(_ov_cascade_srs)}"

            if not _el_gen.empty:
                st.markdown(f"#### 📋 Overall Role Summary — {_ov_scope_label}")
                st.caption(
                    "Matches the Required HC breakdown shown in the Capacity Overview. "
                    "Scoped by the active Step 3 cascade filter (POD / Sr. Accountant). "
                    "Actual HC column requires an HC report upload."
                )
                _el_mcols_ov = [f"M{_ei+1} ({meses_proyeccion[_ei]})" for _ei in range(6)]

                # ── Build required FTEs per role & month, scoped by cascade ──
                _ov_req_by_role = {r: {c: 0.0 for c in _el_mcols_ov} for r in _el_role_fte_map_ov}
                if _ov_cascade_pods or _ov_cascade_srs:
                    # Scope Required FTEs from the client dashboard (hours → FTEs)
                    for _ei in range(6):
                        _mc_o = _el_mcols_ov[_ei]
                        _fcol_ov = f"M{_ei+1} ({meses_proyeccion[_ei]}) - Final Hours"
                        _av_ov   = _el_hrs_fte.get(_ei, 150.0)
                        if not _el_cli.empty and _fcol_ov in _el_cli.columns:
                            _scope_df = _el_cli
                            if _ov_cascade_pods and 'POD' in _scope_df.columns:
                                _scope_df = _scope_df[_scope_df['POD'].astype(str).isin(_ov_cascade_pods)]
                            if _ov_cascade_srs and 'Sr. Accountant' in _scope_df.columns:
                                _scope_df = _scope_df[_scope_df['Sr. Accountant'].astype(str).isin(_ov_cascade_srs)]
                            for _r_ov in _el_role_fte_map_ov:
                                _h = float(_scope_df[_scope_df['Required Role'] == _r_ov][_fcol_ov].sum())
                                _ov_req_by_role[_r_ov][_mc_o] = round(_h / _av_ov, 2) if _av_ov > 0 else 0.0
                else:
                    for _r_ov, _cn_ov in _el_role_fte_map_ov.items():
                        for _ei in range(6):
                            _mc_o = _el_mcols_ov[_ei]
                            _ov_req_by_role[_r_ov][_mc_o] = round(
                                float(_el_gen.iloc[_ei].get(_cn_ov, 0) or 0) if _ei < len(_el_gen) else 0.0, 2
                            )

                # ── Actual HC per role, scoped by cascade ──
                _ov_act_by_role = {r: None for r in _el_role_fte_map_ov}
                if _hc_el:
                    if _ov_cascade_pods:
                        _bpr_ov = _hc_el.get('by_pod_role', pd.DataFrame())
                        if not _bpr_ov.empty:
                            _m_ov = _bpr_ov['POD'].astype(str).isin(_ov_cascade_pods)
                            _bpr_ov = _bpr_ov[_m_ov]
                            for _r_ov in _el_role_fte_map_ov:
                                _rsum = float(_bpr_ov[_bpr_ov['Capacity Role'] == _r_ov]['HC'].sum())
                                _ov_act_by_role[_r_ov] = _rsum
                    elif _ov_cascade_srs:
                        _by_sr_lu = _hc_el.get('by_sr', {})
                        _agg = {r: 0.0 for r in _el_role_fte_map_ov}
                        for _sr_n in _ov_cascade_srs:
                            _sd = _by_sr_lu.get(_sr_n, {})
                            _rd = _sd.get('by_role', {}) if _sd else {}
                            for _r_ov in _el_role_fte_map_ov:
                                _agg[_r_ov] += float(_rd.get(_r_ov, 0) or 0)
                        _ov_act_by_role = {r: _agg[r] for r in _el_role_fte_map_ov}
                    else:
                        _br = _hc_el.get('by_role', {})
                        _ov_act_by_role = {r: float(_br.get(r, 0) or 0) for r in _el_role_fte_map_ov}

                # ── Totals ──
                _ov_total_req = {}
                for _ei in range(6):
                    _mc = _el_mcols_ov[_ei]
                    _ov_total_req[_mc] = round(sum(_ov_req_by_role[r][_mc] for r in _el_role_fte_map_ov), 2)
                _ov_total_act = (sum(v for v in _ov_act_by_role.values() if v is not None)
                                 if _hc_el else None)

                _ov_rows = {}
                _ov_rows['━ Required HC (FTEs)'] = _ov_total_req
                if _ov_total_act is not None:
                    _ov_rows['━ Actual HC (Report)'] = {c: _ov_total_act for c in _el_mcols_ov}
                    _ov_rows['━ Δ (Actual − Required)'] = {
                        c: round(_ov_total_act - _ov_total_req[c], 2) for c in _el_mcols_ov
                    }

                for _r_ov in _el_role_fte_map_ov:
                    _r_req = _ov_req_by_role[_r_ov]
                    _r_act = _ov_act_by_role.get(_r_ov)
                    _ov_rows[f'  · {_r_ov} — Req. FTEs'] = _r_req
                    if _r_act is not None:
                        _ov_rows[f'  · {_r_ov} — Actual HC'] = {c: _r_act for c in _el_mcols_ov}
                        _ov_rows[f'  · {_r_ov} — Δ']         = {c: round(_r_act - _r_req[c], 2) for c in _el_mcols_ov}

                _df_ov_sum = pd.DataFrame(_ov_rows).T
                _df_ov_sum.index.name = "Metric"

                def _ov_style_fn(df):
                    sty = pd.DataFrame('', index=df.index, columns=df.columns)
                    for _idx in df.index:
                        if 'Δ' in _idx:
                            for _col in df.columns:
                                try:
                                    _v = float(df.loc[_idx, _col])
                                    if _v > 0:
                                        sty.loc[_idx, _col] = 'color: green; font-weight: bold'
                                    elif _v < 0:
                                        sty.loc[_idx, _col] = 'color: red; font-weight: bold'
                                except (ValueError, TypeError):
                                    pass
                    return sty

                st.dataframe(
                    _df_ov_sum.style.apply(_ov_style_fn, axis=None).format("{:.2f}"),
                    use_container_width=True,
                    height=min(700, len(_df_ov_sum) * 35 + 60),
                )
                st.divider()

            if _hc_el is None:
                st.info("📂 Upload an HC Weekly Report in the sidebar to enable Employee Level analysis.")
            else:
                _el_has_emails = (
                    not _el_bl.empty
                    and 'Processor Email' in _el_bl.columns
                    and _el_bl['Processor Email'].replace('', pd.NA).notna().any()
                )
                if not _el_has_emails:
                    st.caption(
                        "ℹ️ Re-run **Apply Cascade** to load assignment-based busy hours. "
                        "Showing active HC employees with available hours only."
                    )

                # ── Build HC roster: email → {name, role, pod, attrited} ──
                # Active employees first; then append attrited (terminated) employees
                # with role suffixed by " Att" so capacity planners can see who
                # previously carried load that now needs redistribution.
                _hc_roster = {}
                _hc_det2 = _hc_el.get('detail', pd.DataFrame())
                if not _hc_det2.empty:
                    for _, _hr in _hc_det2[_hc_det2['Capacity Role'].isin(roles_permitidos)].iterrows():
                        _hem = str(_hr.get('Work Email', '')).strip()
                        if _hem and _hem.lower() not in ('nan', 'none', ''):
                            _hc_roster[_hem.lower()] = {
                                'name':     str(_hr.get('Full name', '')).strip(),
                                'email':    _hem,
                                'role':     str(_hr.get('Capacity Role', '')).strip(),
                                'pod':      str(_hr.get('POD', '')).strip(),
                                'attrited': False,
                            }

                # Attrited employees — append with " Att" suffix on role
                _hc_att_det = _hc_el.get('attrited_detail', pd.DataFrame())
                if not _hc_att_det.empty:
                    for _, _har in _hc_att_det.iterrows():
                        _aem = str(_har.get('Work Email', '')).strip()
                        if not _aem or _aem.lower() in ('nan', 'none', ''):
                            continue
                        _aem_k = _aem.lower()
                        # If the person somehow also appears as active, keep the active row
                        if _aem_k in _hc_roster:
                            continue
                        _base_role = str(_har.get('Capacity Role', '')).strip() or 'Other'
                        _hc_roster[_aem_k] = {
                            'name':     str(_har.get('Full name', '')).strip(),
                            'email':    _aem,
                            'role':     f"{_base_role} Att",
                            'pod':      str(_har.get('POD', '')).strip(),
                            'attrited': True,
                        }

                # ── Build ratio cache: (client, role, month_idx) → ratio vs M1 ─
                # Captures fixed-days (flat/learning-curve) vs network-days (proportional)
                _m1_fcol = f"M1 ({meses_proyeccion[0]}) - Final Hours"
                _el_ratio = {}  # (client, role, ei) → ratio
                if not _el_cli.empty and _m1_fcol in _el_cli.columns:
                    for _, _cr in _el_cli.iterrows():
                        _cc = str(_cr.get('Client', '')).strip()
                        _cr2 = str(_cr.get('Required Role', '')).strip()
                        _m1h = float(_cr.get(_m1_fcol, 0) or 0)
                        for _ei in range(6):
                            _mxc = f"M{_ei+1} ({meses_proyeccion[_ei]}) - Final Hours"
                            _mxh = float(_cr.get(_mxc, 0) or 0) if _mxc in _cr.index else _m1h
                            _el_ratio[(_cc, _cr2, _ei)] = _mxh / _m1h if _m1h > 0 else 1.0

                # ── Build assignment dict: email → [(type, client, role, m1_hrs, pod, name)] ─
                _el_assign = {}
                for _, _brow in _el_bl.iterrows():
                    _bpod = str(_brow.get('POD', '')).strip()
                    _bcli = str(_brow.get('Client', '')).strip()

                    _pe  = str(_brow.get('Processor Email', '') or '').strip().lower()
                    _pn  = str(_brow.get('Processor', '') or '').strip()
                    _pr  = str(_brow.get('Processor Role', '') or '').strip()
                    _tp  = float(_brow.get('Total Hrs Proc w/ Shrinkage', 0) or 0)
                    if _pe and _pe not in ('nan', 'none', '') and _tp > 0:
                        _el_assign.setdefault(_pe, []).append(('proc', _bcli, _pr, _tp, _bpod, _pn))

                    _re  = str(_brow.get('Reviewer Email', '') or '').strip().lower()
                    _rn  = str(_brow.get('Reviewer', '') or '').strip()
                    _rr  = str(_brow.get('Reviewer Role', '') or '').strip()
                    _tr  = float(_brow.get('Total Hrs Rev w/ Shrinkage', 0) or 0)
                    if _re and _re not in ('nan', 'none', '') and _tr > 0:
                        _el_assign.setdefault(_re, []).append(('rev', _bcli, _rr, _tr, _bpod, _rn))

                # ── POD selector (built from union of HC + input assignments) ──
                _el_pods_set = set()
                for _inf in _hc_roster.values():
                    if _inf['pod']: _el_pods_set.add(_inf['pod'])
                for _asns in _el_assign.values():
                    for _a in _asns:
                        if _a[4]: _el_pods_set.add(_a[4])
                _el_pods_sorted = ["Overall"] + sorted(_el_pods_set)
                # Auto-select the cascade POD when exactly one POD was filtered —
                # initialises the key only once so the user can still change it later.
                _el_cascade_pods = st.session_state.get('_dash_sel_pods', [])
                if len(_el_cascade_pods) == 1 and _el_cascade_pods[0] in _el_pods_set:
                    if 'el_pod_sel' not in st.session_state:
                        st.session_state['el_pod_sel'] = _el_cascade_pods[0]
                _el_pod = st.selectbox(
                    "Filter by POD (or Overall)",
                    _el_pods_sorted,
                    key="el_pod_sel"
                )

                # ── Role FTE column mapping ─────────────────────────────────────
                _el_role_fte_col = {
                    "Accountant I":       "FTEs Accountant I",
                    "Accountant II":      "FTEs Accountant II",
                    "General Accountant": "FTEs General Acc.",
                    "Sr. Accountant":     "FTEs Sr. Accountant",
                }

                # ── Role Summary — Required FTEs vs Actual HC ──────────────────
                st.markdown("#### Role Summary — Required FTEs vs Actual HC")
                _el_by_pr = _hc_el.get('by_pod_role', pd.DataFrame()) if _hc_el else pd.DataFrame()
                if _el_pod == "Overall":
                    _el_act = {r: float(_hc_el['by_role'].get(r, 0)) for r in roles_permitidos} if _hc_el else {}
                else:
                    _el_act = {}
                    if not _el_by_pr.empty:
                        _bpr_f = _el_by_pr[_el_by_pr['POD'].astype(str) == _el_pod]
                        for r in roles_permitidos:
                            _m = _bpr_f['Capacity Role'] == r
                            _el_act[r] = float(_bpr_f.loc[_m, 'HC'].sum()) if _m.any() else 0.0

                _el_sum_rows = {}
                for r in roles_permitidos:
                    _el_req_r = {}
                    for _ei in range(6):
                        _mc = _el_month_cols[_ei]
                        if _el_pod == "Overall":
                            _el_req_r[_mc] = float(_el_gen.iloc[_ei].get(_el_role_fte_col[r], 0) or 0) \
                                if not _el_gen.empty and _ei < len(_el_gen) else 0.0
                        else:
                            _el_fcol2 = f"M{_ei+1} ({meses_proyeccion[_ei]}) - Final Hours"
                            _el_av2   = _el_hrs_fte.get(_ei, 150.0)
                            if not _el_cli.empty and _el_fcol2 in _el_cli.columns:
                                _el_pd2 = _el_cli[_el_cli['POD'].astype(str) == _el_pod]
                                _el_h2  = float(_el_pd2[_el_pd2['Required Role'] == r][_el_fcol2].sum())
                                _el_req_r[_mc] = round(_el_h2 / _el_av2, 2) if _el_av2 > 0 else 0.0
                            else:
                                _el_req_r[_mc] = 0.0
                    _el_act_v = _el_act.get(r, 0.0)
                    _el_sum_rows[f"{r} — Required FTEs"] = _el_req_r
                    _el_sum_rows[f"{r} — Actual HC"]      = {c: _el_act_v for c in _el_month_cols}
                    _el_sum_rows[f"{r} — Δ (Actual−Req)"] = {
                        c: round(_el_act_v - _el_req_r[c], 2) for c in _el_month_cols
                    }

                if _el_sum_rows:
                    _df_el_sum = pd.DataFrame(_el_sum_rows).T
                    _df_el_sum.index.name = "Role / Metric"

                    def _el_style_fn(df):
                        sty = pd.DataFrame('', index=df.index, columns=df.columns)
                        for idx in df.index:
                            if 'Δ' in idx:
                                for col in df.columns:
                                    try:
                                        v = float(df.loc[idx, col])
                                        if v > 0:
                                            sty.loc[idx, col] = 'color: green; font-weight: bold'
                                        elif v < 0:
                                            sty.loc[idx, col] = 'color: red; font-weight: bold'
                                    except (ValueError, TypeError):
                                        pass
                        return sty

                    st.dataframe(
                        _df_el_sum.style.apply(_el_style_fn, axis=None).format("{:.2f}"),
                        use_container_width=True,
                        height=len(_df_el_sum) * 35 + 38,
                    )

                # ── Employee Hours Forecast ────────────────────────────────────
                st.markdown("#### Employee Hours Forecast")
                st.caption(
                    "Busy Hours = sum of each employee's assigned hours (Processor + Reviewer), "
                    "scaled month-by-month using the same logic as the cascade "
                    "(flat for fixed-days old clients, learning curve for new clients, "
                    "proportional for network-days). "
                    f"**Productive Capacity = available hours × role productivity goal** "
                    f"(Acct I/II = {int(util_acc1*100)}%, Gen. Acc. = {int(util_gen*100)}%, "
                    f"Sr. Acc. = {int(util_sr*100)}%). "
                    "**Hrs Left = productive capacity − busy hours** — negative values flag roles "
                    "that are already over-utilised vs. their productivity target. Employees with "
                    "a role ending in ` Att` are attrited (hours left = 0, all load is unassigned)."
                )

                # Active HC employees (not attrited) + all volume file employees
                _hc_active_set = {e for e, inf in _hc_roster.items() if not inf.get('attrited', False)}
                _all_em = set(_el_assign.keys()) | _hc_active_set

                # Apply POD filter
                if _el_pod != "Overall":
                    _filtered_em = set()
                    for _em in _all_em:
                        _in_hc_pod = _hc_roster.get(_em, {}).get('pod', '') == _el_pod
                        _in_asn_pod = any(a[4] == _el_pod for a in _el_assign.get(_em, []))
                        if _in_hc_pod or _in_asn_pod:
                            _filtered_em.add(_em)
                    _all_em = _filtered_em

                # Role → utilization goal lookup — strips " Att" and " - Alert Relocate" suffixes
                def _role_util(_r):
                    _rk = (_r or '').replace(' Att', '').replace(' - Alert Relocate', '').strip()
                    return utilization_map.get(_rk, util_acc1)

                _el_emp_rows = []
                for _em in sorted(_all_em):
                    _info    = _hc_roster.get(_em, {})
                    _email_d = _info.get('email', _em)
                    _asns    = _el_assign.get(_em, [])

                    # Determine if this employee is active in HC
                    _is_active_hc = _em in _hc_active_set

                    if _is_active_hc:
                        # Role always from HC report for active employees
                        _role  = _info.get('role', '')
                        _pod_d = _info.get('pod', '')
                        _no_active = False
                    else:
                        # Volume-only or not currently active in HC —
                        # derive role from assignments (most common role used)
                        _asn_roles = [a[2] for a in _asns if a[2]]
                        _base_role = max(set(_asn_roles), key=_asn_roles.count) if _asn_roles else 'Unknown'
                        _role  = f"{_base_role} - Alert Relocate"
                        _pod_d = _asns[0][4] if _asns else ''
                        _no_active = True

                    # POD fallback to assignment if missing
                    if not _pod_d and _asns:
                        _pod_d = _asns[0][4]

                    # Role-specific productivity target (85 / 80 / 50%)
                    _util_goal = _role_util(_role)

                    _rd = {'Email': _email_d, 'Role': _role, 'POD': _pod_d}
                    for _ei in range(6):
                        _avail    = _el_hrs_fte.get(_ei, 150.0)
                        _prod_cap = _avail * _util_goal
                        _busy     = 0.0
                        # No Active employees: still show their assigned load so planners
                        # can see hours that need redistribution, but no productive capacity.
                        if not _no_active:
                            for _atype, _acli, _arole, _am1, _apod, _aname in _asns:
                                if _el_pod != "Overall" and _apod != _el_pod:
                                    continue
                                _ratio = _el_ratio.get((_acli, _arole, _ei), 1.0)
                                _busy += _am1 * _ratio
                        _busy = round(_busy, 1)
                        _mc   = _el_month_cols[_ei]
                        _rd[f"{_mc} Busy Hrs"] = _busy
                        if _no_active:
                            # No productive capacity; Util % left blank
                            _rd[f"{_mc} Hrs Left"] = 0.0
                            _rd[f"{_mc} Util %"]   = None
                        else:
                            _rd[f"{_mc} Hrs Left"] = round(_prod_cap - _busy, 1)
                            _rd[f"{_mc} Util %"]   = round(_busy / _prod_cap * 100, 1) if _prod_cap > 0 else 0.0
                    _el_emp_rows.append(_rd)

                if _el_emp_rows:
                    _df_el_emp = pd.DataFrame(_el_emp_rows).sort_values(['POD', 'Role', 'Email'])
                    st.session_state['_s3_emp_level_df'] = _df_el_emp

                    st.dataframe(
                        _df_el_emp,
                        use_container_width=True,
                        height=min(700, len(_df_el_emp) * 35 + 38),
                    )

                    _el_buf = BytesIO()
                    _df_el_emp.to_excel(_el_buf, index=False)
                    st.download_button(
                        label="📥 Download Employee Level Data",
                        data=_el_buf.getvalue(),
                        file_name=f"Employee_Level_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_emp_level"
                    )
                else:
                    st.info("No employee data found. Ensure the HC report is uploaded and cascade has been run.")

        # Helper: sanitize a string into a valid Excel sheet name (max 31 chars)
        def _xl_sheet(name, prefix=''):
            import re
            _s = re.sub(r'[/\\?\*\[\]:]', '_', str(name))
            _s = (prefix + _s) if prefix else _s
            return _s[:31]

        from io import BytesIO
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            # Tab 1: Capacity Overview — Waterfall (transposed: metrics as rows, months as columns)
            _wf_exp = st.session_state.get('_wf_overall_export', pd.DataFrame())
            if not _wf_exp.empty:
                _wf_exp.to_excel(writer, sheet_name='Capacity_Overview_Waterfall')
            # Tab 1b: By POD waterfalls — one sheet per POD
            _wf_pods_exp = st.session_state.get('_wf_pod_all_export', {})
            for _pn_exp, _df_pw_exp in sorted(_wf_pods_exp.items()):
                if not _df_pw_exp.empty:
                    _df_pw_exp.to_excel(writer, sheet_name=_xl_sheet(_pn_exp, 'WF_'))

            # Tab 1c: All Sr. Accountant waterfalls — full cascade format, stacked in one sheet
            _cli_exp   = st.session_state.final_dashboards.get('cliente', pd.DataFrame())
            _hc_xp     = st.session_state.get('hc_data', None)
            _duc_xp    = st.session_state.get('df_clients_unique', pd.DataFrame())
            _raw_xp    = st.session_state.get('df_clean', pd.DataFrame())
            _rb_xp     = st.session_state.get('calc_data', {}).get('df_resumen_base', pd.DataFrame())
            _wdx       = st.session_state.get('calc_data', {}).get('dict_workable_days', {})
            if not _cli_exp.empty and 'Sr. Accountant' in _cli_exp.columns:
                _all_srs_exp = sorted(
                    _cli_exp['Sr. Accountant'].dropna().astype(str)
                    .replace('', pd.NA).dropna().unique().tolist()
                )
                def _srvx(df, col):
                    return float(df[col].sum()) if col in df.columns else 0.0
                def _srole_x(df, role, col):
                    _s = df[df['Required Role'] == role] if 'Required Role' in df.columns else pd.DataFrame()
                    return float(_s[col].sum()) if col in _s.columns else 0.0
                def _safe_nx(x):
                    try: return pd.to_numeric(x, errors='coerce').fillna(0.0)
                    except: return pd.Series([0.0])
                _sr_stacked_frames = []
                for _sr_nm in _all_srs_exp:
                    _sr_df   = _cli_exp[_cli_exp['Sr. Accountant'].astype(str) == _sr_nm].copy()
                    # HC for this Sr.
                    _sr_hcd  = {}
                    if _hc_xp:
                        _sr_hcd = (
                            _hc_xp.get('by_sr_email', {}).get(str(_sr_nm).strip().lower()) or
                            _hc_xp.get('by_sr', {}).get(_sr_nm) or {}
                        )
                    _hcx_tot  = _sr_hcd.get('total', None)
                    _hcx_rl   = _sr_hcd.get('by_role', {})
                    _hcx_acc1 = float(_hcx_rl.get('Accountant I', 0))       if _hcx_tot else None
                    _hcx_acc2 = float(_hcx_rl.get('Accountant II', 0))      if _hcx_tot else None
                    _hcx_gen  = float(_hcx_rl.get('General Accountant', 0)) if _hcx_tot else None
                    _hcx_sr1  = 1.0                                           if _hcx_tot else None
                    # Clients for this Sr.
                    _sr_cli_set   = set(_sr_df['Client'].dropna().astype(str).str.strip().unique()) if 'Client' in _sr_df.columns else set()
                    _sr_cli_lower = {c.lower() for c in _sr_cli_set}
                    _sr_rw = (
                        _raw_xp[_raw_xp['client_name'].astype(str).str.strip().str.lower().isin(_sr_cli_lower)]
                        if not _raw_xp.empty and 'client_name' in _raw_xp.columns else pd.DataFrame()
                    )
                    # Property / door count (static)
                    if not _sr_rw.empty:
                        _sr_snap = _sr_rw.groupby('client_name', as_index=False).agg({
                            c: 'first' for c in ['Res Prop','Commercial Properties','Res doors','Commercial Doors']
                            if c in _sr_rw.columns
                        })
                        _prop_cx = int(_safe_nx(_sr_snap.get('Res Prop', 0)).sum() + _safe_nx(_sr_snap.get('Commercial Properties', 0)).sum())
                        _door_cx = int(_safe_nx(_sr_snap.get('Res doors', 0)).sum() + _safe_nx(_sr_snap.get('Commercial Doors', 0)).sum())
                    else:
                        _prop_cx = _door_cx = 0
                    _rows_x = {}
                    for _ii, _ms in enumerate(meses_proyeccion):
                        _c_fin  = f"M{_ii+1} ({_ms}) - Final Hours"
                        _c_fte  = f"M{_ii+1} ({_ms}) - Final FTEs"
                        _c_base = f"M{_ii+1} ({_ms}) - Base Hours"
                        _c_prod = f"M{_ii+1} ({_ms}) - Productive Hours"
                        _c_save = f"M{_ii+1} ({_ms}) - Auto Saving (Hrs)"
                        _c_plus = f"M{_ii+1} ({_ms}) - Adjustments (+) Hrs"
                        _c_minus= f"M{_ii+1} ({_ms}) - Adjustments (-) Hrs"
                        _fin_h  = _srvx(_sr_df, _c_fin)
                        _fte_t  = _srvx(_sr_df, _c_fte)
                        _base_h = _srvx(_sr_df, _c_base)
                        _save_h = _srvx(_sr_df, _c_save)
                        _adj_h  = _srvx(_sr_df, _c_plus) - _srvx(_sr_df, _c_minus)
                        # Productive hours
                        if not _rb_xp.empty and 'Client' in _rb_xp.columns and _c_prod in _rb_xp.columns:
                            _prod_h = float(_rb_xp[_rb_xp['Client'].astype(str).str.strip().str.lower().isin(_sr_cli_lower)][_c_prod].sum())
                        else:
                            _prod_h = _base_h
                        _shrink_h = max(0.0, _base_h - _prod_h)
                        # New / churn MRR + hours
                        _sr_new_h = _sr_churn_h = _sr_mrr = _sr_new_mrr = _sr_churn_mrr = 0.0
                        if not _duc_xp.empty and 'client_name' in _duc_xp.columns:
                            try:
                                _mes_d = today + relativedelta(months=_month_offsets[_ii])
                                _s_m   = pd.Timestamp(_mes_d.replace(day=1).date())
                                _e_m   = pd.Timestamp((_s_m + relativedelta(months=1) - relativedelta(days=1)).date())
                                _gl_d  = pd.to_datetime(_duc_xp.get('Go Live',            pd.Series(dtype='datetime64[ns]')), errors='coerce')
                                _fsd_d = pd.to_datetime(_duc_xp.get('Final Service Date', pd.Series(dtype='datetime64[ns]')), errors='coerce')
                                _mask_act = ((_gl_d.isna() | (_gl_d <= _e_m)) & (_fsd_d.isna() | (_fsd_d >= _s_m)) & _duc_xp['client_name'].isin(_sr_cli_set))
                                _sr_mrr = float(_duc_xp.loc[_mask_act, 'MRR'].sum()) if 'MRR' in _duc_xp.columns else 0.0
                                _mask_new = (_gl_d >= _s_m) & (_gl_d <= _e_m) & _duc_xp['client_name'].isin(_sr_cli_set)
                                _new_cn  = set(_duc_xp.loc[_mask_new, 'client_name'].astype(str).str.strip().str.lower())
                                _new_r   = _sr_df[_sr_df['Client'].astype(str).str.strip().str.lower().isin(_new_cn)]
                                _sr_new_h   = float(_new_r[_c_base].sum()) if _c_base in _new_r.columns else 0.0
                                _sr_new_mrr = float(_duc_xp.loc[_mask_new, 'MRR'].sum()) if 'MRR' in _duc_xp.columns else 0.0
                                _mask_churn = (_fsd_d >= _s_m) & (_fsd_d <= _e_m) & _duc_xp['client_name'].isin(_sr_cli_set)
                                _churn_cn   = set(_duc_xp.loc[_mask_churn, 'client_name'].astype(str).str.strip().str.lower())
                                _churn_r    = _sr_df[_sr_df['Client'].astype(str).str.strip().str.lower().isin(_churn_cn)]
                                _sr_churn_h   = float(_churn_r[_c_base].sum()) if _c_base in _churn_r.columns else 0.0
                                _sr_churn_mrr = float(_duc_xp.loc[_mask_churn, 'MRR'].sum()) if 'MRR' in _duc_xp.columns else 0.0
                            except Exception:
                                pass
                        # HC delta
                        _dx_tot  = round(_hcx_tot  - _fte_t,                                   2) if _hcx_tot  is not None else None
                        _dx_acc1 = round(_hcx_acc1 - _srole_x(_sr_df,'Accountant I',       _c_fte), 2) if _hcx_acc1 is not None else None
                        _dx_acc2 = round(_hcx_acc2 - _srole_x(_sr_df,'Accountant II',      _c_fte), 2) if _hcx_acc2 is not None else None
                        _dx_gen  = round(_hcx_gen  - _srole_x(_sr_df,'General Accountant', _c_fte), 2) if _hcx_gen  is not None else None
                        _dx_sr   = round(_hcx_sr1  - _srole_x(_sr_df,'Sr. Accountant',     _c_fte), 2) if _hcx_sr1  is not None else None
                        # Tickets / AHT (active clients this month)
                        _s_m2 = pd.Timestamp((today + relativedelta(months=_month_offsets[_ii])).replace(day=1).date())
                        _e_m2 = pd.Timestamp((_s_m2 + relativedelta(months=1) - relativedelta(days=1)).date())
                        if not _sr_rw.empty:
                            _gl_r  = pd.to_datetime(_sr_rw.get('Go Live',            pd.Series(dtype='datetime64[ns]')), errors='coerce')
                            _fsd_r = pd.to_datetime(_sr_rw.get('Final Service Date', pd.Series(dtype='datetime64[ns]')), errors='coerce')
                            _sr_rw_m = _sr_rw[(_gl_r.isna() | (_gl_r <= _e_m2)) & (_fsd_r.isna() | (_fsd_r >= _s_m2))]
                        else:
                            _sr_rw_m = _sr_rw
                        _ptix = float(_safe_nx(_sr_rw_m.get('Closed tickets with Proc time', 0)).sum()) if not _sr_rw_m.empty else 0.0
                        _rtix = float(_safe_nx(_sr_rw_m.get('Closed tickets with rev time',  0)).sum()) if not _sr_rw_m.empty else 0.0
                        _paht = float(_safe_nx(_sr_rw_m.get('>>> FINAL Capacity Proc AHT',   0)).sum()) if not _sr_rw_m.empty else 0.0
                        _raht = float(_safe_nx(_sr_rw_m.get('>>> FINAL Capacity Rev AHT',    0)).sum()) if not _sr_rw_m.empty else 0.0
                        _tot_tix = _ptix + _rtix
                        _avg_aht = (_ptix * _paht + _rtix * _raht) / _tot_tix if _tot_tix > 0 else 0.0
                        _cap_prod_x = round(_prod_h / _fin_h * 100, 2) if _fin_h > 0 else None
                        _col_x = _ms   # month name as column key
                        _rows_x.setdefault("━ Required Hours",               {})[_col_x] = round(_fin_h, 1)
                        _rows_x.setdefault("  Current Customer Hours",       {})[_col_x] = round(_prod_h, 1) or None
                        _rows_x.setdefault("  Shrinkage (Hrs)",              {})[_col_x] = round(_shrink_h, 1) if _shrink_h > 0 else None
                        _rows_x.setdefault("  (+) New Customer Hours",       {})[_col_x] = round(_sr_new_h, 1) if _sr_new_h > 0 else None
                        _rows_x.setdefault("  (-) Confirmed Churn (Hrs)",    {})[_col_x] = round(_sr_churn_h, 1) if _sr_churn_h > 0 else None
                        _rows_x.setdefault("  (-) Automations",              {})[_col_x] = round(_save_h, 1) if _save_h > 0 else None
                        _rows_x.setdefault("  (+) Manual Adjustments",       {})[_col_x] = round(_adj_h, 1) if _adj_h != 0 else None
                        _rows_x.setdefault("(/) Capacity Productivity",      {})[_col_x] = _cap_prod_x
                        _rows_x.setdefault("━ Required HC (FTEs)",           {})[_col_x] = round(_fte_t, 2)
                        _rows_x.setdefault("  · Accountant I",               {})[_col_x] = round(_srole_x(_sr_df,'Accountant I',       _c_fte), 2)
                        _rows_x.setdefault("  · Accountant II",              {})[_col_x] = round(_srole_x(_sr_df,'Accountant II',      _c_fte), 2)
                        _rows_x.setdefault("  · General Accountant",         {})[_col_x] = round(_srole_x(_sr_df,'General Accountant', _c_fte), 2)
                        _rows_x.setdefault("  · Sr. Accountant",             {})[_col_x] = round(_srole_x(_sr_df,'Sr. Accountant',     _c_fte), 2)
                        _rows_x.setdefault("━ Actual HC (Report)",           {})[_col_x] = _hcx_tot
                        _rows_x.setdefault("  · Accountant I (actual)",      {})[_col_x] = _hcx_acc1
                        _rows_x.setdefault("  · Accountant II (actual)",     {})[_col_x] = _hcx_acc2
                        _rows_x.setdefault("  · General Acc. (actual)",      {})[_col_x] = _hcx_gen
                        _rows_x.setdefault("  · Sr. Accountant (actual)",    {})[_col_x] = _hcx_sr1
                        _rows_x.setdefault("━ HC Δ (Actual − Required)",     {})[_col_x] = _dx_tot
                        _rows_x.setdefault("  · Δ Accountant I",             {})[_col_x] = _dx_acc1
                        _rows_x.setdefault("  · Δ Accountant II",            {})[_col_x] = _dx_acc2
                        _rows_x.setdefault("  · Δ General Accountant",       {})[_col_x] = _dx_gen
                        _rows_x.setdefault("  · Δ Sr. Accountant",           {})[_col_x] = _dx_sr
                        _rows_x.setdefault("━ MRR ($)",                      {})[_col_x] = round(_sr_mrr, 2) if _sr_mrr else None
                        _rows_x.setdefault("  (+) New MRR ($)",              {})[_col_x] = round(_sr_new_mrr, 2) if _sr_new_mrr else None
                        _rows_x.setdefault("  (-) Churn MRR ($)",            {})[_col_x] = round(_sr_churn_mrr, 2) if _sr_churn_mrr else None
                        _rows_x.setdefault("━ Property Count",               {})[_col_x] = _prop_cx or None
                        _rows_x.setdefault("  Doors",                        {})[_col_x] = _door_cx or None
                        _rows_x.setdefault("  Tickets to Process",           {})[_col_x] = int(_ptix) if _ptix else None
                        _rows_x.setdefault("  Tickets to Review",            {})[_col_x] = int(_rtix) if _rtix else None
                        _rows_x.setdefault("  AHT (min)",                    {})[_col_x] = round(_avg_aht, 1) if _avg_aht else None
                        _rows_x.setdefault("━ Working Days",                 {})[_col_x] = _wdx.get(_ii, 21)
                        # Per-client FTE breakdown
                        for _cn in sorted(_sr_cli_set):
                            _cli_sub_x = _sr_df[_sr_df['Client'].astype(str).str.strip() == _cn]
                            _rows_x.setdefault(f"  · {_cn}", {})[_col_x] = round(float(_cli_sub_x[_c_fte].sum()) if _c_fte in _cli_sub_x.columns else 0.0, 2)
                    _df_sr_blk = pd.DataFrame(_rows_x, index=meses_proyeccion).T
                    _hdr_frame  = pd.DataFrame({_ms: [_sr_nm] for _ms in meses_proyeccion}, index=[f'▶ {_sr_nm}'])
                    _blank_frame = pd.DataFrame({_ms: [None]   for _ms in meses_proyeccion}, index=[''])
                    _sr_stacked_frames.extend([_hdr_frame, _df_sr_blk, _blank_frame])
                if _sr_stacked_frames:
                    _combined_sr = pd.concat(_sr_stacked_frames)
                    _combined_sr.index.name = 'Sr. Accountant / Metric'
                    _combined_sr.to_excel(writer, sheet_name='Sr_Accountant_Waterfalls')
            # Tab 2: General Waterfall Summary (raw month-level data)
            st.session_state.final_dashboards['general'].to_excel(writer, index=False, sheet_name='General_Summary')
            # Tab 3: Summary by POD
            st.session_state.final_dashboards['pod'].to_excel(writer, index=False, sheet_name='Summary_by_POD')
            # Tab 4: POD x Sr. Accountant — aggregated by POD + Sr. (no Client column)
            _cli_pod_sr = st.session_state.final_dashboards['cliente'].copy()
            _psr_grp_cols = [c for c in ['POD', 'Sr. Accountant', 'Required Role'] if c in _cli_pod_sr.columns]
            _psr_num_cols = [c for c in _cli_pod_sr.columns if c not in _psr_grp_cols and c != 'Client']
            if _psr_grp_cols:
                _cli_pod_sr = _cli_pod_sr.groupby(_psr_grp_cols, as_index=False)[_psr_num_cols].sum()
                _cli_pod_sr = _cli_pod_sr.sort_values(_psr_grp_cols)
            _cli_pod_sr.to_excel(writer, index=False, sheet_name='POD_x_SrAccountant')
            # Tab 5: Client & Role Summary (Cascade) — full detail with CLIENT TOTAL rows
            _cli_raw_exp = st.session_state.final_dashboards['cliente'].copy()
            _cli_grp_exp = [c for c in ['POD', 'Sr. Accountant', 'Client', 'Required Role'] if c in _cli_raw_exp.columns]
            _cli_id_exp  = [c for c in ['POD', 'Sr. Accountant', 'Client'] if c in _cli_raw_exp.columns]
            _cli_num_exp = [c for c in _cli_raw_exp.columns if c not in _cli_grp_exp]
            _cli_exp_frames = []
            _cli_grp_by_exp = [c for c in ['POD', 'Sr. Accountant', 'Client'] if c in _cli_raw_exp.columns]
            for _ck_exp, _cs_exp in _cli_raw_exp.groupby(_cli_grp_by_exp, sort=False):
                _cli_exp_frames.append(_cs_exp)
                _tot_exp = {c: '' for c in _cli_raw_exp.columns}
                for _ic in _cli_id_exp:
                    _tot_exp[_ic] = _cs_exp[_ic].iloc[0] if not _cs_exp.empty else ''
                if 'Required Role' in _tot_exp:
                    _tot_exp['Required Role'] = '>>> CLIENT TOTAL'
                for _nc in _cli_num_exp:
                    try:    _tot_exp[_nc] = _cs_exp[_nc].sum()
                    except: _tot_exp[_nc] = ''
                _cli_exp_frames.append(pd.DataFrame([_tot_exp]))
            _cli_detail_exp = pd.concat(_cli_exp_frames, ignore_index=True) if _cli_exp_frames else _cli_raw_exp
            _cli_detail_exp.to_excel(writer, index=False, sheet_name='Client_Role_Detail')
            # Tab 6: Baseline Audit — most detailed level (drop internal email cols)
            if not st.session_state.final_dashboards.get('baseline', pd.DataFrame()).empty:
                _bl_exp = st.session_state.final_dashboards['baseline'].copy()
                _bl_exp = _bl_exp[[c for c in _bl_exp.columns if c not in {'Processor Email', 'Reviewer Email'}]]
                _bl_exp.to_excel(writer, index=False, sheet_name='Baseline_Audit')
            # Tab 7: Employee Level
            _el_df_exp = st.session_state.get('_s3_emp_level_df', pd.DataFrame())
            if not _el_df_exp.empty:
                _el_df_exp.to_excel(writer, index=False, sheet_name='Employee_Level')
            # Tab 8: Client MRR by Month
            _cmrr_exp = st.session_state.final_dashboards.get('client_mrr', pd.DataFrame())
            if not _cmrr_exp.empty:
                _cmrr_exp.to_excel(writer, index=False, sheet_name='Client_MRR')

        _n_pod_sheets = len(st.session_state.get('_wf_pod_all_export', {}))
        _n_srs_exp = len(
            st.session_state.final_dashboards.get('cliente', pd.DataFrame())
            ['Sr. Accountant'].dropna().astype(str).replace('', pd.NA).dropna().unique()
        ) if 'cliente' in st.session_state.get('final_dashboards', {}) else 0
        # ── Cache the export buffer keyed by role mode ───────────────────────────
        _mode_tag_exp   = st.session_state.get('_cascade_role_mode', 'ideal')
        _other_tag_exp  = 'real' if _mode_tag_exp == 'ideal' else 'ideal'
        st.session_state[f'_cascade_export_buf_{_mode_tag_exp}'] = output.getvalue()

        # ── Mode-mismatch warning ─────────────────────────────────────────────────
        # Check if the user has changed the radio since the last cascade run
        _cur_radio_real = st.session_state.get('s3_role_mode_radio', '').startswith("👥")
        _cur_radio_tag  = 'real' if _cur_radio_real else 'ideal'
        _mode_lbl_map   = {'ideal': '🎯 Ideal Pairs', 'real': '👥 Real Roles'}
        if _cur_radio_tag != _mode_tag_exp:
            st.warning(
                f"⚠️ **Mode mismatch** — the last cascade ran in **{_mode_lbl_map[_mode_tag_exp]}** mode, "
                f"but Step 3 is now set to **{_mode_lbl_map[_cur_radio_tag]}**. "
                f"Expand **Step 3 → Apply Cascade** to regenerate results in {_mode_lbl_map[_cur_radio_tag]} mode."
            )
        else:
            st.caption(
                f"Last cascade: **{_mode_lbl_map[_mode_tag_exp]}** mode · "
                f"Includes: Overall waterfall · {_n_pod_sheets} POD waterfall sheet(s) · "
                f"1 Sr. Accountant sheet ({_n_srs_exp} Srs. stacked) · "
                f"General Summary · POD×Sr. Summary · Client Detail · Client MRR · Baseline Audit · Employee Level."
            )
        # ── Two download buttons: current mode + cached other mode ────────────────
        _this_lbl  = _mode_lbl_map[_mode_tag_exp]
        _other_lbl = _mode_lbl_map[_other_tag_exp]
        _exp_c1, _exp_c2 = st.columns(2)
        with _exp_c1:
            # Stable key — does NOT change between ideal/real so Streamlit never
            # destroys/recreates the widget; the label and file name carry mode info.
            st.download_button(
                label=f"📥 Download — {_this_lbl}",
                data=output.getvalue(),
                file_name=f"Capacity_Projection_Cascade_ROI_{datetime.now().strftime('%Y%m%d')}_{_mode_tag_exp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                key="_dl_btn_primary",
            )
        with _exp_c2:
            _other_buf = st.session_state.get(f'_cascade_export_buf_{_other_tag_exp}')
            if _other_buf:
                st.download_button(
                    label=f"📥 Download — {_other_lbl}",
                    data=_other_buf,
                    file_name=f"Capacity_Projection_Cascade_ROI_{datetime.now().strftime('%Y%m%d')}_{_other_tag_exp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="_dl_btn_secondary",
                )
            else:
                _other_mode_label = _mode_lbl_map[_other_tag_exp]
                st.info(f"Run cascade in **{_other_mode_label}** mode to enable this download.")

        # ── Volume Input Export ────────────────────────────────────────────────
        # Exports the fully-updated volume input (Master DB + AI clients, with
        # all HubSpot / reconciliation overrides applied) so the same run can be
        # re-loaded without repeating the HubSpot sync and reconciliation steps.
        _df_vol_exp = st.session_state.get('df_vol_export', pd.DataFrame())
        if not _df_vol_exp.empty:
            _vol_export_cols = [
                'client_name', 'type', 'subtype',
                'processor', 'Proc Role', 'reviewer', 'Rev Role',
                'Closed tickets with Proc time', 'Closed tickets with rev time',
                '>>> FINAL Capacity Proc AHT', '>>> FINAL Capacity Rev AHT',
                'Capacity Processing Hours', 'Capacity reviewing hours', 'Capacity Hours spent',
                'Ideal Proc', 'Ideal Rev', 'Volume Variation %',
                'Res doors', 'Res Prop', 'Commercial Properties', 'Commercial Doors',
                'SQFT Commercial', 'Corp Books', 'PMS',
                'MRR', 'Status', 'Go Live', 'Final Service Date',
                'POD', 'Sr. Accountant',
            ]
            # Keep only columns that actually exist in the exported df
            _vol_cols_present = [c for c in _vol_export_cols if c in _df_vol_exp.columns]
            _df_vol_out = _df_vol_exp[_vol_cols_present].copy()
            # Sort for readability: POD → Sr. → client → process
            _vol_sort = [c for c in ['POD', 'Sr. Accountant', 'client_name', 'type', 'subtype'] if c in _df_vol_out.columns]
            if _vol_sort:
                _df_vol_out = _df_vol_out.sort_values(_vol_sort).reset_index(drop=True)
            # Format date columns as YYYY-MM-DD strings for clean Excel output
            for _dc in ['Go Live', 'Final Service Date']:
                if _dc in _df_vol_out.columns:
                    _df_vol_out[_dc] = pd.to_datetime(_df_vol_out[_dc], errors='coerce').dt.strftime('%Y-%m-%d')

            _vol_buf = BytesIO()
            with pd.ExcelWriter(_vol_buf, engine='xlsxwriter') as _vol_xw:
                _df_vol_out.to_excel(_vol_xw, sheet_name='Volume_Input', index=False)
            st.download_button(
                label="📤 Export Volume Input (with all updates)",
                data=_vol_buf.getvalue(),
                file_name=f"Volume_Input_Updated_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Re-loadable volume file with all HubSpot, reconciliation, and AI-client updates baked in. "
                     "Upload this as the Master DB next time to skip the sync steps.",
            )


# ==========================================
# STEP 4: CAPACITY SCENARIO PLANNER
# ==========================================
if (
    "final_dashboards" in st.session_state
    and "calc_data" in st.session_state
    and st.session_state.get("view_mode_radio") == "🎮 Complete + Playground"
):
    st.divider()
    st.subheader("📈 Step 4: Capacity Scenario Planner")
    st.markdown(
        "Select **Overall** or a specific **POD**, then adjust Actual HC by role (Ramp Up / Down) "
        "and MRR (New / Churn) to explore how the metrics change. "
        "All adjustments **forward-fill** to upcoming months."
    )

    # Init scenario params from current globals (runs after Global Parameters tab sets the values)
    if "s4v2_params_df" not in st.session_state:
        st.session_state.s4v2_params_df = pd.DataFrame([
            {"Role": "Accountant I",       "Utilization (%)": round(util_acc1 * 100, 1), "Cost / Month ($)": cost_acc1},
            {"Role": "Accountant II",      "Utilization (%)": round(util_acc1 * 100, 1), "Cost / Month ($)": cost_acc2},
            {"Role": "General Accountant", "Utilization (%)": round(util_gen  * 100, 1), "Cost / Month ($)": cost_gen},
            {"Role": "Sr. Accountant",     "Utilization (%)": round(util_sr   * 100, 1), "Cost / Month ($)": cost_sr},
        ])

    @st.fragment
    def _s4v2_fragment():
        # ── Scope selector ──────────────────────────────────────────────────
        # Build Sr. Accountant list from final_dashboards['cliente']
        _fd_cli = st.session_state.final_dashboards.get('cliente', pd.DataFrame())
        _lista_srs_s4 = []
        if not _fd_cli.empty and 'Sr. Accountant' in _fd_cli.columns:
            _lista_srs_s4 = sorted(
                _fd_cli['Sr. Accountant'].dropna().astype(str)
                .replace('', pd.NA).dropna().unique().tolist()
            )
        with st.expander("📍 Scope & Parameters", expanded=True):
            # Always derive available PODs from the cascade result so the selector
            # only shows PODs that were actually computed — respects the POD filter
            # selected at Step 0.  Fall back to the full list only if no cascade data.
            _fd_pod_tmp = st.session_state.get('final_dashboards', {}).get('pod', pd.DataFrame())
            if not _fd_pod_tmp.empty and 'POD' in _fd_pod_tmp.columns:
                _s4_lista_pods = sorted(
                    _fd_pod_tmp['POD'].fillna('').astype(str).str.strip()
                    .where(lambda s: ~s.str.lower().isin({'nan', 'none', ''}))
                    .dropna().unique().tolist()
                )
            else:
                # Fallback: full pod list (no cascade data yet)
                _s4_lista_pods = lista_pods or st.session_state.get('_lista_pods', [])

            # ── Respect sidebar POD / Sr. filters ───────────────────────────
            # When a POD or Sr. filter is active from Data Load & Filters,
            # "Overall" is not meaningful — only show the matching scope(s).
            _s4_filt_pods = st.session_state.get('_filt_pods', [])
            _s4_filt_srs  = st.session_state.get('_filt_srs',  [])
            if _s4_filt_pods:
                _s4_scope_opts = [p for p in _s4_lista_pods if p in _s4_filt_pods] or _s4_lista_pods
                _s4_scope_hint = "Filtered to selected POD(s) from sidebar"
            elif _s4_filt_srs:
                _s4_scope_opts = [s for s in _lista_srs_s4 if s in _s4_filt_srs] or _lista_srs_s4
                _s4_scope_hint = "Filtered to selected Sr. Accountant(s) from sidebar"
            else:
                _s4_scope_opts = ["Overall"] + _s4_lista_pods + (_lista_srs_s4 if _lista_srs_s4 else [])
                _s4_scope_hint = "Select Overall, a specific POD, or a Sr. Accountant to scope the scenario"

            # If the previously stored scope is no longer valid (e.g. filter changed), reset it
            _s4_cur_scope = st.session_state.get('s4v2_scope', _s4_scope_opts[0])
            if _s4_cur_scope not in _s4_scope_opts:
                st.session_state['s4v2_scope'] = _s4_scope_opts[0]

            _scope = st.selectbox(
                "📊 Scope",
                _s4_scope_opts,
                key="s4v2_scope",
                help=_s4_scope_hint,
            )
            # Determine scope type
            _scope_is_pod = _scope in _s4_lista_pods
            _scope_is_sr  = _scope in _lista_srs_s4 and not _scope_is_pod

            # ── Scenario Parameters editor ───────────────────────────────────────
            st.markdown("#### ⚙️ Scenario Parameters — Utilization & Cost per Role")
            st.caption(
                "Initialized from Global Parameters. Edit freely for this scenario — "
                "changes here only affect Step 4 calculations, not the main cascade."
            )
            _prst_col, _ = st.columns([1, 4])
            if _prst_col.button("↩ Reset to Global Parameters", key="s4v2_reset_params", use_container_width=True):
                st.session_state.s4v2_params_df = pd.DataFrame([
                    {"Role": "Accountant I",       "Utilization (%)": round(util_acc1 * 100, 1), "Cost / Month ($)": cost_acc1},
                    {"Role": "Accountant II",      "Utilization (%)": round(util_acc1 * 100, 1), "Cost / Month ($)": cost_acc2},
                    {"Role": "General Accountant", "Utilization (%)": round(util_gen  * 100, 1), "Cost / Month ($)": cost_gen},
                    {"Role": "Sr. Accountant",     "Utilization (%)": round(util_sr   * 100, 1), "Cost / Month ($)": cost_sr},
                ])
                st.rerun(scope="fragment")
            st.session_state.s4v2_params_df = st.data_editor(
                st.session_state.s4v2_params_df,
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                key="s4v2_params_ed",
                column_config={
                    "Role":             st.column_config.TextColumn("Role", disabled=True, width="medium"),
                    "Utilization (%)":  st.column_config.NumberColumn("Utilization (%)",  min_value=0.1, max_value=100.0, format="%.1f"),
                    "Cost / Month ($)": st.column_config.NumberColumn("Cost / Month ($)", min_value=0.0, format="$%.0f"),
                }
            )
            # Extract scenario-specific util & cost helpers
            _s4p = st.session_state.s4v2_params_df.set_index("Role")
            def _s4_util(role):
                return float(_s4p.loc[role, "Utilization (%)"] / 100.0) if role in _s4p.index else 0.85
            def _s4_cost_val(role):
                return float(_s4p.loc[role, "Cost / Month ($)"]) if role in _s4p.index else 1140.0


        # ── Base data references ─────────────────────────────────────────────
        _fd     = st.session_state.final_dashboards
        _exec   = _fd.get('general', pd.DataFrame())
        _pod_df = _fd.get('pod', pd.DataFrame())
        _cli_df = _fd.get('cliente', pd.DataFrame())
        _hc     = st.session_state.get('hc_data', None)
        _df_raw = st.session_state.get('df_clean', pd.DataFrame())
        _duc    = st.session_state.get('df_clients_unique', pd.DataFrame())
        _wdays_d     = st.session_state.get('calc_data', {}).get('dict_workable_days', {})
        _dict_hrs_fte = st.session_state.get('calc_data', {}).get('dict_hrs_per_fte', {})

        # ── Ensure session state adj dfs have correct structure ──────────────
        _expected_roles = set(roles_permitidos)
        if "s4v2_hc_adj_df" not in st.session_state or \
                set(st.session_state.s4v2_hc_adj_df.get("Role", pd.Series()).unique()) != _expected_roles:
            _hc_init = []
            for _rl in roles_permitidos:
                _hc_init.append({"Confirmed": False, "Direction": "↑ Ramp Up",   "Role": _rl, **{c: 0.0 for c in _s4v2_mc}})
                _hc_init.append({"Confirmed": False, "Direction": "↓ Ramp Down", "Role": _rl, **{c: 0.0 for c in _s4v2_mc}})
            st.session_state.s4v2_hc_adj_df = pd.DataFrame(_hc_init)
        elif "Confirmed" not in st.session_state.s4v2_hc_adj_df.columns:
            st.session_state.s4v2_hc_adj_df.insert(0, "Confirmed", False)
        if "s4v2_mrr_adj_df" not in st.session_state or \
                set(st.session_state.s4v2_mrr_adj_df.get("Adjustment", pd.Series()).unique()) != {"⊕ New MRR", "⊖ Churn MRR"}:
            st.session_state.s4v2_mrr_adj_df = pd.DataFrame([
                {"Confirmed": False, "Adjustment": "⊕ New MRR",   **{c: 0.0 for c in _s4v2_mc}},
                {"Confirmed": False, "Adjustment": "⊖ Churn MRR", **{c: 0.0 for c in _s4v2_mc}},
            ])
        elif "Confirmed" not in st.session_state.s4v2_mrr_adj_df.columns:
            st.session_state.s4v2_mrr_adj_df.insert(0, "Confirmed", False)

        # ── Base HC snapshot for scope ───────────────────────────────────────
        if _hc is None:
            _base_hc_tot = None
            _base_hc_by_role = {_rl: None for _rl in roles_permitidos}
            _base_hc_other = None
            _base_hc_mgr   = None
        elif _scope == "Overall":
            _base_hc_tot = _hc.get('total')
            _by_r = _hc.get('by_role', {})
            _base_hc_by_role = {
                'Accountant I':       float(_by_r.get('Accountant I', 0) or 0),
                'Accountant II':      float(_by_r.get('Accountant II', 0) or 0),
                'General Accountant': float(_by_r.get('General Accountant', 0) or 0),
                'Sr. Accountant':     float(_by_r.get('Sr. Accountant', 0) or 0),
            }
            _base_hc_other = float(_by_r.get('Other', 0) or 0)
            _base_hc_mgr   = float(_hc.get('mgr_total', 0) or 0)
        elif _scope_is_sr:
            _sr_hc_s4 = (
                (
                    _hc.get('by_sr_email', {}).get(str(_scope).strip().lower())
                    or _hc.get('by_sr', {}).get(_scope)
                    or _hc.get('by_sr_norm', {}).get(_norm_name(_scope))
                )
                if _hc else {}
            ) or {}
            _sr_roles_s4 = _sr_hc_s4.get('by_role', {})
            _sr_tot_s4   = _sr_hc_s4.get('total', 0) or None
            _base_hc_tot = _sr_tot_s4
            _base_hc_by_role = {
                'Accountant I':       float(_sr_roles_s4.get('Accountant I', 0)),
                'Accountant II':      float(_sr_roles_s4.get('Accountant II', 0)),
                'General Accountant': float(_sr_roles_s4.get('General Accountant', 0)),
                'Sr. Accountant':     float(_sr_roles_s4.get('Sr. Accountant', 0)),
            }
            _base_hc_other = None
            _base_hc_mgr   = float(_sr_hc_s4.get('managers', 0) or 0)
        else:
            _pod_hc_raw = {}
            _hbp = _hc.get('by_pod_role', pd.DataFrame()) if isinstance(_hc, dict) else pd.DataFrame()
            if not _hbp.empty and 'POD' in _hbp.columns:
                def _nhp(s): return str(s).lower().replace(' ', '').strip()
                _hbp_pnorm = _hbp['POD'].apply(_nhp)
                for _, _hr in _hbp[_hbp_pnorm == _nhp(_scope)].iterrows():
                    _pod_hc_raw[_hr['Capacity Role']] = int(_hr.get('HC', 0))
            # Total HC must count only productive roles (exclude 'Other' — managers
            # are tracked separately in their own row).
            _base_hc_tot = sum(
                v for k, v in _pod_hc_raw.items() if k != 'Other'
            ) or None
            _base_hc_by_role = {
                'Accountant I':       float(_pod_hc_raw.get('Accountant I', 0)),
                'Accountant II':      float(_pod_hc_raw.get('Accountant II', 0)),
                'General Accountant': float(_pod_hc_raw.get('General Accountant', 0)),
                'Sr. Accountant':     float(_pod_hc_raw.get('Sr. Accountant', 0)),
            }
            _base_hc_other = None
            # Manager count for this POD (key-matching same normalization as by_pod_role lookup)
            _mgr_by_pod_raw = _hc.get('mgr_by_pod', {}) if isinstance(_hc, dict) else {}
            _mgr_norm_map = {
                str(k).lower().replace(' ', '').strip(): int(v)
                for k, v in _mgr_by_pod_raw.items()
            }
            _base_hc_mgr = float(_mgr_norm_map.get(str(_scope).lower().replace(' ', '').strip(), 0))

        # ── Scope-specific setup (POD or Sr. Accountant) ─────────────────────
        _pod_clients_lower = set()
        _pm_all     = pd.DataFrame()
        _proles_all = pd.DataFrame()
        _pdf_raw    = pd.DataFrame()
        if _scope_is_sr:
            # Sr. scope: filter by clients assigned to this Sr. Accountant
            if 'Sr. Accountant' in _cli_df.columns and 'Client' in _cli_df.columns:
                _pod_clients_lower = set(
                    _cli_df[_cli_df['Sr. Accountant'].astype(str) == _scope]['Client']
                    .dropna().astype(str).str.strip().str.lower().unique()
                )
            if not _df_raw.empty and 'client_name' in _df_raw.columns:
                _pdf_raw = _df_raw[
                    _df_raw['client_name'].astype(str).str.strip().str.lower()
                    .isin(_pod_clients_lower)
                ]
        elif _scope_is_pod:
            if 'POD' in _cli_df.columns and 'Client' in _cli_df.columns:
                _cli_pod_norm = (
                    _cli_df['POD'].fillna('').astype(str).str.strip()
                    .where(lambda s: ~s.str.lower().isin({'nan', 'none', ''}), 'No POD')
                )
                _pod_clients_lower = set(
                    _cli_df[_cli_pod_norm == _scope]['Client']
                    .dropna().astype(str).str.strip().str.lower().unique()
                )
            if not _pod_df.empty and 'POD' in _pod_df.columns:
                _pm_all     = _pod_df[(_pod_df['POD'] == _scope) & (_pod_df['Required Role'] == '>>> POD TOTAL')]
                _proles_all = _pod_df[(_pod_df['POD'] == _scope) & (_pod_df['Required Role'] != '>>> POD TOTAL')]
            if not _df_raw.empty and 'POD' in _df_raw.columns:
                _raw_pod_norm2 = (
                    _df_raw['POD'].fillna('').astype(str).str.strip()
                    .where(lambda s: ~s.str.lower().isin({'nan', 'none', ''}), 'No POD')
                )
                _pdf_raw = _df_raw[_raw_pod_norm2 == _scope]

        # ── Property / ticket metrics for scope ──────────────────────────────
        _src_raw = _pdf_raw if ((_scope_is_pod or _scope_is_sr) and not _pdf_raw.empty) else _df_raw
        if not _src_raw.empty:
            _csn = _src_raw.groupby('client_name', as_index=False).agg({
                c: 'first' for c in ['Res Prop', 'Commercial Properties', 'Res doors', 'Commercial Doors', 'SQFT Commercial']
                if c in _src_raw.columns
            })
            def _safe_num_s4(s): return pd.to_numeric(s, errors='coerce').fillna(0)
            _s_res_prop  = int(_safe_num_s4(_csn.get('Res Prop', 0)).sum())
            _s_comm_prop = int(_safe_num_s4(_csn.get('Commercial Properties', 0)).sum())
            _s_prop      = _s_res_prop + _s_comm_prop
            _s_res_door  = int(_safe_num_s4(_csn.get('Res doors', 0)).sum())
            _s_comm_door = int(_safe_num_s4(_csn.get('Commercial Doors', 0)).sum())
            _s_door      = _s_res_door + _s_comm_door
            _s_sqft      = int(_safe_num_s4(_csn.get('SQFT Commercial', 0)).sum())
        else:
            _s_res_prop = _s_comm_prop = _s_prop = _s_res_door = _s_comm_door = _s_door = _s_sqft = 0; _s_aht = 0.0

        # ── Read adjustments (only confirmed rows) ───────────────────────────
        _hc_adj_df  = st.session_state.s4v2_hc_adj_df
        _mrr_adj_df = st.session_state.s4v2_mrr_adj_df

        # Only apply confirmed HC rows
        _hc_adj_conf = _hc_adj_df[_hc_adj_df.get("Confirmed", pd.Series(False, index=_hc_adj_df.index)).eq(True)]
        _net_hc = {}
        for _rl in roles_permitidos:
            _up_r   = _hc_adj_conf[(_hc_adj_conf["Direction"] == "↑ Ramp Up")   & (_hc_adj_conf["Role"] == _rl)]
            _dn_r   = _hc_adj_conf[(_hc_adj_conf["Direction"] == "↓ Ramp Down") & (_hc_adj_conf["Role"] == _rl)]
            _net = []
            for _mc in _s4v2_mc:
                _uv = float(pd.to_numeric(_up_r[_mc].values[0] if not _up_r.empty else 0, errors='coerce') or 0)
                _dv = float(pd.to_numeric(_dn_r[_mc].values[0] if not _dn_r.empty else 0, errors='coerce') or 0)
                _net.append(_uv - _dv)
            _net_hc[_rl] = _net

        # Only apply confirmed MRR rows
        _mrr_adj_conf = _mrr_adj_df[_mrr_adj_df.get("Confirmed", pd.Series(False, index=_mrr_adj_df.index)).eq(True)]
        _new_mrr_r   = _mrr_adj_conf[_mrr_adj_conf["Adjustment"] == "⊕ New MRR"]
        _chrn_mrr_r  = _mrr_adj_conf[_mrr_adj_conf["Adjustment"] == "⊖ Churn MRR"]
        _adj_new_mrr   = [float(pd.to_numeric(_new_mrr_r[_mc].values[0]  if not _new_mrr_r.empty  else 0, errors='coerce') or 0) for _mc in _s4v2_mc]
        _adj_chrn_mrr  = [float(pd.to_numeric(_chrn_mrr_r[_mc].values[0] if not _chrn_mrr_r.empty else 0, errors='coerce') or 0) for _mc in _s4v2_mc]

        # ── Pre-compute Step 4: base hours + automation savings ──────────────
        # These depend only on scope, underlying data, automation rules, and month
        # params — NOT on HC / MRR / Hours adjustment tables.
        # Cache the result in session state so that editing the adjustment tables
        # is instant: the expensive loops are skipped on every cache hit.
        _s4a_raw = st.session_state.s4v2_auto_df
        _s4a_src = (
            _s4a_raw[_s4a_raw.get("Confirmed", pd.Series(False, index=_s4a_raw.index)).eq(True)].copy()
            if not _s4a_raw.empty else pd.DataFrame()
        )

        # Cache key: changes only when scope, data shapes, automation rules, or
        # month-calculation parameters change.
        try:
            _auto_hash = int(pd.util.hash_pandas_object(_s4a_raw).sum()) if not _s4a_raw.empty else 0
        except Exception:
            _auto_hash = 0
        _s4_cache_key = hash((
            _scope,
            calc_mode, fixed_days,
            _cli_df.shape[0]  if not _cli_df.empty  else 0,
            _src_raw.shape[0] if not _src_raw.empty else 0,
            _auto_hash,
        ))

        _s4_precomp = st.session_state.get('_s4v2_precomp', {})
        if _s4_precomp.get('key') == _s4_cache_key:
            # ── CACHE HIT: reuse previous result, skip all heavy loops ────────
            _s4_base_hrs_role  = _s4_precomp['bhr']
            _rb_scope          = _s4_precomp['rbs']
            _s4_auto_sav_role  = _s4_precomp['asr']
            _s4_auto_sav_total = _s4_precomp['ast']
        else:
            # ── CACHE MISS: run the full precomputation ───────────────────────
            _s4_base_hrs_role  = {_rl: [0.0]*6 for _rl in roles_permitidos}
            _s4_auto_sav_role  = {_rl: [0.0]*6 for _rl in roles_permitidos}
            _s4_auto_sav_total = [0.0]*6

            # 1. Base hours per role from Step 3 Final Hours (post Step-2+3 adj)
            _rb_scope = pd.DataFrame()
            if not _cli_df.empty:
                if _scope_is_sr and 'Client' in _cli_df.columns:
                    _rb_scope = _cli_df[
                        _cli_df['Client'].astype(str).str.strip().str.lower().isin(_pod_clients_lower)
                    ]
                elif _scope_is_pod and 'POD' in _cli_df.columns:
                    _rb_scope = _cli_df[_cli_df['POD'].astype(str).str.strip() == _scope]
                else:
                    _rb_scope = _cli_df
                for _bi, _bms in enumerate(meses_proyeccion):
                    _bcol = f"M{_bi+1} ({_bms}) - Final Hours"
                    if _bcol in _rb_scope.columns:
                        for _brl in roles_permitidos:
                            _s4_base_hrs_role[_brl][_bi] = float(
                                _rb_scope[_rb_scope['Required Role'] == _brl][_bcol].sum())

            # 2. Automation cascade (same logic as Step 3, using s4v2_auto_df)
            if not _s4a_src.empty and not _src_raw.empty and \
                    'type' in _src_raw.columns and 'subtype' in _src_raw.columns:
                _s4df = _src_raw.copy().reset_index(drop=True)
                _s4n  = len(_s4df)

                # Month params
                _s4mp = []
                for _mj, _msj in enumerate(meses_proyeccion):
                    _mdj = today + relativedelta(months=_mj)
                    _smj = pd.Timestamp(_mdj.replace(day=1).date())
                    _emj = pd.Timestamp((_smj + relativedelta(months=1) - relativedelta(days=1)).date())
                    _ndj = (fixed_days if calc_mode == "Fixed days per month"
                            else int(np.busday_count(_smj.strftime('%Y-%m-%d'),
                                                     (_emj + pd.Timedelta(days=1)).strftime('%Y-%m-%d'))))
                    _s4mp.append((_mj, _msj, _smj, _emj, _ndj))

                # Attendance % and learning-curve arrays (vectorised, shape: n_rows)
                _s4gl  = _s4df['Go Live'].values if 'Go Live' in _s4df.columns else [pd.NaT]*_s4n
                _s4fsd = _s4df['Final Service Date'].values if 'Final Service Date' in _s4df.columns else [pd.NaT]*_s4n
                _s4ap = {}; _s4lc = {}
                for _mj, _msj, _smj, _emj, _ndj in _s4mp:
                    _gls  = pd.to_datetime(pd.Series(_s4gl),  errors='coerce').fillna(_smj).clip(lower=_smj, upper=_emj)
                    _fsds = pd.to_datetime(pd.Series(_s4fsd), errors='coerce').fillna(_emj).clip(lower=_smj, upper=_emj)
                    _dias = np.busday_count(_gls.values.astype('datetime64[D]'),
                                            (_fsds + pd.Timedelta(days=1)).values.astype('datetime64[D]'))
                    _s4ap[_mj] = np.clip(np.maximum(_dias.astype(float), 0) / _ndj, 0.0, 1.0) if _ndj > 0 else np.zeros(_s4n)
                    # Zero out clients whose FSD fell BEFORE this month's start (same fix as Step 3)
                    _s4fsd_ser = pd.to_datetime(pd.Series(_s4fsd), errors='coerce')
                    _s4_churned = _s4fsd_ser.notna() & (_s4fsd_ser < _smj)
                    if _s4_churned.any():
                        _s4ap[_mj] = np.where(_s4_churned.values, 0.0, _s4ap[_mj])
                    _hgl  = pd.notna(pd.Series(_s4gl)).values
                    _glts = pd.DatetimeIndex(pd.to_datetime(pd.Series(_s4gl), errors='coerce').fillna(_smj))
                    _mdf  = np.where(_hgl, (_smj.year-_glts.year.values)*12+(_smj.month-_glts.month.values), 999).astype(int)
                    _ap2  = _s4ap[_mj]
                    _s4lc[_mj] = np.select(
                        [~_hgl|(_ap2==0), (_mdf==0)&_hgl&(_ap2>0), (_mdf==1)&_hgl&(_ap2>0), (_mdf==2)&_hgl&(_ap2>0)],
                        [1.0, 1.17, 0.86, 0.99], default=1.0)

                # Build per-combo automation efficiency lookup dict
                def _s4am(v): return pd.isna(v) or str(v).strip() in ('', 'All')
                _s4ac = {}
                _s4up = (_s4df.assign(_tk=_s4df['type'].astype(str)+' - '+_s4df['subtype'].astype(str))
                              .assign(_pk=_s4df.get('POD', pd.Series('', index=_s4df.index)).fillna('').astype(str).str.strip())
                              .assign(_pmk=_s4df.get('PMS', pd.Series('', index=_s4df.index)).fillna('').astype(str).str.strip())
                              [['client_name','_tk','_pk','_pmk']].drop_duplicates())
                for _, _up in _s4up.iterrows():
                    _ck,_tk,_pk,_pmk = str(_up['client_name']).strip(), str(_up['_tk']).strip(), str(_up['_pk']).strip(), str(_up['_pmk']).strip()
                    _mc  = (_s4a_src["Client"].apply(_s4am)) | (_s4a_src["Client"]==_ck)
                    _mt  = (_s4a_src["Task (Type - Subtype)"].apply(_s4am)) | (_s4a_src["Task (Type - Subtype)"]==_tk)
                    _mpd = ((_s4a_src["POD"].apply(_s4am)) | (_s4a_src["POD"]==_pk)) if "POD" in _s4a_src.columns else pd.Series(True, index=_s4a_src.index)
                    _mpm = ((_s4a_src["PMS"].apply(_s4am)) | (_s4a_src["PMS"]==_pmk)) if "PMS" in _s4a_src.columns else pd.Series(True, index=_s4a_src.index)
                    _arows = _s4a_src[_mc & _mt & _mpd & _mpm]
                    _effs = []
                    for _mj2 in range(6):
                        _evp=_evr=_eap=_ear=0.0
                        if not _arows.empty:
                            for _,_au in _arows.iterrows():
                                _v = (pd.to_numeric(_au.get(f"M{_mj2+1} (%)", 0), errors='coerce') or 0.0)/100.0
                                _af = str(_au.get("Affects",""))
                                _ia = _af=="All (Vol + AHT)"
                                if "Vol Proc" in _af or _ia: _evp+=_v
                                if "Vol Rev"  in _af or _ia: _evr+=_v
                                if "AHT Proc" in _af or _ia: _eap+=_v
                                if "AHT Rev"  in _af or _ia: _ear+=_v
                        _effs.append((min(1.0,_evp),min(1.0,_evr),min(1.0,_eap),min(1.0,_ear)))
                    _s4ac[(_ck,_tk,_pk,_pmk)] = _effs

                # ── Vectorised savings computation (replaces iterrows loop) ───
                # Extract per-row arrays once, then compute all 6 months with numpy
                _ip2_arr  = np.array([
                    (lambda v: v if v not in ('nan','None','') else 'Accountant I')(
                        str(_s4df.at[_ri, 'Ideal Proc'] if 'Ideal Proc' in _s4df.columns
                            else _s4df.at[_ri, 'Proc Role'] if 'Proc Role' in _s4df.columns
                            else 'Accountant I').strip())
                    for _ri in range(_s4n)])
                _ir2_arr  = np.array([
                    (lambda v: v if v not in ('nan','None','') else 'Sr. Accountant')(
                        str(_s4df.at[_ri, 'Ideal Rev'] if 'Ideal Rev' in _s4df.columns
                            else _s4df.at[_ri, 'Rev Role'] if 'Rev Role' in _s4df.columns
                            else 'Sr. Accountant').strip())
                    for _ri in range(_s4n)])
                _ptx_arr  = pd.to_numeric(_s4df.get('Closed tickets with Proc time', 0), errors='coerce').fillna(0).values
                _rtx_arr  = pd.to_numeric(_s4df.get('Closed tickets with rev time',  0), errors='coerce').fillna(0).values
                _pa_arr   = pd.to_numeric(_s4df.get('>>> FINAL Capacity Proc AHT',   0), errors='coerce').fillna(0).values
                _ra_arr   = pd.to_numeric(_s4df.get('>>> FINAL Capacity Rev AHT',    0), errors='coerce').fillna(0).values
                _up_arr   = np.array([utilization_map.get(_r, util_acc1) for _r in _ip2_arr])
                _ur_arr   = np.array([utilization_map.get(_r, util_sr)   for _r in _ir2_arr])
                _ck_arr   = _s4df.get('client_name', pd.Series('', index=_s4df.index)).astype(str).str.strip().values
                _tk_arr   = (_s4df.get('type',   pd.Series('', index=_s4df.index)).astype(str) + ' - ' +
                             _s4df.get('subtype', pd.Series('', index=_s4df.index)).astype(str)).values
                _pk_arr   = _s4df.get('POD', pd.Series('', index=_s4df.index)).fillna('').astype(str).str.strip().values
                _pmk_arr  = _s4df.get('PMS', pd.Series('', index=_s4df.index)).fillna('').astype(str).str.strip().values

                # Build efficiency tensor (n_rows × 6 months × 4 factors) with one Python pass
                _eff_t = np.zeros((_s4n, 6, 4))
                for _ri in range(_s4n):
                    _ef2 = _s4ac.get((_ck_arr[_ri], _tk_arr[_ri], _pk_arr[_ri], _pmk_arr[_ri]), None)
                    if _ef2:
                        for _mj in range(6):
                            _eff_t[_ri, _mj, :] = _ef2[_mj]

                # Per-month vectorised computation
                for _mj3 in range(6):
                    _ac3 = _s4ap[_mj3]          # shape (n_rows,)
                    _lc3 = _s4lc[_mj3]          # shape (n_rows,)
                    _vp3 = _eff_t[:, _mj3, 0]
                    _vr3 = _eff_t[:, _mj3, 1]
                    _ap3 = _eff_t[:, _mj3, 2]
                    _ar3 = _eff_t[:, _mj3, 3]
                    _bp_sav = (_ptx_arr * _ac3 * _pa_arr * _lc3 / 60) * (_vp3 + _ap3 - _vp3 * _ap3)
                    _br_sav = (_rtx_arr * _ac3 * _ra_arr * _lc3 / 60) * (_vr3 + _ar3 - _vr3 * _ar3)
                    _sp3_arr = _bp_sav * (1 + (1 - _up_arr) + absenteeism + attrition)
                    _sr3_arr = _br_sav * (1 + (1 - _ur_arr) + absenteeism + attrition)
                    for _rl in roles_permitidos:
                        _pm = (_ip2_arr == _rl)
                        _rm = (_ir2_arr == _rl)
                        _s4_auto_sav_role[_rl][_mj3] += float(_sp3_arr[_pm].sum() + _sr3_arr[_rm].sum())
                    _s4_auto_sav_total[_mj3] += float(_sp3_arr.sum() + _sr3_arr.sum())

            # Store result in session-state cache
            st.session_state['_s4v2_precomp'] = {
                'key': _s4_cache_key,
                'bhr': _s4_base_hrs_role,
                'rbs': _rb_scope,
                'asr': _s4_auto_sav_role,
                'ast': _s4_auto_sav_total,
            }

        # ── Build scenario rows ──────────────────────────────────────────────
        _scen_rows = {}
        _scen_nums = {}   # raw numeric values for Before vs. After comparison
        _base_nums = {}   # Step-3 baseline numeric values (same keys)
        _months    = []

        # ── PERF: Pre-cache role costs/util and shrinkage once (was looked up
        # dozens of times per month via _s4p.loc[role, ...] = slow DataFrame ops).
        _cost_cache = {_rl: _s4_cost_val(_rl) for _rl in roles_permitidos}
        _util_cache = {_rl: _s4_util(_rl)     for _rl in roles_permitidos}
        _shrink_cache = {_rl: 2.0 - _util_cache[_rl] + absenteeism + attrition
                         for _rl in roles_permitidos}
        _c_a1 = _cost_cache['Accountant I']
        _c_a2 = _cost_cache['Accountant II']
        _c_gn = _cost_cache['General Accountant']
        _c_sr = _cost_cache['Sr. Accountant']

        # ── PERF: Pre-aggregate _rb_scope column sums per month (was summed
        # 3-4 times per month inside the loop on the same column).
        _rb_empty = _rb_scope.empty if hasattr(_rb_scope, 'empty') else True
        _rb_cols  = set(_rb_scope.columns) if not _rb_empty else set()
        _rb_sum_cache = {}
        _rb_rolefte_cache = {}  # (month_idx, role) -> float
        for _i_pc in range(len(meses_proyeccion)):
            _ms_pc = meses_proyeccion[_i_pc]
            for _suffix in ("Final Hours", "Base Hours", "Adjustments (+) Hrs",
                             "Adjustments (-) Hrs", "Productive Hours", "Final FTEs"):
                _c = f"M{_i_pc+1} ({_ms_pc}) - {_suffix}"
                if (not _rb_empty) and _c in _rb_cols:
                    _rb_sum_cache[_c] = float(_rb_scope[_c].sum())
            # Per-role FTE sums (avoids repeated .loc + .sum inside the baseline block)
            _cfte_pc = f"M{_i_pc+1} ({_ms_pc}) - Final FTEs"
            if (not _rb_empty) and _cfte_pc in _rb_cols:
                _gb_fte = _rb_scope.groupby('Required Role')[_cfte_pc].sum()
                for _brl in roles_permitidos:
                    _rb_rolefte_cache[(_i_pc, _brl)] = float(_gb_fte.get(_brl, 0.0))

        # ── PERF: Pre-parse _duc date columns + pre-compute per-month MRR masks
        # (was done every iteration via pd.to_datetime / date comparisons).
        _duc_mrr_by_month = {}
        if (not _duc.empty) and 'MRR' in _duc.columns:
            _duc_gl  = pd.to_datetime(_duc.get('Go Live',            pd.Series(dtype='datetime64[ns]')), errors='coerce')
            _duc_fsd = pd.to_datetime(_duc.get('Final Service Date', pd.Series(dtype='datetime64[ns]')), errors='coerce')
            _duc_name_lower = _duc['client_name'].astype(str).str.strip().str.lower() if 'client_name' in _duc.columns else pd.Series([''] * len(_duc))
            _duc_mrr_vals = pd.to_numeric(_duc['MRR'], errors='coerce').fillna(0.0)
            _pod_clients_lower_set = set(_pod_clients_lower) if _pod_clients_lower else set()
            for _i_pc in range(len(meses_proyeccion)):
                _mdt_pc = today + relativedelta(months=_month_offsets[_i_pc])
                _sm_pc  = pd.Timestamp(_mdt_pc.replace(day=1).date())
                _em_pc  = pd.Timestamp((_sm_pc + relativedelta(months=1) - relativedelta(days=1)).date())
                _active_pc = (_duc_gl.isna() | (_duc_gl <= _em_pc)) & (_duc_fsd.isna() | (_duc_fsd >= _sm_pc))
                _duc_mrr_by_month[_i_pc] = (_sm_pc, _em_pc, _active_pc)

        # ── PERF: Pre-parse _src_raw date columns + precompute active masks
        # per month (for the AHT block — avoids pd.to_datetime on every iter).
        _src_empty = _src_raw.empty if hasattr(_src_raw, 'empty') else True
        if not _src_empty:
            _src_gl_all  = pd.to_datetime(_src_raw.get('Go Live',            pd.Series(dtype='datetime64[ns]')), errors='coerce')
            _src_fsd_all = pd.to_datetime(_src_raw.get('Final Service Date', pd.Series(dtype='datetime64[ns]')), errors='coerce')
        else:
            _src_gl_all  = pd.Series(dtype='datetime64[ns]')
            _src_fsd_all = pd.Series(dtype='datetime64[ns]')
        _src_active_by_month = {}
        _src_sm_em_by_month  = {}
        for _i_pc in range(len(meses_proyeccion)):
            _sm_s4 = pd.Timestamp((today + relativedelta(months=_month_offsets[_i_pc])).replace(day=1).date())
            _em_s4 = pd.Timestamp((_sm_s4 + relativedelta(months=1) - relativedelta(days=1)).date())
            _src_sm_em_by_month[_i_pc] = (_sm_s4, _em_s4)
            if not _src_empty:
                _src_active_by_month[_i_pc] = (
                    (_src_gl_all.isna()  | (_src_gl_all  <= _em_s4)) &
                    (_src_fsd_all.isna() | (_src_fsd_all >= _sm_s4))
                )

        for _i, _ms in enumerate(meses_proyeccion):
            if _i >= len(_exec): break
            _r   = _exec.iloc[_i]
            _months.append(_ms)
            _col = _ms

            _mi_new_mrr  = _adj_new_mrr[_i]  if _i < len(_adj_new_mrr)  else 0.0
            _mi_chrn_mrr = _adj_chrn_mrr[_i] if _i < len(_adj_chrn_mrr) else 0.0
            _net_adj_by_role = {_rl: (_net_hc[_rl][_i] if _i < len(_net_hc.get(_rl, [])) else 0) for _rl in roles_permitidos}
            _net_adj_total   = sum(_net_adj_by_role.values())

            if _scope == "Overall":
                # Use Step 3 final hours (post all adjustments) as base so Step 4 = Step 3 when no changes
                _b_curr = float(_r.get("7. Total Required Hours (Final)", 0) or 0)
                _b_base = float(_r.get("3. Total Hours (Pre-Auto)", 0) or 0)   # for shrinkage/productivity display
                _b_new  = float(_r.get("New Clients Hours", 0) or 0)
                _b_chrn = float(_r.get("Confirmed Churn (Hrs)", 0) or 0)
                _b_ap   = float(_r.get("5. Manual Adjustments (+) Hrs", 0) or 0)
                _b_am   = float(_r.get("6. Manual Adjustments (-) Hrs", 0) or 0)
                _b_prod = float(_r.get("1. Productive Hours (Pure Base)", 0) or 0)
                _b_mrr  = float(_r.get("Total MRR ($)", 0) or 0)
                _b_newm = float(_r.get("New MRR ($)", 0) or 0)
                _b_chnm = float(_r.get("Lost MRR (Churn) ($)", 0) or 0)
                _b_wday = float(_r.get("Working Days (Used)", _wdays_d.get(_i, 21)) or _wdays_d.get(_i, 21))
            else:
                _cfin  = f"M{_i+1} ({_ms}) - Final Hours"
                _cbase = f"M{_i+1} ({_ms}) - Base Hours"
                _cplus = f"M{_i+1} ({_ms}) - Adjustments (+) Hrs"
                _cmins = f"M{_i+1} ({_ms}) - Adjustments (-) Hrs"
                if _scope_is_sr:
                    # Sr. scope: sum Final Hours per role from pre-computed breakdown
                    _b_curr = sum(_s4_base_hrs_role.get(_rl, [0.0]*6)[_i] for _rl in roles_permitidos)
                    _b_base = _rb_sum_cache.get(_cbase, _b_curr)
                    _b_ap   = 0.0
                    _b_am   = 0.0
                else:
                    # POD scope: use Final Hours from _rb_scope (post Step-2 + Step-3 adjusted)
                    _b_curr = _rb_sum_cache.get(_cfin,  0.0)
                    _b_base = _rb_sum_cache.get(_cbase, _b_curr)
                    _b_ap   = _rb_sum_cache.get(_cplus, 0.0)
                    _b_am   = _rb_sum_cache.get(_cmins, 0.0)
                # Productive hours (no shrinkage) from scoped data.
                # If the column doesn't exist in _cli_df, derive from base hours ÷ shrinkage
                # so productivity reads the correct utilization target instead of 100%.
                _cprod_s4 = f"M{_i+1} ({_ms}) - Productive Hours"
                _b_prod_raw = _rb_sum_cache.get(_cprod_s4)
                if _b_prod_raw is not None:
                    _b_prod = _b_prod_raw
                else:
                    # productive_hrs = base_hrs / shrink_factor  (shrink = 2-util+abs+att)
                    _b_prod = sum(
                        _s4_base_hrs_role.get(_rl, [0.0]*6)[_i] / max(_shrink_cache.get(_rl, 1.0), 0.01)
                        for _rl in roles_permitidos
                    )
                # Recalculate _b_base from role breakdown to keep consistent with _b_prod
                _b_base_from_roles = sum(_s4_base_hrs_role.get(_rl, [0.0]*6)[_i] for _rl in roles_permitidos)
                if _b_base_from_roles > 0:
                    _b_base = _b_base_from_roles
                _b_new  = _fd.get('pod_new_hrs',   {}).get(_ms, {}).get(_scope, 0.0)
                _b_chrn = _fd.get('pod_churn',     {}).get(_ms, {}).get(_scope, 0.0)
                _b_newm = _fd.get('pod_new_mrr',   {}).get(_ms, {}).get(_scope, 0.0)
                _b_chnm = _fd.get('pod_churn_mrr', {}).get(_ms, {}).get(_scope, 0.0)
                _b_wday = float(_wdays_d.get(_i, 21))
                _b_mrr  = 0.0
                # Use cached per-month date mask (date ops done once, above the loop)
                if _i in _duc_mrr_by_month and _pod_clients_lower_set:
                    _sm, _em, _active_mask = _duc_mrr_by_month[_i]
                    _mk = _active_mask & _duc_name_lower.isin(_pod_clients_lower_set)
                    _b_mrr = float(_duc_mrr_vals[_mk].sum())

            # ── FTEs from pre-auto base + Step-4 automation savings ──────────
            _d_fte_m = float(_dict_hrs_fte.get(_i, 157.5) or 157.5)
            _req_a1 = max(0.0, (_s4_base_hrs_role.get('Accountant I',       [0.0]*6)[_i] - _s4_auto_sav_role.get('Accountant I',       [0.0]*6)[_i]) / _d_fte_m)
            _req_a2 = max(0.0, (_s4_base_hrs_role.get('Accountant II',      [0.0]*6)[_i] - _s4_auto_sav_role.get('Accountant II',      [0.0]*6)[_i]) / _d_fte_m)
            _req_gn = max(0.0, (_s4_base_hrs_role.get('General Accountant', [0.0]*6)[_i] - _s4_auto_sav_role.get('General Accountant', [0.0]*6)[_i]) / _d_fte_m)
            _req_sr = max(0.0, (_s4_base_hrs_role.get('Sr. Accountant',     [0.0]*6)[_i] - _s4_auto_sav_role.get('Sr. Accountant',     [0.0]*6)[_i]) / _d_fte_m)

            # ── Hours adjustments per role ────────────────────────────────────
            _hrs_role_df   = st.session_state.s4v2_hrs_role_df
            _hrs_role_conf = _hrs_role_df[_hrs_role_df.get("Confirmed", pd.Series(False, index=_hrs_role_df.index)).eq(True)]
            _mc_key        = _s4v2_mc[_i] if _i < len(_s4v2_mc) else _s4v2_mc[-1]

            _hrs_adj_by_role = {}
            for _rl in roles_permitidos:
                _rrow = _hrs_role_conf[_hrs_role_conf["Role"] == _rl]
                _hrs_adj_by_role[_rl] = float(pd.to_numeric(
                    _rrow[_mc_key].values[0] if not _rrow.empty else 0, errors='coerce') or 0)

            # PERF: pre-cached shrinkage per role (computed once before the loop)
            _shrink_a1 = _shrink_cache['Accountant I']
            _shrink_a2 = _shrink_cache['Accountant II']
            _shrink_gn = _shrink_cache['General Accountant']
            _shrink_sr = _shrink_cache['Sr. Accountant']

            _adj_fte_a1 = (_hrs_adj_by_role.get('Accountant I',       0) * _shrink_a1) / _d_fte_m
            _adj_fte_a2 = (_hrs_adj_by_role.get('Accountant II',      0) * _shrink_a2) / _d_fte_m
            _adj_fte_gn = (_hrs_adj_by_role.get('General Accountant', 0) * _shrink_gn) / _d_fte_m
            _adj_fte_sr = (_hrs_adj_by_role.get('Sr. Accountant',     0) * _shrink_sr) / _d_fte_m

            _hrs_total_adj = (
                _hrs_adj_by_role.get('Accountant I',       0) * _shrink_a1 +
                _hrs_adj_by_role.get('Accountant II',      0) * _shrink_a2 +
                _hrs_adj_by_role.get('General Accountant', 0) * _shrink_gn +
                _hrs_adj_by_role.get('Sr. Accountant',     0) * _shrink_sr
            )

            # ── Add Hours (Scenario) per scope ───────────────────────────────
            _s4_mhrs_col = f"M{_i+1} (Hrs)"
            _s4_hist_raw = st.session_state.get('s4v2_hist_df', pd.DataFrame())
            _s4_hist_conf = _s4_hist_raw[
                _s4_hist_raw.get('Confirmed', pd.Series(False, index=_s4_hist_raw.index)).eq(True)
            ] if not _s4_hist_raw.empty else pd.DataFrame()
            # Scope filter
            if not _s4_hist_conf.empty:
                if _scope_is_pod:
                    _s4h_pod_match = _s4_hist_conf['POD'].astype(str).str.strip() == _scope
                    _s4h_cli_match = (_s4_hist_conf.get('Client', pd.Series('', index=_s4_hist_conf.index))
                                      .astype(str).str.strip().str.lower().isin(_pod_clients_lower))
                    _s4_hist_conf = _s4_hist_conf[_s4h_pod_match | _s4h_cli_match]
                elif _scope_is_sr:
                    _s4_hist_conf = _s4_hist_conf[
                        _s4_hist_conf.get('Client', pd.Series('', index=_s4_hist_conf.index))
                        .astype(str).str.strip().str.lower().isin(_pod_clients_lower)
                    ]
                # else Overall: use all confirmed rows
            _s4_add_hrs_by_role = {}
            _s4_add_hrs_total   = 0.0
            if not _s4_hist_conf.empty and _s4_mhrs_col in _s4_hist_conf.columns:
                for _rl in roles_permitidos:
                    _rl_mask = _s4_hist_conf['Required Role'].astype(str).str.strip() == _rl
                    _rl_hrs  = float(pd.to_numeric(_s4_hist_conf.loc[_rl_mask, _s4_mhrs_col], errors='coerce').fillna(0.0).sum())
                    _s4_add_hrs_by_role[_rl] = _rl_hrs
                    _s4_add_hrs_total += _rl_hrs * _shrink_cache.get(_rl, 1.0)

            # ── Reduce Hours (Scenario) per scope ────────────────────────────
            _s4_red_raw = st.session_state.get('s4v2_red_df', pd.DataFrame())
            _s4_red_conf = _s4_red_raw[
                _s4_red_raw.get('Confirmed', pd.Series(False, index=_s4_red_raw.index)).eq(True)
            ] if not _s4_red_raw.empty else pd.DataFrame()
            if not _s4_red_conf.empty:
                if _scope_is_pod:
                    _s4r_pod_match = _s4_red_conf['POD'].astype(str).str.strip() == _scope
                    _s4r_cli_match = (_s4_red_conf.get('Client', pd.Series('', index=_s4_red_conf.index))
                                      .astype(str).str.strip().str.lower().isin(_pod_clients_lower))
                    _s4_red_conf = _s4_red_conf[_s4r_pod_match | _s4r_cli_match]
                elif _scope_is_sr:
                    _s4_red_conf = _s4_red_conf[
                        _s4_red_conf.get('Client', pd.Series('', index=_s4_red_conf.index))
                        .astype(str).str.strip().str.lower().isin(_pod_clients_lower)
                    ]
            _s4_red_hrs_by_role = {}
            _s4_red_hrs_total   = 0.0
            if not _s4_red_conf.empty and _s4_mhrs_col in _s4_red_conf.columns:
                for _rl in roles_permitidos:
                    _rl_mask = _s4_red_conf['Required Role'].astype(str).str.strip() == _rl
                    _rl_hrs  = float(pd.to_numeric(_s4_red_conf.loc[_rl_mask, _s4_mhrs_col], errors='coerce').fillna(0.0).sum())
                    _s4_red_hrs_by_role[_rl] = _rl_hrs
                    _s4_red_hrs_total += _rl_hrs * _shrink_cache.get(_rl, 1.0)
                # Also count rows without a specific role (prorate across all roles)
                _s4r_no_role = _s4_red_conf[
                    _s4_red_conf.get('Required Role', pd.Series('', index=_s4_red_conf.index))
                    .astype(str).str.strip().isin(['', 'nan', 'None'])
                ]
                if not _s4r_no_role.empty:
                    _s4r_norole_hrs = float(pd.to_numeric(_s4r_no_role[_s4_mhrs_col], errors='coerce').fillna(0.0).sum())
                    _s4_red_hrs_total += _s4r_norole_hrs  # use average shrinkage for generic rows

            # ── Door Count Variation (Scenario) per scope ────────────────────
            _s4_dc_raw_s4 = st.session_state.get('s4v2_doorcount_df', pd.DataFrame())
            _s4_dc_conf   = _s4_dc_raw_s4[
                _s4_dc_raw_s4.get('Confirmed', pd.Series(False, index=_s4_dc_raw_s4.index)).eq(True)
            ] if not _s4_dc_raw_s4.empty else pd.DataFrame()
            _s4_dc_adj_hrs = 0.0
            _s4_mpct_col = f"M{_i+1} (%)"
            if not _s4_dc_conf.empty and _s4_mpct_col in _s4_dc_conf.columns:
                # Build per-client base hours from _rb_scope (post Step-3 Final Hours)
                _s4_rb_noempty = not (_rb_scope.empty if hasattr(_rb_scope, 'empty') else True)
                _s4_bcol = f"M{_i+1} ({_ms}) - Final Hours"
                for _, _dcr in _s4_dc_conf.iterrows():
                    _dc_cli  = str(_dcr.get('Client', '') or '').strip()
                    _dc_pod  = str(_dcr.get('POD', '') or '').strip()
                    _dc_pct  = float(pd.to_numeric(_dcr.get(_s4_mpct_col, 0), errors='coerce') or 0.0)
                    if _dc_pct == 0.0 or not _dc_cli:
                        continue
                    # Scope check: if we're in a POD scope, the client must belong to it
                    if _scope_is_pod and _dc_cli.lower() not in _pod_clients_lower:
                        if _dc_pod and _dc_pod != _scope:
                            continue
                    elif _scope_is_sr and _dc_cli.lower() not in _pod_clients_lower:
                        continue
                    # Get client's base hours from rb_scope
                    if _s4_rb_noempty and 'Client' in _rb_scope.columns and _s4_bcol in _rb_scope.columns:
                        _cli_mask_dc = _rb_scope['Client'].astype(str).str.strip().str.lower() == _dc_cli.lower()
                        _cli_base_hrs = float(_rb_scope.loc[_cli_mask_dc, _s4_bcol].sum())
                    else:
                        _cli_base_hrs = 0.0
                    _s4_dc_adj_hrs += _cli_base_hrs * (_dc_pct / 100.0)

            # ── Per-role FTE adjustments from Add/Remove Hours ───────────────
            _add_fte_a1 = (_s4_add_hrs_by_role.get('Accountant I',       0) * _shrink_a1) / _d_fte_m
            _add_fte_a2 = (_s4_add_hrs_by_role.get('Accountant II',      0) * _shrink_a2) / _d_fte_m
            _add_fte_gn = (_s4_add_hrs_by_role.get('General Accountant', 0) * _shrink_gn) / _d_fte_m
            _add_fte_sr = (_s4_add_hrs_by_role.get('Sr. Accountant',     0) * _shrink_sr) / _d_fte_m
            _red_fte_a1 = (_s4_red_hrs_by_role.get('Accountant I',       0) * _shrink_a1) / _d_fte_m
            _red_fte_a2 = (_s4_red_hrs_by_role.get('Accountant II',      0) * _shrink_a2) / _d_fte_m
            _red_fte_gn = (_s4_red_hrs_by_role.get('General Accountant', 0) * _shrink_gn) / _d_fte_m
            _red_fte_sr = (_s4_red_hrs_by_role.get('Sr. Accountant',     0) * _shrink_sr) / _d_fte_m

            # Final Required FTEs (post-auto base + hours adjustments + add/reduce hours)
            _new_req_a1  = max(0.0, _req_a1 + _adj_fte_a1 + _add_fte_a1 - _red_fte_a1)
            _new_req_a2  = max(0.0, _req_a2 + _adj_fte_a2 + _add_fte_a2 - _red_fte_a2)
            _new_req_gn  = max(0.0, _req_gn + _adj_fte_gn + _add_fte_gn - _red_fte_gn)
            _new_req_sr  = max(0.0, _req_sr + _adj_fte_sr + _add_fte_sr - _red_fte_sr)
            _new_req_tot = _new_req_a1 + _new_req_a2 + _new_req_gn + _new_req_sr

            # Automation savings in total hours (for "(-) Automations" display row)
            _scen_auto_hrs = _s4_auto_sav_total[_i] if _i < len(_s4_auto_sav_total) else 0.0
            _new_b_auto    = _scen_auto_hrs
            _new_b_fin     = max(0.0, _b_curr - _scen_auto_hrs + _hrs_total_adj + _s4_add_hrs_total - _s4_red_hrs_total + _s4_dc_adj_hrs)

            # ── Apply Actual HC adjustments ──────────────────────────────────
            _m_tot  = (float(_base_hc_tot) + _net_adj_total) if _base_hc_tot is not None else None
            _m_acc1 = float(_base_hc_by_role.get('Accountant I', 0) or 0)       + _net_adj_by_role.get('Accountant I', 0)
            _m_acc2 = float(_base_hc_by_role.get('Accountant II', 0) or 0)      + _net_adj_by_role.get('Accountant II', 0)
            _m_gen  = float(_base_hc_by_role.get('General Accountant', 0) or 0) + _net_adj_by_role.get('General Accountant', 0)
            _m_sr   = float(_base_hc_by_role.get('Sr. Accountant', 0) or 0)     + _net_adj_by_role.get('Sr. Accountant', 0)
            _m_mgr  = float(_base_hc_mgr) if _base_hc_mgr is not None else None
            _m_mrr  = _b_mrr + _mi_new_mrr - _mi_chrn_mrr

            _d_tot  = round(_m_tot  - _new_req_tot, 2) if _m_tot  is not None else None
            _d_acc1 = round(_m_acc1 - _new_req_a1,  2) if _base_hc_by_role.get('Accountant I')       is not None else None
            _d_acc2 = round(_m_acc2 - _new_req_a2,  2) if _base_hc_by_role.get('Accountant II')      is not None else None
            _d_gen  = round(_m_gen  - _new_req_gn,  2) if _base_hc_by_role.get('General Accountant') is not None else None
            _d_sr   = round(_m_sr   - _new_req_sr,  2) if _base_hc_by_role.get('Sr. Accountant')     is not None else None
            _rev_hc = _m_mrr / _m_tot if (_m_tot and _m_tot > 0) else None
            _tgt    = (_b_prod / _b_base * 100) if _b_base > 0 else 0
            _hol    = holidays_per_month.get(_ms, 0)

            # Capacity Cost = Required FTEs × cost per role (PERF: cached costs)
            _s4_cap_cost = (
                _new_req_a1 * _c_a1 +
                _new_req_a2 * _c_a2 +
                _new_req_gn * _c_gn +
                _new_req_sr * _c_sr
            )
            _s4_cap_margin     = _m_mrr - _s4_cap_cost
            _s4_cap_margin_pct = (_s4_cap_margin / _m_mrr * 100) if _m_mrr and _m_mrr != 0 else None
            # Expected Cost = Actual HC × cost per role (PERF: cached costs)
            _s4_exp_cost = (
                _m_acc1 * _c_a1 +
                _m_acc2 * _c_a2 +
                _m_gen  * _c_gn +
                _m_sr   * _c_sr
            )
            _s4_exp_margin     = _m_mrr - _s4_exp_cost
            _s4_exp_margin_pct = (_s4_exp_margin / _m_mrr * 100) if _m_mrr and _m_mrr != 0 else None

            _s4_shrinkage_hrs = _b_base - _b_prod
            _scen_rows.setdefault("━ Required Hours",            {})[_col] = _fmt(_new_b_fin,  'n')
            _scen_rows.setdefault("  Current Customer Hours",    {})[_col] = _fmt(_b_prod, 'n')
            _scen_rows.setdefault("  Shrinkage (Hrs)",           {})[_col] = _fmt(_s4_shrinkage_hrs if _s4_shrinkage_hrs > 0 else None, 'n')
            _scen_rows.setdefault("  (+) New Customer Hours",    {})[_col] = _fmt(_b_new  if _b_new  else None, 'n')
            _scen_rows.setdefault("  (-) Confirmed Churn (Hrs)", {})[_col] = _fmt(_b_chrn if _b_chrn else None, 'n')
            _scen_rows.setdefault("  (-) Automations",           {})[_col] = _fmt(_new_b_auto if _new_b_auto else None, 'n')
            _scen_rows.setdefault("  (+) Manual Adjustments",    {})[_col] = _fmt(_b_ap - _b_am if (_b_ap - _b_am) != 0 else None, 'n')
            _scen_rows.setdefault("  (+) Add Hours (Scenario)",  {})[_col] = _fmt(_s4_add_hrs_total if _s4_add_hrs_total else None, 'n')
            _scen_rows.setdefault("  (-) Reduce Hours (Scenario)",{})[_col] = _fmt(_s4_red_hrs_total if _s4_red_hrs_total else None, 'n')
            _scen_rows.setdefault("  (+/-) Door Count (Scenario)",{})[_col] = _fmt(_s4_dc_adj_hrs if _s4_dc_adj_hrs else None, 'n')
            _scen_rows.setdefault("(/) Capacity Productivity",  {})[_col] = _fmt(_tgt, '%')
            _scen_rows.setdefault("(/) Shrinkage (%)",          {})[_col] = _fmt(100 - _tgt if _new_b_fin > 0 else None, '%')
            _s4_act_hc_prod = (
                (_m_acc1 * _util_cache['Accountant I']       +
                 _m_acc2 * _util_cache['Accountant II']      +
                 _m_gen  * _util_cache['General Accountant'] +
                 _m_sr   * _util_cache['Sr. Accountant'])
                / _m_tot * 100
            ) if _m_tot and _m_tot > 0 else None
            _scen_rows.setdefault("(/) Actual HC Productivity", {})[_col] = _fmt(_s4_act_hc_prod, '%')
            _scen_rows.setdefault("━ Required HC (FTEs)",        {})[_col] = _fmt(_new_req_tot,'fte')
            _scen_rows.setdefault("  · Accountant I",            {})[_col] = _fmt(_new_req_a1, 'fte')
            _scen_rows.setdefault("  · Accountant II",           {})[_col] = _fmt(_new_req_a2, 'fte')
            _scen_rows.setdefault("  · General Accountant",      {})[_col] = _fmt(_new_req_gn, 'fte')
            _scen_rows.setdefault("  · Sr. Accountant",          {})[_col] = _fmt(_new_req_sr, 'fte')
            _scen_rows.setdefault("━ Actual HC (Report)",        {})[_col] = _fmt(_m_tot,  'fte')
            _scen_rows.setdefault("  · Accountant I (actual)",   {})[_col] = _fmt(_m_acc1, 'fte')
            _scen_rows.setdefault("  · Accountant II (actual)",  {})[_col] = _fmt(_m_acc2, 'fte')
            _scen_rows.setdefault("  · General Acc. (actual)",   {})[_col] = _fmt(_m_gen,  'fte')
            _scen_rows.setdefault("  · Sr. Accountant (actual)", {})[_col] = _fmt(_m_sr,   'fte')
            _scen_rows.setdefault("  · Managers (actual)",       {})[_col] = _fmt(_m_mgr,  'fte')
            _scen_rows.setdefault("━ HC Δ (Actual − Required)",  {})[_col] = _fmt(_d_tot,  'dec')
            _scen_rows.setdefault("  · Δ Accountant I",          {})[_col] = _fmt(_d_acc1, 'dec')
            _scen_rows.setdefault("  · Δ Accountant II",         {})[_col] = _fmt(_d_acc2, 'dec')
            _scen_rows.setdefault("  · Δ General Accountant",    {})[_col] = _fmt(_d_gen,  'dec')
            _scen_rows.setdefault("  · Δ Sr. Accountant",        {})[_col] = _fmt(_d_sr,   'dec')
            _scen_rows.setdefault("━ MRR ($)",                   {})[_col] = _fmt(_m_mrr,  '$')
            _m_new_mrr_disp = _b_newm + _mi_new_mrr
            _m_chrn_mrr_disp= _b_chnm + _mi_chrn_mrr
            _scen_rows.setdefault("  (+) New MRR ($)",           {})[_col] = _fmt(_m_new_mrr_disp  if _m_new_mrr_disp  else None, '$')
            _scen_rows.setdefault("  (-) Churn MRR ($)",         {})[_col] = _fmt(_m_chrn_mrr_disp if _m_chrn_mrr_disp else None, '$')
            _scen_rows.setdefault("  Revenue / HC ($)",          {})[_col] = _fmt(_rev_hc, '$')
            _scen_rows.setdefault("━ Cost & Margin",             {})[_col] = _fmt(_s4_exp_margin_pct, '%')
            _scen_rows.setdefault("  Capacity Cost ($)",         {})[_col] = _fmt(_s4_cap_cost, '$')
            _scen_rows.setdefault("  Capacity Margin ($)",       {})[_col] = _fmt(_s4_cap_margin, '$')
            _scen_rows.setdefault("  Capacity Margin (%)",       {})[_col] = _fmt(_s4_cap_margin_pct, '%')
            _scen_rows.setdefault("  Expected Cost ($)",         {})[_col] = _fmt(_s4_exp_cost, '$')
            _scen_rows.setdefault("  Expected Margin ($)",       {})[_col] = _fmt(_s4_exp_margin, '$')
            _scen_rows.setdefault("  Expected Margin (%)",       {})[_col] = _fmt(_s4_exp_margin_pct, '%')
            # Per-month learning-curve-aware AHT and split ticket counts
            # (PERF: date columns parsed once above the loop; active masks cached.)
            _s4aht_sm, _s4aht_em = _src_sm_em_by_month[_i]
            if not _src_empty:
                _src_m = _src_raw[_src_active_by_month[_i]].copy()
            else:
                _src_m = _src_raw
            if not _src_m.empty and 'Go Live' in _src_m.columns:
                _s4aht_gl  = pd.to_datetime(_src_m['Go Live'], errors='coerce')
                _s4aht_hgl = _s4aht_gl.notna()
                _s4aht_glf = _s4aht_gl.fillna(_s4aht_sm)
                _s4aht_md  = np.where(_s4aht_hgl,
                    (_s4aht_sm.year  - _s4aht_glf.dt.year)  * 12 +
                    (_s4aht_sm.month - _s4aht_glf.dt.month), 999)
                _s4aht_lc  = np.select(
                    [~_s4aht_hgl, (_s4aht_md==0), (_s4aht_md==1), (_s4aht_md==2)],
                    [1.0, 1.17, 0.86, 0.99], default=1.0)
            else:
                _s4aht_lc = np.ones(len(_src_m)) if not _src_m.empty else np.array([1.0])
            _s4_ptix = _safe_num_s4(_src_m.get('Closed tickets with Proc time', pd.Series(dtype=float))) if not _src_m.empty else pd.Series([0.0])
            _s4_rtix = _safe_num_s4(_src_m.get('Closed tickets with rev time',  pd.Series(dtype=float))) if not _src_m.empty else pd.Series([0.0])
            _s4_paht = _safe_num_s4(_src_m.get('>>> FINAL Capacity Proc AHT',   pd.Series(dtype=float))) if not _src_m.empty else pd.Series([0.0])
            _s4_raht = _safe_num_s4(_src_m.get('>>> FINAL Capacity Rev AHT',    pd.Series(dtype=float))) if not _src_m.empty else pd.Series([0.0])
            _s4_proc_tix = int(_s4_ptix.sum())
            _s4_rev_tix  = int(_s4_rtix.sum())
            _s4_tot_tix  = _s4_proc_tix + _s4_rev_tix
            _s_aht = (
                (_s4_ptix * _s4_paht * _s4aht_lc).sum() +
                (_s4_rtix * _s4_raht * _s4aht_lc).sum()
            ) / _s4_tot_tix if _s4_tot_tix > 0 else 0.0
            # Active client count for this scope this month (PERF: reuse cached masks)
            _s4_cli_count = 0
            if _i in _duc_mrr_by_month and not _duc.empty:
                _sm_c, _em_c, _active_c = _duc_mrr_by_month[_i]
                if _scope == "Overall":
                    _s4_cli_count = int(_active_c.sum())
                elif _pod_clients_lower_set:
                    _s4_cli_count = int((_active_c & _duc_name_lower.isin(_pod_clients_lower_set)).sum())
            # ── Per-month properties/doors/sqft (active clients only) ──────
            if _i in _duc_mrr_by_month and not _duc.empty and 'client_name' in _duc.columns and not _csn.empty:
                _sm_s4p, _em_s4p, _act_s4p = _duc_mrr_by_month[_i]
                _act_names_s4 = set(_duc.loc[_act_s4p, 'client_name'].astype(str).str.strip().str.lower())
                if _pod_clients_lower_set:
                    _act_names_s4 &= _pod_clients_lower_set
                _csn_m = _csn[_csn['client_name'].astype(str).str.strip().str.lower().isin(_act_names_s4)]
                _s_res_prop  = int(_safe_num_s4(_csn_m.get('Res Prop', pd.Series(dtype=float))).sum())
                _s_comm_prop = int(_safe_num_s4(_csn_m.get('Commercial Properties', pd.Series(dtype=float))).sum())
                _s_prop      = _s_res_prop + _s_comm_prop
                _s_res_door  = int(_safe_num_s4(_csn_m.get('Res doors', pd.Series(dtype=float))).sum())
                _s_comm_door = int(_safe_num_s4(_csn_m.get('Commercial Doors', pd.Series(dtype=float))).sum())
                _s_door      = _s_res_door + _s_comm_door
                _s_sqft      = int(_safe_num_s4(_csn_m.get('SQFT Commercial', pd.Series(dtype=float))).sum())
            _scen_rows.setdefault("━ Property Count",            {})[_col] = _fmt(_s_prop, 'n')
            _scen_rows.setdefault("  Res Properties",            {})[_col] = _fmt(_s_res_prop if _s_res_prop else None, 'n')
            _scen_rows.setdefault("  Comm Properties",           {})[_col] = _fmt(_s_comm_prop if _s_comm_prop else None, 'n')
            _scen_rows.setdefault("  Client Count",              {})[_col] = _fmt(_s4_cli_count, 'n')
            _scen_rows.setdefault("  Res Doors",                 {})[_col] = _fmt(_s_res_door if _s_res_door else None, 'n')
            _scen_rows.setdefault("  Comm Doors",                {})[_col] = _fmt(_s_comm_door if _s_comm_door else None, 'n')
            _scen_rows.setdefault("  SQFT (Comm)",               {})[_col] = _fmt(_s_sqft if _s_sqft else None, 'n')
            _scen_rows.setdefault("  Tickets to Process",        {})[_col] = _fmt(_s4_proc_tix, 'n')
            _scen_rows.setdefault("  Tickets to Review",         {})[_col] = _fmt(_s4_rev_tix,  'n')
            _scen_rows.setdefault("  AHT (min)",                 {})[_col] = _fmt(_s_aht,        'dec')
            _scen_rows.setdefault("━ Working Days",              {})[_col] = _fmt(_b_wday, 'n')
            _scen_rows.setdefault("  Holidays",                  {})[_col] = _fmt(_hol,    'n')

            # ── Capture raw numerics for Before vs. After summary ────────────
            _scen_nums.setdefault("Required Hours (Hrs)",        {})[_col] = _new_b_fin
            _scen_nums.setdefault("Required HC (FTEs)",          {})[_col] = _new_req_tot
            _scen_nums.setdefault("  · Accountant I",            {})[_col] = _new_req_a1
            _scen_nums.setdefault("  · Accountant II",           {})[_col] = _new_req_a2
            _scen_nums.setdefault("  · General Accountant",      {})[_col] = _new_req_gn
            _scen_nums.setdefault("  · Sr. Accountant",          {})[_col] = _new_req_sr
            _scen_nums.setdefault("HC Δ (Actual − Required)",    {})[_col] = _d_tot   if _d_tot   is not None else float('nan')
            _scen_nums.setdefault("MRR ($)",                     {})[_col] = _m_mrr
            _scen_nums.setdefault("Expected Margin (%)",         {})[_col] = _s4_exp_margin_pct if _s4_exp_margin_pct is not None else float('nan')
            _scen_nums.setdefault("Expected Cost ($)",           {})[_col] = _s4_exp_cost
            _scen_nums.setdefault("Capacity Cost ($)",           {})[_col] = _s4_cap_cost
            _scen_nums.setdefault("Capacity Margin (%)",         {})[_col] = _s4_cap_margin_pct if _s4_cap_margin_pct is not None else float('nan')
            # Step-3 baseline values for same month (from _exec / _rb_scope)
            # NOTE: FTEs are recomputed from _s4_base_hrs_role / _d_fte_m — the SAME
            # formula the scenario uses (minus Step-4 adjustments). This guarantees
            # baseline ≡ scenario when no Step-4 changes are applied, eliminating
            # rounding-boundary phantom Δs (e.g. 383.91 vs 383.90).
            if _scope == "Overall":
                _b3_fte_a1  = max(0.0, float(_s4_base_hrs_role.get('Accountant I',       [0.0]*6)[_i]) / _d_fte_m)
                _b3_fte_a2  = max(0.0, float(_s4_base_hrs_role.get('Accountant II',      [0.0]*6)[_i]) / _d_fte_m)
                _b3_fte_gn  = max(0.0, float(_s4_base_hrs_role.get('General Accountant', [0.0]*6)[_i]) / _d_fte_m)
                _b3_fte_sr  = max(0.0, float(_s4_base_hrs_role.get('Sr. Accountant',     [0.0]*6)[_i]) / _d_fte_m)
                _b3_fte_tot = _b3_fte_a1 + _b3_fte_a2 + _b3_fte_gn + _b3_fte_sr
                _b3_mrr     = float(_r.get("Total MRR ($)", 0) or 0)
                _base_nums.setdefault("Required Hours (Hrs)",    {})[_col] = float(_r.get("7. Total Required Hours (Final)", 0) or 0)
                _base_nums.setdefault("Required HC (FTEs)",      {})[_col] = _b3_fte_tot
                _base_nums.setdefault("  · Accountant I",        {})[_col] = _b3_fte_a1
                _base_nums.setdefault("  · Accountant II",       {})[_col] = _b3_fte_a2
                _base_nums.setdefault("  · General Accountant",  {})[_col] = _b3_fte_gn
                _base_nums.setdefault("  · Sr. Accountant",      {})[_col] = _b3_fte_sr
                _base_nums.setdefault("MRR ($)",                 {})[_col] = _b3_mrr
                # HC Δ derived from Actual HC minus baseline FTE total (matches scenario's formula)
                _base_hc_tot_b3 = float(_base_hc_tot) if _base_hc_tot is not None else 0.0
                _base_nums.setdefault("HC Δ (Actual − Required)", {})[_col] = round(_base_hc_tot_b3 - _b3_fte_tot, 2)
                # Capacity Cost / Margin using Step-3 Required FTEs × role cost (PERF: cached)
                _b3_cap_cost = (
                    _b3_fte_a1 * _c_a1 +
                    _b3_fte_a2 * _c_a2 +
                    _b3_fte_gn * _c_gn +
                    _b3_fte_sr * _c_sr
                )
                _b3_cap_margin_pct = ((_b3_mrr - _b3_cap_cost) / _b3_mrr * 100) if _b3_mrr else float('nan')
                # Expected Cost / Margin using base Actual HC × role cost (PERF: cached)
                _b3_exp_cost = (
                    float(_base_hc_by_role.get('Accountant I', 0) or 0)       * _c_a1 +
                    float(_base_hc_by_role.get('Accountant II', 0) or 0)      * _c_a2 +
                    float(_base_hc_by_role.get('General Accountant', 0) or 0) * _c_gn +
                    float(_base_hc_by_role.get('Sr. Accountant', 0) or 0)     * _c_sr
                )
                _b3_exp_margin_pct = ((_b3_mrr - _b3_exp_cost) / _b3_mrr * 100) if _b3_mrr else float('nan')
                _base_nums.setdefault("Expected Margin (%)",     {})[_col] = _b3_exp_margin_pct
                _base_nums.setdefault("Expected Cost ($)",       {})[_col] = _b3_exp_cost
                _base_nums.setdefault("Capacity Cost ($)",       {})[_col] = _b3_cap_cost
                _base_nums.setdefault("Capacity Margin (%)",     {})[_col] = _b3_cap_margin_pct
            else:
                _cfin_b3   = f"M{_i+1} ({_ms}) - Final Hours"
                # Baseline FTEs recomputed from _s4_base_hrs_role / _d_fte_m
                # (same formula as scenario → zero phantom Δ when no changes).
                _b3_fte_by_role = {
                    _brl: max(0.0, float(_s4_base_hrs_role.get(_brl, [0.0]*6)[_i]) / _d_fte_m)
                    for _brl in roles_permitidos
                }
                _b3_fte_tot = sum(_b3_fte_by_role.values())
                # Required Hours: mirror scenario's _b_curr exactly.
                # Sr scope → sum of per-role base hours; POD scope → _rb_scope Final Hours sum.
                if _scope_is_sr:
                    _b3_req_hrs = sum(float(_s4_base_hrs_role.get(_rl, [0.0]*6)[_i]) for _rl in roles_permitidos)
                else:
                    _b3_req_hrs = _rb_sum_cache.get(_cfin_b3, 0.0)
                _base_nums.setdefault("Required Hours (Hrs)",    {})[_col] = _b3_req_hrs
                _base_nums.setdefault("Required HC (FTEs)",      {})[_col] = _b3_fte_tot
                for _brl in roles_permitidos:
                    _base_nums.setdefault(f"  · {_brl}", {})[_col] = _b3_fte_by_role[_brl]
                _base_nums.setdefault("HC Δ (Actual − Required)", {})[_col] = float('nan')
                _base_nums.setdefault("MRR ($)",                 {})[_col] = _b_mrr
                # Capacity Cost / Margin using scoped Step-3 Required FTEs × role cost (PERF: cached)
                _b3_cap_cost = (
                    _b3_fte_by_role.get('Accountant I', 0.0)       * _c_a1 +
                    _b3_fte_by_role.get('Accountant II', 0.0)      * _c_a2 +
                    _b3_fte_by_role.get('General Accountant', 0.0) * _c_gn +
                    _b3_fte_by_role.get('Sr. Accountant', 0.0)     * _c_sr
                )
                _b3_cap_margin_pct = ((_b_mrr - _b3_cap_cost) / _b_mrr * 100) if _b_mrr else float('nan')
                # Expected Cost / Margin using base Actual HC × role cost (PERF: cached)
                _b3_exp_cost = (
                    float(_base_hc_by_role.get('Accountant I', 0) or 0)       * _c_a1 +
                    float(_base_hc_by_role.get('Accountant II', 0) or 0)      * _c_a2 +
                    float(_base_hc_by_role.get('General Accountant', 0) or 0) * _c_gn +
                    float(_base_hc_by_role.get('Sr. Accountant', 0) or 0)     * _c_sr
                )
                _b3_exp_margin_pct = ((_b_mrr - _b3_exp_cost) / _b_mrr * 100) if _b_mrr else float('nan')
                _base_nums.setdefault("Expected Margin (%)",     {})[_col] = _b3_exp_margin_pct
                _base_nums.setdefault("Expected Cost ($)",       {})[_col] = _b3_exp_cost
                _base_nums.setdefault("Capacity Cost ($)",       {})[_col] = _b3_cap_cost
                _base_nums.setdefault("Capacity Margin (%)",     {})[_col] = _b3_cap_margin_pct

        if not _months:
            st.info("ℹ️ Run Steps 1–3 first to populate baseline data.")
            return

        _df_scen = pd.DataFrame(_scen_rows, index=_months).T
        _df_scen.index.name = ""

        # ── Collapsible groups (same structure as Capacity Overview) ─────────
        _s4_groups = {
            "━ Required Hours":           ["  Current Customer Hours", "  Shrinkage (Hrs)", "  (+) New Customer Hours", "  (-) Confirmed Churn (Hrs)", "  (-) Automations", "  (+) Manual Adjustments"],
            "━ Required HC (FTEs)":       ["  · Accountant I", "  · Accountant II", "  · General Accountant", "  · Sr. Accountant"],
            "━ Actual HC (Report)":       ["  · Accountant I (actual)", "  · Accountant II (actual)", "  · General Acc. (actual)", "  · Sr. Accountant (actual)", "  · Managers (actual)"],
            "━ HC Δ (Actual − Required)": ["  · Δ Accountant I", "  · Δ Accountant II", "  · Δ General Accountant", "  · Δ Sr. Accountant"],
            "━ MRR ($)":                  ["  (+) New MRR ($)", "  (-) Churn MRR ($)", "  Revenue / HC ($)"],
            "━ Cost & Margin":            ["  Capacity Cost ($)", "  Capacity Margin ($)", "  Capacity Margin (%)",
                                           "  Expected Cost ($)", "  Expected Margin ($)", "  Expected Margin (%)"],
            "━ Property Count":           ["  Res Properties", "  Comm Properties", "  Client Count", "  Res Doors", "  Comm Doors", "  SQFT (Comm)"],
            "━ Working Days":             ["  Holidays"],
        }
        for _gh in _s4_groups:
            if f"_s4v2_exp_{_gh}" not in st.session_state:
                st.session_state[f"_s4v2_exp_{_gh}"] = False
        _short_s4 = lambda s: s.replace("━ ", "").replace(" (FTEs)", "").replace(" ($)", "").replace(" (Report)", "")
        _ca_c, _ea_c, *_gcols = st.columns([1, 1] + [1] * len(_s4_groups))
        if _ca_c.button("▶ Collapse All", key="_s4v2_ca", use_container_width=True):
            for _gh in _s4_groups: st.session_state[f"_s4v2_exp_{_gh}"] = False
            st.rerun(scope="fragment")
        if _ea_c.button("► Expand All", key="_s4v2_ea", use_container_width=True):
            for _gh in _s4_groups: st.session_state[f"_s4v2_exp_{_gh}"] = True
            st.rerun(scope="fragment")
        for _ci, (_gh, _) in enumerate(_s4_groups.items()):
            _ek  = f"_s4v2_exp_{_gh}"
            _ico = "▼" if st.session_state.get(_ek, False) else "▶"
            if _gcols[_ci].button(f"{_ico} {_short_s4(_gh)}", key=f"_s4v2btn_{_ci}", use_container_width=True):
                st.session_state[_ek] = not st.session_state.get(_ek, False)
                st.rerun(scope="fragment")

        _all_det = {_rr for _drs in _s4_groups.values() for _rr in _drs}
        _vis = []
        for _rr in _df_scen.index:
            if _rr in _all_det:
                for _gh, _drs in _s4_groups.items():
                    if _rr in _drs and st.session_state.get(f"_s4v2_exp_{_gh}", False):
                        _vis.append(_rr); break
            else:
                _vis.append(_rr)
        _h = min(900, len(_vis) * 35 + 42)
        st.dataframe(_df_scen.loc[_vis], use_container_width=True, height=_h)

        # ── Before vs. After Summary ─────────────────────────────────────────
        with st.expander("📊 Before vs. After Summary (Step 3 → Step 4)", expanded=False):
            st.caption(
                "Compares the **Step 3 baseline** (before any Step 4 changes) against the **Step 4 scenario**. "
                "Green Δ = improvement (lower hours/FTEs required or better margin). "
                "Red Δ = higher demand or worse margin."
            )
            if _scen_nums and _base_nums:
                # Build a flat comparison DataFrame:
                # rows = metric, columns = M1 Base / M1 Scen / M1 Δ / M2 Base / ...
                _cmp_rows = {}
                _cmp_months = list(_months)
                # Metrics to display and whether lower delta is better ("lower") or higher ("higher") or neutral ("")
                _cmp_metrics = [
                    ("Required Hours (Hrs)",     "lower",  "{:,.0f}"),
                    ("Required HC (FTEs)",       "lower",  "{:.2f}"),
                    ("  · Accountant I",         "lower",  "{:.2f}"),
                    ("  · Accountant II",        "lower",  "{:.2f}"),
                    ("  · General Accountant",   "lower",  "{:.2f}"),
                    ("  · Sr. Accountant",       "lower",  "{:.2f}"),
                    ("HC Δ (Actual − Required)", "higher", "{:.2f}"),
                    ("MRR ($)",                  "higher", "${:,.0f}"),
                    ("Capacity Cost ($)",        "lower",  "${:,.0f}"),
                    ("Capacity Margin (%)",      "higher", "{:.1f}%"),
                    ("Expected Cost ($)",        "lower",  "${:,.0f}"),
                    ("Expected Margin (%)",      "higher", "{:.1f}%"),
                ]
                def _fmtv(v, fs):
                    if isinstance(v, float) and v != v: return "—"
                    try: return fs.format(v)
                    except Exception: return str(round(v, 2))

                _cmp_cols = []
                for _cm in _cmp_months:
                    _cmp_cols += [f"{_cm} Baseline", f"{_cm} Scenario", f"{_cm} Δ"]

                for _mname, _dir, _fmt_s in _cmp_metrics:
                    _row_vals = []
                    for _cm in _cmp_months:
                        _bv = _base_nums.get(_mname, {}).get(_cm, float('nan'))
                        _sv = _scen_nums.get(_mname, {}).get(_cm, float('nan'))
                        _dv = (_sv - _bv) if (not (isinstance(_bv, float) and _bv != _bv) and
                                               not (isinstance(_sv, float) and _sv != _sv)) else float('nan')
                        _row_vals += [_fmtv(_bv, _fmt_s), _fmtv(_sv, _fmt_s), _fmtv(_dv, _fmt_s)]
                    _cmp_rows[_mname] = _row_vals

                _df_cmp = pd.DataFrame(_cmp_rows, index=_cmp_cols).T
                _df_cmp.index.name = "Metric"

                def _highlight_delta(row):
                    styles = [""] * len(row)
                    # Find which columns are "Δ" columns (every 3rd starting at index 2)
                    for _ci, _cv in enumerate(row):
                        if _ci % 3 == 2:  # Δ column
                            _metric = row.name
                            _dir = next((_d for _mn, _d, _ in _cmp_metrics if _mn == _metric), "")
                            try:
                                _num = float(str(_cv).replace("$", "").replace(",", "").replace("%", ""))
                                if _dir == "lower":
                                    styles[_ci] = "color: green; font-weight: bold" if _num < -0.001 else (
                                        "color: red; font-weight: bold" if _num > 0.001 else "color: gray")
                                elif _dir == "higher":
                                    styles[_ci] = "color: green; font-weight: bold" if _num > 0.001 else (
                                        "color: red; font-weight: bold" if _num < -0.001 else "color: gray")
                            except Exception:
                                pass
                    return styles

                _cmp_h = len(_cmp_metrics) * 35 + 38
                st.dataframe(
                    _df_cmp.style.apply(_highlight_delta, axis=1),
                    use_container_width=True,
                    height=_cmp_h,
                )
            else:
                st.info("Run the scenario to see the Before vs. After comparison.")

        # Persist _df_scen so the separate inputs fragment can use it for downloads
        st.session_state['_s4v2_df_scen'] = _df_scen

        # NOTE: editors (HC / MRR / Hours / Automations / Downloads) are rendered
        # by _s4v2_inputs_frag() below — separated so value-typing never reruns
        # this heavy display fragment.

    # ── END of _s4v2_fragment ────────────────────────────────────────────────────

    # ===========================================================================
    # INPUTS FRAGMENT — editors only; Confirmed checkbox triggers display update
    # ===========================================================================
    @st.fragment
    def _s4v2_inputs_frag():
        # ── Snapshot Confirmed state from PREVIOUS run ───────────────────────────
        # We only trigger a full-page rerun (which refreshes the display fragment)
        # when a Confirmed checkbox changes, NOT when a numeric value is typed.
        def _conf_snap():
            try:
                return hash((
                    tuple(st.session_state.s4v2_hc_adj_df.get("Confirmed",     pd.Series()).tolist()),
                    tuple(st.session_state.s4v2_mrr_adj_df.get("Confirmed",    pd.Series()).tolist()),
                    tuple(st.session_state.s4v2_hrs_role_df.get("Confirmed",   pd.Series()).tolist()),
                    tuple(st.session_state.s4v2_auto_df.get("Confirmed",       pd.Series()).tolist())
                        if not st.session_state.s4v2_auto_df.empty else (),
                    tuple(st.session_state.s4v2_hist_df.get("Confirmed",       pd.Series()).tolist())
                        if not st.session_state.s4v2_hist_df.empty else (),
                    tuple(st.session_state.s4v2_red_df.get("Confirmed",        pd.Series()).tolist())
                        if not st.session_state.s4v2_red_df.empty else (),
                    tuple(st.session_state.s4v2_doorcount_df.get("Confirmed",  pd.Series()).tolist())
                        if not st.session_state.s4v2_doorcount_df.empty else (),
                ))
            except Exception:
                return None

        _conf_before = st.session_state.get('_s4_conf_snap')

        with st.expander("👥 HC · MRR · Hours Adjustments", expanded=False):
            # ── HC Adjustment inputs ─────────────────────────────────────────────
            st.markdown("#### 👥 HC Adjustments — Actual Headcount")
            st.caption(
                "Enter FTEs to **ramp up ↑** or **ramp down ↓** per role. "
                "A value entered in any month forward-fills to all later months. "
                "Changes update the table above in real time."
            )
            st.caption("☑️ Check **Confirmed** on a row to include it in the scenario calculation. Unconfirmed rows are ignored.")
            _hc_cc = {
                "Confirmed":  st.column_config.CheckboxColumn("✅", default=False, help="Only confirmed rows affect the scenario"),
                "Direction": st.column_config.TextColumn("Direction", disabled=True, width="small"),
                "Role":      st.column_config.TextColumn("Role",      disabled=True, width="medium"),
                **{_mc: st.column_config.NumberColumn(f"M{_mc[1:]}", default=0.0, format="%.2f", min_value=0.0) for _mc in _s4v2_mc},
            }
            st.session_state.s4v2_hc_adj_df = st.data_editor(
                st.session_state.s4v2_hc_adj_df,
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                key="s4v2_hc_ed",
                column_config=_hc_cc,
            )
            _hc_conf = st.session_state.s4v2_hc_adj_df["Confirmed"].eq(True)
            _hc_num = st.session_state.s4v2_hc_adj_df.loc[_hc_conf, _s4v2_mc].apply(pd.to_numeric, errors='coerce').fillna(0.0) if _hc_conf.any() else pd.DataFrame(columns=_s4v2_mc)
            _hc_ff  = _hc_num.replace(0.0, np.nan).ffill(axis=1).fillna(0.0)
            if not _hc_num.equals(_hc_ff):
                st.session_state.s4v2_hc_adj_df.loc[_hc_conf, _s4v2_mc] = _hc_ff.values
                try:
                    st.rerun(scope="fragment")
                except Exception:
                    st.rerun()

            # ── MRR Adjustment inputs ────────────────────────────────────────────
            st.markdown("#### 💰 MRR Adjustments")
            st.caption(
                "Enter **New MRR ⊕** to add or **Churn MRR ⊖** to subtract each month. "
                "A value entered in any month forward-fills to all later months."
            )
            st.caption("☑️ Check **Confirmed** on a row to include it in the scenario calculation. Unconfirmed rows are ignored.")
            _mrr_cc = {
                "Confirmed":  st.column_config.CheckboxColumn("✅", default=False, help="Only confirmed rows affect the scenario"),
                "Adjustment": st.column_config.TextColumn("Adjustment", disabled=True, width="medium"),
                **{_mc: st.column_config.NumberColumn(f"M{_mc[1:]}", default=0.0, format="$%.0f", min_value=0.0) for _mc in _s4v2_mc},
            }
            st.session_state.s4v2_mrr_adj_df = st.data_editor(
                st.session_state.s4v2_mrr_adj_df,
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                key="s4v2_mrr_ed",
                column_config=_mrr_cc,
            )
            _mrr_conf = st.session_state.s4v2_mrr_adj_df["Confirmed"].eq(True)
            _mrr_num = st.session_state.s4v2_mrr_adj_df.loc[_mrr_conf, _s4v2_mc].apply(pd.to_numeric, errors='coerce').fillna(0.0) if _mrr_conf.any() else pd.DataFrame(columns=_s4v2_mc)
            _mrr_ff  = _mrr_num.replace(0.0, np.nan).ffill(axis=1).fillna(0.0)
            if not _mrr_num.equals(_mrr_ff):
                st.session_state.s4v2_mrr_adj_df.loc[_mrr_conf, _s4v2_mc] = _mrr_ff.values
                try:
                    st.rerun(scope="fragment")
                except Exception:
                    st.rerun()

            # ── Required Hours adjustments per Role ──────────────────────────────
            st.markdown("#### ⏱️ Required Hours Adjustments per Role")
            st.caption(
                "Enter **productive hours** to add (+) or reduce (−) per role per month. "
                "Shrinkage is applied automatically — productive hours × (2 − utilization + absenteeism + attrition) — "
                "then converted to FTEs via working-day hours. Values forward-fill to later months."
            )
            st.caption("☑️ Check **Confirmed** on a row to include it in the scenario calculation. Unconfirmed rows are ignored.")
            _hrs_role_cc = {
                "Confirmed": st.column_config.CheckboxColumn("✅", default=False, help="Only confirmed rows affect the scenario"),
                "Role": st.column_config.TextColumn("Role", disabled=True, width="medium"),
                **{_mc: st.column_config.NumberColumn(f"M{_mc[1:]}", default=0.0, format="%.0f") for _mc in _s4v2_mc},
            }
            st.session_state.s4v2_hrs_role_df = st.data_editor(
                st.session_state.s4v2_hrs_role_df,
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                key="s4v2_hrs_ed",
                column_config=_hrs_role_cc,
            )
            _hrs_conf = st.session_state.s4v2_hrs_role_df["Confirmed"].eq(True)
            _hrs_num = st.session_state.s4v2_hrs_role_df.loc[_hrs_conf, _s4v2_mc].apply(pd.to_numeric, errors='coerce').fillna(0.0) if _hrs_conf.any() else pd.DataFrame(columns=_s4v2_mc)
            _hrs_ff  = _hrs_num.replace(0.0, np.nan).ffill(axis=1).fillna(0.0)
            if not _hrs_num.equals(_hrs_ff):
                st.session_state.s4v2_hrs_role_df.loc[_hrs_conf, _s4v2_mc] = _hrs_ff.values
                try:
                    st.rerun(scope="fragment")
                except Exception:
                    st.rerun()

        # ── Scenario Automation (full replica of Step 2 Automations tab) ────
        with st.expander("⚙️ Scenario Automations", expanded=False):
            st.markdown(
                "**Choose Client, Task, and the metric(s) this automation reduces (%).** "
                "Use the **Affects** dropdown to select which factors are impacted. "
                "Values are applied from the pre-automation base (independent of Step 2). "
                "Cumulative roll-over: each month's % cannot be less than the previous month."
            )
            _s4a_pods_opts = ["", "All"] + (lista_pods or st.session_state.get('_lista_pods', []))
            _s4a_pms_opts  = _get_pms_opts(include_all=True)
            _s4a_cli_opts  = lista_clientes
            _s4a_tsk_opts  = lista_tareas

            # ── Template download ─────────────────────────────────────────────
            _s4a_tmpl = pd.DataFrame([
                {"Initiative Name": "AP Automation", "Client": "All",
                 "Task (Type - Subtype)": "AP - Invoice Processing",
                 "Affects": "Vol Proc + Vol Rev",
                 "M1 (%)": 10, "M2 (%)": 15, "M3 (%)": 20,
                 "M4 (%)": 20, "M5 (%)": 20, "M6 (%)": 20},
                {"Initiative Name": "AI Coding", "Client": "Acme Corp",
                 "Task (Type - Subtype)": "All",
                 "Affects": "AHT Proc",
                 "M1 (%)": 5, "M2 (%)": 5, "M3 (%)": 10,
                 "M4 (%)": 10, "M5 (%)": 10, "M6 (%)": 10},
            ])
            _s4a_tmpl_buf = BytesIO()
            _s4a_tmpl.to_excel(_s4a_tmpl_buf, index=False)
            _s4a_dl_col, _s4a_ul_col = st.columns([1, 1])
            with _s4a_dl_col:
                st.download_button(
                    "📄 Download Automations Template",
                    _s4a_tmpl_buf.getvalue(),
                    file_name="S4_Automations_Template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_s4a_tmpl",
                )
            with _s4a_ul_col:
                _s4a_upload = st.file_uploader(
                    "📂 Load automations from Excel (optional)",
                    type=["xlsx", "xls"],
                    key="s4a_file_load",
                    label_visibility="collapsed",
                )
                st.caption("📂 Load automations from Excel (optional)")
                if _s4a_upload:
                    try:
                        _s4a_loaded = pd.read_excel(_s4a_upload)
                        _s4a_need_cols = [
                            "Initiative Name", "POD", "PMS", "Client",
                            "Task (Type - Subtype)", "Affects",
                            "M1 (%)", "M2 (%)", "M3 (%)", "M4 (%)", "M5 (%)", "M6 (%)",
                        ]
                        for _s4nc in _s4a_need_cols:
                            if _s4nc not in _s4a_loaded.columns:
                                _s4a_loaded[_s4nc] = None
                        _s4a_loaded = _s4a_loaded[_s4a_need_cols].copy()
                        # Normalise % columns
                        for _s4pc in ["M1 (%)", "M2 (%)", "M3 (%)", "M4 (%)", "M5 (%)", "M6 (%)"]:
                            _s4a_loaded[_s4pc] = pd.to_numeric(_s4a_loaded[_s4pc], errors='coerce').fillna(0.0)
                        # All uploaded rows confirmed by default
                        _s4a_loaded.insert(0, "Confirmed", True)
                        st.session_state.s4v2_auto_df = _s4a_loaded
                        st.success(f"✅ Loaded {len(_s4a_loaded)} automation row(s) — all marked Confirmed.")
                    except Exception as _s4ae:
                        st.error(f"❌ Could not load file: {_s4ae}")

            # ── Ensure required columns exist ─────────────────────────────────
            for _s4col, _s4pos, _s4def in [
                ("Confirmed",      0, True),
                ("POD",            2, ""),
                ("PMS",            3, ""),
            ]:
                if _s4col not in st.session_state.s4v2_auto_df.columns:
                    st.session_state.s4v2_auto_df.insert(_s4pos, _s4col, _s4def)

            st.caption("☑️ Check **Confirmed** to include a row. Leave **POD / PMS / Client / Task** blank or 'All' to apply broadly.")
            st.session_state.s4v2_auto_df = st.data_editor(
                st.session_state.s4v2_auto_df,
                num_rows="dynamic",
                use_container_width=True,
                key="s4v2_auto_ed",
                column_config={
                    "Confirmed":             st.column_config.CheckboxColumn("✅", default=True, help="Only confirmed rows are applied"),
                    "Initiative Name":       st.column_config.TextColumn("Initiative Name"),
                    "POD":                   st.column_config.SelectboxColumn("POD",    options=_s4a_pods_opts, default=""),
                    "PMS":                   st.column_config.SelectboxColumn("PMS",    options=_s4a_pms_opts,  default=""),
                    "Client":                st.column_config.SelectboxColumn("Client", options=_s4a_cli_opts,  default="All"),
                    "Task (Type - Subtype)": st.column_config.SelectboxColumn("Task",   options=_s4a_tsk_opts,  default="All"),
                    "Affects":               st.column_config.SelectboxColumn("Affects", options=AFFECTS_OPTIONS, default="All (Vol + AHT)",
                        help="Vol Proc = Processing Volume | Vol Rev = Review Volume | AHT Proc = Processing Handle Time | AHT Rev = Review Handle Time"),
                    "M1 (%)": st.column_config.NumberColumn("M1 (%)", min_value=0.0, max_value=100.0, default=0.0, format="%.1f%%"),
                    "M2 (%)": st.column_config.NumberColumn("M2 (%)", min_value=0.0, max_value=100.0, default=0.0, format="%.1f%%"),
                    "M3 (%)": st.column_config.NumberColumn("M3 (%)", min_value=0.0, max_value=100.0, default=0.0, format="%.1f%%"),
                    "M4 (%)": st.column_config.NumberColumn("M4 (%)", min_value=0.0, max_value=100.0, default=0.0, format="%.1f%%"),
                    "M5 (%)": st.column_config.NumberColumn("M5 (%)", min_value=0.0, max_value=100.0, default=0.0, format="%.1f%%"),
                    "M6 (%)": st.column_config.NumberColumn("M6 (%)", min_value=0.0, max_value=100.0, default=0.0, format="%.1f%%"),
                },
            )
            # Enforce cumulative roll-over (same as Step 2)
            _s4a_pct_cols = ["M1 (%)", "M2 (%)", "M3 (%)", "M4 (%)", "M5 (%)", "M6 (%)"]
            if not st.session_state.s4v2_auto_df.empty:
                _s4a_exist = [c for c in _s4a_pct_cols if c in st.session_state.s4v2_auto_df.columns]
                if _s4a_exist:
                    _s4a_before = st.session_state.s4v2_auto_df[_s4a_exist].apply(pd.to_numeric, errors='coerce').fillna(0.0)
                    _s4a_after  = _s4a_before.cummax(axis=1)
                    if not _s4a_before.equals(_s4a_after):
                        st.session_state.s4v2_auto_df[_s4a_exist] = _s4a_after
                        try:
                            st.rerun(scope="fragment")
                        except Exception:
                            st.rerun()

        # ── Add Hours (Scenario) ─────────────────────────────────────────────
        with st.expander("➕ Add Hours (Scenario)", expanded=False):
            st.markdown("**Extra hours to add per client/role per month (e.g. onboarding, historical accounting, new clients).** Include the POD for new clients not yet in the cascade.")
            _s4h_pods_opts = [""] + (lista_pods or st.session_state.get('_lista_pods', []))
            _s4h_cli_opts  = [""] + lista_clientes
            _s4h_rol_opts  = roles_permitidos

            # Template download
            _s4h_tmpl = pd.DataFrame([
                {"POD": "POD A", "Client": "Acme Corp", "Required Role": "Accountant I",
                 "M1 (Hrs)": 20, "M2 (Hrs)": 20, "M3 (Hrs)": 20,
                 "M4 (Hrs)": 0,  "M5 (Hrs)": 0,  "M6 (Hrs)": 0},
            ])
            _s4h_buf = BytesIO()
            _s4h_tmpl.to_excel(_s4h_buf, index=False)
            _s4h_dl_col, _s4h_ul_col = st.columns([1, 1])
            with _s4h_dl_col:
                st.download_button("📄 Download Add Hours Template", _s4h_buf.getvalue(),
                                   file_name="S4_AddHours_Template.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key="dl_s4h_tmpl")
            with _s4h_ul_col:
                _s4h_upload = st.file_uploader("📂 Upload Add Hours file", type=["xlsx"],
                                               key="fu_s4h_tmpl", label_visibility="collapsed")
                st.caption("📂 Upload Add Hours file (.xlsx)")
                if _s4h_upload is not None:
                    try:
                        _s4h_up_df = pd.read_excel(_s4h_upload)
                        if 'Role' in _s4h_up_df.columns and 'Required Role' not in _s4h_up_df.columns:
                            _s4h_up_df = _s4h_up_df.rename(columns={'Role': 'Required Role'})
                        _s4h_expected = ['Confirmed', 'POD', 'Client', 'Required Role',
                                         'M1 (Hrs)', 'M2 (Hrs)', 'M3 (Hrs)', 'M4 (Hrs)', 'M5 (Hrs)', 'M6 (Hrs)']
                        for _s4hc in _s4h_expected:
                            if _s4hc not in _s4h_up_df.columns:
                                _s4h_up_df[_s4hc] = True if _s4hc == 'Confirmed' else (0.0 if '(Hrs)' in _s4hc else '')
                        _s4h_up_df = _s4h_up_df[_s4h_expected]
                        _s4h_up_df['Confirmed'] = True
                        for _s4hc in ['M1 (Hrs)', 'M2 (Hrs)', 'M3 (Hrs)', 'M4 (Hrs)', 'M5 (Hrs)', 'M6 (Hrs)']:
                            _s4h_up_df[_s4hc] = pd.to_numeric(_s4h_up_df[_s4hc], errors='coerce').fillna(0.0)
                        st.session_state.s4v2_hist_df = _s4h_up_df
                        st.success(f"✅ Loaded {len(_s4h_up_df)} row(s) — all marked Confirmed.")
                    except Exception as _s4he:
                        st.error(f"❌ Could not load file: {_s4he}")

            if "Confirmed" not in st.session_state.s4v2_hist_df.columns:
                st.session_state.s4v2_hist_df.insert(0, "Confirmed", False)
            if "POD" not in st.session_state.s4v2_hist_df.columns:
                st.session_state.s4v2_hist_df.insert(1, "POD", "")
            st.caption("☑️ **Confirmed** rows are applied in the scenario. All uploaded rows are confirmed by default.")
            st.session_state.s4v2_hist_df = st.data_editor(
                st.session_state.s4v2_hist_df,
                num_rows="dynamic",
                use_container_width=True,
                key="s4v2_hist_ed",
                column_config={
                    "Confirmed": st.column_config.CheckboxColumn("✅", default=True, help="Only confirmed rows are applied"),
                    "POD": st.column_config.SelectboxColumn("POD", options=_s4h_pods_opts, default="", help="Required for new clients"),
                    "Client": st.column_config.SelectboxColumn("Client", options=_s4h_cli_opts, required=True),
                    "Required Role": st.column_config.SelectboxColumn("Required Role", options=_s4h_rol_opts, required=True),
                    "M1 (Hrs)": st.column_config.NumberColumn("M1 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M2 (Hrs)": st.column_config.NumberColumn("M2 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M3 (Hrs)": st.column_config.NumberColumn("M3 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M4 (Hrs)": st.column_config.NumberColumn("M4 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M5 (Hrs)": st.column_config.NumberColumn("M5 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M6 (Hrs)": st.column_config.NumberColumn("M6 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                }
            )

        # ── Reduce Hours (Scenario) ──────────────────────────────────────────
        with st.expander("➖ Reduce Hours (Scenario)", expanded=False):
            st.markdown(
                "**Manual hours to subtract from the scenario.** Targeting is hierarchical:\n\n"
                "- **POD only** → prorates across all clients & roles in that POD\n"
                "- **POD + Role** → prorates across all clients in that POD for that role\n"
                "- **POD + Client + Role** → applies to that exact combination\n"
                "- **Client + Role** → applies to that client/role (no POD filter)"
            )
            _s4r_pods_opts = [""] + (lista_pods or st.session_state.get('_lista_pods', []))
            _s4r_cli_opts  = [""] + lista_clientes

            # Template download
            _s4r_tmpl = pd.DataFrame([
                {"Confirmed": True, "POD": "POD A", "Client": "", "Required Role": "",
                 "M1 (Hrs)": 10, "M2 (Hrs)": 10, "M3 (Hrs)": 0, "M4 (Hrs)": 0, "M5 (Hrs)": 0, "M6 (Hrs)": 0},
                {"Confirmed": True, "POD": "", "Client": "Acme Corp", "Required Role": "Accountant I",
                 "M1 (Hrs)": 8, "M2 (Hrs)": 0, "M3 (Hrs)": 0, "M4 (Hrs)": 0, "M5 (Hrs)": 0, "M6 (Hrs)": 0},
            ])
            _s4r_buf = BytesIO()
            _s4r_tmpl.to_excel(_s4r_buf, index=False)
            _s4r_dl_col, _s4r_ul_col = st.columns([1, 1])
            with _s4r_dl_col:
                st.download_button("📄 Download Reduce Hours Template", _s4r_buf.getvalue(),
                                   file_name="S4_ReduceHours_Template.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key="dl_s4r_tmpl")
            with _s4r_ul_col:
                _s4r_upload = st.file_uploader("📂 Upload Reduce Hours file", type=["xlsx"],
                                               key="fu_s4r_tmpl", label_visibility="collapsed")
                st.caption("📂 Upload Reduce Hours file (.xlsx)")
                if _s4r_upload is not None:
                    try:
                        _s4r_up_df = pd.read_excel(_s4r_upload)
                        if 'Role' in _s4r_up_df.columns and 'Required Role' not in _s4r_up_df.columns:
                            _s4r_up_df = _s4r_up_df.rename(columns={'Role': 'Required Role'})
                        _s4r_expected = ['Confirmed', 'POD', 'Client', 'Required Role',
                                         'M1 (Hrs)', 'M2 (Hrs)', 'M3 (Hrs)', 'M4 (Hrs)', 'M5 (Hrs)', 'M6 (Hrs)']
                        for _s4rc in _s4r_expected:
                            if _s4rc not in _s4r_up_df.columns:
                                _s4r_up_df[_s4rc] = True if _s4rc == 'Confirmed' else (0.0 if '(Hrs)' in _s4rc else '')
                        _s4r_up_df = _s4r_up_df[_s4r_expected]
                        _s4r_up_df['Confirmed'] = True
                        for _s4rc in ['M1 (Hrs)', 'M2 (Hrs)', 'M3 (Hrs)', 'M4 (Hrs)', 'M5 (Hrs)', 'M6 (Hrs)']:
                            _s4r_up_df[_s4rc] = pd.to_numeric(_s4r_up_df[_s4rc], errors='coerce').fillna(0.0)
                        st.session_state.s4v2_red_df = _s4r_up_df
                        st.success(f"✅ Loaded {len(_s4r_up_df)} row(s) — all marked Confirmed.")
                    except Exception as _s4re:
                        st.error(f"❌ Could not load file: {_s4re}")

            if "Confirmed" not in st.session_state.s4v2_red_df.columns:
                st.session_state.s4v2_red_df.insert(0, "Confirmed", False)
            if "POD" not in st.session_state.s4v2_red_df.columns:
                st.session_state.s4v2_red_df.insert(1, "POD", "")
            st.caption("☑️ Check **Confirmed** on each row to include it in the scenario. Unconfirmed rows are ignored.")
            st.session_state.s4v2_red_df = st.data_editor(
                st.session_state.s4v2_red_df,
                num_rows="dynamic",
                use_container_width=True,
                key="s4v2_red_ed",
                column_config={
                    "Confirmed": st.column_config.CheckboxColumn("✅", default=False, help="Only confirmed rows are applied"),
                    "POD": st.column_config.SelectboxColumn("POD", options=_s4r_pods_opts, default="", help="Leave blank to target by Client only"),
                    "Client": st.column_config.SelectboxColumn("Client", options=_s4r_cli_opts, default="", help="Leave blank to target entire POD"),
                    "Required Role": st.column_config.SelectboxColumn("Role", options=[""] + roles_permitidos, default="", help="Leave blank to prorate across all roles"),
                    "M1 (Hrs)": st.column_config.NumberColumn("M1 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M2 (Hrs)": st.column_config.NumberColumn("M2 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M3 (Hrs)": st.column_config.NumberColumn("M3 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M4 (Hrs)": st.column_config.NumberColumn("M4 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M5 (Hrs)": st.column_config.NumberColumn("M5 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                    "M6 (Hrs)": st.column_config.NumberColumn("M6 (Hrs)", min_value=0.0, default=0.0, format="%.1f"),
                }
            )

        # ── Door Count Variation (Scenario) ─────────────────────────────────
        with st.expander("🚪 Door Count Variation (Scenario)", expanded=False):
            st.markdown(
                "**Door / property count variation per client.** "
                "Applies a percentage change to that client's capacity hours for the selected months — "
                "positive % = increase (more doors), negative % = decrease. "
                "Only ✅ Confirmed rows are applied in the scenario."
            )
            _s4dc_cli_opts  = [""] + lista_clientes
            _s4dc_pods_opts = [""] + (lista_pods or st.session_state.get('_lista_pods', []))

            # Template download
            _s4dc_tmpl = pd.DataFrame([
                {"Client": "Acme Corp", "POD": "POD A",
                 "M1 (%)": 5.0, "M2 (%)": 5.0, "M3 (%)": 0.0,
                 "M4 (%)": 0.0, "M5 (%)": 0.0, "M6 (%)": 0.0},
                {"Client": "Beta LLC",  "POD": "",
                 "M1 (%)": -10.0, "M2 (%)": -10.0, "M3 (%)": 0.0,
                 "M4 (%)": 0.0,  "M5 (%)": 0.0,    "M6 (%)": 0.0},
            ])
            _s4dc_buf = BytesIO()
            _s4dc_tmpl.to_excel(_s4dc_buf, index=False)
            _s4dc_dl_col, _s4dc_ul_col = st.columns([1, 1])
            with _s4dc_dl_col:
                st.download_button(
                    "📄 Download Door Count Template", _s4dc_buf.getvalue(),
                    file_name="S4_DoorCount_Template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_s4dc_tmpl",
                )
            with _s4dc_ul_col:
                _s4dc_upload = st.file_uploader("📂 Upload Door Count file", type=["xlsx"],
                                                key="fu_s4dc_tmpl", label_visibility="collapsed")
                st.caption("📂 Upload Door Count file (.xlsx)")
                if _s4dc_upload is not None:
                    try:
                        _s4dc_up_df = pd.read_excel(_s4dc_upload)
                        _s4dc_expected = ['Confirmed', 'Client', 'POD'] + meses_pct_cols
                        for _s4dcc in _s4dc_expected:
                            if _s4dcc not in _s4dc_up_df.columns:
                                _s4dc_up_df[_s4dcc] = True if _s4dcc == 'Confirmed' else (0.0 if '(%)' in _s4dcc else '')
                        _s4dc_up_df = _s4dc_up_df[_s4dc_expected]
                        _s4dc_up_df['Confirmed'] = True
                        for _s4dcc in meses_pct_cols:
                            _s4dc_up_df[_s4dcc] = pd.to_numeric(_s4dc_up_df[_s4dcc], errors='coerce').fillna(0.0)
                        st.session_state.s4v2_doorcount_df = _s4dc_up_df
                        st.success(f"✅ Loaded {len(_s4dc_up_df)} row(s) — all marked Confirmed.")
                    except Exception as _s4dce:
                        st.error(f"❌ Could not load file: {_s4dce}")

            if "Confirmed" not in st.session_state.s4v2_doorcount_df.columns:
                st.session_state.s4v2_doorcount_df.insert(0, "Confirmed", True)
            if "POD" not in st.session_state.s4v2_doorcount_df.columns:
                st.session_state.s4v2_doorcount_df.insert(2, "POD", "")

            st.caption("☑️ **Confirmed** rows are applied in the scenario. Positive % = hours increase, negative % = decrease.")
            st.session_state.s4v2_doorcount_df = st.data_editor(
                st.session_state.s4v2_doorcount_df,
                num_rows="dynamic",
                use_container_width=True,
                key="s4v2_dc_ed",
                column_config={
                    "Confirmed": st.column_config.CheckboxColumn("✅", default=True, help="Only confirmed rows are applied"),
                    "Client": st.column_config.SelectboxColumn("Client", options=_s4dc_cli_opts, required=True),
                    "POD": st.column_config.SelectboxColumn("POD", options=_s4dc_pods_opts, default=""),
                    "M1 (%)": st.column_config.NumberColumn("M1 (%)", format="%.1f%%", help="% change for month 1"),
                    "M2 (%)": st.column_config.NumberColumn("M2 (%)", format="%.1f%%"),
                    "M3 (%)": st.column_config.NumberColumn("M3 (%)", format="%.1f%%"),
                    "M4 (%)": st.column_config.NumberColumn("M4 (%)", format="%.1f%%"),
                    "M5 (%)": st.column_config.NumberColumn("M5 (%)", format="%.1f%%"),
                    "M6 (%)": st.column_config.NumberColumn("M6 (%)", format="%.1f%%"),
                },
            )

        # ── Confirmed-change detection ───────────────────────────────────────
        # Compare Confirmed state after all editors have rendered.
        # If only numeric values changed → scope="fragment" rerun (fast, display untouched).
        # If any Confirmed checkbox changed → full rerun so display fragment updates.
        _conf_after = _conf_snap()
        st.session_state['_s4_conf_snap'] = _conf_after
        if _conf_before is not None and _conf_after != _conf_before:
            st.rerun()   # full page → display fragment also recalculates

        # ── Downloads ────────────────────────────────────────────────────────
        with st.expander("💾 Downloads & Save / Load Scenario", expanded=False):
            _dl_c1, _dl_c2 = st.columns(2)

            # Scenario-only download (with detailed sheets)
            _dl_buf = BytesIO()
            _fd_s4  = st.session_state.get('final_dashboards', {})
            # Use the latest _df_scen produced by the display fragment
            _df_scen = st.session_state.get('_s4v2_df_scen', pd.DataFrame())
            with pd.ExcelWriter(_dl_buf, engine='openpyxl') as _xw:
                # Scenario waterfall summary
                _df_scen.to_excel(_xw, sheet_name='Scenario')
                # Capacity Overview — Waterfall (from Step 3)
                _wf_s4 = st.session_state.get('_wf_overall_export', pd.DataFrame())
                if not _wf_s4.empty:
                    _wf_s4.to_excel(_xw, sheet_name='Capacity_Overview')
                # Detailed level — Step 3 underlying data
                if not _fd_s4.get('cliente', pd.DataFrame()).empty:
                    _fd_s4['cliente'].to_excel(_xw, sheet_name='Client_Role_Detail', index=False)
                if not _fd_s4.get('baseline', pd.DataFrame()).empty:
                    _fd_s4['baseline'].to_excel(_xw, sheet_name='Baseline_Audit', index=False)
                _df_resbase_s4 = st.session_state.get('calc_data', {}).get('df_resumen_base', pd.DataFrame())
                if not _df_resbase_s4.empty:
                    _df_resbase_s4.to_excel(_xw, sheet_name='Base_Hours_by_Role', index=False)
                # Scenario inputs
                st.session_state.s4v2_hc_adj_df.to_excel(_xw, sheet_name='HC_Adjustments', index=False)
                st.session_state.s4v2_mrr_adj_df.to_excel(_xw, sheet_name='MRR_Adjustments', index=False)
                st.session_state.s4v2_hrs_role_df.to_excel(_xw, sheet_name='Hours_Adjustments', index=False)
                st.session_state.s4v2_auto_df.to_excel(_xw, sheet_name='Auto_Savings', index=False)
                st.session_state.s4v2_hist_df.to_excel(_xw, sheet_name='Add_Hours', index=False)
                st.session_state.s4v2_red_df.to_excel(_xw, sheet_name='Reduce_Hours', index=False)
                st.session_state.s4v2_doorcount_df.to_excel(_xw, sheet_name='DoorCount_Variation', index=False)
            _dl_c1.download_button(
                "📥 Download Scenario",
                _dl_buf.getvalue(),
                file_name=f"Step4_Scenario_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="s4v2_dl",
                use_container_width=True,
            )

            # Download ALL — Steps 2 + 3 + 4 in one file (all tabs, full detail)
            _dl_all_buf = BytesIO()
            with pd.ExcelWriter(_dl_all_buf, engine='openpyxl') as _xw_all:
                _fd_all = st.session_state.get('final_dashboards', {})
                # ── Step 3: all tabs ──────────────────────────────────────────
                # Tab: Capacity Overview — Waterfall (transposed view)
                _wf_all = st.session_state.get('_wf_overall_export', pd.DataFrame())
                if not _wf_all.empty:
                    _wf_all.to_excel(_xw_all, sheet_name='3_Capacity_Overview')
                # Tab: General Waterfall Summary (raw month rows)
                if not _fd_all.get('general', pd.DataFrame()).empty:
                    _fd_all['general'].to_excel(_xw_all, sheet_name='3_General_Summary', index=False)
                # Tab: Summary by POD
                if not _fd_all.get('pod', pd.DataFrame()).empty:
                    _fd_all['pod'].to_excel(_xw_all, sheet_name='3_Summary_by_POD', index=False)
                # Tab: POD x Sr. Accountant — aggregated by POD + Sr. (no Client column)
                if not _fd_all.get('cliente', pd.DataFrame()).empty:
                    _pod_sr_all = _fd_all['cliente'].copy()
                    _psr_grp_all = [c for c in ['POD', 'Sr. Accountant', 'Required Role'] if c in _pod_sr_all.columns]
                    _psr_num_all = [c for c in _pod_sr_all.columns if c not in _psr_grp_all and c != 'Client']
                    if _psr_grp_all:
                        _pod_sr_all = _pod_sr_all.groupby(_psr_grp_all, as_index=False)[_psr_num_all].sum()
                        _pod_sr_all = _pod_sr_all.sort_values(_psr_grp_all)
                    _pod_sr_all.to_excel(_xw_all, sheet_name='3_POD_x_SrAccountant', index=False)
                # Tab: Client & Role Summary — full detail
                if not _fd_all.get('cliente', pd.DataFrame()).empty:
                    _fd_all['cliente'].to_excel(_xw_all, sheet_name='3_Client_Role_Detail', index=False)
                # Tab: Baseline Audit — most detailed level (drop internal email cols)
                if not _fd_all.get('baseline', pd.DataFrame()).empty:
                    _bl_all = _fd_all['baseline'].copy()
                    _bl_all = _bl_all[[c for c in _bl_all.columns if c not in {'Processor Email', 'Reviewer Email'}]]
                    _bl_all.to_excel(_xw_all, sheet_name='3_Baseline_Audit', index=False)
                # Tab: Employee Level
                _el_df_all = st.session_state.get('_s3_emp_level_df', pd.DataFrame())
                if not _el_df_all.empty:
                    _el_df_all.to_excel(_xw_all, sheet_name='3_Employee_Level', index=False)
                # ── Step 2 inputs ─────────────────────────────────────────────
                if not st.session_state.automations_df.empty:
                    st.session_state.automations_df.to_excel(_xw_all, sheet_name='2_Automations', index=False)
                if not st.session_state.historical_df.empty:
                    st.session_state.historical_df.to_excel(_xw_all, sheet_name='2_Add_Hours', index=False)
                if not st.session_state.reductions_df.empty:
                    st.session_state.reductions_df.to_excel(_xw_all, sheet_name='2_Reductions', index=False)
                _dc_exp = st.session_state.get('doorcount_df', pd.DataFrame())
                if not _dc_exp.empty:
                    _dc_exp.to_excel(_xw_all, sheet_name='2_DoorCount_Variation', index=False)
                # ── Step 4 scenario ───────────────────────────────────────────
                _df_scen.to_excel(_xw_all, sheet_name='4_Scenario')
                st.session_state.s4v2_params_df.to_excel(_xw_all, sheet_name='4_Global_Params', index=False)
                # Detailed level for scenario
                if not _fd_all.get('cliente', pd.DataFrame()).empty:
                    _fd_all['cliente'].to_excel(_xw_all, sheet_name='4_Client_Role_Detail', index=False)
                if not _fd_all.get('baseline', pd.DataFrame()).empty:
                    _fd_all['baseline'].to_excel(_xw_all, sheet_name='4_Baseline_Audit', index=False)
                _df_resbase_all = st.session_state.get('calc_data', {}).get('df_resumen_base', pd.DataFrame())
                if not _df_resbase_all.empty:
                    _df_resbase_all.to_excel(_xw_all, sheet_name='4_Base_Hours_by_Role', index=False)
                # Scenario inputs
                st.session_state.s4v2_hc_adj_df.to_excel(_xw_all, sheet_name='4_HC_Adjustments', index=False)
                st.session_state.s4v2_mrr_adj_df.to_excel(_xw_all, sheet_name='4_MRR_Adjustments', index=False)
                st.session_state.s4v2_hrs_role_df.to_excel(_xw_all, sheet_name='4_Hours_Adjustments', index=False)
                st.session_state.s4v2_auto_df.to_excel(_xw_all, sheet_name='4_Auto_Savings', index=False)
                st.session_state.s4v2_hist_df.to_excel(_xw_all, sheet_name='4_Add_Hours', index=False)
                st.session_state.s4v2_red_df.to_excel(_xw_all, sheet_name='4_Reduce_Hours', index=False)
                st.session_state.s4v2_doorcount_df.to_excel(_xw_all, sheet_name='4_DoorCount_Variation', index=False)
            _dl_c2.download_button(
                "📥 Download All (Steps 2 + 3 + 4)",
                _dl_all_buf.getvalue(),
                file_name=f"CapacityOnline_Full_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="s4v2_dl_all",
                use_container_width=True,
                type="primary",
            )

            # ── Save & Load Scenario ─────────────────────────────────────────────
            st.divider()
            st.markdown("#### 💾 Save / Load Scenario")
            st.caption(
                "Save the full Step 4 scenario (results + all inputs) to reuse as a starting point later. "
                "Loading a saved scenario restores all inputs so you can continue from where you left off."
            )

            # Save
            _sv_c1, _sv_c2, _sv_c3 = st.columns([2, 3, 1])
            _sv_name = _sv_c1.text_input("Scenario name", value=f"Plan {datetime.now().strftime('%b %Y')}", key="s4v2_sv_name")
            _sv_desc = _sv_c2.text_input("Description (optional)", key="s4v2_sv_desc")
            if _sv_c3.button("💾 Save", key="s4v2_sv_btn", use_container_width=True):
                _sv_payload = {
                    'df_scen':       _df_scen,
                    'params_df':     st.session_state.s4v2_params_df.copy(),
                    'hc_adj_df':     st.session_state.s4v2_hc_adj_df.copy(),
                    'mrr_adj_df':    st.session_state.s4v2_mrr_adj_df.copy(),
                    'hrs_role_df':   st.session_state.s4v2_hrs_role_df.copy(),
                    'auto_df':       st.session_state.s4v2_auto_df.copy(),
                    'hist_df':       st.session_state.s4v2_hist_df.copy(),
                    'red_df':        st.session_state.s4v2_red_df.copy(),
                    'doorcount_df':  st.session_state.s4v2_doorcount_df.copy(),
                    'scope':         st.session_state.get('s4v2_scope', 'Overall'),
                    # Raw pipeline data so the scenario can be reloaded as a base
                    'vol_merged':    st.session_state.get('df_clean', pd.DataFrame()),
                    'new_clients':   st.session_state.get('pipeline_new_clients', pd.DataFrame()),
                    # Step 3 results snapshot
                    'final_dashboards': {k: v.copy() if isinstance(v, pd.DataFrame) else v
                                         for k, v in st.session_state.get('final_dashboards', {}).items()},
                }
                _sv_path = _save_step4_scenario(_sv_name, _sv_desc, _sv_payload)
                st.success(f"Saved: `{os.path.basename(_sv_path)}`")
                st.rerun(scope="fragment")

            # Load
            _s4_saved = _list_step4_scenarios()
            if _s4_saved:
                st.markdown("**Load a saved scenario:**")
                _ld_opts = {f"{s['name']}  ({s['saved_at'][:10]})  — {s['description']}": s for s in _s4_saved}
                _ld_col1, _ld_col2 = st.columns([4, 1])
                _ld_sel  = _ld_col1.selectbox("Select scenario", options=list(_ld_opts.keys()), key="s4v2_ld_sel", label_visibility="collapsed")
                if _ld_col2.button("📂 Load", key="s4v2_ld_btn", use_container_width=True):
                    _ld = _load_scenario(_ld_opts[_ld_sel]['path'])
                    if _ld.get('params_df')    is not None: st.session_state.s4v2_params_df   = _ld['params_df']
                    if _ld.get('hc_adj_df')    is not None: st.session_state.s4v2_hc_adj_df   = _ld['hc_adj_df']
                    if _ld.get('mrr_adj_df')   is not None: st.session_state.s4v2_mrr_adj_df  = _ld['mrr_adj_df']
                    if _ld.get('hrs_role_df')  is not None: st.session_state.s4v2_hrs_role_df = _ld['hrs_role_df']
                    if _ld.get('auto_df')      is not None: st.session_state.s4v2_auto_df      = _ld['auto_df']
                    if _ld.get('hist_df')      is not None: st.session_state.s4v2_hist_df      = _ld['hist_df']
                    if _ld.get('red_df')       is not None: st.session_state.s4v2_red_df       = _ld['red_df']
                    if _ld.get('doorcount_df') is not None: st.session_state.s4v2_doorcount_df = _ld['doorcount_df']
                    st.success(f"Loaded: **{_ld.get('name', '')}**")
                    st.rerun()   # full rerun so display fragment also refreshes

    # ── end of _s4v2_inputs_frag ────────────────────────────────────────────────

    # ── Step 4 progress indicator ────────────────────────────────────────────────
    # Full-screen modal overlay that blocks interaction while Step 4 builds.
    # Injected into the parent document via components.html; auto-dismisses when done.
    if not st.session_state.get('_s4_ready', False):
        import time as _time_s4
        import streamlit.components.v1 as _stc_s4prog
        _stc_s4prog.html("""
<script>
(function(){
    var doc = window.parent.document;
    if (doc.getElementById('cap-s4-overlay')) return;

    // ── Full-screen blocking overlay ──────────────────────────────────────
    var ov = doc.createElement('div');
    ov.id = 'cap-s4-overlay';
    ov.style.cssText = [
        'position:fixed','top:0','left:0','width:100%','height:100%',
        'background:rgba(10,12,20,0.82)',
        'z-index:2147483647',
        'display:flex','align-items:center','justify-content:center',
        'backdrop-filter:blur(3px)',
        '-webkit-backdrop-filter:blur(3px)',
    ].join(';');

    ov.innerHTML = [
        '<div style="background:#1a1e2e;border-radius:18px;padding:48px 64px;',
        'text-align:center;box-shadow:0 16px 48px rgba(0,0,0,0.7);min-width:340px;max-width:90vw;">',
        '<div style="font-size:2.6rem;margin-bottom:14px;">⚙️</div>',
        '<div style="color:#e8ecff;font-size:1.15rem;font-weight:700;margin-bottom:6px;">',
        'Loading Step 4 Scenario Planner</div>',
        '<div style="color:#7a85aa;font-size:0.85rem;margin-bottom:24px;">',
        'Please wait while the scenario is being prepared…</div>',
        '<div style="background:#252a3d;border-radius:100px;height:10px;overflow:hidden;margin-bottom:10px;">',
        '<div id="cap-s4-bar" style="background:linear-gradient(90deg,#ff4b4b,#ff7b5e);',
        'height:100%;width:0%;border-radius:100px;transition:width 80ms linear;"></div>',
        '</div>',
        '<div id="cap-s4-pct" style="color:#9ea8c7;font-size:0.82rem;font-variant-numeric:tabular-nums;">0%</div>',
        '</div>',
    ].join('');
    doc.body.appendChild(ov);

    // ── Animate: 0 → 100% over ~3 s, then fade out ───────────────────────
    var pct = 0;
    var bar = doc.getElementById('cap-s4-bar');
    var lbl = doc.getElementById('cap-s4-pct');
    var iv  = setInterval(function(){
        pct = Math.min(pct + 1.4, 100);
        bar.style.width = pct + '%';
        lbl.textContent = Math.round(pct) + '%';
        if (pct >= 100){
            clearInterval(iv);
            setTimeout(function(){
                ov.style.transition = 'opacity 0.5s ease';
                ov.style.opacity = '0';
                setTimeout(function(){ if (ov.parentNode) ov.remove(); }, 520);
            }, 400);
        }
    }, 42);   // 42 ms × ~72 steps ≈ 3 s total
})();
</script>
""", height=0)
        st.session_state['_s4_ready'] = True

    _s4v2_fragment()          # display: scope, params, computation, table, comparison
    _s4v2_inputs_frag()       # inputs:  editors + downloads (isolated, fast reruns)

    if "df_clean" in st.session_state and "calc_data" in st.session_state:
        with st.expander("🧠 New Clients AI Prediction", expanded=False):
            _make_ai_prediction_fragment(pfx="s4", add_to_scenario=True)()

# ── Close the Data Load & Filters tab context that was re-entered above
# so the remaining sibling tabs render their own isolated content.
tab1.__exit__(None, None, None)

# ==========================================
with tab_predict:
    st.header("🧠 New Client Prediction")
    if "df_clean" not in st.session_state or "calc_data" not in st.session_state:
        st.info("⬅️ Upload both the capacity data file and the HC Weekly Report in **Data Load & Filters** first.")
    else:
        _make_ai_prediction_fragment(pfx="qt", add_to_scenario=False)()

# ==========================================
# TAB: NEW CLIENT PER VOLUME & AHT
# ==========================================
with tab_vol_aht:
    st.header("📊 New Client per Volume & AHT")

    if "df_clean" not in st.session_state:
        st.info("⬅️ Upload the volume data file in **Data Load & Filters** first.")
    else:
        @st.fragment
        def _vol_aht_fragment():
            # ── Example template download ────────────────────────────────────
            with st.expander("📄 Template — Download / Upload", expanded=False):
                _ex_tasks = _build_vol_aht_task_df(pms_filter=None)
                _tmpl_buf = BytesIO()
                if not _ex_tasks.empty:
                    _tmpl_df = _ex_tasks[['Type', 'Subtype', 'Proc AHT (min)', 'Rev AHT (min)', 'Volume', 'QC %']].copy()
                    _tmpl_df.insert(0, 'Client Name', 'Example Client')
                    _tmpl_df.insert(1, 'PMS', '')
                    _tmpl_df.insert(2, 'MRR', 3000)
                    with pd.ExcelWriter(_tmpl_buf, engine='openpyxl') as _tw:
                        _tmpl_df.to_excel(_tw, sheet_name='Volume Input', index=False)
                    st.download_button(
                        "📥 Download Example Template",
                        _tmpl_buf.getvalue(),
                        file_name="NewClient_VolumeTemplate.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="va_tmpl_dl"
                    )

                _up_file = st.file_uploader("Upload filled template", type=["xlsx", "xls"], key="va_upload")
                if _up_file is not None:
                    try:
                        _up_df = pd.read_excel(_up_file, sheet_name='Volume Input')
                        _up_df.columns = _up_df.columns.str.strip()
                        # Extract client meta from first row
                        if not _up_df.empty:
                            st.session_state['va_client_name'] = str(_up_df.get('Client Name', pd.Series([''])).iloc[0])
                            st.session_state['va_mrr']         = float(pd.to_numeric(_up_df.get('MRR', pd.Series([0])).iloc[0], errors='coerce') or 0)
                            _pms_up = str(_up_df.get('PMS', pd.Series([''])).iloc[0]).strip()
                            if _pms_up and _pms_up.lower() not in ('nan', 'none', ''):
                                st.session_state['va_pms'] = _pms_up
                            # Build task df from upload
                            _task_cols = ['Type', 'Subtype', 'Proc AHT (min)', 'Rev AHT (min)', 'Volume', 'QC %']
                            _up_tasks  = _up_df[[c for c in _task_cols if c in _up_df.columns]].copy()
                            _base_df   = _build_vol_aht_task_df(st.session_state.get('va_pms'))
                            if not _base_df.empty and not _up_tasks.empty:
                                # Merge uploaded volumes/AHTs onto base task list
                                _merged = _base_df.merge(
                                    _up_tasks[['Type', 'Subtype'] + [c for c in ['Volume', 'QC %', 'Proc AHT (min)', 'Rev AHT (min)'] if c in _up_tasks.columns]],
                                    on=['Type', 'Subtype'], how='left', suffixes=('', '_up')
                                )
                                for _uc in ['Volume', 'QC %', 'Proc AHT (min)', 'Rev AHT (min)']:
                                    _uc_up = _uc + '_up'
                                    if _uc_up in _merged.columns:
                                        _merged[_uc] = _merged[_uc_up].fillna(_merged[_uc])
                                        _merged.drop(columns=[_uc_up], inplace=True)
                                st.session_state['va_task_df'] = _merged
                            st.success("✅ Template loaded successfully.")
                            st.rerun(scope="fragment")
                    except Exception as _upe:
                        st.error(f"Upload error: {_upe}")

            # ── General client info ──────────────────────────────────────────
            st.subheader("📋 Client Details")
            _c1, _c2, _c3, _c4 = st.columns([2, 1, 1, 1])

            _va_client = _c1.text_input(
                "Client Name", value=st.session_state.get('va_client_name', ''), key="va_client_name_inp"
            )
            _va_mrr = _c2.number_input(
                "MRR ($)", min_value=0.0, value=float(st.session_state.get('va_mrr', 0.0)),
                format="%.2f", key="va_mrr_inp"
            )
            _pms_list_va = _get_pms_opts(include_unknown=False)
            _va_pms_default = st.session_state.get('va_pms', _pms_list_va[0] if _pms_list_va else '')
            _va_pms_idx     = _pms_list_va.index(_va_pms_default) if _va_pms_default in _pms_list_va else 0
            _va_pms = _c3.selectbox("PMS", options=_pms_list_va, index=_va_pms_idx, key="va_pms_sel")

            # Month selector (for FTE conversion)
            _va_month_idx = _c4.selectbox(
                "Month", options=list(range(6)),
                format_func=lambda i: meses_proyeccion[i],
                key="va_month_sel"
            )

            # Rebuild task df when PMS changes
            _cur_pms = st.session_state.get('_va_last_pms')
            if _cur_pms != _va_pms or 'va_task_df' not in st.session_state:
                st.session_state['va_task_df']    = _build_vol_aht_task_df(_va_pms)
                st.session_state['_va_last_pms']  = _va_pms

            st.session_state['va_client_name'] = _va_client
            st.session_state['va_mrr']         = _va_mrr
            st.session_state['va_pms']         = _va_pms

            # ── Volume & AHT editor ──────────────────────────────────────────
            st.subheader("📝 Volume & AHT per Task")
            st.caption(
                "Enter volumes per task. **Proc AHT**, **Rev AHT**, and **QC %** are editable — "
                "click any cell to modify. Changing the PMS above will refresh AHTs to that PMS's averages."
            )

            _va_df_edit = st.data_editor(
                st.session_state['va_task_df'],
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                disabled=["Type", "Subtype", "Proc Role", "Rev Role"],
                column_config={
                    "Type":           st.column_config.TextColumn("Type"),
                    "Subtype":        st.column_config.TextColumn("Subtype"),
                    "Proc Role":      st.column_config.TextColumn("Proc Role"),
                    "Rev Role":       st.column_config.TextColumn("Rev Role"),
                    "Proc AHT (min)": st.column_config.NumberColumn("Proc AHT (min)", min_value=0.0, format="%.2f",
                                        help="Processing AHT in minutes — editable"),
                    "Rev AHT (min)":  st.column_config.NumberColumn("Rev AHT (min)",  min_value=0.0, format="%.2f",
                                        help="Review/QA AHT in minutes — editable"),
                    "Volume":         st.column_config.NumberColumn("Volume", min_value=0, format="%d"),
                    "QC %":           st.column_config.NumberColumn("QC %", min_value=0.0, max_value=100.0, format="%.1f%%"),
                },
                key="va_task_editor"
            )
            # Persist edits
            st.session_state['va_task_df'] = _va_df_edit

            # ── Live calculation ─────────────────────────────────────────────
            _active_rows = _va_df_edit[pd.to_numeric(_va_df_edit['Volume'], errors='coerce').fillna(0) > 0].copy()

            if _active_rows.empty:
                st.info("Enter volumes above to see the staffing summary.")
                return

            st.divider()
            st.subheader("📊 Staffing & Cost Summary")

            _hrs_fte = (
                st.session_state.get('calc_data', {})
                .get('dict_hrs_per_fte', {})
                .get(_va_month_idx, 157.5)
            )

            # ── Compute productive & full hours per row ───────────────────────
            _res_rows = []
            for _, row in _active_rows.iterrows():
                _vol    = float(pd.to_numeric(row['Volume'],        errors='coerce') or 0)
                _qc_pct = float(pd.to_numeric(row['QC %'],          errors='coerce') or 100) / 100.0
                _p_aht  = float(pd.to_numeric(row['Proc AHT (min)'],errors='coerce') or 0)
                _r_aht  = float(pd.to_numeric(row['Rev AHT (min)'], errors='coerce') or 0)
                _p_role = str(row.get('Proc Role', 'Accountant I'))
                _r_role = str(row.get('Rev Role',  'Sr. Accountant'))

                _prod_p = (_vol * _p_aht) / 60.0
                _prod_r = (_vol * _r_aht * _qc_pct) / 60.0

                _util_p = utilization_map.get(_p_role, 0.85)
                _util_r = utilization_map.get(_r_role, 0.50)

                _full_p = _prod_p * (1 + (1 - _util_p) + absenteeism + attrition)
                _full_r = _prod_r * (1 + (1 - _util_r) + absenteeism + attrition)

                _cost_p_hr = cost_map.get(_p_role, cost_acc1) / _hrs_fte if _hrs_fte > 0 else 0
                _cost_r_hr = cost_map.get(_r_role, cost_sr)   / _hrs_fte if _hrs_fte > 0 else 0

                _res_rows.append({
                    'Type':        row['Type'],
                    'Subtype':     row['Subtype'],
                    'Proc Role':   _p_role,
                    'Rev Role':    _r_role,
                    'Volume':      _vol,
                    'QC %':        _qc_pct * 100,
                    'Prod Proc Hrs':  round(_prod_p, 2),
                    'Prod Rev Hrs':   round(_prod_r, 2),
                    'Full Proc Hrs':  round(_full_p, 2),
                    'Full Rev Hrs':   round(_full_r, 2),
                    'Prod Proc Cost': round(_prod_p * _cost_p_hr, 2),
                    'Prod Rev Cost':  round(_prod_r * _cost_r_hr, 2),
                    'Full Proc Cost': round(_full_p * _cost_p_hr, 2),
                    'Full Rev Cost':  round(_full_r * _cost_r_hr, 2),
                })
            _res_df = pd.DataFrame(_res_rows)

            # ── Per-Role Summary ─────────────────────────────────────────────
            # Aggregate by role (proc + rev sides)
            _role_summary = {}
            for _, row in _res_df.iterrows():
                for side, role_key, prod_hrs_key, full_hrs_key, prod_cost_key, full_cost_key in [
                    ('proc', 'Proc Role', 'Prod Proc Hrs', 'Full Proc Hrs', 'Prod Proc Cost', 'Full Proc Cost'),
                    ('rev',  'Rev Role',  'Prod Rev Hrs',  'Full Rev Hrs',  'Prod Rev Cost',  'Full Rev Cost'),
                ]:
                    _role = row[role_key]
                    if _role not in _role_summary:
                        _role_summary[_role] = {'Productive Hrs': 0, 'Full Hrs': 0, 'Productive Cost': 0, 'Full Cost': 0}
                    _role_summary[_role]['Productive Hrs']  += row[prod_hrs_key]
                    _role_summary[_role]['Full Hrs']         += row[full_hrs_key]
                    _role_summary[_role]['Productive Cost']  += row[prod_cost_key]
                    _role_summary[_role]['Full Cost']         += row[full_cost_key]

            _summary_rows = []
            for _role in roles_permitidos:
                if _role in _role_summary:
                    d = _role_summary[_role]
                    _prod_fte = d['Productive Hrs'] / _hrs_fte if _hrs_fte > 0 else 0
                    _full_fte = d['Full Hrs']        / _hrs_fte if _hrs_fte > 0 else 0
                    _summary_rows.append({
                        'Role':              _role,
                        'Productive Hrs':    round(d['Productive Hrs'], 2),
                        'Productive FTE':    round(_prod_fte, 3),
                        'Required FTE':      round(_full_fte, 3),
                        'Productive Cost':   round(d['Productive Cost'], 2),
                        'Required Cost':     round(d['Full Cost'], 2),
                    })

            _sum_df = pd.DataFrame(_summary_rows) if _summary_rows else pd.DataFrame()

            # Totals
            _total_prod_hrs  = _res_df['Prod Proc Hrs'].sum() + _res_df['Prod Rev Hrs'].sum()
            _total_full_hrs  = _res_df['Full Proc Hrs'].sum() + _res_df['Full Rev Hrs'].sum()
            _total_prod_cost = _res_df['Prod Proc Cost'].sum() + _res_df['Prod Rev Cost'].sum()
            _total_full_cost = _res_df['Full Proc Cost'].sum() + _res_df['Full Rev Cost'].sum()
            _total_prod_fte  = _total_prod_hrs / _hrs_fte if _hrs_fte > 0 else 0
            _total_req_fte   = _total_full_hrs / _hrs_fte if _hrs_fte > 0 else 0
            _mrr_val         = float(_va_mrr or 0)
            _prod_margin     = _mrr_val - _total_prod_cost
            _req_margin      = _mrr_val - _total_full_cost
            _prod_margin_pct = (_prod_margin / _mrr_val * 100) if _mrr_val > 0 else None
            _req_margin_pct  = (_req_margin  / _mrr_val * 100) if _mrr_val > 0 else None

            # ── Display ──────────────────────────────────────────────────────
            # KPI strip
            _k1, _k2, _k3, _k4, _k5, _k6 = st.columns(6)
            _k1.metric("Prod. Hours",    f"{_total_prod_hrs:,.1f}")
            _k2.metric("Prod. FTE",      f"{_total_prod_fte:.2f}")
            _k3.metric("Required FTE",   f"{_total_req_fte:.2f}")
            _k4.metric("Prod. Cost",     f"${_total_prod_cost:,.0f}")
            _k5.metric("Required Cost",  f"${_total_full_cost:,.0f}")
            if _mrr_val > 0:
                _k6.metric(
                    "Expected Margin",
                    f"${_req_margin:,.0f}",
                    delta=f"{_req_margin_pct:.1f}%" if _req_margin_pct is not None else None
                )

            st.divider()

            # Per-role table
            if not _sum_df.empty:
                st.markdown("#### Per-Role Breakdown")
                st.dataframe(
                    _sum_df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Productive Cost":  st.column_config.NumberColumn("Productive Cost",  format="$%.2f"),
                        "Required Cost":    st.column_config.NumberColumn("Required Cost",    format="$%.2f"),
                        "Productive FTE":   st.column_config.NumberColumn("Productive FTE",   format="%.3f"),
                        "Required FTE":     st.column_config.NumberColumn("Required FTE",     format="%.3f"),
                        "Productive Hrs":   st.column_config.NumberColumn("Productive Hrs",   format="%.2f"),
                    }
                )

            # Margin summary
            if _mrr_val > 0:
                st.markdown("#### Margin Summary")
                _mg1, _mg2, _mg3 = st.columns(3)
                _mg1.metric("MRR",                  f"${_mrr_val:,.2f}")
                _mg2.metric("Productive Margin",
                             f"${_prod_margin:,.2f}",
                             delta=f"{_prod_margin_pct:.1f}%" if _prod_margin_pct is not None else None)
                _mg3.metric("Expected Margin (req.)",
                             f"${_req_margin:,.2f}",
                             delta=f"{_req_margin_pct:.1f}%" if _req_margin_pct is not None else None)

            # Task detail expander
            with st.expander("🔍 Task Detail", expanded=False):
                st.dataframe(
                    _res_df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        c: st.column_config.NumberColumn(c, format="$%.2f")
                        for c in ['Prod Proc Cost', 'Prod Rev Cost', 'Full Proc Cost', 'Full Rev Cost']
                    }
                )

            # Download
            _va_dl_buf = BytesIO()
            with pd.ExcelWriter(_va_dl_buf, engine='openpyxl') as _va_xw:
                _sum_df.to_excel(_va_xw, sheet_name='Role Summary', index=False)
                _res_df.to_excel(_va_xw, sheet_name='Task Detail', index=False)
                pd.DataFrame([{
                    'Client': _va_client, 'PMS': _va_pms, 'MRR': _mrr_val,
                    'Month': meses_proyeccion[_va_month_idx],
                    'Productive Hours': round(_total_prod_hrs, 2),
                    'Productive FTE': round(_total_prod_fte, 3),
                    'Required FTE': round(_total_req_fte, 3),
                    'Productive Cost': round(_total_prod_cost, 2),
                    'Required Cost': round(_total_full_cost, 2),
                    'Productive Margin': round(_prod_margin, 2),
                    'Expected Margin': round(_req_margin, 2),
                    'Prod Margin %': round(_prod_margin_pct, 2) if _prod_margin_pct else None,
                    'Req Margin %': round(_req_margin_pct, 2) if _req_margin_pct else None,
                }]).to_excel(_va_xw, sheet_name='Client Summary', index=False)
            st.download_button(
                "📥 Download Summary",
                _va_dl_buf.getvalue(),
                file_name=f"VolumeAHT_{_va_client.replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="va_dl_btn"
            )

        _vol_aht_fragment()

# ==========================================
# TAB: RECON
# ==========================================
with tab_recon:
    st.header("🔄 Recon")
    st.caption("Upload two versions of a Final Dashboards export or a Volume Input export to reconcile differences.")

    @st.fragment
    def _recon_fragment():
        # ── Mode selector ────────────────────────────────────────────────────
        _recon_mode = st.radio(
            "Reconciliation source",
            ["📊 Final Dashboards (V1 vs V2)", "📋 Volume Input (V1 vs V2)"],
            horizontal=True,
            key="recon_mode_radio",
        )

        _rc1, _rc2 = st.columns(2)
        with _rc1:
            _rv1 = st.file_uploader("📂 Upload V1 (baseline)", type=["xlsx"], key="recon_v1")
        with _rc2:
            _rv2 = st.file_uploader("📂 Upload V2 (updated)", type=["xlsx"], key="recon_v2")

        if st.button("▶ Run Reconciliation", key="recon_run_btn", type="primary",
                     disabled=(_rv1 is None or _rv2 is None)):
            try:
                # ── Helper: outer-join + delta (interleaved column layout) ──
                def _merge_delta(df1, df2, key_cols, val_cols):
                    _d1 = df1[[c for c in key_cols + val_cols if c in df1.columns]].copy()
                    _d2 = df2[[c for c in key_cols + val_cols if c in df2.columns]].copy()
                    for _vc in val_cols:
                        if _vc not in _d1.columns:
                            _d1[_vc] = 0.0
                        if _vc not in _d2.columns:
                            _d2[_vc] = 0.0
                        _d1[_vc] = pd.to_numeric(_d1[_vc], errors='coerce').fillna(0.0)
                        _d2[_vc] = pd.to_numeric(_d2[_vc], errors='coerce').fillna(0.0)
                    _merged = pd.merge(_d1, _d2, on=key_cols, how='outer', suffixes=(' [V1]', ' [V2]'))
                    for _vc in val_cols:
                        _c1 = f'{_vc} [V1]'
                        _c2 = f'{_vc} [V2]'
                        if _c1 not in _merged.columns:
                            _merged[_c1] = 0.0
                        if _c2 not in _merged.columns:
                            _merged[_c2] = 0.0
                        _merged[_c1] = _merged[_c1].fillna(0.0)
                        _merged[_c2] = _merged[_c2].fillna(0.0)
                        _merged[f'Δ {_vc}'] = _merged[_c2] - _merged[_c1]

                    # Status flag
                    def _status(row):
                        _in1 = any(row.get(f'{vc} [V1]', 0) != 0 for vc in val_cols)
                        _in2 = any(row.get(f'{vc} [V2]', 0) != 0 for vc in val_cols)
                        if _in1 and not _in2:
                            return 'Removed'
                        if not _in1 and _in2:
                            return 'New'
                        _changed = any(abs(row.get(f'Δ {vc}', 0)) > 0.001 for vc in val_cols)
                        return 'Changed' if _changed else 'Same'

                    _merged['Status'] = _merged.apply(_status, axis=1)

                    # Interleaved column order: keys | metric [V1] | metric [V2] | Δ metric | ... | Status
                    _col_order = list(key_cols)
                    for _vc in val_cols:
                        _col_order += [f'{_vc} [V1]', f'{_vc} [V2]', f'Δ {_vc}']
                    _col_order += ['Status']
                    _col_order = [c for c in _col_order if c in _merged.columns]
                    return _merged[_col_order]

                # ── Load files ───────────────────────────────────────────────
                _is_fd = _recon_mode.startswith("📊")

                if _is_fd:
                    # ── FINAL DASHBOARDS mode ─────────────────────────────────
                    try:
                        _fd1 = pd.read_excel(_rv1, sheet_name='Client_Role_Detail')
                    except Exception:
                        _fd1 = pd.read_excel(_rv1, sheet_name=0)
                    try:
                        _fd2 = pd.read_excel(_rv2, sheet_name='Client_Role_Detail')
                    except Exception:
                        _fd2 = pd.read_excel(_rv2, sheet_name=0)

                    _fd1.columns = _fd1.columns.str.strip()
                    _fd2.columns = _fd2.columns.str.strip()

                    # Read MRR from Baseline_Audit sheet — Client_Role_Detail has no MRR column.
                    # MRR is stored per-row in Baseline_Audit as 'MRR ($)'; deduplicate by Client.
                    def _extract_mrr(file_obj, fd_df):
                        try:
                            _ba = pd.read_excel(file_obj, sheet_name='Baseline_Audit')
                            _ba.columns = _ba.columns.str.strip()
                            # Find MRR column (could be 'MRR ($)' or 'MRR')
                            _mcol = next((c for c in _ba.columns if c.strip().upper().startswith('MRR')), None)
                            if _mcol and 'Client' in _ba.columns and 'Client' in fd_df.columns:
                                _mrr_lkp = (
                                    _ba.drop_duplicates('Client')[['Client', _mcol]]
                                    .rename(columns={_mcol: 'MRR'})
                                )
                                return pd.merge(fd_df, _mrr_lkp, on='Client', how='left')
                        except Exception:
                            pass
                        return fd_df

                    _fd1 = _extract_mrr(_rv1, _fd1)
                    _fd2 = _extract_mrr(_rv2, _fd2)

                    # Non-metric / key columns
                    _fd_non_metric = {'POD', 'Sr. Accountant', 'Client', 'Required Role',
                                      'Type', 'Sub-Type', 'Subtype', 'Sub Type', 'MRR'}

                    # Month columns = numeric cols that are not metadata
                    _month_cols_1 = [c for c in _fd1.columns if c not in _fd_non_metric
                                     and pd.api.types.is_numeric_dtype(_fd1[c])]
                    _month_cols_2 = [c for c in _fd2.columns if c not in _fd_non_metric
                                     and pd.api.types.is_numeric_dtype(_fd2[c])]
                    _all_month_cols = list(dict.fromkeys(_month_cols_1 + _month_cols_2))

                    # MRR now joined from Baseline_Audit
                    _has_mrr_fd = 'MRR' in _fd1.columns or 'MRR' in _fd2.columns
                    # Val cols: MRR first (scalar per client), then month hours
                    _val_cols_hrs = (['MRR'] if _has_mrr_fd else []) + _all_month_cols

                    def _safe_cols(df, wanted):
                        return [c for c in wanted if c in df.columns]

                    def _agg_fd(df, key_cols):
                        _kc = [c for c in key_cols if c in df.columns]
                        _vc_months = [c for c in _all_month_cols if c in df.columns]
                        if not _kc:
                            return pd.DataFrame()
                        # Hours cols — simple sum
                        _base = df.groupby(_kc, as_index=False, dropna=False)[_vc_months].sum() if _vc_months else \
                                df[_kc].drop_duplicates().reset_index(drop=True)
                        # MRR — deduplicate per client within each group, then sum
                        if _has_mrr_fd and 'MRR' in df.columns and 'Client' in df.columns:
                            _cli_key = list(dict.fromkeys(_kc + (['Client'] if 'Client' not in _kc else [])))
                            _cli_key = [c for c in _cli_key if c in df.columns]
                            _mrr_cli = df.groupby(_cli_key, dropna=False)['MRR'].first().reset_index()
                            _mrr_grp_key = [c for c in _kc if c in _mrr_cli.columns]
                            _mrr_sum = _mrr_cli.groupby(_mrr_grp_key, as_index=False, dropna=False)['MRR'].sum()
                            _base = pd.merge(_base, _mrr_sum, on=_mrr_grp_key, how='left')
                        return _base

                    def _client_count_fd(df, grp_cols):
                        _kc = [c for c in grp_cols if c in df.columns and c != 'Required Role' and c != 'Client']
                        if not _kc or 'Client' not in df.columns:
                            return pd.DataFrame()
                        _cnt = df.groupby(_kc, as_index=False, dropna=False)['Client'].nunique()
                        return _cnt.rename(columns={'Client': 'Client Count'})

                    _levels = {}

                    # — Overall —
                    _ov1_d = {'Level': ['Overall']}
                    _ov2_d = {'Level': ['Overall']}
                    for _vc in _val_cols_hrs:
                        if _vc == 'MRR' and 'Client' in _fd1.columns and 'MRR' in _fd1.columns:
                            _ov1_d[_vc] = [_fd1.drop_duplicates('Client')['MRR'].sum()]
                            _ov2_d[_vc] = [_fd2.drop_duplicates('Client')['MRR'].sum() if 'MRR' in _fd2.columns else 0]
                        else:
                            _ov1_d[_vc] = [_fd1[_vc].sum() if _vc in _fd1.columns else 0]
                            _ov2_d[_vc] = [_fd2[_vc].sum() if _vc in _fd2.columns else 0]
                    _ov_merged = _merge_delta(pd.DataFrame(_ov1_d), pd.DataFrame(_ov2_d), ['Level'], _val_cols_hrs)
                    _cli_col_fd = 'Client' if 'Client' in _fd1.columns else None
                    _ov_merged.insert(1, 'Clients V1', _fd1[_cli_col_fd].nunique() if _cli_col_fd else 0)
                    _ov_merged.insert(2, 'Clients V2', _fd2[_cli_col_fd].nunique() if _cli_col_fd else 0)
                    _ov_merged.insert(3, 'Δ Clients', _ov_merged['Clients V2'] - _ov_merged['Clients V1'])
                    _levels['Overall'] = _ov_merged

                    # — By POD —
                    _p1 = _agg_fd(_fd1, ['POD'])
                    _p2 = _agg_fd(_fd2, ['POD'])
                    if not _p1.empty and not _p2.empty:
                        _vc_pod = [c for c in _val_cols_hrs if c in _p1.columns or c in _p2.columns]
                        _pod_m = _merge_delta(_p1, _p2, ['POD'], _vc_pod)
                        _pc1 = _client_count_fd(_fd1, ['POD'])
                        _pc2 = _client_count_fd(_fd2, ['POD'])
                        if not _pc1.empty and not _pc2.empty:
                            _pc_m = pd.merge(_pc1, _pc2, on=['POD'], how='outer', suffixes=(' V1', ' V2')).fillna(0)
                            _pc_m['Δ Clients'] = _pc_m['Client Count V2'] - _pc_m['Client Count V1']
                            _pod_m = pd.merge(_pod_m, _pc_m, on=['POD'], how='left')
                        _levels['By_POD'] = _pod_m

                    # — By Sr. Accountant —
                    _sr_key_fd = _safe_cols(_fd1, ['POD', 'Sr. Accountant'])
                    if _sr_key_fd:
                        _s1 = _agg_fd(_fd1, _sr_key_fd)
                        _s2 = _agg_fd(_fd2, _sr_key_fd)
                        if not _s1.empty and not _s2.empty:
                            _vc_sr = [c for c in _val_cols_hrs if c in _s1.columns or c in _s2.columns]
                            _sr_m = _merge_delta(_s1, _s2, _sr_key_fd, _vc_sr)
                            _sc1 = _client_count_fd(_fd1, _sr_key_fd)
                            _sc2 = _client_count_fd(_fd2, _sr_key_fd)
                            if not _sc1.empty and not _sc2.empty:
                                _sc_m = pd.merge(_sc1, _sc2, on=_sr_key_fd, how='outer', suffixes=(' V1', ' V2')).fillna(0)
                                _sc_m['Δ Clients'] = _sc_m['Client Count V2'] - _sc_m['Client Count V1']
                                _sr_m = pd.merge(_sr_m, _sc_m, on=_sr_key_fd, how='left')
                            _levels['By_Sr_Accountant'] = _sr_m

                    # — By Client —
                    _cli_key_fd = _safe_cols(_fd1, ['POD', 'Sr. Accountant', 'Client'])
                    if _cli_key_fd:
                        _c1 = _agg_fd(_fd1, _cli_key_fd)
                        _c2 = _agg_fd(_fd2, _cli_key_fd)
                        if not _c1.empty and not _c2.empty:
                            _ck = [c for c in _cli_key_fd if c in _c1.columns and c in _c2.columns]
                            _vc_cli = [c for c in _val_cols_hrs if c in _c1.columns or c in _c2.columns]
                            _levels['By_Client'] = _merge_delta(_c1, _c2, _ck, _vc_cli)

                    # — By Type & Subtype —
                    _type_key_fd = _safe_cols(_fd1, ['POD', 'Sr. Accountant', 'Client',
                                                      'Type', 'Sub-Type', 'Subtype', 'Sub Type'])
                    if _type_key_fd:
                        _t1 = _agg_fd(_fd1, _type_key_fd)
                        _t2 = _agg_fd(_fd2, _type_key_fd)
                        if not _t1.empty and not _t2.empty:
                            _tk = [c for c in _type_key_fd if c in _t1.columns and c in _t2.columns]
                            _vc_typ = [c for c in _val_cols_hrs if c in _t1.columns or c in _t2.columns]
                            _levels['By_Type_Subtype'] = _merge_delta(_t1, _t2, _tk, _vc_typ)

                else:
                    # ── VOLUME INPUT mode ─────────────────────────────────────
                    _vi1 = pd.read_excel(_rv1, sheet_name=0)
                    _vi2 = pd.read_excel(_rv2, sheet_name=0)
                    _vi1.columns = _vi1.columns.str.strip()
                    _vi2.columns = _vi2.columns.str.strip()

                    # Normalise column names to standard form
                    _vi_col_map = {
                        'client_name': 'Client',
                        'type':        'Type',
                        'subtype':     'Subtype',
                    }
                    _vi1 = _vi1.rename(columns={k: v for k, v in _vi_col_map.items() if k in _vi1.columns})
                    _vi2 = _vi2.rename(columns={k: v for k, v in _vi_col_map.items() if k in _vi2.columns})

                    # Non-summed / identity columns
                    _vi_id_cols = {'POD', 'Sr. Accountant', 'Client', 'Type', 'Subtype',
                                   'processor', 'Proc Role', 'reviewer', 'Rev Role',
                                   'PMS', 'Status', 'Go Live', 'Final Service Date'}

                    # Explicit value columns (ordered for readability)
                    _vi_val_priority = [
                        'Capacity Hours spent',
                        'Capacity Processing Hours',
                        'Capacity reviewing hours',
                        'MRR',
                        'Closed tickets with Proc time',
                        'Closed tickets with rev time',
                        '>>> FINAL Capacity Proc AHT',
                        '>>> FINAL Capacity Rev AHT',
                        'Res doors', 'Res Prop',
                        'Commercial Properties', 'Commercial Doors',
                    ]
                    _vi_vals = [c for c in _vi_val_priority if c in _vi1.columns or c in _vi2.columns]

                    def _agg_vi(df, key_cols):
                        _kc = [c for c in key_cols if c in df.columns]
                        _vc = [c for c in _vi_vals if c in df.columns]
                        if not _kc or not _vc:
                            return pd.DataFrame()
                        # MRR: deduplicate per client before summing
                        _vc_no_mrr = [c for c in _vc if c != 'MRR']
                        _base = df.groupby(_kc, as_index=False, dropna=False)[_vc_no_mrr].sum() if _vc_no_mrr else \
                                df[_kc].drop_duplicates().reset_index(drop=True)
                        if 'MRR' in _vc and 'Client' in df.columns:
                            _mrr_key = list(dict.fromkeys(_kc + (['Client'] if 'Client' not in _kc else [])))
                            _mrr_key = [c for c in _mrr_key if c in df.columns]
                            _mrr_cli = df.groupby(_mrr_key, dropna=False)['MRR'].first().reset_index()
                            _mrr_grp_key = [c for c in _kc if c in _mrr_cli.columns]
                            _mrr_sum = _mrr_cli.groupby(_mrr_grp_key, as_index=False, dropna=False)['MRR'].sum()
                            _base = pd.merge(_base, _mrr_sum, on=[c for c in _kc if c in _mrr_sum.columns], how='left')
                        return _base

                    _client_col_vi = 'Client' if 'Client' in _vi1.columns else None

                    _levels = {}

                    # — Overall —
                    _ov1_d = {'Level': ['Overall']}
                    _ov2_d = {'Level': ['Overall']}
                    for _vc in _vi_vals:
                        if _vc == 'MRR' and _client_col_vi:
                            _ov1_d[_vc] = [_vi1.drop_duplicates(_client_col_vi)['MRR'].sum() if 'MRR' in _vi1.columns else 0]
                            _ov2_d[_vc] = [_vi2.drop_duplicates(_client_col_vi)['MRR'].sum() if 'MRR' in _vi2.columns else 0]
                        else:
                            _ov1_d[_vc] = [_vi1[_vc].sum() if _vc in _vi1.columns else 0]
                            _ov2_d[_vc] = [_vi2[_vc].sum() if _vc in _vi2.columns else 0]
                    _ov_m = _merge_delta(pd.DataFrame(_ov1_d), pd.DataFrame(_ov2_d), ['Level'], _vi_vals)
                    _ov_m.insert(1, 'Clients V1', _vi1[_client_col_vi].nunique() if _client_col_vi and _client_col_vi in _vi1.columns else 0)
                    _ov_m.insert(2, 'Clients V2', _vi2[_client_col_vi].nunique() if _client_col_vi and _client_col_vi in _vi2.columns else 0)
                    _ov_m.insert(3, 'Δ Clients', _ov_m['Clients V2'] - _ov_m['Clients V1'])
                    _levels['Overall'] = _ov_m

                    # — By POD —
                    if 'POD' in _vi1.columns:
                        _p1 = _agg_vi(_vi1, ['POD'])
                        _p2 = _agg_vi(_vi2, ['POD'])
                        if not _p1.empty and not _p2.empty:
                            _vc_p = [c for c in _vi_vals if c in _p1.columns or c in _p2.columns]
                            _pod_vm = _merge_delta(_p1, _p2, ['POD'], _vc_p)
                            if _client_col_vi and _client_col_vi in _vi1.columns:
                                _pc1 = _vi1.groupby('POD', dropna=False)[_client_col_vi].nunique().reset_index(name='Client Count V1')
                                _pc2 = _vi2.groupby('POD', dropna=False)[_client_col_vi].nunique().reset_index(name='Client Count V2')
                                _pc_m = pd.merge(_pc1, _pc2, on='POD', how='outer').fillna(0)
                                _pc_m['Δ Clients'] = _pc_m['Client Count V2'] - _pc_m['Client Count V1']
                                _pod_vm = pd.merge(_pod_vm, _pc_m, on='POD', how='left')
                            _levels['By_POD'] = _pod_vm

                    # — By Sr. Accountant —
                    _vi_sr_key = [c for c in ['POD', 'Sr. Accountant'] if c in _vi1.columns]
                    if _vi_sr_key:
                        _s1 = _agg_vi(_vi1, _vi_sr_key)
                        _s2 = _agg_vi(_vi2, _vi_sr_key)
                        if not _s1.empty and not _s2.empty:
                            _vc_s = [c for c in _vi_vals if c in _s1.columns or c in _s2.columns]
                            _sr_vm = _merge_delta(_s1, _s2, _vi_sr_key, _vc_s)
                            if _client_col_vi and _client_col_vi in _vi1.columns:
                                _sc1 = _vi1.groupby(_vi_sr_key, dropna=False)[_client_col_vi].nunique().reset_index(name='Client Count V1')
                                _sc2 = _vi2.groupby(_vi_sr_key, dropna=False)[_client_col_vi].nunique().reset_index(name='Client Count V2')
                                _sc_m = pd.merge(_sc1, _sc2, on=_vi_sr_key, how='outer').fillna(0)
                                _sc_m['Δ Clients'] = _sc_m['Client Count V2'] - _sc_m['Client Count V1']
                                _sr_vm = pd.merge(_sr_vm, _sc_m, on=_vi_sr_key, how='left')
                            _levels['By_Sr_Accountant'] = _sr_vm

                    # — By Client —
                    _vi_cli_key = [c for c in ['POD', 'Sr. Accountant', 'Client'] if c in _vi1.columns]
                    if _vi_cli_key:
                        _c1 = _agg_vi(_vi1, _vi_cli_key)
                        _c2 = _agg_vi(_vi2, _vi_cli_key)
                        if not _c1.empty and not _c2.empty:
                            _ck = [c for c in _vi_cli_key if c in _c1.columns and c in _c2.columns]
                            _vc_c = [c for c in _vi_vals if c in _c1.columns or c in _c2.columns]
                            _levels['By_Client'] = _merge_delta(_c1, _c2, _ck, _vc_c)

                    # — By Type & Subtype per Client —
                    _vi_type_key = [c for c in ['POD', 'Sr. Accountant', 'Client', 'Type', 'Subtype',
                                                 'Sub-Type', 'Sub Type'] if c in _vi1.columns]
                    if _vi_type_key:
                        _t1 = _agg_vi(_vi1, _vi_type_key)
                        _t2 = _agg_vi(_vi2, _vi_type_key)
                        if not _t1.empty and not _t2.empty:
                            _tk = [c for c in _vi_type_key if c in _t1.columns and c in _t2.columns]
                            _vc_t = [c for c in _vi_vals if c in _t1.columns or c in _t2.columns]
                            _levels['By_Type_Subtype'] = _merge_delta(_t1, _t2, _tk, _vc_t)

                # ── Build Excel export ───────────────────────────────────────
                _recon_buf = BytesIO()
                with pd.ExcelWriter(_recon_buf, engine='openpyxl') as _rw:
                    from openpyxl.styles import PatternFill, Font, Alignment
                    from openpyxl.utils import get_column_letter

                    _green_fill  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    _red_fill    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    _yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    _hdr_font    = Font(bold=True)

                    _sheet_order = ['Overall', 'By_POD', 'By_Sr_Accountant', 'By_Client', 'By_Type_Subtype']
                    for _sname in _sheet_order:
                        if _sname not in _levels:
                            continue
                        _df_sh = _levels[_sname]
                        if _df_sh is None or _df_sh.empty:
                            continue
                        _df_sh.to_excel(_rw, sheet_name=_sname, index=False)
                        _ws = _rw.sheets[_sname]
                        # Bold + centre header row
                        for _cell in _ws[1]:
                            _cell.font = _hdr_font
                            _cell.alignment = Alignment(horizontal='center')
                        # Colour delta columns
                        for _ci, _col_hdr in enumerate(_df_sh.columns, start=1):
                            if str(_col_hdr).startswith('Δ'):
                                for _ri in range(2, _ws.max_row + 1):
                                    _cell = _ws.cell(row=_ri, column=_ci)
                                    try:
                                        _v = float(_cell.value) if _cell.value not in (None, '') else 0.0
                                        if _v > 0.001:
                                            _cell.fill = _green_fill
                                        elif _v < -0.001:
                                            _cell.fill = _red_fill
                                    except (TypeError, ValueError):
                                        pass
                        # Colour Status column
                        _status_col_idx = None
                        for _ci, _col_hdr in enumerate(_df_sh.columns, start=1):
                            if str(_col_hdr) == 'Status':
                                _status_col_idx = _ci
                                break
                        if _status_col_idx:
                            _status_fill_map = {
                                'New':     _green_fill,
                                'Removed': _red_fill,
                                'Changed': _yellow_fill,
                            }
                            for _ri in range(2, _ws.max_row + 1):
                                _cell = _ws.cell(row=_ri, column=_status_col_idx)
                                _sf = _status_fill_map.get(str(_cell.value))
                                if _sf:
                                    _cell.fill = _sf
                        # Auto-width columns
                        for _ci_w, _col_cells in enumerate(_ws.columns, start=1):
                            _max_w = max(
                                (len(str(c.value)) if c.value is not None else 0) for c in _col_cells
                            )
                            _ws.column_dimensions[get_column_letter(_ci_w)].width = min(max(_max_w + 2, 10), 50)

                # ── Show overall on page ─────────────────────────────────────
                st.success("✅ Reconciliation complete!")
                st.subheader("Overall Summary")
                if 'Overall' in _levels and not _levels['Overall'].empty:
                    st.dataframe(_levels['Overall'], use_container_width=True)

                    _ov_display = _levels['Overall']
                    _delta_cols = [c for c in _ov_display.columns if c.startswith('Δ')]
                    _pos_parts, _neg_parts = [], []
                    for _dc in _delta_cols:
                        try:
                            _dv = float(_ov_display[_dc].iloc[0])
                            if _dv > 0.001:
                                _pos_parts.append(f'**{_dc}**: +{_dv:,.2f}')
                            elif _dv < -0.001:
                                _neg_parts.append(f'**{_dc}**: {_dv:,.2f}')
                        except Exception:
                            pass
                    if _pos_parts:
                        st.markdown('🟢 ' + '  |  '.join(_pos_parts))
                    if _neg_parts:
                        st.markdown('🔴 ' + '  |  '.join(_neg_parts))

                # Download button
                st.download_button(
                    "📥 Download Full Reconciliation",
                    _recon_buf.getvalue(),
                    file_name=f"Reconciliation_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="recon_dl_btn",
                )

            except Exception as _recon_err:
                st.error(f"❌ Reconciliation failed: {_recon_err}")
                import traceback
                st.code(traceback.format_exc())

    _recon_fragment()

# ==========================================
# TAB: ACTUAL HOURS DISTRIBUTION
# ==========================================
with tab_actual_hours:
    st.header("📊 Actual Hours Distribution")
    st.caption(
        "Upload a Volume Input file to see actual capacity hours distributed by role, "
        "based on **Proc Role** (processing hours) and **Rev Role** (reviewing hours)."
    )

    @st.fragment
    def _actual_hours_fragment():
        _ah_file = st.file_uploader(
            "📂 Upload Volume Input file (.xlsx)",
            type=["xlsx"],
            key="ah_vi_upload",
            label_visibility="collapsed",
        )
        st.caption("📂 Upload Volume Input file (.xlsx)")

        if _ah_file is None:
            st.info("Upload a Volume Input file above to generate the hours distribution.")
            return

        try:
            _ah_raw = pd.read_excel(_ah_file, sheet_name=0)
        except Exception as _ae:
            st.error(f"❌ Could not read file: {_ae}")
            return

        _ah_raw.columns = _ah_raw.columns.str.strip()

        # Normalise column names to standard form
        _ah_raw = _ah_raw.rename(columns={
            'client_name': 'Client',
            'type':        'Type',
            'subtype':     'Subtype',
        })

        # Validate required columns
        _ah_required = ['Proc Role', 'Rev Role', 'Capacity Processing Hours', 'Capacity reviewing hours']
        _ah_missing = [c for c in _ah_required if c not in _ah_raw.columns]
        if _ah_missing:
            st.error(f"❌ Missing required columns: {_ah_missing}")
            return

        # Coerce numeric
        for _nc in ['Capacity Processing Hours', 'Capacity reviewing hours', 'MRR']:
            if _nc in _ah_raw.columns:
                _ah_raw[_nc] = pd.to_numeric(_ah_raw[_nc], errors='coerce').fillna(0.0)

        _ah_key_cols = [c for c in ['POD', 'Sr. Accountant', 'Client'] if c in _ah_raw.columns]
        _ah_has_mrr  = 'MRR' in _ah_raw.columns and 'Client' in _ah_raw.columns

        # ── Build attribution: split each row into two — one for proc, one for rev ──
        # Proc side: role = Proc Role, hours = Capacity Processing Hours
        _proc_cols = _ah_key_cols + ['Proc Role', 'Capacity Processing Hours']
        _proc_side = _ah_raw[[c for c in _proc_cols if c in _ah_raw.columns]].copy()
        _proc_side = _proc_side.rename(columns={
            'Proc Role': 'Role',
            'Capacity Processing Hours': 'Hours',
        })
        _proc_side['Source'] = 'Processing'

        # Rev side: role = Rev Role, hours = Capacity reviewing hours (zero-rev rows still included)
        _rev_cols = _ah_key_cols + ['Rev Role', 'Capacity reviewing hours']
        _rev_side = _ah_raw[[c for c in _rev_cols if c in _ah_raw.columns]].copy()
        _rev_side  = _rev_side.rename(columns={
            'Rev Role': 'Role',
            'Capacity reviewing hours': 'Hours',
        })
        _rev_side['Source'] = 'Reviewing'

        _combined = pd.concat([_proc_side, _rev_side], ignore_index=True)
        _combined['Hours'] = pd.to_numeric(_combined['Hours'], errors='coerce').fillna(0.0)
        # Drop zero-hour rows to keep aggregations clean
        _combined = _combined[_combined['Hours'] > 0].copy()

        # Role order
        _all_roles = sorted(_combined['Role'].dropna().unique().tolist())

        # MRR lookup: one MRR per client (not duplicated across rows)
        if _ah_has_mrr:
            _mrr_lkp = (
                _ah_raw
                .drop_duplicates('Client')
                [[c for c in _ah_key_cols + ['MRR'] if c in _ah_raw.columns]]
                .copy()
            )
        else:
            _mrr_lkp = pd.DataFrame()

        # FTE base hours / month
        _fte_base = 160.0

        # ── Helper: build pivot (group_cols × Role) with totals ─────────────────
        def _build_role_pivot(df, group_cols):
            _gc = [c for c in group_cols if c in df.columns]
            if not _gc:
                return pd.DataFrame()

            # If a role value shares a name with a group key (e.g. 'Sr. Accountant' is both
            # a key column and a role type), rename the role value to 'X (Role)' to avoid
            # duplicate columns after pivot_table + reset_index.
            _colliding = set(df['Role'].dropna().unique()) & set(_gc)
            _role_rename = {r: f'{r} (Role)' for r in _colliding}
            _df_piv = df.copy()
            if _role_rename:
                _df_piv['Role'] = _df_piv['Role'].replace(_role_rename)

            # Updated role list for this pivot
            _all_roles_piv = [_role_rename.get(r, r) for r in _all_roles]

            _grp = (
                _df_piv.groupby(_gc + ['Role'], as_index=False, dropna=False)['Hours']
                .sum()
            )
            _piv = _grp.pivot_table(
                index=_gc, columns='Role', values='Hours',
                aggfunc='sum', fill_value=0,
            ).reset_index()
            _piv.columns.name = None

            _role_cols_piv = [c for c in _piv.columns if c not in _gc]
            # Ensure every known role has a column (fill absent ones with 0)
            for _r in _all_roles_piv:
                if _r not in _piv.columns:
                    _piv[_r] = 0.0
            _role_cols_piv = [c for c in _all_roles_piv if c in _piv.columns]

            _piv['Total Hours'] = _piv[_role_cols_piv].sum(axis=1)
            _piv['Total FTEs']  = (_piv['Total Hours'] / _fte_base).round(3)
            for _r in _role_cols_piv:
                _piv[f'{_r} FTE'] = (_piv[_r] / _fte_base).round(3)

            # Column order: group keys | role hours | Total Hours | Total FTEs | role FTEs
            _col_order = (
                _gc
                + _role_cols_piv
                + ['Total Hours', 'Total FTEs']
                + [f'{_r} FTE' for _r in _role_cols_piv]
            )
            _piv = _piv[[c for c in _col_order if c in _piv.columns]]

            # Append a TOTAL row
            _tot = {c: '' for c in _piv.columns}
            for _nc in _piv.columns:
                if _nc not in _gc:
                    try:
                        _tot[_nc] = round(_piv[_nc].sum(), 3)
                    except Exception:
                        _tot[_nc] = ''
            if _gc:
                _tot[_gc[0]] = '>>> TOTAL'
            _piv = pd.concat([_piv, pd.DataFrame([_tot])], ignore_index=True)
            return _piv

        # ── Top-level metrics ────────────────────────────────────────────────────
        _total_hrs      = _combined['Hours'].sum()
        _total_clients  = _ah_raw['Client'].nunique() if 'Client' in _ah_raw.columns else 0
        _total_mrr      = _mrr_lkp['MRR'].sum() if not _mrr_lkp.empty else 0
        _total_fte      = _total_hrs / _fte_base

        _cm1, _cm2, _cm3, _cm4 = st.columns(4)
        _cm1.metric("Total Hours",     f"{_total_hrs:,.1f}")
        _cm2.metric("Total FTEs",      f"{_total_fte:,.2f}")
        _cm3.metric("Unique Clients",  f"{_total_clients:,}")
        if _ah_has_mrr:
            _cm4.metric("Total MRR",   f"${_total_mrr:,.0f}")

        st.divider()

        # ── Sub-tabs ─────────────────────────────────────────────────────────────
        _ah_t_overall, _ah_t_pod, _ah_t_cli = st.tabs([
            "🌎 Overall", "🏷️ By POD", "🏢 By Client",
        ])

        # ── Overall ──────────────────────────────────────────────────────────────
        with _ah_t_overall:
            st.write("### Hours & FTEs by Role")

            _ov_role = (
                _combined.groupby('Role', as_index=False, dropna=False)['Hours']
                .sum()
                .sort_values('Hours', ascending=False)
            )
            _ov_role['FTEs']       = (_ov_role['Hours'] / _fte_base).round(3)
            _ov_role['% of Total'] = (_ov_role['Hours'] / _ov_role['Hours'].sum() * 100).round(1)

            _ov_total_row = pd.DataFrame([{
                'Role':        '>>> TOTAL',
                'Hours':       round(_ov_role['Hours'].sum(), 2),
                'FTEs':        round(_ov_role['FTEs'].sum(), 3),
                '% of Total':  100.0,
            }])
            _ov_display = pd.concat([_ov_role, _ov_total_row], ignore_index=True)
            st.dataframe(_ov_display, use_container_width=True, hide_index=True)

            st.write("### Processing vs Reviewing Split by Role")
            _src_grp = (
                _combined.groupby(['Source', 'Role'], as_index=False, dropna=False)['Hours']
                .sum()
            )
            _src_piv = _src_grp.pivot_table(
                index='Role', columns='Source', values='Hours',
                aggfunc='sum', fill_value=0,
            ).reset_index()
            _src_piv.columns.name = None
            _proc_c = 'Processing' if 'Processing' in _src_piv.columns else None
            _rev_c  = 'Reviewing'  if 'Reviewing'  in _src_piv.columns else None
            _src_piv['Total'] = (
                (_src_piv[_proc_c] if _proc_c else 0)
                + (_src_piv[_rev_c] if _rev_c else 0)
            )
            _src_piv = _src_piv.sort_values('Total', ascending=False)
            _src_tot = {c: '' for c in _src_piv.columns}
            _src_tot['Role'] = '>>> TOTAL'
            for _sc in ([_proc_c] if _proc_c else []) + ([_rev_c] if _rev_c else []) + ['Total']:
                try:
                    _src_tot[_sc] = round(_src_piv[_sc].sum(), 2)
                except Exception:
                    pass
            _src_piv = pd.concat([_src_piv, pd.DataFrame([_src_tot])], ignore_index=True)
            st.dataframe(_src_piv, use_container_width=True, hide_index=True)

        # ── By POD ───────────────────────────────────────────────────────────────
        with _ah_t_pod:
            if 'POD' not in _combined.columns:
                st.info("No POD column found in the file.")
            else:
                st.write("### Hours by POD × Role")
                _pod_piv = _build_role_pivot(_combined, ['POD'])
                if not _pod_piv.empty:
                    st.dataframe(_pod_piv, use_container_width=True, hide_index=True)

        # ── By Client ─────────────────────────────────────────────────────────────
        with _ah_t_cli:
            _cli_key_ah = [c for c in ['POD', 'Sr. Accountant', 'Client'] if c in _combined.columns]
            if not _cli_key_ah:
                st.info("No Client column found in the file.")
            else:
                st.write("### Hours by Client × Role")
                _cli_piv = _build_role_pivot(_combined, _cli_key_ah)

                # Join MRR at client level
                if not _mrr_lkp.empty and not _cli_piv.empty:
                    _mrr_join_key = [c for c in _cli_key_ah if c in _mrr_lkp.columns]
                    _cli_piv = pd.merge(_cli_piv, _mrr_lkp[_mrr_join_key + ['MRR']], on=_mrr_join_key, how='left')

                # Reorder columns using actual pivot output — avoids duplicate-column issues
                # when a role name (e.g. 'Sr. Accountant') matches a group key column name.
                if not _cli_piv.empty:
                    _fixed = set(_cli_key_ah) | {'MRR', 'Total Hours', 'Total FTEs'}
                    _cli_role_hrs = [c for c in _cli_piv.columns if c not in _fixed and not c.endswith(' FTE')]
                    _cli_fte_cols = [c for c in _cli_piv.columns if c.endswith(' FTE') and c not in _fixed]
                    _cli_col_ord  = (
                        _cli_key_ah
                        + (['MRR'] if 'MRR' in _cli_piv.columns else [])
                        + _cli_role_hrs
                        + ['Total Hours', 'Total FTEs']
                        + _cli_fte_cols
                    )
                    _cli_piv = _cli_piv[[c for c in _cli_col_ord if c in _cli_piv.columns]]

                if not _cli_piv.empty:
                    st.dataframe(_cli_piv, use_container_width=True, hide_index=True,
                                 height=min(800, len(_cli_piv) * 35 + 38))

        # ── Export ────────────────────────────────────────────────────────────────
        st.divider()
        if st.button("📥 Export to Excel", key="ah_export_btn", type="secondary"):
            _ah_exp_buf = BytesIO()
            with pd.ExcelWriter(_ah_exp_buf, engine='openpyxl') as _ah_xw:
                from openpyxl.styles import PatternFill, Font, Alignment
                from openpyxl.utils import get_column_letter

                _hdr_font   = Font(bold=True)
                _total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                _fte_fill   = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid')

                def _write_sheet(df, sheet_name):
                    if df is None or df.empty:
                        return
                    df.to_excel(_ah_xw, sheet_name=sheet_name, index=False)
                    _ws = _ah_xw.sheets[sheet_name]
                    # Bold header
                    for _cell in _ws[1]:
                        _cell.font = _hdr_font
                        _cell.alignment = Alignment(horizontal='center')
                    # Highlight TOTAL rows
                    for _ri in range(2, _ws.max_row + 1):
                        _cv = _ws.cell(row=_ri, column=1).value
                        if str(_cv).startswith('>>>'):
                            for _ci in range(1, _ws.max_column + 1):
                                _ws.cell(row=_ri, column=_ci).fill = _total_fill
                                _ws.cell(row=_ri, column=_ci).font = Font(bold=True)
                    # Highlight FTE columns
                    for _ci, _ch in enumerate(df.columns, start=1):
                        if 'FTE' in str(_ch):
                            for _ri in range(1, _ws.max_row + 1):
                                _ws.cell(row=_ri, column=_ci).fill = _fte_fill
                    # Auto-width
                    for _ci_w, _col_cells in enumerate(_ws.columns, start=1):
                        _max_w = max((len(str(c.value)) if c.value is not None else 0) for c in _col_cells)
                        _ws.column_dimensions[get_column_letter(_ci_w)].width = min(max(_max_w + 2, 10), 40)

                # Overall role summary
                _write_sheet(_ov_display, 'Overall_by_Role')

                # Proc vs Rev split
                _write_sheet(_src_piv, 'Proc_vs_Rev_Split')

                # By POD
                if 'POD' in _combined.columns:
                    _write_sheet(_build_role_pivot(_combined, ['POD']), 'By_POD')

                # By Client (with MRR)
                _cli_k2 = [c for c in ['POD', 'Sr. Accountant', 'Client'] if c in _combined.columns]
                if _cli_k2:
                    _cli_exp = _build_role_pivot(_combined, _cli_k2)
                    if not _mrr_lkp.empty and not _cli_exp.empty:
                        _mj2 = [c for c in _cli_k2 if c in _mrr_lkp.columns]
                        _cli_exp = pd.merge(_cli_exp, _mrr_lkp[_mj2 + ['MRR']], on=_mj2, how='left')
                    _write_sheet(_cli_exp, 'By_Client')

                # Raw attribution (proc + rev rows)
                _raw_export = _combined.copy()
                _write_sheet(_raw_export, 'Raw_Attribution')

            st.download_button(
                "📥 Download Hours Distribution",
                _ah_exp_buf.getvalue(),
                file_name=f"ActualHoursDistribution_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="ah_dl_btn",
            )

    _actual_hours_fragment()
